import time
# Bỏ import urlparse vì không cần tách domain nữa
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# ==============================================================================
# CẤU HÌNH (User Config)
# ==============================================================================

# Giữ nguyên đường dẫn như file cũ của bạn
EXCEL_PATH = r"D:\Thuc_tap\Tool_other\DS_LocBaiTrungLap.xlsx"

# Thông tin đăng nhập Admin
ADMIN_USER = "adminvtk"
ADMIN_PASS = "Khanhkh@nh9999"

# Đường dẫn Chrome Driver
CHROMEDRIVER_PATH = r"D:\Thuc_tap\chromedriver-win64\chromedriver-win64\chromedriver.exe"
CHROME_BINARY     = r"D:\Thuc_tap\chrome-win64\chrome-win64\chrome.exe"

# ==============================================================================
# LOGIC KHÔI PHỤC (RESTORER CLASS)
# ==============================================================================

class WordPressRestorer:
    def __init__(self):
        self.driver = self.setup_driver()
        # Giữ nguyên thời gian chờ 15s để ổn định việc đăng nhập
        self.wait = WebDriverWait(self.driver, 15) 

    def setup_driver(self):
        chrome_options = Options()
        chrome_options.binary_location = CHROME_BINARY
        # Chạy ẩn (headless) nếu muốn nhanh hơn, nhưng nên để hiện (False) để theo dõi
        chrome_options.add_argument("--headless=new") 
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--window-size=1280,800")
        
        service = Service(CHROMEDRIVER_PATH)
        driver = webdriver.Chrome(service=service, options=chrome_options)
        return driver

    # Đã xóa hàm extract_base_url theo yêu cầu

    def login_wordpress(self, base_url):
        """Đăng nhập vào wp-admin - Phiên bản ổn định"""
        # Đảm bảo URL không có dấu / thừa ở cuối khi ghép chuỗi
        if base_url.endswith("/"):
            base_url = base_url[:-1]
            
        login_url = f"{base_url}/wp-login.php"
        try:
            self.driver.get(login_url)
            
            # Kiểm tra nhanh nếu đã đăng nhập rồi (có wp-admin trong URL)
            if "wp-admin" in self.driver.current_url:
                return True

            # 1. NHẬP USERNAME
            user_field = self.wait.until(EC.element_to_be_clickable((By.ID, "user_login")))
            user_field.click() 
            user_field.clear()
            user_field.send_keys(ADMIN_USER)
            
            # Nghỉ 1 giây
            time.sleep(1)

            # 2. NHẬP PASSWORD
            pass_field = self.wait.until(EC.element_to_be_clickable((By.ID, "user_pass")))
            pass_field.click()
            pass_field.clear()
            pass_field.send_keys(ADMIN_PASS)

            # Nghỉ 1 giây
            time.sleep(1)

            # 3. CLICK LOGIN
            submit_btn = self.wait.until(EC.element_to_be_clickable((By.ID, "wp-submit")))
            submit_btn.click()

            # 4. CHỜ KẾT QUẢ
            self.wait.until(EC.url_contains("wp-admin"))
            return True

        except Exception as e:
            print(f"    ❌ Lỗi đăng nhập: {e}")
            return False

    def restore_all_trash(self, base_url):
        """Vào thùng rác và khôi phục TOÀN BỘ bài viết"""
        if base_url.endswith("/"):
            base_url = base_url[:-1]

        # URL truy cập thẳng vào thùng rác
        trash_url = f"{base_url}/wp-admin/edit.php?post_status=trash&post_type=post"
        
        total_restored = 0
        
        while True:
            try:
                self.driver.get(trash_url)
                
                # Kiểm tra thùng rác trống
                try:
                    self.driver.find_element(By.ID, "cb-select-all-1")
                except NoSuchElementException:
                    print("    ✅ Thùng rác trống (hoặc đã hết bài).")
                    break

                posts = self.driver.find_elements(By.CSS_SELECTOR, "#the-list tr")
                count_on_page = len(posts)
                
                if count_on_page == 0 or "no-items" in posts[0].get_attribute("class"):
                    print("    ✅ Thùng rác trống.")
                    break

                print(f"    ...Đang khôi phục {count_on_page} bài ở trang hiện tại...")

                # THỰC HIỆN BULK ACTION
                select_all = self.driver.find_element(By.ID, "cb-select-all-1")
                if not select_all.is_selected():
                    select_all.click()

                # Chọn "Restore"
                select_element = Select(self.driver.find_element(By.NAME, "action"))
                try:
                    select_element.select_by_value("untrash")
                except:
                    print("    ⚠️ Không tìm thấy tùy chọn 'Phục hồi/Restore'.")
                    break

                # Click Apply
                apply_btn = self.driver.find_element(By.ID, "doaction")
                apply_btn.click()

                self.wait.until(EC.staleness_of(apply_btn))
                
                total_restored += count_on_page
                time.sleep(2)

            except Exception as e:
                print(f"    ❌ Lỗi trong quá trình khôi phục: {e}")
                break
        
        return total_restored

    def close(self):
        if self.driver:
            self.driver.quit()

# ==============================================================================
# MAIN EXECUTION
# ==============================================================================

def main():
    print("🚀 BẮT ĐẦU QUÁ TRÌNH KHÔI PHỤC BÀI VIẾT (Từ Base URL Excel)...")
    
    # 1. Load Excel
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    except Exception as e:
        print(f"❌ Không mở được file Excel: {e}")
        return

    # 2. Khởi tạo tool
    restorer = WordPressRestorer()

    try:
        # Duyệt qua các dòng trong Excel (Bắt đầu từ dòng 2)
        # Giả định cột A (index 0) bây giờ chứa trực tiếp Base URL 
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            url_cell = row[0]  # Cột A
            
            if not url_cell.value:
                continue

            # Lấy trực tiếp giá trị từ Excel
            base_url = str(url_cell.value).strip()
            
            # --- UPDATE: Tự động thêm https:// nếu thiếu ---
            if not base_url.startswith("http"):
                base_url = "https://" + base_url

            # Xử lý nhẹ: nếu có dấu / ở cuối thì bỏ đi
            if base_url.endswith("/"):
                base_url = base_url[:-1]
            
            print(f"\n{'='*60}")
            print(f" 🏫 TRƯỜNG: {base_url}")
            print(f"    (Dòng Excel: {row_idx})")
            print(f"{'='*60}")
            
            # Thử đăng nhập
            login_success = restorer.login_wordpress(base_url)
            if not login_success:
                print("    ⚠️ Đăng nhập lần 1 thất bại, thử lại lần 2...")
                time.sleep(2)
                login_success = restorer.login_wordpress(base_url)

            if login_success:
                # Thực hiện khôi phục
                count = restorer.restore_all_trash(base_url)
                
                # Ghi log vào Excel
                # Cột E (index 4)
                row[4].value = f"Restored {count} items" 
                
                print(f"    ✅ Đã khôi phục thành công: {count} bài viết.")
                
                wb.save(EXCEL_PATH) 
            else:
                row[4].value = "Login Fail"
                print("    ⏭️ Bỏ qua do lỗi đăng nhập.")

    except Exception as e:
        print(f"❌ [LỖI CHƯƠNG TRÌNH] {e}")
    finally:
        restorer.close()
        wb.save(EXCEL_PATH)
        print("\n🏁 ĐÃ HOÀN THÀNH TOÀN BỘ.")

if __name__ == "__main__":
    main()