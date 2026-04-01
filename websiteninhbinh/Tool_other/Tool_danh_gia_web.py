import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import os
import time
from urllib.parse import urlparse # [MỚI] Thư viện bóc tách URL an toàn

# Khai báo Selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By 
from selenium.common.exceptions import WebDriverException, TimeoutException

# ==============================================================================
# CẤU HÌNH ĐƯỜNG DẪN 
# ==============================================================================
BASE_DIR = r"D:\D_Document\VS Code\Python\Thuc_tap"
CHROMEDRIVER_PATH = rf"{BASE_DIR}\chromedriver-win64\chromedriver-win64\chromedriver.exe"
CHROME_BINARY     = rf"{BASE_DIR}\chrome-win64\chrome-win64\chrome.exe"

class WebAuditorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Công cụ Đánh giá Website - Tích hợp Selenium (Bản chuẩn)")
        self.root.geometry("600x480") 
        
        # --- Các biến trạng thái quản lý Excel ---
        self.excel_file_path = ""
        self.workbook = None
        self.sheet = None
        self.headers = {} 
        self.current_row = 2 
        self.max_row = 0
        
        self.driver = None 
        
        # --- BIẾN KIỂM SOÁT RAM VÀ TIẾN TRÌNH ---
        self.is_auto_running = False
        self.processed_count = 0 # Đếm số trang đã xử lý để xả RAM
        
        # --- Các biến lưu trữ giá trị UI ---
        self.var_trang_thai = tk.StringVar(value="bỏ qua")
        self.var_banner = tk.StringVar(value="bỏ qua")
        self.var_bai_viet = tk.StringVar(value="bỏ qua")
        
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.root.bind("<Key>", self.handle_keypress)
        
        self.create_widgets()

    # ==============================================================================
    # HÀM CHỐNG ĐƠ GIAO DIỆN (Thay thế time.sleep)
    # ==============================================================================
    def safe_sleep(self, seconds):
        """Hàm chờ an toàn: Giúp đếm ngược thời gian mà không làm đơ Tkinter (Not Responding)"""
        end_time = time.time() + seconds
        while time.time() < end_time:
            self.root.update() # Giữ cho giao diện luôn phản hồi
            time.sleep(0.1)    # Nghỉ cực ngắn để không ăn CPU

    def handle_keypress(self, event):
        char = event.char.lower()
        
        # [SỬA LỖI TIỀM TÀNG]: Khóa mọi phím bấm khi đang Auto (chỉ cho phép phím 4 để dừng)
        # Tránh việc user vô tình gõ phím làm sai lệch kết quả đang auto test
        if self.is_auto_running and char != '4':
            return
            
        if char == '1': self.var_trang_thai.set("1")
        elif char == '2': self.var_trang_thai.set("Web trắng")
        elif char == '3': self.auto_check_web_trang()
        elif char == '4': self.toggle_auto_run()
        elif char == 'q': self.var_banner.set("mới")
        elif char == 'w': self.var_banner.set("cũ")
        elif char == 'a': self.var_bai_viet.set("1")
        elif char == 's': self.var_bai_viet.set("0")

    def create_widgets(self):
        frame_top = tk.Frame(self.root, pady=10)
        frame_top.pack(fill="x")
        
        self.btn_load_excel = tk.Button(frame_top, text="1. Chọn File Excel", command=self.load_excel, bg="lightblue", font=("Arial", 10, "bold"))
        self.btn_load_excel.pack()
        
        self.lbl_file_info = tk.Label(frame_top, text="Chưa chọn file", fg="gray")
        self.lbl_file_info.pack()

        frame_info = tk.Frame(self.root, pady=10)
        frame_info.pack(fill="x")
        
        self.lbl_progress = tk.Label(frame_info, text="Tiến độ: 0/0", font=("Arial", 10, "bold"))
        self.lbl_progress.pack()
        
        self.lbl_url = tk.Label(frame_info, text="URL: Đang chờ...", fg="blue", wraplength=450, justify="center")
        self.lbl_url.pack()
        
        self.frame_options = tk.LabelFrame(self.root, text="Bảng Đánh Giá", padx=10, pady=10)
        self.frame_options.pack(fill="both", expand="yes", padx=15, pady=5)
        
        lbl_tt = tk.Label(self.frame_options, text="Trạng Thái (Phím 1, 2):", font=("Arial", 10, "bold"))
        lbl_tt.grid(row=0, column=0, sticky="w", pady=5)
        tk.Radiobutton(self.frame_options, text="Web trắng", variable=self.var_trang_thai, value="Web trắng").grid(row=0, column=1, sticky="w")
        tk.Radiobutton(self.frame_options, text="1", variable=self.var_trang_thai, value="1").grid(row=0, column=2, sticky="w")
        tk.Radiobutton(self.frame_options, text="0", variable=self.var_trang_thai, value="0").grid(row=0, column=3, sticky="w")
        tk.Radiobutton(self.frame_options, text="Bỏ qua", variable=self.var_trang_thai, value="bỏ qua").grid(row=0, column=4, sticky="w")
        
        self.btn_auto_check = tk.Button(self.frame_options, text="🤖 Auto Check (Phím 3)", command=self.auto_check_web_trang, bg="#a29bfe", font=("Arial", 8, "bold"))
        self.btn_auto_check.grid(row=0, column=5, padx=(10, 0))

        lbl_bn = tk.Label(self.frame_options, text="Banner (Phím Q, W):", font=("Arial", 10, "bold"))
        lbl_bn.grid(row=1, column=0, sticky="w", pady=5)
        tk.Radiobutton(self.frame_options, text="Mới", variable=self.var_banner, value="mới").grid(row=1, column=1, sticky="w")
        tk.Radiobutton(self.frame_options, text="Cũ", variable=self.var_banner, value="cũ").grid(row=1, column=2, sticky="w")
        tk.Radiobutton(self.frame_options, text="Bỏ qua", variable=self.var_banner, value="bỏ qua").grid(row=1, column=3, sticky="w")

        lbl_bv = tk.Label(self.frame_options, text="Bài viết (Phím A, S):", font=("Arial", 10, "bold"))
        lbl_bv.grid(row=2, column=0, sticky="w", pady=5)
        tk.Radiobutton(self.frame_options, text="1", variable=self.var_bai_viet, value="1").grid(row=2, column=1, sticky="w")
        tk.Radiobutton(self.frame_options, text="0", variable=self.var_bai_viet, value="0").grid(row=2, column=2, sticky="w")
        tk.Radiobutton(self.frame_options, text="Bỏ qua", variable=self.var_bai_viet, value="bỏ qua").grid(row=2, column=3, sticky="w")

        frame_nav = tk.Frame(self.root)
        frame_nav.pack(fill="x", padx=15, pady=15)
        
        self.btn_back = tk.Button(frame_nav, text="⏮ LÙI", command=self.go_back, bg="#ffeaa7", font=("Arial", 9, "bold"))
        self.btn_back.pack(side="left", fill="x", expand=True, padx=2)
        self.btn_back.config(state="disabled")

        self.btn_auto_run = tk.Button(frame_nav, text="▶ CHẠY TỰ ĐỘNG (Phím 4)", command=self.toggle_auto_run, bg="#74b9ff", font=("Arial", 9, "bold"))
        self.btn_auto_run.pack(side="left", fill="x", expand=True, padx=2)
        self.btn_auto_run.config(state="disabled")

        self.btn_next = tk.Button(frame_nav, text="LƯU & CHUYỂN ⏭", command=self.save_and_next, bg="lightgreen", font=("Arial", 9, "bold"))
        self.btn_next.pack(side="right", fill="x", expand=True, padx=2)
        self.btn_next.config(state="disabled")

    def toggle_auto_run(self):
        if not self.driver:
            messagebox.showwarning("Cảnh báo", "Chưa tải file hoặc chưa kết nối trình duyệt!")
            return
            
        self.is_auto_running = not self.is_auto_running 
        
        if self.is_auto_running:
            self.btn_auto_run.config(text="⏹ DỪNG TỰ ĐỘNG (Phím 4)", bg="#ff7675")
            self.btn_auto_check.config(state="disabled") 
            print(f"\n[Hệ thống] BẮT ĐẦU CHẠY TỰ ĐỘNG TỪ DÒNG {self.current_row}...")
            self.auto_check_web_trang(called_by_auto=True) # Đánh dấu là hệ thống gọi
        else:
            self.btn_auto_run.config(text="▶ CHẠY TỰ ĐỘNG (Phím 4)", bg="#74b9ff")
            self.btn_auto_check.config(state="normal")
            print("\n[Hệ thống] ĐÃ DỪNG CHẠY TỰ ĐỘNG.")

    def auto_check_web_trang(self, called_by_auto=False):
        if not self.driver: return

        if not self.is_auto_running:
            self.btn_auto_check.config(text="⏳ Đang test...", state="disabled")
        self.root.update()

        try:
            title = self.driver.title.lower()
            src = self.driver.page_source.lower()
            try:
                body_elem = self.driver.find_element(By.TAG_NAME, "body")
                body_text = body_elem.text.lower()
            except:
                body_text = ""

            # CASE 1: Lỗi cơ bản (GHI RÕ TEXT VÀO EXCEL)
            if "403 forbidden" in title or "403 forbidden" in src or "403 - forbidden" in title:
                self._finish_auto_check("Lỗi 403 Forbidden", "Lỗi 403 Forbidden", called_by_auto)
                return
            
            # [CẬP NHẬT LOGIC 404]: Kết hợp linh hoạt Tiêu đề, URL và Nội dung
            current_url_lower = self.driver.current_url.lower()
            
            # 1. Blank trong tiêu đề (Title rỗng, hoặc bị hiển thị thành URL, hoặc bắt đầu bằng http)
            is_title_blank = (title.strip() == "" or title.startswith("http") or current_url_lower in title)
            # 2. 404 trong địa chỉ
            has_404_in_url = ("404" in current_url_lower or "error" in current_url_lower or "aspxerrorpath" in current_url_lower)
            # 3. 404 trong nội dung
            has_404_in_body = ("404" in body_text and ("không tìm thấy" in body_text or "not found" in body_text or "đã bị xóa" in body_text))
            
            # Đánh giá: Chắc chắn là lỗi nếu (URL có 404 VÀ Nội dung có 404) 
            # HOẶC (Tiêu đề blank VÀ Nội dung có 404) 
            # HOẶC mang URL đặc trưng của VNPT Portal (aspxerrorpath)
            # HOẶC chuỗi chính xác "404 - không tìm thấy trang"
            if (has_404_in_url and has_404_in_body) or (is_title_blank and has_404_in_body) or "aspxerrorpath" in current_url_lower or "404 - không tìm thấy trang" in body_text:
                self._finish_auto_check("Web trắng / Lỗi 404", "Phát hiện Lỗi 404 (Trang không tồn tại)", called_by_auto)
                return

            # CASE 2: Web trắng tinh hoặc Quả cầu lửa
            if len(body_text.strip()) < 50 or "cgi-sys/defaultwebpage.cgi" in src or "apache is functioning normally" in src:
                self._finish_auto_check("Web trắng / Quả cầu lửa", "Web trắng / Quả cầu lửa", called_by_auto)
                return

            # CASE 3: Click thử chuyên mục
            initial_len = len(body_text.strip())
            if initial_len > 50:
                links = self.driver.find_elements(By.TAG_NAME, "a")
                clicked = False
                
                # Phân tích tên miền bằng thư viện chuẩn
                current_domain = urlparse(self.driver.current_url).netloc
                
                for link in links:
                    try:
                        href = link.get_attribute("href")
                        if href and current_domain in href and href != self.driver.current_url:
                            if link.is_displayed() and len(link.text.strip()) > 2:
                                # Dùng get(href) thay vì click vật lý 
                                # Né mọi lỗi mở Tab mới và lỗi Banner đè lên link
                                self.driver.get(href)
                                self.safe_sleep(3.0)  # Chờ load chuyên mục
                                clicked = True
                                break
                    except:
                        continue 
                        
                if clicked:
                    self.driver.back() 
                    self.safe_sleep(3.0)  # Chờ load lại trang chủ
                    try:
                        new_body_text = self.driver.find_element(By.TAG_NAME, "body").text.strip()
                        if len(new_body_text) < 50 or len(new_body_text) < initial_len * 0.1:
                            self._finish_auto_check("Mất nội dung khi click", "Mất toàn bộ nội dung khi quay lại trang chủ", called_by_auto)
                            return
                    except:
                        self._finish_auto_check("Lỗi tải HTML khi click", "Mất HTML khi quay lại trang chủ", called_by_auto)
                        return

            self._finish_auto_check("Bình thường", "Bình thường", called_by_auto)

        except Exception as e:
            self._finish_auto_check(None, f"Lỗi quét: {str(e).splitlines()[0]}", called_by_auto)

    def _finish_auto_check(self, result, msg, called_by_auto=False):
        if result:
            self.var_trang_thai.set(result)
            
        if called_by_auto:
            # Đảm bảo user chưa bấm nút Dừng giữa chừng khi đang delay
            if self.is_auto_running: 
                print(f" -> Dòng {self.current_row}: {msg}")
                self.root.after(1000, self.save_and_next)
        else:
            messagebox.showinfo("Kết quả", msg)
            self.btn_auto_check.config(text="🤖 Auto Check (Phím 3)", state="normal")

    def setup_driver(self):
        try:
            chrome_options = Options()
            chrome_options.binary_location = CHROME_BINARY
            chrome_options.add_argument("--ignore-certificate-errors")
            chrome_options.add_argument("--start-maximized")
            chrome_options.add_argument("--incognito") 
            chrome_options.add_argument("--disable-application-cache")
            
            # Tối ưu hóa thêm để giảm thiểu ngốn RAM của Chrome
            chrome_options.add_argument("--disable-gpu")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            
            service = Service(executable_path=CHROMEDRIVER_PATH)
            driver = webdriver.Chrome(service=service, options=chrome_options)
            driver.set_page_load_timeout(30)
            return driver
        except Exception as e:
            messagebox.showerror("Lỗi Cài Đặt", f"Lỗi khởi tạo Chrome: {e}")
            return None

    def clear_browser_cache(self):
        if self.driver:
            try:
                self.driver.delete_all_cookies()
                self.driver.execute_cdp_cmd('Network.clearBrowserCache', {})
                self.driver.execute_cdp_cmd('Network.clearBrowserCookies', {})
            except:
                pass

    def init_excel_headers(self):
        self.headers = {cell.value: cell.column for cell in self.sheet[1] if cell.value}
        required_new_cols = ['Trạng Thái', 'Banner', 'Bài Viết']
        modified = False
        for col_name in required_new_cols:
            if col_name not in self.headers:
                new_col_idx = self.sheet.max_column + 1
                self.sheet.cell(row=1, column=new_col_idx).value = col_name
                self.headers[col_name] = new_col_idx
                modified = True
        if modified:
            try:
                self.workbook.save(self.excel_file_path)
            except PermissionError:
                raise PermissionError("File Excel đang mở.")

    def load_excel(self):
        filepath = filedialog.askopenfilename(title="Chọn file", filetypes=(("Excel", "*.xlsx"), ("All", "*.*")))
        if not filepath: return
        try:
            self.excel_file_path = filepath
            self.workbook = openpyxl.load_workbook(self.excel_file_path)
            self.sheet = self.workbook.active
            self.max_row = self.sheet.max_row
            
            self.init_excel_headers()
            
            self.lbl_file_info.config(text=f"Đã tải: {os.path.basename(filepath)}", fg="green")
            self.btn_next.config(state="normal")
            self.btn_auto_run.config(state="normal")
            
            self.driver = self.setup_driver()
            if self.driver:
                self.current_row = 2 
                self.process_current_row()
        except PermissionError:
            messagebox.showerror("Lỗi Excel", "ĐÓNG file Excel trước khi chạy.")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Khởi tạo lỗi: {e}")

    def process_current_row(self):
        # 1. KIỂM TRA & GIẢI PHÓNG RAM CHROME TRƯỚC KHI CHẠY (Mỗi 15 URL)
        if self.processed_count > 0 and self.processed_count % 15 == 0:
            print(f"\n[Hệ thống RAM] Đã xử lý {self.processed_count} trang. Đang khởi động lại Chrome để xả RAM...")
            try:
                self.driver.quit()
            except:
                pass
            self.safe_sleep(1) # Chờ 1s cho tiến trình cũ chết hẳn
            self.driver = self.setup_driver()
            
            # Ngăn crash nếu khởi động lại Chrome thất bại
            if not self.driver:
                self.is_auto_running = False
                self.btn_auto_run.config(text="▶ CHẠY TỰ ĐỘNG (Phím 4)", bg="#74b9ff", state="disabled")
                return

        # 2. XỬ LÝ DÒNG TIẾP THEO
        if self.current_row > self.max_row:
            self.is_auto_running = False 
            self.btn_auto_run.config(text="▶ CHẠY TỰ ĐỘNG (Phím 4)", bg="#74b9ff", state="disabled")
            messagebox.showinfo("Hoàn thành", "Đã duyệt hết danh sách!")
            self.lbl_url.config(text="Hoàn thành!")
            self.btn_next.config(state="disabled")
            if self.driver: self.driver.quit()
            return

        self.btn_back.config(state="normal" if self.current_row > 2 else "disabled")

        url = self.sheet.cell(row=self.current_row, column=2).value
        self.lbl_progress.config(text=f"Tiến độ: Dòng {self.current_row} / {self.max_row}")
        
        url_str = str(url).strip() if url is not None else ""
        
        # KIỂM TRA URL HỢP LỆ (Fix lỗi crash ở dòng "Không tìm thấy" hoặc ô trống)
        if not url_str or " " in url_str or "." not in url_str:
            self.lbl_url.config(text=f"Bỏ qua vì không phải link: '{url_str}'")
            self.current_row += 1
            self.root.after(100, self.process_current_row) # Nhảy dòng an toàn
            return
            
        if not url_str.startswith("http"): url_str = "http://" + url_str
            
        self.lbl_url.config(text=url_str)
        
        # Khôi phục dữ liệu
        col_tt = self.headers.get('Trạng Thái')
        col_bn = self.headers.get('Banner')
        col_bv = self.headers.get('Bài Viết')

        val_tt = self.sheet.cell(row=self.current_row, column=col_tt).value if col_tt else None
        val_bn = self.sheet.cell(row=self.current_row, column=col_bn).value if col_bn else None
        val_bv = self.sheet.cell(row=self.current_row, column=col_bv).value if col_bv else None

        str_tt = str(val_tt).strip() if val_tt is not None else "bỏ qua"
        str_bn = str(val_bn).strip().lower() if val_bn is not None else "bỏ qua"
        str_bv = str(val_bv).strip().lower() if val_bv is not None else "bỏ qua"

        # Đảm bảo nếu Excel có các chữ chi tiết thì UI vẫn giữ nguyên để save lại đúng khi lùi/tới
        if str_tt.lower() in ["1", "0", "bỏ qua", "web trắng"]:
            self.var_trang_thai.set(str_tt.lower() if str_tt.lower() != "web trắng" else "Web trắng")
        else:
            self.var_trang_thai.set(str_tt) # Cho phép lưu giữ các câu dài ("Bình thường", "Lỗi 403 Forbidden"...)

        self.var_banner.set(str_bn if str_bn in ["mới", "cũ", "bỏ qua"] else "bỏ qua")
        self.var_bai_viet.set(str_bv if str_bv in ["1", "0", "bỏ qua"] else "bỏ qua")
        
        try:
            self.driver.title 
            self.clear_browser_cache()
            
            # Cập nhật UI ngay trước khi gọi hàm get() gây block
            self.root.update() 
            self.driver.get(url_str)
            
            # Tăng bộ đếm RAM
            self.processed_count += 1
            
            if self.is_auto_running:
                # Chuyển đổi gọi hàm với tham số called_by_auto=True
                self.root.after(1000, lambda: self.auto_check_web_trang(called_by_auto=True))
                
        except TimeoutException:
            self.driver.execute_script("window.stop();") 
            self.processed_count += 1
            if self.is_auto_running:
                self.root.after(1000, lambda: self.auto_check_web_trang(called_by_auto=True))
                
        except WebDriverException:
            # KIỂM TRA: Chrome sập hay chỉ là link sai / web chết?
            try:
                self.driver.current_window_handle # Thử gọi Chrome
                # Chrome vẫn sống, đây là lỗi ko tải được link (VD: ERR_NAME_NOT_RESOLVED)
                self.var_trang_thai.set("Web trắng / Không truy cập được")
                if self.is_auto_running:
                    self.root.after(1000, self.save_and_next)
            except:
                # Chrome đã sập thực sự
                self.is_auto_running = False 
                self.btn_auto_run.config(text="▶ CHẠY TỰ ĐỘNG (Phím 4)", bg="#74b9ff")
                if messagebox.askyesno("Lỗi", "Trình duyệt đã đóng hoặc bị tràn RAM sập. Mở lại?"):
                    self.driver = self.setup_driver()
                    self.process_current_row() 
                else:
                    self.root.destroy()

    def save_and_next(self):
        col_tt = self.headers.get('Trạng Thái')
        col_bn = self.headers.get('Banner')
        col_bv = self.headers.get('Bài Viết')
        
        if col_tt: self.sheet.cell(row=self.current_row, column=col_tt, value=self.var_trang_thai.get())
        if col_bn: self.sheet.cell(row=self.current_row, column=col_bn, value=self.var_banner.get())
        if col_bv: self.sheet.cell(row=self.current_row, column=col_bv, value=self.var_bai_viet.get())
        
        try:
            self.workbook.save(self.excel_file_path)
        except PermissionError:
            self.is_auto_running = False 
            self.btn_auto_run.config(text="▶ CHẠY TỰ ĐỘNG (Phím 4)", bg="#74b9ff")
            messagebox.showerror("Lỗi Lưu", "ĐÓNG file Excel lại rồi bấm 'LƯU & CHUYỂN' một lần nữa.")
            return 
            
        self.clear_browser_cache()
        self.current_row += 1
        self.process_current_row()

    def go_back(self):
        if self.current_row > 2:
            self.is_auto_running = False 
            self.btn_auto_run.config(text="▶ CHẠY TỰ ĐỘNG (Phím 4)", bg="#74b9ff")
            self.current_row -= 1
            self.process_current_row()

    def on_closing(self):
        if self.driver:
            try: self.driver.quit()
            except: pass
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = WebAuditorApp(root)
    root.attributes('-topmost', True) 
    root.mainloop()