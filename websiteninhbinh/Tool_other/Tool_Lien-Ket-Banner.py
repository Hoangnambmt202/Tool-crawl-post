# -*- coding: utf-8 -*-
"""
TOOL QUẢN LÝ BANNER TOÀN DIỆN - WRITTEN FROM SCRATCH
-----------------------------------------------------------------------------
TÍNH NĂNG: Nuclear Clean, Smart Media ID, Exclusive Buckets, Detail Logging, Excel Update.
"""

import os
import time
import random
import threading
import openpyxl
from datetime import datetime
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException, NoSuchElementException, StaleElementReferenceException

# ==============================================================================
# 1. CẤU HÌNH
# ==============================================================================

# HỆ THỐNG FILES
CHROMEDRIVER_PATH = r"D:\Thuc_tap\chromedriver-win64\chromedriver-win64\chromedriver.exe"
CHROME_BINARY     = r"D:\Thuc_tap\chrome-win64\chrome-win64\chrome.exe"
EXCEL_PATH        = r"D:\Thuc_tap\Tool_other\DS_lien_ket_banner.xlsx"

# TÀI KHOẢN ADMIN
WP_USER = "adminvtk"
WP_PASS = "Khanhkh@nh9999"

# THAM SỐ LUỒNG
MAX_RETRIES     = 3
TOTAL_WORKERS   = 10 # Số luồng thực hiện (Luồng_1 -> Luồng_5)

# 5 BANNER MỤC TIÊU
LIST_OF_BANNERS = [
    {"title": "Thống kê giáo dục", "link": "http://thongkegiaoduc.viettechkey.com", "order": "1", "img_path": r"D:\Thuc_tap\File\thong-ke-giao-duc.jpg"},
    {"title": "Kho học liệu số", "link": "http://khls.viettechkey.com", "order": "1", "img_path": r"D:\Thuc_tap\File\kho-học-liệu-số.jpg"},
    {"title": "Kiểm định chất lượng giáo dục", "link": "http://kdcl.viettechkey.com", "order": "1", "img_path": r"D:\Thuc_tap\File\KĐCL.jpg"},
    {"title": "Đánh giá mức độ chuyển đổi số", "link": "http://dti.viettechkey.com", "order": "1", "img_path": r"D:\Thuc_tap\File\DTI-dgmdcds.png"},
    {"title": "Phổ cập giáo dục", "link": "https://pmpcgd.daklak.edu.vn/auth/google/login", "order": "1", "img_path": r"D:\Thuc_tap\File\PCGD.jpg"}
]

# Tên ô Input (Lấy từ HTML thực tế - Types/Toolset)
INPUT_ORDER_NAME = "wpcf[lkb-thu-tu]"
INPUT_LINK_NAME  = "wpcf[lkb-link-website]"

# Định nghĩa màu LOG cho rõ ràng chuyên nghiệp
class TermStyle:
    BOLD   = '\033[1m'
    RESET  = '\033[0m'
    CYAN   = '\033[96m'
    GREEN  = '\033[92m'
    YELLOW = '\033[93m'
    RED    = '\033[91m'
    BLUE   = '\033[94m'

# Biến khóa đồng bộ
excel_sync_lock = threading.Lock()
print_sync_lock = threading.Lock()

# ==============================================================================
# 2. MODULE TRỢ GIÚP (LOGGING & STANDARDIZATION)
# ==============================================================================

def write_detail_log(t_id, url, step, status, color=TermStyle.RESET):
    """Log format chuẩn: [Luồng_X.]: [link trường] - (đang làm gì): (trạng thái)"""
    with print_sync_lock:
        tm = datetime.now().strftime("%H:%M:%S")
        thread_label = f"Luồng_{t_id}."
        print(f"[{TermStyle.YELLOW}{tm}{TermStyle.RESET}][{TermStyle.BLUE}{thread_label}{TermStyle.RESET}]: "
              f"{url} - {TermStyle.CYAN}({step}){TermStyle.RESET}: {color}{status}{TermStyle.RESET}")

def standardize_link_simple(link):
    """Rút gọn link để so khớp (bỏ https, http, dấu /)"""
    if not link: return ""
    return str(link).lower().replace("https://","").replace("http://","").strip("/")

# ==============================================================================
# 3. WORKER - THỰC THI CHO TỪNG WEBSITE (CHI TIẾT TỪNG BƯỚC)
# ==============================================================================

class SchoolBot:
    def __init__(self, target_url, excel_row_idx, t_id):
        self.site_url = self.prep_url(target_url)
        self.excel_row = excel_row_idx
        self.t_id = t_id
        self.driver = None

    def prep_url(self, url):
        if not url: return ""
        if not url.startswith("http"): url = "https://" + url
        p = urlparse(url)
        return f"{p.scheme}://{p.netloc}/"

    def init_browser(self):
        """Khởi chạy trình duyệt thu nhỏ xuống Taskbar"""
        if self.driver: 
            try: self.driver.quit()
            except: pass
        opts = Options()
        opts.binary_location = CHROME_BINARY
        opts.add_argument("--disable-gpu")
        opts.add_argument("--no-sandbox")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        
        self.driver = webdriver.Chrome(service=Service(CHROMEDRIVER_PATH), options=opts)
        self.driver.minimize_window() # [YÊU CẦU] Luôn thu nhỏ
        self.wait = WebDriverWait(self.driver, 20)

    # --------------------------------------------------------------------------
    # Excel Updater: Cập nhật Done (cột D) và Kết quả (cột E)
    # --------------------------------------------------------------------------
    def update_status_to_excel(self, is_done=False, result_msg=""):
        with excel_sync_lock:
            try:
                wb = openpyxl.load_workbook(EXCEL_PATH)
                ws = wb.active
                # Cập nhật Done (Cột 4 - D)
                if is_done: ws.cell(row=self.excel_row, column=4).value = 1
                # Cập nhật Kết quả (Cột 5 - E)
                ws.cell(row=self.excel_row, column=5).value = result_msg
                wb.save(EXCEL_PATH)
            except Exception as e:
                write_detail_log(self.t_id, self.site_url, "Excel Update", f"Lỗi: {e}", TermStyle.RED)

    # --------------------------------------------------------------------------
    # BƯỚC 1: ĐĂNG NHẬP & BỎ QUA POPUP
    # --------------------------------------------------------------------------
    def login_process(self):
        """Logic vượt 'Xác minh Email' theo hình ảnh cung cấp"""
        last_error = ""
        for i in range(1, MAX_RETRIES + 1):
            try:
                self.driver.get(self.site_url + "wp-login.php")
                # SendKeys thuần không dùng clipboard
                u = self.wait.until(EC.presence_of_element_located((By.ID, "user_login")))
                u.clear(); u.send_keys(WP_USER)
                p = self.driver.find_element(By.ID, "user_pass")
                p.clear(); p.send_keys(WP_PASS)
                self.driver.find_element(By.ID, "wp-submit").click()
                
                time.sleep(4)
                # BẮT POPUP XÁC MINH (Update: Nhắc tôi sau)
                try:
                    btns = self.driver.find_elements(By.XPATH, "//a[contains(text(), 'Nhắc tôi sau')] | //a[contains(@href, 'remind_me_later')]")
                    if btns:
                        write_detail_log(self.t_id, self.site_url, "Login", "Nhấn: Nhắc tôi sau", TermStyle.YELLOW)
                        btns[0].click()
                except: pass

                # Verify Dashboard
                self.wait.until(EC.presence_of_element_located((By.ID, "wpadminbar")))
                write_detail_log(self.t_id, self.site_url, "Đăng nhập", "THÀNH CÔNG", TermStyle.GREEN)
                return True, ""
            except Exception as ex:
                last_error = f"(Đăng nhập): Lỗi (lần {i}): Web lag hoặc Timeout."
                write_detail_log(self.t_id, self.site_url, "Đăng nhập", f"Lỗi lần {i}, thử lại...", TermStyle.RED)
                self.init_browser() # Reset trình duyệt nếu lỗi văng ra ngoài
        
        return False, last_error

    # --------------------------------------------------------------------------
    # BƯỚC 2: QUÉT SẠCH RÁC / TRÙNG / NHÁP (CLEANER NUCLEAR)
    # --------------------------------------------------------------------------
    def nuclear_cleaning(self):
        """Xóa Draft, Xóa bài (không có tiêu đề), Xóa bài Trùng link/title"""
        write_detail_log(self.t_id, self.site_url, "Dọn dẹp", "Quét sạch liên kết Bản nháp/Trùng lặp...", TermStyle.CYAN)
        target_t = [b['title'].lower() for b in LIST_OF_BANNERS]
        target_l = [standardize_link_simple(b['link']) for b in LIST_OF_BANNERS]

        try:
            # Lặp xóa từng hàng cho đến khi sạch sẽ để đảm bảo lấy link mã Nonce chính xác
            while True:
                self.driver.get(self.site_url + "wp-admin/edit.php?post_type=lien-ket-banner")
                # Wait table list
                self.wait.until(EC.presence_of_element_located((By.ID, "the-list")))
                rows = self.driver.find_elements(By.CSS_SELECTOR, "#the-list tr")
                
                if not rows or "no-items" in rows[0].get_attribute("class"): break
                
                kill_link = None
                kill_reason = ""

                for row in rows:
                    try:
                        title_cell = row.find_element(By.CSS_SELECTOR, ".row-title")
                        full_name  = title_cell.text.strip().lower()
                        
                        link_raw = ""
                        # Quét tất cả td tìm URL
                        for td in row.find_elements(By.TAG_NAME, "td"):
                            if "http" in td.text: link_raw = td.text.strip(); break
                        
                        link_clean = standardize_link_simple(link_raw)
                        has_featured_img = len(row.find_elements(By.CSS_SELECTOR, "td img")) > 0

                        # LOGIC DIỆT:
                        diệt = False
                        if "— bản nháp" in full_name: 
                            diệt = True; kill_reason = "Xóa Bản nháp"
                        elif "không có tiêu đề" in full_name or not full_name:
                            # rác thực sự là không tên, không link, không ảnh
                            if not link_clean and not has_featured_img: diệt = True; kill_reason = "Xóa rác trống"
                            else: diệt = True; kill_reason = "Tiêu đề trống"
                        elif any(t in full_name for t in target_t):
                            diệt = True; kill_reason = f"Xóa bài trùng Title"
                        elif link_clean in target_l and link_clean != "":
                            diệt = True; kill_reason = f"Xóa bài trùng Link"

                        if diệt:
                            # Săn link Xóa bài chứa mã bảo mật Nonce (_wpnonce)
                            del_btn = row.find_elements(By.CSS_SELECTOR, "a.submitdelete")
                            if del_btn:
                                kill_link = del_btn[0].get_attribute("href")
                                break
                    except: continue

                if kill_link:
                    write_detail_log(self.t_id, self.site_url, "Cleaner", f"Tiêu diệt: {kill_reason}", TermStyle.YELLOW)
                    self.driver.get(kill_link)
                    time.sleep(1)
                else: break # Trang đã sạch
            write_detail_log(self.t_id, self.site_url, "Cleaner", "Website đã vệ sinh hoàn hảo.", TermStyle.GREEN)
        except Exception as e:
            write_detail_log(self.t_id, self.site_url, "Cleaner", "Lỗi nhẹ bảng dọn rác, bỏ qua...", TermStyle.YELLOW)

    # --------------------------------------------------------------------------
    # BƯỚC 3: XỬ LÝ MEDIA (Reuse logic Dang_bai_le.py)
    # --------------------------------------------------------------------------
    def get_id_or_upload_img(self, img_path):
        """Search Media cũ để dùng lại tệp (tránh trùng) - Tải tệp mới từ disk nếu ko thấy"""
        fn = os.path.basename(img_path)
        search_key = os.path.splitext(fn)[0]
        try:
            # Truy cập upload tìm tệp có sẵn
            self.driver.get(self.site_url + f"wp-admin/upload.php?mode=list&s={search_key}")
            rows = self.driver.find_elements(By.CSS_SELECTOR, "#the-list tr")
            if rows and "không tìm thấy" not in rows[0].text.lower():
                m_id = rows[0].get_attribute("id").split("-")[-1]
                write_detail_log(self.t_id, self.site_url, "Media", f"Dùng ID {m_id} (Đã tồn tại)", TermStyle.GREEN)
                return m_id
            
            # Phải tải tệp mới lên
            write_detail_log(self.t_id, self.site_url, "Media", "Tải mới từ ổ đĩa máy tính...", TermStyle.YELLOW)
            self.driver.get(self.site_url + "wp-admin/post-new.php?post_type=lien-ket-banner")
            # Cuộn để tìm nút ảnh đại diện
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            self.driver.find_element(By.ID, "set-post-thumbnail").click()
            time.sleep(2)
            # Chọn tab Tải tập tin
            self.driver.execute_script("document.querySelectorAll('.media-menu-item')[0].click();")
            # Gửi path tệp cục bộ
            self.driver.find_element(By.CSS_SELECTOR, "input[type='file']").send_keys(img_path)
            # Chờ hoàn tất -> Bấm Insert
            ok_btn = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, ".media-toolbar-primary button")))
            ok_btn.click()
            self.wait.until(EC.invisibility_of_element_located((By.CLASS_NAME, "media-modal")))
            return "ALREADY_UPLOADED"
        except Exception as e:
            write_detail_log(self.t_id, self.site_url, "Media", f"Lỗi bước Ảnh: {str(e)[:30]}", TermStyle.RED)
            return None

    # --------------------------------------------------------------------------
    # BƯỚC 4: QUY TRÌNH ĐĂNG 1 BANNER
    # --------------------------------------------------------------------------
    def single_post_routine(self, b_data):
        for retry in range(MAX_RETRIES):
            try:
                # 1. Image Check
                stt_img = self.get_id_or_upload_img(b_data['img_path'])
                if not stt_img: raise Exception("Bỏ qua Banner do lỗi Media")

                # 2. Form Page
                self.driver.get(self.site_url + "wp-admin/post-new.php?post_type=lien-ket-banner")
                self.wait.until(EC.presence_of_element_located((By.ID, "title"))).send_keys(b_data['title'])
                
                # Link & Order
                self.driver.find_element(By.NAME, INPUT_LINK_NAME).send_keys(b_data['link'])
                f_ord = self.driver.find_element(By.NAME, INPUT_ORDER_NAME)
                f_ord.clear(); f_ord.send_keys(b_data['order'])
                
                # Injection Image nếu search thấy ID
                if str(stt_img).isdigit():
                    self.driver.execute_script(f"document.getElementById('_thumbnail_id').value = '{stt_img}';")

                # BẤM ĐĂNG
                self.driver.execute_script("window.scrollTo(0, 0);")
                self.driver.find_element(By.ID, "publish").click()
                time.sleep(3)
                
                # Verification thành công
                if self.driver.find_elements(By.ID, "message") or "post.php" in self.driver.current_url:
                    write_detail_log(self.t_id, self.site_url, "Xử lý", f"Đã xong: {b_data['title']}", TermStyle.GREEN)
                    return True
            except Exception as e:
                write_detail_log(self.t_id, self.site_url, "Xử lý", f"Lỗi ({b_data['title']}) thử lại {retry+1}...", TermStyle.RED)
                time.sleep(2)
        return False

    # --------------------------------------------------------------------------
    # THỰC THI CHIẾN DỊCH
    # --------------------------------------------------------------------------
    def start_job(self):
        try:
            self.init_browser()
            # Bắt đầu Đăng nhập
            log_stt, log_msg = self.login_process()
            if not log_stt:
                self.update_status_to_excel(False, log_msg if log_msg else "(FATAL): Trình duyệt kẹt không đăng nhập được.")
                return

            # Bước Clean rác
            self.nuclear_cleaning()
            
            # Bước Đăng
            overall_ok = True
            succ_count = 0
            for b in LIST_OF_BANNERS:
                if self.single_post_routine(b): succ_count += 1
                else: overall_ok = False
            
            # Kết luận
            if overall_ok and succ_count == 5:
                self.update_status_to_excel(True, "thành công")
                write_detail_log(self.t_id, self.site_url, "OVERALL", "TRƯỜNG HOÀN THÀNH XONG XUÔI.", TermStyle.BOLD + TermStyle.GREEN)
            else:
                self.update_status_to_excel(False, f"(FINISH): Không đủ 5 banner ({succ_count}/5), sẽ bỏ qua lưu trạng thái Excel.")
                write_detail_log(self.t_id, self.site_url, "OVERALL", f"Xử lý xong nhưng thiếu bài ({succ_count}/5)", TermStyle.YELLOW)
                
        except Exception as big_crash:
            write_detail_log(self.t_id, self.site_url, "SẬP LUỒNG", f"Lỗi: {big_crash}", TermStyle.RED)
            self.update_status_to_excel(False, f"(FATAL): Crash hệ thống: {str(big_crash)[:50]}")
        finally:
            if self.driver: self.driver.quit()

# ==============================================================================
# 4. CHIA GIỎ CÔNG VIỆC VÀ VẬN HÀNH (EXCLUSIVE BUCKETS)
# ==============================================================================

def worker_thread_pool(thread_number, task_list):
    """Luồng cầm túi việc riêng: Chạy tuần tự trong danh sách đã được chỉ định"""
    for u, r_idx in task_list:
        bot = SchoolBot(u, r_idx, thread_number)
        bot.start_job()

def engine_init():
    if not os.path.exists(EXCEL_PATH):
        print(f"{TermStyle.RED}LỖI: Tệp Excel không thấy ở: {EXCEL_PATH}{TermStyle.RESET}"); return

    # Bước 1: Quét tất cả site chưa làm
    main_wb = openpyxl.load_workbook(EXCEL_PATH)
    all_rows = []
    for r_num, val in enumerate(main_wb.active.iter_rows(min_row=2, values_only=True), start=2):
        u_target = val[2] # URL (Cột C)
        stt_done = val[3] # Done (Cột D)
        if u_target and str(stt_done) != "1":
            all_rows.append((u_target, r_num))
            
    if not all_rows:
        print(f"{TermStyle.GREEN}--- Không có website nào cần đăng banner mới ---{TermStyle.RESET}"); return

    # Bước 2: CHIA CÔNG VIỆC RIÊNG BIỆT (KHÔNG CHO ĐĂNG TRÙNG TRƯỜNG)
    # Chia 100 site cho 5 Luồng -> Luồng 1 nhận mảng rổ 1, Luồng 2 mảng rổ 2.
    job_buckets = [[] for _ in range(TOTAL_WORKERS)]
    for i, data in enumerate(all_rows):
        job_buckets[i % TOTAL_WORKERS].append(data)

    print(f"{TermStyle.BOLD}{TermStyle.BLUE}--- KHỞI CHẠY CHIẾN DỊCH TỔNG LỰC: {len(all_rows)} TRƯỜNG, {TOTAL_WORKERS} LUỒNG ---{TermStyle.RESET}\n")

    # Bước 3: Phóng nạp ThreadPool cho các Bucket rạch ròi
    with ThreadPoolExecutor(max_workers=TOTAL_WORKERS) as pool:
        for idx, bucket in enumerate(job_buckets):
            if bucket:
                # Mỗi pool.submit tương ứng 1 ID Luồng duy nhất làm rổ việc riêng
                pool.submit(worker_thread_pool, idx + 1, bucket)

    print(f"\n{TermStyle.BOLD}{TermStyle.GREEN}--- ĐÃ XONG TOÀN BỘ DANH SÁCH CHIẾN DỊCH ---{TermStyle.RESET}")

if __name__ == "__main__":
    engine_init()