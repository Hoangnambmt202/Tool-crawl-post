import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
import time
import os
import re

# ================= 1. CẤU HÌNH ĐƯỜNG DẪN & TÀI KHOẢN =================
BASE_DIR = r"D:\D_Document\VS Code\Python\Thuc_tap"
CHROMEDRIVER_PATH = rf"{BASE_DIR}\chromedriver-win64\chromedriver-win64\chromedriver.exe"
CHROME_BINARY = rf"{BASE_DIR}\chrome-win64\chrome-win64\chrome.exe"

EXCEL_PATH = rf"{BASE_DIR}\Tool_other\DS_tool_plugin.xlsx"

# ĐƯỜNG DẪN FILE PLUGIN CẦN CÀI ĐẶT
PLUGIN_ZIP_PATH = rf"{BASE_DIR}\File\mammoth-docx-converter.1.22.0.zip"

DA_LOGIN_URL = "https://sv1126.viettechkey.com:2222"

WP_USER = "adminvtk"
WP_PASS = "Khanhkh@nh9999"

# ================= 2. TÙY CHỌN TÍNH NĂNG =================
SKIP_DONE_ROWS         = True   # Bỏ qua các dòng trong file Excel đã ghi chữ "Thành công" ở cột Done.
USE_SMART_DOMAIN       = True   # Tự động đoán tên miền (VD: đổi c1... thành th...) nếu tên gốc bị sai.
CHECK_HOMEPAGE_DOMAINS = True   # Quét danh sách các tên miền đang có sẵn trên trang chủ DirectAdmin để đối chiếu.
HEADLESS_MODE          = False  # Đặt thành True nếu muốn tool chạy ngầm, không mở cửa sổ Chrome lên làm phiền.
AUTO_LOGOUT_DA         = True   # Tự động đăng xuất tài khoản DirectAdmin sau khi hoàn thành để đảm bảo bảo mật.

# [TÙY CHỌN MỚI 1] Chế độ quét tên miền ở Nhiệm vụ 1
SCAN_ALL_DA_DOMAINS    = True   
# True: Quét và mở khóa wp-config của TẤT CẢ thư mục tên miền có trên host (Đề phòng web chạy Alias/Pointer).
# False: CHỈ ưu tiên sửa đúng tên miền lấy từ Cột B Excel (và tên miền thông minh nếu bật). Sửa được 1 cái là dừng để tiết kiệm thời gian.

# [TÙY CHỌN MỚI 2] Tính năng hoàn trả trạng thái bảo mật
REVERT_FILE_MODS       = True   
# True: Sau khi cài xong plugin, tự động quay lại DirectAdmin đổi false thành true để khóa web lại cho an toàn.

# [TÙY CHỌN MỚI 3] Xử lý khi Plugin đã tồn tại
OVERWRITE_EXISTING_PLUGIN = False  
# True: Nếu phát hiện plugin đã tồn tại, tự động bấm "Thay thế hiện tại bằng bản tải lên" để xóa và ghi đè.
# False: Bỏ qua và giữ nguyên bản cũ nếu phát hiện plugin đã tồn tại.

if not os.path.exists(PLUGIN_ZIP_PATH):
    print(f"LỖI: Không tìm thấy file Plugin tại: {PLUGIN_ZIP_PATH}")
    exit()

# ================= 3. HÀM KHỞI TẠO =================
def setup_driver():
    chrome_options = Options()
    chrome_options.binary_location = CHROME_BINARY
    chrome_options.add_argument("--ignore-certificate-errors")
    if HEADLESS_MODE:
        chrome_options.add_argument("--headless=new")
    else:
        chrome_options.add_argument("--start-maximized")
    
    service = Service(executable_path=CHROMEDRIVER_PATH)
    driver = webdriver.Chrome(service=service, options=chrome_options)
    return driver

def get_target_domain(excel_domain):
    domain = str(excel_domain).strip().lower()
    if domain.startswith('c3'): return domain.replace('c3', 'thpt', 1)
    elif domain.startswith('c2'): return domain.replace('c2', 'thcs', 1)
    elif domain.startswith('c1'): return domain.replace('c1', 'th', 1)
    elif domain.startswith('c0'): return domain.replace('c0', 'mn', 1)
    return domain

# ================= 4. LUỒNG CHÍNH =================
def main():
    print("Đang đọc dữ liệu từ file Excel (Giữ nguyên định dạng)...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        sheet = wb.active
        headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
        
        required_cols = ['Tên miền', 'User host', 'Mật khẩu host']
        for col in required_cols:
            if col not in headers:
                print(f"LỖI: Không tìm thấy cột '{col}' trong dòng đầu tiên.")
                return

        if 'Done' not in headers:
            new_col = sheet.max_column + 1
            sheet.cell(row=1, column=new_col).value = 'Done'
            headers['Done'] = new_col
        
        if 'Domain thực tế' not in headers:
            new_col = sheet.max_column + 1
            sheet.cell(row=1, column=new_col).value = 'Domain thực tế'
            headers['Domain thực tế'] = new_col
            
        try:
            wb.save(EXCEL_PATH)
        except PermissionError:
            print("\n[!] LỖI TỚI HẠN: File Excel đang được mở! Vui lòng ĐÓNG file Excel trước khi chạy Tool.\n")
            return

    except Exception as e:
        print(f"Không thể đọc file Excel. Lỗi: {e}")
        return

    driver = setup_driver()
    wait = WebDriverWait(driver, 15)
    wait_short = WebDriverWait(driver, 5)
    wait_long = WebDriverWait(driver, 60)

    for row_idx in range(2, sheet.max_row + 1):
        stt_col = headers.get('STT')
        stt = sheet.cell(row=row_idx, column=stt_col).value if stt_col else row_idx - 1
        
        domain_val = sheet.cell(row=row_idx, column=headers['Tên miền']).value
        if domain_val is None:
            continue
            
        domain = str(domain_val).strip().lower()
        if domain == '' or domain == 'nan':
            continue
            
        da_user = str(sheet.cell(row=row_idx, column=headers['User host']).value or "").strip()
        da_pass = str(sheet.cell(row=row_idx, column=headers['Mật khẩu host']).value or "").strip()
        ket_qua = str(sheet.cell(row=row_idx, column=headers['Done']).value or "").strip().lower()

        if not da_user or not da_pass:
            print(f"\n[{stt}] Bỏ qua {domain} vì KHÔNG CÓ tài khoản hoặc mật khẩu host.")
            sheet.cell(row=row_idx, column=headers['Done']).value = "Lỗi: Thiếu tài khoản host"
            continue

        if SKIP_DONE_ROWS and "thành công" in ket_qua:
            print(f"\n[{stt}] Bỏ qua {domain} vì đã làm xong trước đó.")
            continue
            
        print(f"\n[{stt}] ================= ĐANG XỬ LÝ: {domain} =================")

        try:
            # ---------------- NHIỆM VỤ 1: SỬA WP-CONFIG TRONG DA ----------------
            print("  -> Bắt đầu Nhiệm vụ 1: Đăng nhập DirectAdmin...")
            driver.delete_all_cookies()
            driver.get(DA_LOGIN_URL)

            user_input = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@name='username']")))
            pass_input = driver.find_element(By.XPATH, "//input[@name='password']")
            user_input.clear()
            user_input.send_keys(da_user)
            pass_input.clear()
            pass_input.send_keys(da_pass)
            pass_input.send_keys(Keys.RETURN)

            wait.until(EC.presence_of_element_located((By.XPATH, "//a[contains(@href, 'CMD_LOGOUT')]")))
            
            available_domains = []
            if CHECK_HOMEPAGE_DOMAINS:
                domain_elements = driver.find_elements(By.XPATH, "//a[contains(@href, '/CMD_SHOW_DOMAIN?domain=')]")
                available_domains = [el.text.strip().lower() for el in domain_elements]
            
            target_domain = get_target_domain(domain) if USE_SMART_DOMAIN else domain
            
            domains_to_try = []
            
            if domain in available_domains or not CHECK_HOMEPAGE_DOMAINS:
                domains_to_try.append(domain)
                
            if target_domain not in domains_to_try and (target_domain in available_domains or not CHECK_HOMEPAGE_DOMAINS):
                domains_to_try.append(target_domain)
                
            if SCAN_ALL_DA_DOMAINS:
                for avail_d in available_domains:
                    if avail_d not in domains_to_try:
                        domains_to_try.append(avail_d)
            
            if not domains_to_try:
                domains_to_try = [domain]
                
            successful_da_domains = []
            last_da_error = ""
            
            print(f"  -> Sẽ đi qua {len(domains_to_try)} thư mục tên miền để mở khóa (false): {domains_to_try}")
            
            for attempt_domain in domains_to_try:
                print(f"    -> Đang thử sửa wp-config.php ở thư mục: {attempt_domain}")
                try:
                    edit_url = f"{DA_LOGIN_URL}/CMD_FILE_MANAGER/domains/{attempt_domain}/public_html/wp-config.php?action=edit"
                    driver.get(edit_url)
                    
                    textarea = wait_short.until(EC.presence_of_element_located((By.NAME, "text")))
                    current_code = textarea.get_attribute("value")
                    
                    new_code = re.sub(r"define\(\s*'DISALLOW_FILE_MODS'\s*,\s*true\s*\);", 
                                      "define( 'DISALLOW_FILE_MODS', false );", 
                                      current_code)
                    
                    if new_code != current_code:
                        driver.execute_script("arguments[0].value = arguments[1];", textarea, new_code)
                        btn_save = driver.find_element(By.XPATH, "//input[@value='Save As']")
                        btn_save.click()
                        time.sleep(2.5) 
                        print(f"    -> [THÀNH CÔNG] Đã mở khóa (false) file wp-config.php cho {attempt_domain}.")
                    else:
                        print(f"    -> [BỎ QUA] File của {attempt_domain} đã mở sẵn hoặc không có biến này.")
                    
                    successful_da_domains.append(attempt_domain)
                    
                    if not SCAN_ALL_DA_DOMAINS:
                        break
                    
                except TimeoutException:
                    last_da_error = "Không tìm thấy file wp-config.php (Sai thư mục)"
                    print(f"    -> Bỏ qua do: {last_da_error}")
                except Exception as e:
                    last_da_error = str(e).split('\n')[0]
                    print(f"    -> Bỏ qua do lỗi: {last_da_error}")

            if not successful_da_domains:
                print("  -> LỖI NHIỆM VỤ 1: Không thể sửa wp-config.php ở bất kỳ thư mục nào.")
                sheet.cell(row=row_idx, column=headers['Done']).value = f"Lỗi wp-config: {last_da_error}"
                if AUTO_LOGOUT_DA:
                    try: driver.get(f"{DA_LOGIN_URL}/CMD_LOGOUT")
                    except: pass
                continue

            # ---------------- NHIỆM VỤ 2: ĐĂNG NHẬP WP & CÀI PLUGIN ----------------
            wp_domain = domain 
            
            print(f"  -> Bắt đầu Nhiệm vụ 2: Đăng nhập WP Admin ({wp_domain})...")
            driver.delete_all_cookies()
            
            try:
                driver.get(f"https://{wp_domain}/wp-login.php")
                
                page_src = driver.page_source.lower()
                if "no input file specified" in page_src:
                    raise Exception("Web bị lỗi 'No input file specified' (Lỗi Server)")
                elif "error establishing a database connection" in page_src:
                    raise Exception("Web bị lỗi Database")
                elif "404" in driver.title or "not found" in page_src:
                    raise Exception("Không tìm thấy link wp-login.php")
                
                try:
                    wp_user_input = wait.until(EC.presence_of_element_located((By.ID, "user_login")))
                except TimeoutException:
                    raise Exception("Không thể tải khung đăng nhập WordPress (Hết 15s chờ).")

                wp_user_input.send_keys(WP_USER)
                driver.find_element(By.ID, "user_pass").send_keys(WP_PASS)
                driver.find_element(By.ID, "wp-submit").click()
                
                try:
                    wait.until(EC.presence_of_element_located((By.ID, "wpadminbar")))
                except TimeoutException:
                    raise Exception("Sai tài khoản/mật khẩu WP hoặc web có Captcha bảo vệ.")
                    
                print("    -> Đăng nhập WordPress thành công.")

                driver.get(f"https://{wp_domain}/wp-admin/plugin-install.php")
                
                upload_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[contains(@class, 'upload-view-toggle')]")))
                upload_btn.click()
                
                file_input = wait.until(EC.presence_of_element_located((By.ID, "pluginzip")))
                file_input.send_keys(PLUGIN_ZIP_PATH)
                
                print("    -> Đang tải file lên và chờ xử lý cài đặt...")
                install_btn = driver.find_element(By.ID, "install-plugin-submit")
                install_btn.click()
                
                # --- CHỜ XỬ LÝ THÔNG MINH (Không bị treo 60s) ---
                timeout_limit = 60
                start_time = time.time()
                install_status = "TIMEOUT"
                action_btn = None
                
                while time.time() - start_time < timeout_limit:
                    try:
                        page_src = driver.page_source.lower()
                    except:
                        # Bỏ qua lỗi trong lúc DOM đang load tải trang
                        time.sleep(1)
                        continue
                    
                    # 1. Quét tìm nút Kích hoạt / Thay thế
                    btns = driver.find_elements(By.XPATH, "//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'kích hoạt') or contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'activate plugin') or contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'thay thế') or contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'replace current')]")
                    
                    if btns:
                        action_btn = btns[0]
                        btn_text = action_btn.text.lower()
                        if "thay" in btn_text or "replace" in btn_text:
                            install_status = "EXIST_REPLACE_BTN"
                        else:
                            install_status = "SUCCESS_ACTIVATE_BTN"
                        break
                        
                    # 2. Quét tìm chữ báo lỗi Tồn tại (Dành riêng cho WP bản cũ không có nút Thay thế)
                    if "đã tồn tại" in page_src or "already exists" in page_src or "không thành công" in page_src:
                        install_status = "EXIST_OLD_WP"
                        break
                        
                    time.sleep(1) # Nghỉ 1s rồi quét tiếp
                    
                # --- THỰC THI THEO TRẠNG THÁI ---
                if install_status == "SUCCESS_ACTIVATE_BTN":
                    action_btn.click()
                    print(f"  -> Đã KÍCH HOẠT PLUGIN thành công cho {wp_domain}!")
                    sheet.cell(row=row_idx, column=headers['Done']).value = "Thành công"
                    
                elif install_status == "EXIST_REPLACE_BTN":
                    if OVERWRITE_EXISTING_PLUGIN:
                        print("    -> Phát hiện plugin đã tồn tại. Đang tiến hành XÓA VÀ GHI ĐÈ...")
                        action_btn.click()
                        time.sleep(3) # Nghỉ chờ trang update load
                        try:
                            activate_link = driver.find_element(By.XPATH, "//a[contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'kích hoạt') or contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'activate plugin')]")
                            activate_link.click()
                            print(f"    -> Đã GHI ĐÈ & KÍCH HOẠT lại plugin cho {wp_domain}!")
                        except:
                            print(f"    -> Đã GHI ĐÈ thành công (Plugin đã được kích hoạt sẵn).")
                        sheet.cell(row=row_idx, column=headers['Done']).value = "Thành công (Đã ghi đè plugin)"
                    else:
                        print(f"    -> CẢNH BÁO: Plugin đã tồn tại sẵn. BỎ QUA không ghi đè theo tùy chọn.")
                        sheet.cell(row=row_idx, column=headers['Done']).value = "Thành công (Plugin đã tồn tại sẵn)"
                        
                elif install_status == "EXIST_OLD_WP":
                    print(f"  -> CẢNH BÁO: Plugin đã tồn tại nhưng bản WP cũ không có nút Thay thế. BỎ QUA an toàn.")
                    sheet.cell(row=row_idx, column=headers['Done']).value = "Thành công (Plugin tồn tại sẵn - WP cũ)"
                    
                else: # TIMEOUT
                    raise Exception("Quá 60s chờ cài đặt (Mạng chậm hoặc host chặn tải lên).")
                    
                sheet.cell(row=row_idx, column=headers['Domain thực tế']).value = wp_domain
                
            except Exception as e:
                last_nv2_error = str(e).split('\n')[0]
                print(f"  -> LỖI TỔNG NHIỆM VỤ 2: Thất bại do: {last_nv2_error}")
                sheet.cell(row=row_idx, column=headers['Done']).value = f"Lỗi WP: {last_nv2_error}"

            # ---------------- NHIỆM VỤ 3: TRẢ LẠI TRẠNG THÁI KHÓA (REVERT) ----------------
            if REVERT_FILE_MODS and successful_da_domains:
                print(f"  -> Bắt đầu Nhiệm vụ 3: Đăng nhập lại DirectAdmin để khóa {len(successful_da_domains)} file (Revert to true)...")
                try:
                    driver.delete_all_cookies()
                    driver.get(DA_LOGIN_URL)
                    
                    user_input = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@name='username']")))
                    pass_input = driver.find_element(By.XPATH, "//input[@name='password']")
                    user_input.clear()
                    user_input.send_keys(da_user)
                    pass_input.clear()
                    pass_input.send_keys(da_pass)
                    pass_input.send_keys(Keys.RETURN)
                    
                    wait.until(EC.presence_of_element_located((By.XPATH, "//a[contains(@href, 'CMD_LOGOUT')]")))
                    
                    for revert_domain in successful_da_domains:
                        print(f"    -> Đang khóa lại wp-config.php ở: {revert_domain}")
                        try:
                            edit_url = f"{DA_LOGIN_URL}/CMD_FILE_MANAGER/domains/{revert_domain}/public_html/wp-config.php?action=edit"
                            driver.get(edit_url)
                            
                            textarea = wait_short.until(EC.presence_of_element_located((By.NAME, "text")))
                            current_code = textarea.get_attribute("value")
                            
                            new_code = re.sub(r"define\(\s*'DISALLOW_FILE_MODS'\s*,\s*false\s*\);", 
                                              "define( 'DISALLOW_FILE_MODS', true );", 
                                              current_code)
                            
                            if new_code != current_code:
                                driver.execute_script("arguments[0].value = arguments[1];", textarea, new_code)
                                btn_save = driver.find_element(By.XPATH, "//input[@value='Save As']")
                                btn_save.click()
                                time.sleep(2.5)
                                print(f"    -> [THÀNH CÔNG] Đã khóa file (true) cho {revert_domain}.")
                            else:
                                print(f"    -> [BỎ QUA] File đã ở trạng thái khóa sẵn.")
                        except Exception as e:
                            print(f"    -> Lỗi khi khóa lại {revert_domain}: {str(e).splitlines()[0]}")
                            
                    if AUTO_LOGOUT_DA:
                        try: driver.get(f"{DA_LOGIN_URL}/CMD_LOGOUT")
                        except: pass
                except Exception as e:
                    print(f"  -> LỖI NHIỆM VỤ 3: Không thể đăng nhập lại DA để khóa file ({str(e).splitlines()[0]})")

        except Exception as e:
            error_msg = str(e).split('\n')[0]
            print(f"  -> LỖI TỔNG: {error_msg}")
            sheet.cell(row=row_idx, column=headers['Done']).value = f"Lỗi: {error_msg}"
        
        finally:
            try:
                wb.save(EXCEL_PATH)
            except PermissionError:
                print("\n  [!] LỖI LƯU EXCEL: File Excel đang được mở. Dữ liệu dòng này chưa được lưu!\n")
            except Exception as e:
                print(f"\n  [!] LỖI LƯU EXCEL KHÔNG XÁC ĐỊNH: {e}\n")
            
            time.sleep(1.5)

    driver.quit()
    print("\n================= HOÀN THÀNH TOÀN BỘ QUÁ TRÌNH =================")

if __name__ == "__main__":
    main()