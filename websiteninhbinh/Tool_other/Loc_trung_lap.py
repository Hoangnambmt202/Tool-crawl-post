import time
import re
from urllib.parse import urlparse, urljoin
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

# ==============================================================================
# CẤU HÌNH (User Config)
# ==============================================================================

# Đường dẫn file Excel đầu vào
EXCEL_PATH = r"D:\D_Document\VS Code\Python\Thuc_tap\Tool_other\DS_LocBaiTrungLap.xlsx"

# Thông tin đăng nhập Admin
ADMIN_USER = "adminvtk"
ADMIN_PASS = "Khanhkh@nh9999"

# Đường dẫn Chrome Driver
CHROMEDRIVER_PATH = r"D:\D_Document\VS Code\Python\Thuc_tap\chromedriver-win64\chromedriver-win64\chromedriver.exe"
CHROME_BINARY     = r"D:\D_Document\VS Code\Python\Thuc_tap\chrome-win64\chrome-win64\chrome.exe"

# --- TÙY CHỌN MỚI ---
DELETE_ALL_DRAFTS = True  # True: Xóa TẤT CẢ bản nháp | False: Chỉ xóa bản nháp nếu bị trùng bài tốt hơn
# --------------------

# ==============================================================================
# DATA STRUCTURE
# ==============================================================================
class PostItem:
    def __init__(self, post_id, title, has_thumb, delete_url, is_published=True):
        self.id = post_id
        self.title = title
        self.has_thumb = has_thumb # True nếu có ảnh đại diện/media
        self.delete_url = delete_url
        self.is_published = is_published  # True = Đã đăng, False = Nháp/Chờ duyệt
        
        # ID thường dạng "post-1234", lấy số để so sánh
        try:
            self.id_num = int(post_id.replace("post-", ""))
        except:
            self.id_num = 0

    def __repr__(self):
        status_icon = "🟢" if self.is_published else "🟡"
        thumb_str = "📸" if self.has_thumb else "⚪"
        return f"{status_icon} ID:{self.id_num} {thumb_str} - {self.title[:30]}..."

# ==============================================================================
# LOGIC XỬ LÝ
# ==============================================================================

class DuplicateCleaner:
    def __init__(self):
        self.driver = self.setup_driver()
        self.wait = WebDriverWait(self.driver, 15)

    def setup_driver(self):
        """Khởi tạo Chrome Driver"""
        options = Options()
        options.binary_location = CHROME_BINARY
        # options.add_argument("--headless") # Bỏ comment nếu muốn chạy ẩn
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--window-size=1280,800")
        options.add_argument("--ignore-certificate-errors")
        
        service = Service(CHROMEDRIVER_PATH)
        driver = webdriver.Chrome(service=service, options=options)
        return driver

    def normalize_url(self, url):
        """Chuẩn hóa Base URL từ Excel."""
        url = str(url).strip().rstrip('/')
        if not url.startswith('http'):
            url = 'https://' + url
        return url

    def login_wordpress(self, base_url):
        """Đăng nhập vào trang Admin"""
        login_url = urljoin(base_url + '/', "wp-login.php")
        
        try:
            print(f" 🔑 Đang đăng nhập: {login_url}")
            self.driver.get(login_url)
            
            if "wp-admin" in self.driver.current_url:
                print("    -> Đã đăng nhập sẵn.")
                return True

            try:
                user_field = self.wait.until(EC.element_to_be_clickable((By.ID, "user_login")))
                user_field.click()
                user_field.clear()
                user_field.send_keys(ADMIN_USER)
                time.sleep(0.5) 

                pass_field = self.wait.until(EC.element_to_be_clickable((By.ID, "user_pass")))
                pass_field.click()
                pass_field.clear()
                pass_field.send_keys(ADMIN_PASS)
                time.sleep(0.5)

                submit_btn = self.wait.until(EC.element_to_be_clickable((By.ID, "wp-submit")))
                submit_btn.click()
                
                self.wait.until(EC.url_contains("wp-admin"))
                print("    -> Đăng nhập thành công.")
                return True
            except TimeoutException:
                if "wp-admin" in self.driver.current_url:
                     print("    -> Đăng nhập thành công (Re-check).")
                     return True
                
                print("    ❌ Lỗi: Không thể đăng nhập (Timeout hoặc sai pass).")
                return False
                
        except Exception as e:
            print(f"    ❌ Lỗi kết nối: {e}")
            return False

    def collect_posts(self, base_url):
        """Bước 1: Quét và thu thập thông tin bài viết (TẤT CẢ CÁC TRANG)"""
        admin_list_url = urljoin(base_url + '/', "wp-admin/edit.php")
        self.driver.get(admin_list_url)
        
        collected_posts = []
        page_num = 1
        print(f" 🔍 Bắt đầu quét TOÀN BỘ bài viết...")

        while True:
            print(f"    -> Đang quét trang {page_num}...", end="\r")
            rows = self.driver.find_elements(By.CSS_SELECTOR, "#the-list tr")
            
            # Nếu không tìm thấy dòng nào hoặc trang trống
            if not rows or "no-posts" in rows[0].get_attribute("class"):
                break
            
            for row in rows:
                try:
                    # 1. Lấy Tiêu đề
                    title_elem = row.find_element(By.CSS_SELECTOR, ".row-title")
                    title_text = title_elem.text.strip()
                    post_id = row.get_attribute("id")
                    
                    if not title_text: continue

                    # 2. Kiểm tra trạng thái (Published vs Draft)
                    is_published = True
                    try:
                        state_elem = row.find_elements(By.CSS_SELECTOR, ".post-state")
                        if state_elem:
                            state_text = state_elem[0].text.strip()
                            if state_text: # Nếu có text (Nháp, Pending...)
                                is_published = False
                    except:
                        pass

                    # 3. Kiểm tra có Ảnh đại diện
                    has_thumb = False
                    try:
                        thumb_img = row.find_elements(By.CSS_SELECTOR, ".column-featured_image img")
                        if thumb_img:
                            has_thumb = True
                    except:
                        pass

                    # 4. Lấy link xóa (Trash URL)
                    delete_url = None
                    try:
                        del_link_elem = row.find_element(By.CSS_SELECTOR, "a.submitdelete")
                        delete_url = del_link_elem.get_attribute("href")
                    except:
                        continue

                    # Tạo object
                    post_obj = PostItem(post_id, title_text, has_thumb, delete_url, is_published)
                    collected_posts.append(post_obj)

                except Exception:
                    continue

            # Chuyển trang (Tìm nút Next Page)
            try:
                # Tìm nút next page không có class 'disabled'
                next_page_btn = self.driver.find_elements(By.CSS_SELECTOR, ".tablenav-pages .next-page:not(.disabled)")
                if not next_page_btn:
                    break # Hết trang
                
                next_url = next_page_btn[0].get_attribute("href")
                if not next_url or "javascript" in next_url:
                    break
                
                self.driver.get(next_url)
                page_num += 1
                time.sleep(1) # Chờ một chút để trang load
            except:
                break
        
        print(f"\n ✅ Đã quét xong {page_num} trang.")
        return collected_posts

    def analyze_and_execute_delete(self, posts):
        """Bước 2: Phân tích trùng lặp và chọn bài 'thông minh' nhất để giữ"""
        
        grouped = {}
        for p in posts:
            if p.title not in grouped:
                grouped[p.title] = []
            grouped[p.title].append(p)

        total_duplicates = 0
        delete_list = [] 
        unique_drafts_kept = 0

        print(f" 📊 Tổng số bài đã quét: {len(posts)}")
        if DELETE_ALL_DRAFTS:
            print(f" ⚙️  CHẾ ĐỘ: Xóa trùng lặp + Xóa TẤT CẢ bản nháp...")
        else:
            print(f" ⚙️  Đang phân tích và chọn lọc bài tối ưu...")

        has_any_duplicate = False

        for title, items in grouped.items():
            if len(items) > 1:
                has_any_duplicate = True
                total_duplicates += (len(items) - 1)
                
                print(f"\n   ⚠️  TRÙNG LẶP ({len(items)} bài): \"{title}\"")
                print(f"   {'-'*70}")
                
                # Sắp xếp theo tiêu chí ưu tiên
                items.sort(key=lambda x: (not x.is_published, not x.has_thumb, -x.id_num))
                
                winner = items[0]
                losers = items[1:]
                
                def format_status(item):
                    s = "🟢 Đã đăng" if item.is_published else "🟡 Nháp/Chờ"
                    t = "📸 Có ảnh/Media" if item.has_thumb else "⚪ Không ảnh"
                    return f"{s} | {t}"

                # Kiểm tra xem Winner có bị xóa không (nếu là bản nháp và bật chế độ xóa sạch)
                if not winner.is_published and DELETE_ALL_DRAFTS:
                    print(f"      🗑️  XÓA BẢN NHẤP       | ID: {winner.id_num:<6} | {format_status(winner)} | Lý do: Chế độ xóa hết bản nháp")
                    if winner.delete_url:
                        delete_list.append(winner.delete_url)
                else:
                    print(f"      🏆 GIỮ LẠI (Tốt nhất) | ID: {winner.id_num:<6} | {format_status(winner)}")
                
                for loser in losers:
                    reason = ""
                    if loser.is_published != winner.is_published:
                        reason = "Vì là bản Nháp/Chưa duyệt"
                    elif loser.has_thumb != winner.has_thumb:
                        reason = "Vì nội dung sơ sài (thiếu ảnh/media)"
                    else:
                        reason = "Vì cũ hơn bài được giữ"

                    # Nếu đang bật xóa hết nháp thì lý do sẽ ưu tiên nháp
                    if not loser.is_published and DELETE_ALL_DRAFTS:
                        reason = "Chế độ xóa hết bản nháp"

                    print(f"      🗑️  XÓA BỎ             | ID: {loser.id_num:<6} | {format_status(loser)} | Lý do: {reason}")
                    
                    if loser.delete_url:
                        delete_list.append(loser.delete_url)
                print(f"   {'-'*70}")
            else:
                # Trường hợp bài viết không trùng (len == 1)
                item = items[0]
                if not item.is_published:
                    if DELETE_ALL_DRAFTS:
                        print(f"      🗑️  XÓA BẢN NHÁP DUY NHẤT | ID: {item.id_num:<6} | Lý do: Chế độ xóa hết bản nháp")
                        if item.delete_url:
                            delete_list.append(item.delete_url)
                    else:
                        unique_drafts_kept += 1

        if not has_any_duplicate:
            print("   ✅ Không phát hiện bài viết nào trùng lặp.")
        
        if not DELETE_ALL_DRAFTS and unique_drafts_kept > 0:
            print(f"   🛡️  Đã bảo vệ {unique_drafts_kept} bản nháp duy nhất (Không xóa).")

        # Bước 3: Thực hiện xóa
        deleted_count = 0
        if delete_list:
            # Loại bỏ các URL trùng trong danh sách xóa (nếu có)
            delete_list = list(dict.fromkeys(delete_list))
            
            print(f"\n 🚀 [THỰC THI] Đang xóa {len(delete_list)} bài viết...")
            for i, url in enumerate(delete_list, 1):
                try:
                    self.driver.get(url)
                    deleted_count += 1
                    if i % 10 == 0:
                        print(f"      ... đã xóa {i} bài", end="\r")
                except Exception as e:
                    print(f"      ❌ [LỖI XÓA] {e}")
            print(f"    -> Đã dọn dẹp xong.")
        else:
            print("\n 🆗 Sạch sẽ, không cần xóa gì.")

        return len(posts), total_duplicates, deleted_count

    def process_site(self, base_url):
        all_posts = self.collect_posts(base_url)
        return self.analyze_and_execute_delete(all_posts)

    def close(self):
        self.driver.quit()

# ==============================================================================
# MAIN
# ==============================================================================
def main():
    print("================================================================")
    print("   TOOL LỌC BÀI VIẾT TRÙNG LẶP (QUÉT TOÀN BỘ WEBSITE)")
    print("   Tiêu chí giữ bài: 1.ĐÃ ĐĂNG > 2.CÓ ẢNH/MEDIA > 3.MỚI NHẤT")
    if DELETE_ALL_DRAFTS:
        print("   CHẾ ĐỘ HIỆN TẠI: XÓA TẤT CẢ BẢN NHÁP")
    print("================================================================")
    
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    except FileNotFoundError:
        print(f"❌ [LỖI] Không tìm thấy file {EXCEL_PATH}")
        return

    cleaner = DuplicateCleaner()

    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            url_cell = row[0]
            if not url_cell.value:
                continue

            raw_url = str(url_cell.value)
            base_url = cleaner.normalize_url(raw_url)
            
            print(f"\n{'='*70}")
            print(f" 🏫 TRƯỜNG: {base_url}")
            print(f"    (Dòng Excel: {row_idx})")
            print(f"{'='*70}")
            
            if cleaner.login_wordpress(base_url):
                counted, duplicates, deleted = cleaner.process_site(base_url)
                
                # Cập nhật Excel
                row[1].value = counted
                row[2].value = duplicates
                row[3].value = deleted
                row[4].value = "OK"
                
                print(f"\n 📋 TỔNG KẾT: Quét {counted} | Trùng {duplicates} | Đã xóa {deleted}")
                wb.save(EXCEL_PATH)
            else:
                row[4].value = "Login Fail"
                print("    ⏭️ Bỏ qua do lỗi đăng nhập.")

    except Exception as e:
        print(f"❌ [LỖI CHƯƠNG TRÌNH] {e}")
    finally:
        cleaner.close()
        wb.save(EXCEL_PATH)
        print("\n================================================================")
        print("   HOÀN TẤT VÀ ĐÃ LƯU FILE EXCEL")
        print("================================================================")

if __name__ == "__main__":
    main()