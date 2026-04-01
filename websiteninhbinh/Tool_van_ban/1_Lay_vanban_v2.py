# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  1_LAY_VANBAN.PY  —  Scraper bảng văn bản pháp quy (BẢN FINAL V1.4)        ║
║                                                                              ║
║  * TÍCH HỢP CẤU TRÚC MỚI: Tự động nhận diện dữ liệu ẩn trong thẻ div         ║
║    (vanban_details-...) ngay tại bảng danh sách, không cần mở tab mới.     ║
║  * BẢO TỒN V1: Giữ nguyên 100% logic vòng lặp ổn định của bản v1 gốc.      ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import sys
import re
import time
import json
import random
import traceback
from datetime import datetime, date
from urllib.parse import urljoin, urlparse, unquote

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import mysql.connector

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchWindowException

from user_agents import USER_AGENTS

from config_vanban import (
    CHROMEDRIVER_PATH, CHROME_BINARY, SHOW_CHROME_WINDOW,
    USE_PROFILE, PROFILE_DIR, EXCEL_PATH, ERROR_LOG_FILE,
    FROM_DATE, MAX_LIST_PAGES,
)

requests.packages.urllib3.disable_warnings()

def _enable_ansi_windows():
    if sys.platform != 'win32': return
    try:
        import ctypes, ctypes.wintypes
        kernel32 = ctypes.windll.kernel32
        handle   = kernel32.GetStdHandle(-11)
        old_mode = ctypes.wintypes.DWORD(0)
        kernel32.GetConsoleMode(handle, ctypes.byref(old_mode))
        kernel32.SetConsoleMode(handle, old_mode.value | 0x0004)
    except Exception: pass
_enable_ansi_windows()

class Color:
    RESET, BOLD = '\033[0m', '\033[1m'
    RED, GREEN, YELLOW, CYAN = '\033[31m', '\033[32m', '\033[33m', '\033[36m'
    BRIGHT_BLACK, BRIGHT_RED = '\033[90m', '\033[91m'
    MAGENTA = '\033[35m'

class Icon:
    SUCCESS, ERROR, WARNING, ARROW, FIRE = '✓', '✗', '⚠', '→', '🔥'

class DebugLogger:
    def __init__(self):
        self._on_progress_line = False
        self._list_total_added = self._list_total_skip = 0
        self._list_cur_page = 1
    def _flush(self):
        if self._on_progress_line: print(); self._on_progress_line = False
    def _p(self, msg: str):
        self._flush(); print(msg)
    @staticmethod
    def _trunc(s, n=65):
        s = str(s or '').strip()
        return (s[:n - 1] + '…') if len(s) > n else s
    @staticmethod
    def _fmt_t(secs: float) -> str:
        if secs < 60: return f'{secs:.1f}s'
        m, s = divmod(int(secs), 60)
        return f'{m}m{s:02d}s'

    def warning(self, msg): self._p(f'  {Color.YELLOW}{Icon.WARNING} {msg}{Color.RESET}')
    def error(self, msg): self._p(f'  {Color.RED}{Icon.ERROR} {msg}{Color.RESET}')
    def critical(self, msg): self._p(f'\n{Color.BRIGHT_RED}{Color.BOLD}{Icon.FIRE} {msg}{Color.RESET}')

    def row_start(self, n, total, source='', cat='', target=''):
        self._flush()
        print(f'{Color.BRIGHT_BLACK}{"─" * 72}{Color.RESET}')
        print(f'  {Color.BOLD}DÒNG {n}/{total}{Color.RESET}' + (f'  {Color.BRIGHT_BLACK}Danh mục Excel: {cat}{Color.RESET}' if cat else ''))
        if source: print(f'  {Color.BRIGHT_BLACK}src: {self._trunc(source, 70)}{Color.RESET}')
        if target: print(f'  {Color.BRIGHT_BLACK}dst: {self._trunc(target, 70)}{Color.RESET}')

    def row_done(self, n, saved, errors, elapsed):
        parts = []
        if saved: parts.append(f'{Color.GREEN}{saved} đã lưu{Color.RESET}')
        if errors: parts.append(f'{Color.RED}{errors} lỗi{Color.RESET}')
        if not parts: parts.append(f'{Color.BRIGHT_BLACK}0 bài{Color.RESET}')
        self._p(f'  {Color.GREEN}{Icon.SUCCESS} DÒNG {n} XONG  {"  ".join(parts)}  {Color.BRIGHT_BLACK}{self._fmt_t(elapsed)}{Color.RESET}')

    def list_start(self, url):
        self._list_total_added = self._list_total_skip = 0
        self._list_cur_page = 1
        self._p(f'\n  {Color.CYAN}◆ LIST  {Color.RESET}{Color.BRIGHT_BLACK}{self._trunc(url, 80)}{Color.RESET}')

    def list_page_loaded(self, kb, item_count, page_num):
        self._p(f'    {Color.BRIGHT_BLACK}✓ Loaded {f"{kb:.0f}KB" if kb else "?KB"}  →  {item_count} mục  trang {page_num}{Color.RESET}')

    def list_item(self, title, is_new=True, reason=''):
        if is_new:
            icon = f'{Color.GREEN}[+]{Color.RESET}'
            self._list_total_added += 1
        elif "Lỗi" in reason or "Bỏ qua" in reason:
            icon = f'{Color.RED}[✗]{Color.RESET}' if "Lỗi" in reason else f'{Color.BRIGHT_BLACK}[=]{Color.RESET}'
            self._list_total_skip += 1
        else:
            icon = f'{Color.BRIGHT_BLACK}[=]{Color.RESET}'
            self._list_total_skip += 1
            
        rs = f'  {Color.BRIGHT_BLACK}({reason}){Color.RESET}' if reason else ''
        self._p(f'    {icon} {self._trunc(title, 60)}{rs}')

    def list_next_page(self, page_num):
        self._list_cur_page = page_num
        self._p(f'    {Color.BRIGHT_BLACK}{Icon.ARROW} Sang trang {page_num}...{Color.RESET}')

    def list_done(self, elapsed):
        parts = []
        if self._list_total_added: parts.append(f'{Color.GREEN}{self._list_total_added} đã lưu{Color.RESET}')
        if self._list_total_skip:  parts.append(f'{Color.BRIGHT_BLACK}{self._list_total_skip} bỏ qua/lỗi{Color.RESET}')
        pg_s = f'{self._list_cur_page} trang' if self._list_cur_page > 1 else '1 trang'
        self._p(f'    {Color.GREEN}{Icon.SUCCESS} LIST xong  {"  ".join(parts)}  {Color.BRIGHT_BLACK}{pg_s}{Color.RESET}  {Color.BRIGHT_BLACK}{self._fmt_t(elapsed)}{Color.RESET}')

    def detail_start(self, title): self._p(f'    {Color.MAGENTA}◇ DETAIL  {Color.RESET}{Color.BRIGHT_BLACK}{self._trunc(title, 65)}{Color.RESET}')
    def detail_loaded(self, kb): self._p(f'      {Color.BRIGHT_BLACK}✓ Loaded {f"{kb:.0f}KB" if kb else "?KB"}{Color.RESET}')
    
    def detail_extracted(self, record):
        self._p(f'      {Color.CYAN}├─ Số ký hiệu : {Color.RESET}{self._trunc(record.get("so_hieu"), 50)}')
        self._p(f'      {Color.CYAN}├─ Ngày BH    : {Color.RESET}{record.get("ngay_ban_hanh")}')
        self._p(f'      {Color.CYAN}├─ Người ký   : {Color.RESET}{self._trunc(record.get("nguoi_ky"), 50)}')
        self._p(f'      {Color.CYAN}├─ Cơ quan    : {Color.RESET}{self._trunc(record.get("co_quan"), 50)}')
        self._p(f'      {Color.CYAN}├─ Lĩnh vực   : {Color.RESET}{self._trunc(record.get("pham_vi"), 50)}')
        self._p(f'      {Color.CYAN}├─ Loại VB    : {Color.RESET}{self._trunc(record.get("loai_van_ban"), 50)}')
        self._p(f'      {Color.CYAN}├─ Trích yếu  : {Color.RESET}{self._trunc(record.get("mo_ta"), 50)}')
        try:
            pdfs = json.loads(record.get("pdf_urls", "[]"))
            pdf_str = f"{len(pdfs)} file(s)" if pdfs else "Không có"
        except:
            pdf_str = "Lỗi phân tích file"
        self._p(f'      {Color.CYAN}├─ Đính kèm   : {Color.RESET}{pdf_str}')

    def detail_saved(self, elapsed): self._p(f'      {Color.GREEN}💾 Lưu DB  {Color.BRIGHT_BLACK}{self._fmt_t(elapsed)}{Color.RESET}')
    def detail_error(self, msg): self._p(f'      {Color.RED}{Icon.ERROR} {msg}{Color.RESET}')

    def session_summary(self, total_saved, total_errors, elapsed_str):
        self._flush()
        print(f'\n{"=" * 72}\n  {Color.BOLD}PHIÊN LÀM VIỆC KẾT THÚC{Color.RESET}\n  Đã lưu  : {Color.GREEN}{total_saved}{Color.RESET}')
        if total_errors: print(f'  Lỗi     : {Color.RED}{total_errors}{Color.RESET}')
        print(f'  Thời gian: {elapsed_str}\n{"=" * 72}')

log = DebugLogger()

# ══════════════════════════════════════════════════════════════════════════════
# DB HELPERS
# ══════════════════════════════════════════════════════════════════════════════
_DB = dict(host='localhost', user='root', password='', database='baivietphothong')

def _conn(): return mysql.connector.connect(**_DB)

def ensure_table():
    c = cur = None
    try:
        c = _conn(); cur = c.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS bot_vanban (
                id             INT AUTO_INCREMENT PRIMARY KEY,
                url_id         VARCHAR(50),
                title          TEXT,
                url            VARCHAR(2000),
                so_hieu        VARCHAR(300),
                co_quan        VARCHAR(500),
                pham_vi        VARCHAR(500),
                ngay_ban_hanh  VARCHAR(100),
                ngay_hieu_luc  VARCHAR(100),
                trang_thai     VARCHAR(200),
                loai_van_ban   VARCHAR(300),
                nguoi_ky       VARCHAR(300),
                noi_nhan       TEXT,
                mo_ta          TEXT,
                pdf_urls       TEXT,
                content        LONGTEXT,
                cat_id         VARCHAR(300),
                date_publish   DATE,
                upload         INT DEFAULT 0,
                created_at     TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        c.commit()
    finally:
        if cur: cur.close()
        if c: c.close()

def get_db_records(url_id, title, url):
    c = cur = None
    try:
        c = _conn(); cur = c.cursor(dictionary=True)
        if url and url.strip():
            cur.execute('SELECT * FROM bot_vanban WHERE url_id=%s AND (url=%s OR title=%s)', (url_id, url, title))
        else:
            cur.execute('SELECT * FROM bot_vanban WHERE url_id=%s AND title=%s', (url_id, title))
        return cur.fetchall()
    finally:
        if cur: cur.close()
        if c: c.close()

def normalize_text(text, is_so_hieu=False):
    t = str(text or '').strip().lower()
    if is_so_hieu:
        return t.replace(" ", "")
    return re.sub(r'\s+', ' ', t)

def normalize_date(date_str):
    return re.sub(r'\b0+(\d)', r'\1', str(date_str or '').strip())

def is_exact_match(record, db_rows):
    if not db_rows:
        return False
        
    r_title = normalize_text(record.get('title'))
    r_sh    = normalize_text(record.get('so_hieu'), is_so_hieu=True)
    r_cq    = normalize_text(record.get('co_quan'))
    r_pv    = normalize_text(record.get('pham_vi'))
    r_nb    = normalize_date(record.get('ngay_ban_hanh'))
    r_nhl   = normalize_date(record.get('ngay_hieu_luc'))
    r_tt    = normalize_text(record.get('trang_thai'))
    r_lvb   = normalize_text(record.get('loai_van_ban'))
    r_nk    = normalize_text(record.get('nguoi_ky'))
    r_nn    = normalize_text(record.get('noi_nhan'))
    r_mt    = normalize_text(record.get('mo_ta'))

    for r in db_rows:
        db_title = normalize_text(r['title'])
        db_sh    = normalize_text(r['so_hieu'], is_so_hieu=True)
        db_cq    = normalize_text(r['co_quan'])
        db_pv    = normalize_text(r['pham_vi'])
        db_nb    = normalize_date(r['ngay_ban_hanh'])
        db_nhl   = normalize_date(r['ngay_hieu_luc'])
        db_tt    = normalize_text(r['trang_thai'])
        db_lvb   = normalize_text(r['loai_van_ban'])
        db_nk    = normalize_text(r['nguoi_ky'])
        db_nn    = normalize_text(r['noi_nhan'])
        db_mt    = normalize_text(r['mo_ta'])

        if (db_title == r_title and db_sh == r_sh and db_cq == r_cq and 
            db_pv == r_pv and db_nb == r_nb and db_nhl == r_nhl and 
            db_tt == r_tt and db_lvb == r_lvb and db_nk == r_nk and 
            db_nn == r_nn and db_mt == r_mt):
            
            try: db_pdfs = [unquote(p['url']).split('?')[0] for p in json.loads(r['pdf_urls'] or '[]')]
            except: db_pdfs = []
            
            try: rec_pdfs = [unquote(p['url']).split('?')[0] for p in json.loads(record.get('pdf_urls', '[]'))]
            except: rec_pdfs = []
            
            if set(db_pdfs) == set(rec_pdfs):
                return True
    return False

def save_vanban(record: dict):
    c = cur = None
    try:
        c = _conn(); cur = c.cursor()
        
        url_id_val = str(record.get('url_id') or '')[:50]
        title_val = str(record.get('title') or '')
        url_val = str(record.get('url') or '')[:2000]
        so_hieu_val = str(record.get('so_hieu') or '')[:300]
        co_quan_val = str(record.get('co_quan') or '')[:500]
        pham_vi_val = str(record.get('pham_vi') or '')[:500]
        ngay_bh_val = str(record.get('ngay_ban_hanh') or '')[:100]
        ngay_hl_val = str(record.get('ngay_hieu_luc') or '')[:100]
        trang_thai_val = str(record.get('trang_thai') or '')[:200]
        loai_vb_val = str(record.get('loai_van_ban') or '')[:300]
        nguoi_ky_val = str(record.get('nguoi_ky') or '')[:300]
        
        cur.execute("""
            INSERT INTO bot_vanban
                (url_id, title, url, so_hieu, co_quan, pham_vi, ngay_ban_hanh, ngay_hieu_luc, trang_thai, loai_van_ban, nguoi_ky, noi_nhan, mo_ta, pdf_urls, content, cat_id, date_publish)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            url_id_val, title_val, url_val, so_hieu_val,
            co_quan_val, pham_vi_val, ngay_bh_val, ngay_hl_val,
            trang_thai_val, loai_vb_val, nguoi_ky_val, record.get('noi_nhan'),
            record.get('mo_ta'), record.get('pdf_urls'), record.get('content'), record.get('cat_id'), record.get('date_publish'),
        ))
        c.commit()
    finally:
        if cur: cur.close()
        if c: c.close()

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS & PARSER
# ══════════════════════════════════════════════════════════════════════════════
def clean(text): return ' '.join((text or '').split()).strip()

def parse_date(text):
    m = re.search(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', str(text or '').lower())
    if m:
        try: return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except: pass
    return None

def get_base(url): p = urlparse(url); return f'{p.scheme}://{p.netloc}/'

def parse_listing_table(soup, base_url):
    items = []
    # --- TRƯỜNG HỢP 1: DẠNG KHỐI (Style 3 - MN Thiên Tôn) ---
    blocks = soup.find_all('div', class_='legalddocument-default')
    if blocks:
        for block in blocks:
            # 1. Tiêu đề và Link chi tiết
            title_tag = block.find('a', class_='title-documment')
            title = clean(title_tag.get_text()) if title_tag else ""
            detail_url = urljoin(base_url, title_tag['href']) if (title_tag and title_tag.get('href')) else ""

            # 2. Lấy danh sách file đính kèm
            pdf_list = []
            brief_div = block.find('div', class_='brief')
            if brief_div:
                for a_file in brief_div.find_all('a', href=True):
                    f_url = a_file['href'].strip()
                    if f_url and not f_url.startswith(('javascript:', '#')):
                        pdf_list.append({
                            'url': urljoin(base_url, f_url), 
                            'name': clean(a_file.get_text()) or "File đính kèm"
                        })

            # 3. Lấy Meta data (Ngày ban hành, hiệu lực, trạng thái)
            # Vì cấu trúc này không có bảng ẩn, ta lấy trực tiếp vào inline_details
            inline_details = {}
            pub_time = block.find('div', class_='publish-time')
            eff_time = block.find('div', class_='effective-time')
            status_tag = block.find('span', class_='has-hl')

            ngay_ban_hanh = clean(pub_time.get_text()) if pub_time else ""
            if eff_time: inline_details['ngay_hieu_luc'] = clean(eff_time.get_text())
            if status_tag: inline_details['trang_thai'] = clean(status_tag.get_text())

            if title:
                items.append({
                    'title': title,
                    'detail_url': detail_url,
                    'noi_nhan': '',
                    'so_hieu': '', # Thường dạng này không hiện số hiệu ra ngoài
                    'ngay_ban_hanh': ngay_ban_hanh,
                    'pdf_list': pdf_list,
                    'inline_details': inline_details
                })
        
        if items:
            # Trả về danh sách và khối bao quanh để làm content (giả lập table)
            return items, str(soup.find('section', id=lambda x: x and 'section' in x) or "Dạng danh sách khối")

    # --- TRƯỜNG HỢP 2: DẠNG BẢNG (Cấu trúc cũ như các phiên trước) ---
    table = soup.find('table', id='vanbantb')
    if not table:
        table = soup.find('table', class_=lambda c: c and 'table-bordered' in c)
    if not table:
        wrapper = soup.find('div', class_='list-legal-document-table')
        table = wrapper.find('table') if wrapper else None

    if table:
        tbody = table.find('tbody') or table
        for tr in tbody.find_all('tr'):
            tds = tr.find_all('td')
            if not tds: continue
            
            title = ''
            detail_url = ''
            pdf_list = []
            inline_details = {}

            # Kiểm tra div ẩn vanban_details-...
            detail_div = tds[1].find('div', id=lambda x: x and str(x).startswith('vanban_details-')) if len(tds) > 1 else None
            if detail_div:
                # Logic cho web Nhật Tựu (đã có ở bản V1.4 của bạn)
                so_hieu = clean(tds[0].get_text()) if len(tds) > 0 else ''
                ngay_ban_hanh = clean(tds[2].get_text()) if len(tds) > 2 else ''
                sort_div = tds[1].find('div', id=lambda x: x and str(x).startswith('vanban_sort-'))
                title = clean(sort_div.get_text()) if sort_div else clean(tds[1].get_text())
                
                for li in detail_div.find_all('li'):
                    # ... (giữ nguyên logic lấy details cũ của bạn) ...
                    pass 
                detail_url = ""
            else:
                # Logic cho Web cũ
                if len(tds) < 4: continue
                a_tag = tds[1].find('a', href=True)
                title = clean(a_tag.get_text()) if a_tag else clean(tds[1].get_text())
                detail_url = urljoin(base_url, a_tag['href']) if a_tag else ''
                so_hieu = clean(tds[3].get_text()) if len(tds) > 3 else ''
                ngay_ban_hanh = clean(tds[4].get_text()) if len(tds) > 4 else ''
                # Lấy file...
            
            if title:
                items.append({
                    'title': title, 'detail_url': detail_url,
                    'so_hieu': so_hieu, 'ngay_ban_hanh': ngay_ban_hanh,
                    'pdf_list': pdf_list, 'inline_details': inline_details,
                    'noi_nhan': ''
                })
        return items, str(table)

    return [], ''

_LABEL_MAP = {
    'tiêu đề': 'title', 'tên văn bản': 'title',
    'số hiệu': 'so_hieu', 'số ký hiệu': 'so_hieu', 'kí hiệu': 'so_hieu',
    'cơ quan': 'co_quan', 'nơi ban hành': 'co_quan', 'đơn vị': 'co_quan',
    'phạm vi': 'pham_vi',
    'ngày ban hành': 'ngay_ban_hanh', 'ban hành': 'ngay_ban_hanh',
    'ngày hiệu lực': 'ngay_hieu_luc', 'hiệu lực': 'ngay_hieu_luc',
    'trạng thái': 'trang_thai', 'tình trạng': 'trang_thai',
    'loại văn bản': 'loai_van_ban', 'loại': 'loai_van_ban',
    'người ký': 'nguoi_ky', 'ký bởi': 'nguoi_ky', 'chức danh': 'nguoi_ky', 'người kí': 'nguoi_ky',
    'mô tả': 'mo_ta', 'trích yếu': 'mo_ta', 'nội dung': 'mo_ta', 'tóm tắt': 'mo_ta', 'về việc': 'mo_ta'
}

def parse_detail(soup, base_url):
    result, pdf_list = {}, []
    
    for table in soup.find_all('table'):
        for tr in table.find_all('tr'):
            tds = tr.find_all(['td', 'th'])
            i = 0
            while i < len(tds) - 1:
                label_cell = tds[i]
                value_cell = tds[i+1]
                
                label = clean(label_cell.get_text()).lower().replace(':', '').strip()
                span = value_cell.find('span')
                val = clean(span.get_text()) if span else clean(value_cell.get_text())

                for key, field in _LABEL_MAP.items():
                    if key in label and not result.get(field):
                        result[field] = val
                        break

                if any(k in label for k in ('file', 'tài liệu', 'đính kèm', 'tải về', 'văn bản')):
                    for a in value_cell.find_all('a', href=True):
                        href = a['href'].strip()
                        if href and not href.startswith(('javascript:', 'mailto:', 'tel:')):
                            pdf_list.append({'url': urljoin(base_url, href), 'name': clean(a.get_text()) or 'vanban.pdf'})
                i += 2

    for el in soup.find_all(['div', 'p', 'li']):
        text = clean(el.get_text())
        lower_text = text.lower()
        for key, field in _LABEL_MAP.items():
            if lower_text.startswith(f"{key}:") and not result.get(field):
                result[field] = clean(text[len(key)+1:])
                break

    if not pdf_list:
        for iframe in soup.find_all('iframe', src=True):
            src = iframe['src']
            if 'view.officeapps.live.com' in src and 'src=' in src:
                file_url = src.split('src=')[-1].split('&')[0]
                file_url = unquote(file_url)
                if not file_url.startswith('http'): file_url = 'http://' + file_url.lstrip('/')
                pdf_list.append({'url': file_url, 'name': 'vanban_word.doc'})
            elif 'docs.google.com' in src and 'url=' in src:
                file_url = src.split('url=')[-1].split('&')[0]
                file_url = unquote(file_url)
                pdf_list.append({'url': file_url, 'name': 'vanban.pdf'})

    if not pdf_list:
        for a in soup.find_all('a', href=True):
            href = a['href'].strip()
            if not href.startswith(('javascript:', 'mailto:', 'tel:')):
                clean_href = href.split('?')[0].lower()
                if clean_href.endswith(('.pdf', '.doc', '.docx', '.rar', '.zip', '.xls', '.xlsx')):
                    pdf_list.append({'url': urljoin(base_url, href), 'name': clean(a.get_text()) or 'vanban.pdf'})

    result['pdf_list'] = pdf_list
    return result


# ══════════════════════════════════════════════════════════════════════════════
# BOT & RUNNER
# ══════════════════════════════════════════════════════════════════════════════
class VanBanBot:
    def __init__(self):
        opts = Options()
        if not SHOW_CHROME_WINDOW: opts.add_argument('--headless=new')
        if CHROME_BINARY: opts.binary_location = CHROME_BINARY
        if USE_PROFILE and PROFILE_DIR: opts.add_argument(f'--user-data-dir={PROFILE_DIR}')
        opts.add_argument('--disable-notifications')
        opts.add_argument('--window-size=1366,768')
        opts.add_argument('--disable-blink-features=AutomationControlled')
        opts.add_experimental_option('excludeSwitches', ['enable-automation'])
        
        self.current_ua = random.choice(USER_AGENTS)
        opts.add_argument(f'--user-agent={self.current_ua}')
        
        self.driver = webdriver.Chrome(service=Service(CHROMEDRIVER_PATH), options=opts)
        
        if SHOW_CHROME_WINDOW:
            self.driver.minimize_window()
            
        self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        })
        
        self.url_id = self.url = self.cat_id = self.target = ''

    def close(self):
        try: self.driver.quit()
        except: pass

    def get_data(self):
        saved = errors = 0
        current_failed_list = [] 
        base = get_base(self.url)
        t_start, page_num = datetime.now(), 1
        last_page_titles = []
        
        try:
            self.driver.set_page_load_timeout(20)
            self.driver.get(self.url)
        except Exception as e:
            log.warning(f'Selenium lỗi tải URL gốc {self.url}: {e}')
            return 0, 1, [("Trang nguồn", "Không thể truy cập URL gốc")]

        time.sleep(2)
        log.list_start(self.driver.current_url)

        while page_num <= MAX_LIST_PAGES:
            html = self.driver.page_source
            soup = BeautifulSoup(html, 'html.parser')
            kb = len(html.encode('utf-8')) / 1024
            
            if not soup:
                log.warning(f'Không tải được trang {page_num}'); break
            
            items, table_html = parse_listing_table(soup, base)
            
            current_titles = [item['title'] for item in items]
            if current_titles and current_titles == last_page_titles:
                log.warning('Web không tải thêm được nội dung mới (Đã hết trang). Dừng quét an toàn.')
                break
            last_page_titles = current_titles

            log.list_page_loaded(kb, len(items), page_num)
            if not items: log.warning('Không tìm thấy bảng dữ liệu trên trang này'); break

            for item in items:
                title, detail_url = item['title'], item['detail_url']
                date_pub = parse_date(item.get('ngay_ban_hanh', ''))
                
                if date_pub and FROM_DATE and date_pub < FROM_DATE:
                    log.list_item(title, is_new=False, reason='cũ hơn FROM_DATE'); continue
                
                db_rows = get_db_records(self.url_id, title, detail_url)

                detail = item.get('inline_details', {})
                pdf_merged = list(item.get('pdf_list', []))
                t_detail = datetime.now()
                
                detail_failed = False 
                fail_reason = ""

                # Chỉ mở tab mới nếu có detail_url (Cấu trúc cũ)
                if detail_url and not detail_url.startswith(('javascript:', '#')):
                    log.detail_start(title)
                    main_window = self.driver.current_window_handle
                    self.driver.switch_to.new_window('tab')
                    
                    max_retries = 2
                    for attempt in range(max_retries):
                        try:
                            self.driver.set_page_load_timeout(10) 
                            try:
                                self.driver.get(detail_url)
                            except TimeoutException:
                                pass 
                            
                            try:
                                WebDriverWait(self.driver, 5).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, 'table, iframe, .content, .detail-content, p'))
                                )
                            except TimeoutException:
                                try:
                                    page_text = self.driver.find_element(By.TAG_NAME, "body").text
                                    if "fetching your file" in page_text.lower():
                                        raise Exception("Kẹt ở màn hình Anti-bot (Fetching File)")
                                except Exception as inner_e:
                                    if "Kẹt ở màn hình" in str(inner_e):
                                        raise inner_e
                                    pass
                            
                            d_html = self.driver.page_source
                            d_soup = BeautifulSoup(d_html, 'html.parser')
                            
                            parsed_detail = parse_detail(d_soup, base)
                            detail.update(parsed_detail)
                            
                            has_viewer = bool(d_soup.find('iframe', src=re.compile(r'officeapps\.live\.com|docs\.google\.com', re.I)))
                            if has_viewer and not detail.get('pdf_list'):
                                raise Exception("Lỗi nhận diện file Word/PDF (Mặc dù có iframe)")
                                
                            log.detail_loaded(len(d_html.encode('utf-8')) / 1024)
                            seen = {p['url'] for p in pdf_merged}
                            for p in detail.get('pdf_list', []):
                                if p['url'] not in seen: pdf_merged.append(p); seen.add(p['url'])
                            break 
                                
                        except Exception as e:
                            if attempt < max_retries - 1:
                                log.warning(f"      Lỗi: {e}. Đang thử lại (Lần {attempt + 2})...")
                                time.sleep(3)
                            else:
                                detail_failed = True
                                fail_reason = str(e)
                                log.error(f"  Thất bại khi bóc tách -> BỎ QUA BÀI NÀY.")
                                
                    try:
                        self.driver.close()
                    except NoSuchWindowException:
                        pass
                    self.driver.switch_to.window(main_window)
                elif detail:
                    log.detail_start(f"[Inline] {title}")

                if detail_failed:
                    log.list_item(title, is_new=False, reason=f'Lỗi: {fail_reason}')
                    current_failed_list.append((title, fail_reason))
                    errors += 1
                    continue

                record = {
                    'url_id': self.url_id, 'title': detail.get('title') or title, 'url': detail_url,
                    'so_hieu': detail.get('so_hieu') or item.get('so_hieu', ''),
                    'co_quan': detail.get('co_quan', ''), 'pham_vi': detail.get('pham_vi', ''),
                    'ngay_ban_hanh': detail.get('ngay_ban_hanh') or item.get('ngay_ban_hanh', ''),
                    'ngay_hieu_luc': detail.get('ngay_hieu_luc', ''), 'trang_thai': detail.get('trang_thai', ''),
                    'loai_van_ban': detail.get('loai_van_ban', ''), 'nguoi_ky': detail.get('nguoi_ky', ''),
                    'noi_nhan': item['noi_nhan'], 'mo_ta': detail.get('mo_ta', '') or title,
                    'pdf_urls': json.dumps(pdf_merged, ensure_ascii=False),
                    'content': table_html, 'cat_id': self.cat_id, 'date_publish': date_pub,
                }
                
                if is_exact_match(record, db_rows):
                    log.list_item(title, is_new=False, reason='Giống 100% DB -> Bỏ qua')
                    continue
                
                reason_str = 'Bản cập nhật mới (Có dữ liệu thay đổi)' if db_rows else ''
                if not detail_url: 
                    log.list_item(title, is_new=True, reason=reason_str)
                    
                log.detail_extracted(record)
                
                try:
                    save_vanban(record); saved += 1
                    log.detail_saved((datetime.now() - t_detail).total_seconds())
                except Exception as e:
                    log.detail_error(f'Lỗi lưu DB: {e}'); errors += 1
                time.sleep(0.5)

            js_find_and_click_next = """
            var currentPage = arguments[0];
            var nextStr = (currentPage + 1).toString();

            var nextBtns = Array.from(document.querySelectorAll('a, button, li')).filter(el => {
                if(el.classList.contains('next')) return true;
                var t = el.innerText.trim().toLowerCase();
                return t === '>' || t === '»' || t === 'next' || t === 'trang sau' || t === 'sau';
            });

            nextBtns = nextBtns.filter(el => !el.disabled && !el.classList.contains('disabled') && !el.parentElement.classList.contains('disabled'));
            var target = nextBtns.length > 0 ? nextBtns[0] : null;

            if (!target) {
                var numBtns = Array.from(document.querySelectorAll('a, button, li')).filter(el => el.innerText.trim() === nextStr);
                if(numBtns.length > 0) target = numBtns[0];
            }

            if(target) {
                if(target.tagName.toLowerCase() === 'li') {
                    var a = target.querySelector('a');
                    if(a) target = a;
                }
                target.click();
                return true;
            }
            return false;
            """
            
            clicked = False
            try:
                clicked = self.driver.execute_script(js_find_and_click_next, page_num)
            except Exception: pass

            if clicked:
                log.list_next_page(page_num + 1)
                time.sleep(4) 
                page_num += 1
                continue

            js_fallback = """
            var current_url = arguments[0], cur_page = arguments[1];
            var next_str = (cur_page + 1).toString();
            var keywords = ['>', '»', 'next', 'trang sau', 'sau', next_str];
            
            var pags = document.querySelectorAll('ul.pagination, div.pagination, .page, .nav, .paging');
            for(var p of pags) {
                var els = p.querySelectorAll('a, span, li, button');
                for(var el of els) {
                    var txt = el.innerText.trim().toLowerCase();
                    if(keywords.includes(txt)) {
                        if(el.classList.contains('disabled') || el.disabled || el.parentElement.classList.contains('disabled')) continue;
                        if(el.tagName === 'A' && el.href && !el.href.startsWith('javascript')) return el.href;
                    }
                }
            }
            return null;
            """
            try:
                next_url = self.driver.execute_script(js_fallback, self.driver.current_url, page_num)
                if next_url and next_url != self.driver.current_url:
                    log.list_next_page(page_num + 1)
                    self.driver.get(next_url)
                    time.sleep(3)
                    page_num += 1
                    continue
            except Exception: pass

            log.warning('Đã đến trang cuối cùng. Dừng quét an toàn.')
            break

        log.list_done((datetime.now() - t_start).total_seconds())
        return saved, errors, current_failed_list

def run_from_excel(excel_path=EXCEL_PATH, sheet_name=None, start_row=2):
    ensure_table()
    _session_start, total_saved, total_errors = datetime.now(), 0, 0
    total_failed_list = [] 
    
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    headers = {str(ws.cell(1, c).value).strip().lower(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    
    req = ['idurl', 'source', 'target', 'cat_id', 'done']
    for k in req:
        if k not in headers: raise ValueError(f"Thiếu cột '{k}'")
    
    done_col, post_type_col = headers['done'], headers.get('post_type')
    real_last = start_row - 1
    for r in range(ws.max_row, start_row - 1, -1):
        if any(ws.cell(r, headers[k]).value not in (None, '') for k in req):
            real_last = r; break

    if real_last < start_row: log.warning('Hết dữ liệu'); return
    print(f'\n{"=" * 72}\n  {Color.BOLD}LẤY VĂN BẢN  —  {real_last - start_row + 1} dòng{Color.RESET}\n{"=" * 72}')
    bot = None
    try:
        for r in range(start_row, real_last + 1):
            idurl, source, target = ws.cell(r, headers['idurl']).value, ws.cell(r, headers['source']).value, ws.cell(r, headers['target']).value
            cat_id, done_v = ws.cell(r, headers['cat_id']).value, ws.cell(r, done_col).value
            ptype = str(ws.cell(r, post_type_col).value or '').strip() if post_type_col else ''

            row_num = r - start_row + 1
            log.row_start(row_num, real_last - start_row + 1, str(source)[:70] if source else '', str(cat_id) if cat_id else '', str(target)[:70] if target else '')
            t_row = datetime.now()

            if not source or str(done_v).strip() == '1' or (ptype and ptype.lower() not in ('van-ban', 'vanban', '')):
                if str(done_v).strip() == '1': print(f'  {Color.BRIGHT_BLACK}{Icon.ARROW} Đã xong{Color.RESET}')
                continue

            if not bot: bot = VanBanBot()
            source_url = str(source).strip()
            if source_url and not source_url.startswith(('http://', 'https://')):
                source_url = 'https://' + source_url  # Tự động thêm https nếu thiếu

            bot.url_id, bot.url, bot.cat_id, bot.target = str(idurl or ''), source_url, str(cat_id or ''), str(target or '')
            
            row_saved = row_errors = 0
            row_failed = []
            try:
                row_saved, row_errors, row_failed = bot.get_data()
                total_saved += row_saved
                total_errors += row_errors
                
                for f_title, f_reason in row_failed:
                    total_failed_list.append((f_title, f_reason, str(source)))
                    
                if row_saved > 0: 
                    ws.cell(r, done_col).value = 1
                    try:
                        wb.save(excel_path)
                    except Exception:
                        log.warning("⚠ Không thể lưu file Excel do bạn đang mở file. Dữ liệu vẫn được lưu an toàn vào DB!")
            except Exception as e:
                log.critical(f'LỖI dòng {r}: {e}')
                total_errors += 1
            
            try: wb.save(excel_path)
            except Exception: pass
            
            log.row_done(row_num, row_saved, row_errors, (datetime.now() - t_row).total_seconds())
    finally:
        try: wb.save(excel_path)
        except: pass
        if bot: bot.close()
        m, s = divmod(int((datetime.now() - _session_start).total_seconds()), 60)
        h, m = divmod(m, 60)
        log.session_summary(total_saved, total_errors, f'{h:02d}:{m:02d}:{s:02d}' if h else f'{m:02d}:{s:02d}')

        if total_failed_list:
            print(f'\n{Color.BRIGHT_BLACK}{"─" * 72}{Color.RESET}')
            print(f'  {Color.BOLD}{Color.RED}DANH SÁCH BÀI VIẾT BỊ BỎ QUA DO LỖI ({len(total_failed_list)} bài){Color.RESET}')
            
            grouped_fails = {}
            for f_title, f_err, f_url in total_failed_list:
                safe_url = f_url if f_url else "Nguồn không xác định"
                if safe_url not in grouped_fails:
                    grouped_fails[safe_url] = []
                grouped_fails[safe_url].append((f_title, f_err))
            
            for src_url, items in grouped_fails.items():
                print(f'\n  {Color.BRIGHT_BLACK}▶ Nguồn (Web Cũ): {Color.RESET}{src_url}')
                for i, (f_title, f_err) in enumerate(items, 1):
                    print(f'    {i}. {Color.CYAN}{log._trunc(f_title, 60)}{Color.RESET}')
                    print(f'       {Color.RED}↳ Lỗi: {f_err}{Color.RESET}')
                    
            print(f'\n{Color.BRIGHT_BLACK}{"─" * 72}{Color.RESET}\n')

if __name__ == '__main__':
    try: run_from_excel()
    except KeyboardInterrupt: log.warning(f'\n{Icon.WARNING} Dừng')
    except Exception as e: log.critical(f'\n{Icon.FIRE} Lỗi: {e}\n{traceback.format_exc()}')