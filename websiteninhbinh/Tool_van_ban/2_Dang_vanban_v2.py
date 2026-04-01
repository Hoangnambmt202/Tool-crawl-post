# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  2_DANG_VANBAN.PY  —  Đăng văn bản lên WordPress qua Giao Diện (BẢN CHUẨN)  ║
║                                                                              ║
║  * TỔNG HỢP TOÀN BỘ LOGIC: 100% Chuẩn hóa (Cắt khoảng trắng, Xóa Query URL)║
║  * TỐI ƯU CỬA SỔ: Chạy minimize_window() tuân thủ yêu cầu chuẩn mực.       ║
║  * CHỐNG OVERFLOW: Cắt ngắn tên file dưới 100 ký tự chống lỗi Windows 255. ║
║  * FIX GRID MODE: Ép WordPress hiển thị List Mode ngay sau khi Upload.     ║
║  * HOT FIX: Truyền Cookie của Selenium vào Requests để vượt tường lửa 403. ║
║  * ĐỒNG BỘ (V1.3): Nhất quán dữ liệu Lĩnh Vực (pham_vi) từ File 1.         ║
║  * FIX UPDATE (V2.0): Sửa lỗi "Định dạng không hợp lệ" và Logic Cập nhật.  ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import os
import re
import sys
import json
import time
import random
import traceback
from datetime import datetime
from urllib.parse import urlparse, parse_qs, unquote, quote as _quote

import requests
import urllib3
import mysql.connector
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

from user_agents import USER_AGENTS

from config_vanban import (
    CHROMEDRIVER_PATH, CHROME_BINARY, SHOW_CHROME_WINDOW,
    USE_PROFILE, PROFILE_DIR, WP_EMAIL as EMAIL, WP_PASSWORD as PASSWORD,
    DUPLICATE_MODE, XOA_FILE_SAU_KHI_DANG, TMP_DIR, LOG_TXT, EXCEL_PATH
)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class Style:
    RESET    = '\033[0m'
    BOLD     = '\033[1m'
    RED      = '\033[91m'
    GREEN    = '\033[92m'
    YELLOW   = '\033[93m'
    CYAN     = '\033[96m'
    BBLACK   = '\033[90m'
    B_RED    = '\033[1;91m'
    B_GREEN  = '\033[1;92m'

def _enable_ansi_windows():
    if sys.platform != 'win32':
        return
    try:
        import ctypes, ctypes.wintypes
        kernel32 = ctypes.windll.kernel32
        handle   = kernel32.GetStdHandle(-11)
        old_mode = ctypes.wintypes.DWORD(0)
        kernel32.GetConsoleMode(handle, ctypes.byref(old_mode))
        kernel32.SetConsoleMode(handle, old_mode.value | 0x0004)
    except Exception:
        pass

_enable_ansi_windows()

_ANSI_RE = re.compile(r'\033\[[0-9;]*[mABCDEFGHJKSTsuhl]')

def _sa(s) -> str:
    return _ANSI_RE.sub('', str(s or ''))

def _tr(s, n: int) -> str:
    s = str(s or '').strip()
    plain = _sa(s)
    return (plain[:n - 1] + '…') if len(plain) > n else plain

def _now() -> str:
    return datetime.now().strftime('%H:%M:%S')

def _fmt_t(secs: float) -> str:
    if secs < 60:
        return f'{secs:.1f}s'
    m, s = divmod(int(secs), 60)
    return f'{m}m{s:02d}s'

def log_row(n, total, title, wp_url):
    sep = f'{Style.BBLACK}{"─" * 72}{Style.RESET}'
    print(sep)
    print(f'  {Style.BOLD}BÀI {n}/{total}{Style.RESET}  {_tr(title, 55)}')
    print(f'  {Style.BBLACK}dst: {_tr(wp_url, 70)}{Style.RESET}')

def log_ok(wp_link, elapsed):
    print(f'  {Style.B_GREEN}✓ Đăng thành công  {_fmt_t(elapsed)}{Style.RESET}')
    if wp_link:
        print(f'  {Style.BBLACK}{_tr(wp_link, 70)}{Style.RESET}')

def log_skip(reason=''):
    print(f'  {Style.YELLOW}→ Bỏ qua{Style.RESET}' + (f'  {Style.BBLACK}({reason}){Style.RESET}' if reason else ''))

def log_err(msg):
    print(f'  {Style.RED}✗ {msg}{Style.RESET}')

def log_warn(msg):
    print(f'  {Style.YELLOW}⚠ {msg}{Style.RESET}')

def _write_log_txt(line: str):
    if not LOG_TXT: return
    try:
        os.makedirs(os.path.dirname(LOG_TXT), exist_ok=True)
        with open(LOG_TXT, 'a', encoding='utf-8') as f:
            f.write(line + '\n')
    except Exception:
        pass

# ══════════════════════════════════════════════════════════════════════════════
# DB HELPERS (An Toàn)
# ══════════════════════════════════════════════════════════════════════════════

_DB = dict(host='localhost', user='root', password='', database='baivietphothong')

def _conn():
    return mysql.connector.connect(**_DB)

def read_pending():
    c = cur = None
    try:
        c = _conn()
        cur = c.cursor(dictionary=True)
        cur.execute('SELECT * FROM bot_vanban WHERE upload=0 ORDER BY id')
        return cur.fetchall()
    finally:
        if cur: cur.close()
        if c: c.close()

def mark_uploaded(record_id):
    c = cur = None
    try:
        c = _conn()
        cur = c.cursor()
        cur.execute('UPDATE bot_vanban SET upload=1 WHERE id=%s', (record_id,))
        c.commit()
    finally:
        if cur: cur.close()
        if c: c.close()

# ══════════════════════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ══════════════════════════════════════════════════════════════════════════════

_target_cache: dict = {}

def get_wp_target(url_id: str) -> str:
    if url_id in _target_cache: return _target_cache[url_id]
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        headers = {str(ws.cell(1, c).value).strip().lower(): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
        id_col, tgt_col = headers.get('idurl'), headers.get('target')
        if id_col and tgt_col:
            for r in range(2, ws.max_row + 1):
                if str(ws.cell(r, id_col).value or '').strip() == str(url_id).strip():
                    val = str(ws.cell(r, tgt_col).value or '').strip()
                    _target_cache[url_id] = val
                    return val
    except Exception: pass
    _target_cache[url_id] = ''
    return ''

def _get_base_url(url: str) -> str:
    url = url.strip()
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    p = urlparse(url)
    return f'{p.scheme}://{p.netloc}'

# 💡 FIX QUAN TRỌNG 1: Đổi mặc định từ 'vanban' thành 'van-ban' để khớp với định dạng WP
def _get_post_type_from_url(wp_url: str, default='van-ban') -> str:
    try:
        qs = parse_qs(urlparse(wp_url).query)
        if 'post_type' in qs: return str(qs['post_type'][0]).strip()
    except Exception: pass
    return default

# ══════════════════════════════════════════════════════════════════════════════
# HÀM CHUẨN HÓA DỮ LIỆU ĐỂ SO SÁNH 100%
# ══════════════════════════════════════════════════════════════════════════════

def normalize_text(text, is_so_hieu=False):
    t = str(text or '').strip().lower()
    if is_so_hieu:
        return t.replace(" ", "") 
    return re.sub(r'\s+', ' ', t)

def normalize_date(date_str):
    return re.sub(r'\b0+(\d)', r'\1', str(date_str or '').strip())

# ══════════════════════════════════════════════════════════════════════════════
# 5. WORDPRESS UI BOT (SELENIUM)
# ══════════════════════════════════════════════════════════════════════════════

class WordPressUIBot:
    def __init__(self):
        opts = Options()
        if not SHOW_CHROME_WINDOW: 
            opts.add_argument('--headless=new')
        if CHROME_BINARY: 
            opts.binary_location = CHROME_BINARY
        if USE_PROFILE and PROFILE_DIR: 
            opts.add_argument(f'--user-data-dir={PROFILE_DIR}')
            
        opts.add_argument('--disable-notifications')
        opts.add_argument('--window-size=1366,768')
        opts.add_argument('--disable-blink-features=AutomationControlled')
        opts.add_experimental_option('excludeSwitches', ['enable-automation'])
        
        self.current_ua = random.choice(USER_AGENTS)
        opts.add_argument(f'--user-agent={self.current_ua}')
        
        self.driver = webdriver.Chrome(service=Service(CHROMEDRIVER_PATH), options=opts)
        
        if SHOW_CHROME_WINDOW:
            self.driver.minimize_window()
            
        self.driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        })
        
        self.logged_in_domains = set()

    def close(self):
        try: self.driver.quit()
        except Exception: pass

    def login_if_needed(self, base_url: str):
        self.driver.get(f"{base_url}/wp-admin/")
        if "wp-login.php" not in self.driver.current_url and self.driver.find_elements(By.ID, "wpadminbar"):
            self.logged_in_domains.add(base_url)
            return True

        print(f"  {Style.BBLACK}Đang mở trang đăng nhập/Đăng nhập lại...{Style.RESET}")
        self.driver.get(f"{base_url}/wp-login.php")
        try:
            WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "user_login"))).send_keys(EMAIL)
            self.driver.find_element(By.ID, "user_pass").send_keys(PASSWORD)
            self.driver.find_element(By.ID, "wp-submit").click()
            WebDriverWait(self.driver, 15).until(lambda d: "wp-admin" in d.current_url or d.find_elements(By.ID, "wpadminbar"))
            self.logged_in_domains.add(base_url)
            print(f"  {Style.CYAN}✓ Đã đăng nhập vào Admin{Style.RESET}")
            return True
        except Exception as e:
            log_err(f"Lỗi đăng nhập: {e}")
            return False

    # 💡 FIX QUAN TRỌNG 2: Thay đổi giá trị trả về thành Tuple[Trạng Thái, Link Chỉnh Sửa]
    def check_duplicate(self, base_url: str, post_type: str, record: dict) -> tuple[str, str]:
        """
        Kiểm tra trùng lặp 100% tất cả các dữ liệu có trên Form Web.
        Trả về tuple: ("SKIP"|"UPDATE"|"NEW", edit_url_neu_co)
        """
        if DUPLICATE_MODE == 0: 
            return "NEW", ""
        
        raw_title = str(record.get('title') or '').strip()
        r_title = normalize_text(raw_title)
        
        r_sh  = normalize_text(record.get('so_hieu'), is_so_hieu=True)
        r_cq  = normalize_text(record.get('co_quan'))
        r_pv  = normalize_text(record.get('pham_vi'))
        r_nb  = normalize_date(record.get('ngay_ban_hanh'))
        r_nhl = normalize_date(record.get('ngay_hieu_luc'))
        r_nk  = normalize_text(record.get('nguoi_ky'))
        r_mt  = normalize_text(record.get('mo_ta') or raw_title)
        r_lvb = normalize_text(record.get('loai_van_ban'))

        search_term = raw_title[:40].rsplit(' ', 1)[0] if len(raw_title) > 40 else raw_title
        self.driver.get(f"{base_url}/wp-admin/edit.php?post_type={post_type}&s={_quote(search_term)}")
        
        try:
            all_edit_urls = []
            
            while True:
                if self.driver.find_elements(By.CLASS_NAME, "no-items"): 
                    break

                edit_elements = self.driver.find_elements(By.CSS_SELECTOR, f"#the-list tr.type-{post_type} a.row-title")
                for el in edit_elements:
                    href = el.get_attribute("href")
                    if href and href not in all_edit_urls:
                        all_edit_urls.append(href)

                next_btns = self.driver.find_elements(By.CSS_SELECTOR, ".tablenav.bottom .next-page:not(.disabled)")
                if next_btns:
                    next_url = next_btns[0].get_attribute("href")
                    self.driver.get(next_url)
                else:
                    break 

            if not all_edit_urls: 
                return "NEW", ""

            for edit_href in all_edit_urls:
                try:
                    self.driver.get(edit_href)
                    WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "title")))
                    time.sleep(1.5) 

                    js_get_data = """
                    let tt = document.getElementById("title");
                    let sh = document.querySelector('input[name="wpcf[vb-so-ky-hieu]"]');
                    let nb = document.querySelector('input[name="wpcf[vb-ngay-ban-hanh][display-only]"]');
                    let nhl = document.querySelector('input[name="wpcf[vb-ngay-hieu-luc][display-only]"]');
                    let nk = document.querySelector('input[name="wpcf[vb-nguoi-ky]"]');
                    let mt = document.querySelector('textarea[name="wpcf[vb-trich-yeu]"]');
                    
                    let getTax = (taxId) => {
                        let els = document.querySelectorAll('#' + taxId + 'checklist input:checked');
                        return Array.from(els).map(e => e.parentElement.innerText.trim()).join(' ');
                    };
                    let cq = getTax('co-quan-ban-hanh');
                    let lvb = getTax('loai-van-ban');
                    let pv = getTax('linh-vuc');
                    
                    let pdfs = [];
                    let mainFile = document.querySelector('input[name="wpcf[vb-file-dinh-kem]"]');
                    if (mainFile && mainFile.value) pdfs.push(mainFile.value.trim());
                    
                    let extraFiles = document.querySelectorAll('#vb-files-list .vb-file-item a');
                    extraFiles.forEach(a => {
                        if (a.href && !pdfs.includes(a.href)) pdfs.push(a.href.trim());
                    });
                    
                    return {
                        title: tt ? tt.value.trim() : "",
                        so_hieu: sh ? sh.value.trim() : "",
                        ngay_ban_hanh: nb ? nb.value.trim() : "",
                        ngay_hieu_luc: nhl ? nhl.value.trim() : "",
                        nguoi_ky: nk ? nk.value.trim() : "",
                        mo_ta: mt ? mt.value.trim() : "",
                        co_quan: cq,
                        loai_van_ban: lvb,
                        pham_vi: pv,
                        pdfs: pdfs
                    };
                    """
                    w_data = self.driver.execute_script(js_get_data)

                    w_title = normalize_text(w_data.get('title'))
                    if w_title != r_title: 
                        continue 
                    
                    w_sh  = normalize_text(w_data.get('so_hieu'), is_so_hieu=True)
                    w_nb  = normalize_date(w_data.get('ngay_ban_hanh'))
                    w_nhl = normalize_date(w_data.get('ngay_hieu_luc'))
                    w_nk  = normalize_text(w_data.get('nguoi_ky'))
                    w_mt  = normalize_text(w_data.get('mo_ta') or w_data.get('title'))
                    w_cq  = normalize_text(w_data.get('co_quan'))
                    w_lvb = normalize_text(w_data.get('loai_van_ban'))
                    w_pv  = normalize_text(w_data.get('pham_vi'))
                    
                    w_pdfs = w_data.get('pdfs', [])

                    def get_fname(u): return unquote(str(u).split('/')[-1].split('?')[0]).lower()
                    
                    w_filenames = set(get_fname(u) for u in w_pdfs if u)
                    
                    try: r_pdfs_list = json.loads(record.get('pdf_urls') or '[]')
                    except Exception: r_pdfs_list = []
                    
                    r_filenames = set(get_fname(p.get('url','')) for p in r_pdfs_list if p.get('url'))

                    if (w_sh == r_sh and w_nb == r_nb and w_nhl == r_nhl and 
                        w_nk == r_nk and w_mt == r_mt and 
                        (not r_cq or r_cq in w_cq) and 
                        (not r_lvb or r_lvb in w_lvb) and 
                        (not r_pv or r_pv in w_pv)): 
                        
                        if not r_filenames or r_filenames.issubset(w_filenames):
                            print(f"  {Style.YELLOW}→ Bỏ qua (Đã tồn tại 100% trên Web){Style.RESET}")
                            return "SKIP", ""
                        else:
                            print(f"  {Style.CYAN}✦ Phát hiện bài viết trùng nhưng cập nhật FILE MỚI!{Style.RESET}")
                            # Trả về cờ UPDATE kèm theo Link bài viết cần edit
                            return "UPDATE", edit_href 
                            
                except Exception:
                    continue 

            print(f"  {Style.CYAN}✦ Phát hiện bài viết trùng Tên nhưng khác Nội dung (Đăng mới){Style.RESET}")
            return "NEW", "" 

        except Exception: 
            return "NEW", ""

    def download_and_upload_pdf(self, base_url: str, pdf_url: str, pdf_name: str) -> tuple[str, str]:
        safe_name = re.sub(r'[\\/*?:"<>|]', "", pdf_name).replace('\n', '').replace('\r', '').strip()
        
        name_part, ext = os.path.splitext(safe_name)
        if not ext.lower() in ['.pdf', '.doc', '.docx', '.zip', '.rar', '.xls', '.xlsx']: 
            ext = '.pdf'
            
        safe_name = (name_part[:100] + ext) if len(name_part) > 100 else (name_part + ext)
            
        if len(safe_name) < 5 or safe_name.startswith('.'):
            safe_name = f"tailieu_{int(time.time())}.pdf"
            
        print(f'  {Style.BBLACK}🔍 Kiểm tra file Media ({safe_name})...{Style.RESET}')
        try:
            self.driver.get(f"{base_url}/wp-admin/upload.php?mode=list&s={_quote(safe_name)}")
            if not self.driver.find_elements(By.CLASS_NAME, "no-items"):
                items = self.driver.find_elements(By.CSS_SELECTOR, ".wp-list-table.media tbody tr .row-actions .edit a")
                if items:
                    edit_href = items[0].get_attribute("href")
                    self.driver.get(edit_href)
                    val = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "attachment_url"))).get_attribute("value")
                    media_id = parse_qs(urlparse(edit_href).query).get('post', [''])[0]
                    print(f'  {Style.CYAN}✓ Đã có sẵn file trên server: {safe_name}{Style.RESET}')
                    return val, media_id
        except Exception: pass

        print(f'  {Style.BBLACK}⬆ Tải file mới lên server...{Style.RESET}')
        try:
            session = requests.Session()
            for cookie in self.driver.get_cookies():
                session.cookies.set(cookie['name'], cookie['value'])
                
            r = session.get(pdf_url, timeout=30, verify=False, headers={'User-Agent': self.current_ua})
            
            if r.status_code != 200: 
                return "", ""
                
            os.makedirs(TMP_DIR, exist_ok=True)
            tmp_path = os.path.abspath(os.path.join(TMP_DIR, safe_name))
            with open(tmp_path, 'wb') as f: 
                f.write(r.content)
            
            self.driver.get(f"{base_url}/wp-admin/media-new.php?browser-uploader")
            WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.NAME, "async-upload"))).send_keys(tmp_path)
            
            btn = self.driver.find_element(By.NAME, "html-upload")
            btn.click()
            
            try:
                WebDriverWait(self.driver, 60).until(EC.staleness_of(btn))
            except TimeoutException:
                pass
            
            self.driver.get(f"{base_url}/wp-admin/upload.php?mode=list")
            
            edit_link = WebDriverWait(self.driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".wp-list-table.media tbody tr:first-child .row-actions .edit a")))
            edit_href = edit_link.get_attribute("href")
            self.driver.get(edit_href)
            val = WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "attachment_url"))).get_attribute("value")
            media_id = parse_qs(urlparse(edit_href).query).get('post', [''])[0]
            
            if XOA_FILE_SAU_KHI_DANG:
                try: os.remove(tmp_path)
                except Exception: pass
                
            print(f'  {Style.CYAN}✓ Upload thành công!{Style.RESET}')
            return val, media_id
        except Exception: 
            return "", ""

    def set_taxonomy_term(self, tax_slug: str, term_string: str):
        if not term_string: return
        parts = [str(p).strip() for p in str(term_string).split('/') if str(p).strip()]
        parent_name = ""
        
        for part in parts:
            js_script = """
            var taxSlug = arguments[0];
            var termName = arguments[1];
            var parentName = arguments[2];
            
            function findParentValue(selectEl, pName) {
                if (!selectEl || !pName) return "";
                var opts = Array.from(selectEl.options);
                var target = opts.find(o => o.innerText.replace(/\\u00a0/g, '').replace(/\\s+/g, ' ').trim().toLowerCase() === pName.toLowerCase());
                return target ? target.value : "";
            }
            
            var list = document.getElementById(taxSlug + 'checklist');
            if (list) {
                var labels = Array.from(list.querySelectorAll('label'));
                var target = labels.find(l => l.innerText.trim().toLowerCase() === termName.toLowerCase());
                if (target) {
                    var cb = target.querySelector('input[type="checkbox"]');
                    if (cb && !cb.checked) { cb.click(); }
                    return "exists";
                }
            }
            
            var addToggle = document.getElementById(taxSlug + '-add-toggle');
            var addBlock = document.getElementById(taxSlug + '-add');
            if (addToggle && addBlock && addBlock.classList.contains('wp-hidden-child')) {
                addToggle.click();
            }
            
            var input = document.getElementById('new' + taxSlug);
            var submit = document.getElementById(taxSlug + '-add-submit');
            var select = document.getElementById('new' + taxSlug + '_parent');
            
            if (input && submit) {
                input.value = termName;
                if (parentName && select) {
                    var pVal = findParentValue(select, parentName);
                    if (pVal) select.value = pVal;
                }
                submit.click(); 
                return "added";
            }
            return "error";
            """
            try:
                res = self.driver.execute_script(js_script, tax_slug, part, parent_name)
                if res == "added":
                    time.sleep(3) 
                    self.driver.execute_script(f"""
                        var list = document.getElementById('{tax_slug}checklist');
                        if(list){{
                            var labels = Array.from(list.querySelectorAll('label'));
                            var target = labels.find(l => l.innerText.trim().toLowerCase() === '{part}'.toLowerCase());
                            if(target){{
                                var cb = target.querySelector('input[type="checkbox"]');
                                if(cb && !cb.checked) cb.click();
                            }}
                        }}
                    """)
            except Exception:
                pass
            
            parent_name = part

    # 💡 FIX QUAN TRỌNG 3: Bổ sung logic Mở lại bài cũ để Cập Nhật thay vì luôn tạo Bài Mới
    def create_post(self, base_url: str, post_type: str, record: dict, final_pdfs: list, edit_url: str = "") -> tuple[str, str]:
        raw_title = str(record.get('title') or '').strip()
        
        # Nếu có edit_url truyền vào -> Mở bài viết cũ
        if edit_url:
            print(f'  {Style.BBLACK}Đang mở bài viết cũ để CẬP NHẬT thông tin/file...{Style.RESET}')
            self.driver.get(edit_url)
            try:
                WebDriverWait(self.driver, 15).until(EC.presence_of_element_located((By.ID, "title")))
            except TimeoutException:
                pass
        # Nếu không có -> Tạo bài viết mới
        else:
            print(f'  {Style.BBLACK}Đang tạo bài viết MỚI...{Style.RESET}')
            self.driver.get(f"{base_url}/wp-admin/post-new.php?post_type={post_type}")
            try:
                WebDriverWait(self.driver, 15).until(EC.presence_of_element_located((By.ID, "title-prompt-text")))
            except TimeoutException:
                pass
        
        try:
            time.sleep(2) 
            
            cat_id_excel = str(record.get("cat_id") or "").strip()
            loai_vb_web  = str(record.get("loai_van_ban") or "").strip()
            co_quan_web  = str(record.get("co_quan") or "").strip()
            linh_vuc_web = str(record.get("pham_vi") or "").strip()
            
            print(f'  {Style.BBLACK}Đang xử lý Danh mục (Taxonomy)...{Style.RESET}')
            if cat_id_excel: self.set_taxonomy_term('loai-van-ban', cat_id_excel)
            if loai_vb_web: self.set_taxonomy_term('loai-van-ban', loai_vb_web)
            if co_quan_web: self.set_taxonomy_term('co-quan-ban-hanh', co_quan_web)
            if linh_vuc_web: self.set_taxonomy_term('linh-vuc', linh_vuc_web)
            time.sleep(2)
            
            # Ghi đè Title
            title_input = self.driver.find_element(By.NAME, "post_title")
            title_input.clear()
            title_input.send_keys(raw_title)
            
            content = str(record.get('content') or '')
            self.driver.execute_script("""
                if(typeof tinyMCE!=='undefined'&&tinyMCE.activeEditor){tinyMCE.activeEditor.setContent(arguments[0]);}
                else if(document.getElementById('content')){document.getElementById('content').value=arguments[0];}
            """, content)
            
            try:
                WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[name*="wpcf"]')))
            except TimeoutException:
                pass

            meta_data = {
                "so_hieu": str(record.get("so_hieu") or "").strip(),
                "ngay_ban_hanh": str(record.get("ngay_ban_hanh") or "").strip(),
                "ngay_hieu_luc": str(record.get("ngay_hieu_luc") or "").strip(),
                "nguoi_ky": str(record.get("nguoi_ky") or "").strip(),
                "mo_ta": str(record.get("mo_ta") or "").strip() or raw_title,
                "final_pdfs": final_pdfs, 
            }
            
            print(f'  {Style.CYAN}├─ Danh mục (Excel): {Style.RESET}{cat_id_excel}')
            print(f'  {Style.CYAN}├─ Cơ quan ban hành: {Style.RESET}{co_quan_web}')
            print(f'  {Style.CYAN}├─ Loại văn bản: {Style.RESET}{loai_vb_web}')
            print(f'  {Style.CYAN}├─ Lĩnh vực   : {Style.RESET}{linh_vuc_web}')
            print(f'  {Style.CYAN}├─ Số ký hiệu : {Style.RESET}{meta_data["so_hieu"]}')
            print(f'  {Style.CYAN}├─ Ngày BH    : {Style.RESET}{meta_data["ngay_ban_hanh"]}')
            print(f'  {Style.CYAN}├─ Ngày HL    : {Style.RESET}{meta_data["ngay_hieu_luc"]}')
            print(f'  {Style.CYAN}├─ Người ký   : {Style.RESET}{meta_data["nguoi_ky"]}')
            print(f'  {Style.CYAN}├─ Trích yếu  : {Style.RESET}{_tr(meta_data["mo_ta"], 50)}')
            
            if final_pdfs:
                for idx, pdf in enumerate(final_pdfs):
                    print(f'  {Style.CYAN}├─ File đính kèm {idx+1}: {Style.RESET}{_tr(pdf["url"], 50)} (Media ID: {pdf["id"]})')
            else:
                print(f'  {Style.CYAN}├─ File đính kèm  : {Style.RESET}Không có file')

            print(f'  {Style.BBLACK}Đang điền dữ liệu Form và Lưu...{Style.RESET}')
            
            js_fill_basic_fields = """
            const data = arguments[0];
            
            function triggerEvents(el) {
                if(!el) return;
                el.dispatchEvent(new Event('input', {bubbles: true}));
                el.dispatchEvent(new Event('change', {bubbles: true}));
            }

            let fSoKyHieu = document.querySelector('input[name="wpcf[vb-so-ky-hieu]"]');
            if (fSoKyHieu) { fSoKyHieu.value = data.so_hieu; triggerEvents(fSoKyHieu); }

            function fillToolsetDate(fieldName, dateStr) {
                if (!dateStr) return;
                let disp = document.querySelector('input[name="wpcf[' + fieldName + '][display-only]"]');
                let hid = document.querySelector('input[name="wpcf[' + fieldName + '][datepicker]"]');

                if (disp) {
                    disp.removeAttribute('readonly');
                    disp.value = dateStr;
                    triggerEvents(disp);
                    if (typeof jQuery !== 'undefined') jQuery(disp).trigger('change');
                }
                
                if (hid) {
                    let parts = dateStr.split(/[\\/\\-]/);
                    if (parts.length === 3) {
                        let d = parseInt(parts[0], 10);
                        let m = parseInt(parts[1], 10) - 1;
                        let y = parseInt(parts[2], 10);
                        let dateObj = new Date(y, m, d);
                        hid.value = Math.floor(dateObj.getTime() / 1000).toString();
                        triggerEvents(hid);
                    }
                }
            }
            fillToolsetDate('vb-ngay-ban-hanh', data.ngay_ban_hanh);
            fillToolsetDate('vb-ngay-hieu-luc', data.ngay_hieu_luc);

            let fNguoiKy = document.querySelector('input[name="wpcf[vb-nguoi-ky]"]');
            if (fNguoiKy) { fNguoiKy.value = data.nguoi_ky; triggerEvents(fNguoiKy); }

            let fTrichYeu = document.querySelector('textarea[name="wpcf[vb-trich-yeu]"]');
            if (fTrichYeu) { fTrichYeu.value = data.mo_ta; triggerEvents(fTrichYeu); }

            if (data.final_pdfs && data.final_pdfs.length > 0) {
                let fFile = document.querySelector('input[name="wpcf[vb-file-dinh-kem]"]');
                if (fFile) { fFile.value = data.final_pdfs[0].url; triggerEvents(fFile); }

                let filesList = document.getElementById('vb-files-list');
                if (filesList) {
                    let noFiles = filesList.querySelector('.vb-no-files');
                    if (noFiles) noFiles.remove();
                    
                    data.final_pdfs.forEach(pdf => {
                        if (pdf.id && pdf.url) {
                            let parts = pdf.url.split('/');
                            let fileName = decodeURIComponent(parts[parts.length - 1]);

                            if (!filesList.querySelector('input[value="' + pdf.id + '"]')) {
                                let html = '<div class="vb-file-item" data-index="' + pdf.id + '">' +
                                    '<span class="dashicons dashicons-media-document"></span>' +
                                    '<a href="' + pdf.url + '" target="_blank">' + fileName + '</a>' +
                                    '<button type="button" class="button vb-remove-file" data-id="' + pdf.id + '">' +
                                    '<span class="dashicons dashicons-no"></span> Xóa' +
                                    '</button>' +
                                    '<input type="hidden" name="vb_files[]" value="' + pdf.id + '" />' +
                                    '</div>';
                                filesList.insertAdjacentHTML('beforeend', html);
                            }
                        }
                    });
                }
            }

            let rNoiBo = document.querySelector('input[name="wpcf[vb-van-ban-noi-bo]"][value="0"]');
            if (rNoiBo && !rNoiBo.checked) { 
                rNoiBo.click(); 
                triggerEvents(rNoiBo);
            } else if (!rNoiBo) {
                let labels = Array.from(document.querySelectorAll('label'));
                let targetLabel = labels.find(l => l.innerText.trim() === 'Không');
                if (targetLabel) {
                    let rbId = targetLabel.getAttribute('for');
                    let rb = rbId ? document.getElementById(rbId) : targetLabel.querySelector('input[type="radio"]');
                    if (rb && !rb.checked) { rb.click(); triggerEvents(rb); }
                }
            }
            """
            self.driver.execute_script(js_fill_basic_fields, meta_data)
            time.sleep(1.5)
            
            try:
                # Nút Đăng mới (Publish) hay Cập nhật (Update) trong WordPress đều dùng ID="publish"
                publish_btn = WebDriverWait(self.driver, 5).until(EC.presence_of_element_located((By.ID, "publish")))
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", publish_btn)
                time.sleep(1); publish_btn.click()
            except Exception: pass
            
            try:
                # 💡 FIX QUAN TRỌNG 4: Bắt thêm text 'cập nhật' hoặc 'updated' cho trường hợp Cập nhật bài cũ
                WebDriverWait(self.driver, 15).until(
                    EC.presence_of_element_located((By.XPATH, "//*[@id='message']/p[contains(text(), 'published') or contains(text(), 'đã được đăng') or contains(translate(text(), 'C', 'c'), 'cập nhật') or contains(text(), 'updated')] | //*[@id='sample-permalink']/a"))
                )
                time.sleep(1)
            except TimeoutException:
                time.sleep(3)
            
            js_verify_status = """
            let errors = [];
            let errorNodes = document.querySelectorAll('.wpt-form-error, .acf-error-message, .form-invalid, .notice-error, .error');
            errorNodes.forEach(node => {
                let container = node.closest('.wpt-field, .acf-field, tr, .form-item, div.inside');
                if (container) {
                    let label = container.querySelector('label');
                    if (label && label.innerText) {
                        let txt = label.innerText.replace('*', '').trim();
                        if (!errors.includes(txt)) errors.push(txt);
                    }
                }
            });
            
            let statusEl = document.getElementById('post-status-display');
            let statusTxt = statusEl ? statusEl.innerText.trim().toLowerCase() : '';
            
            let linkEl = document.querySelector('#sample-permalink a') || document.querySelector('#message a');
            let link = linkEl ? linkEl.href : '';
            
            return {
                'url': window.location.href,
                'status': statusTxt,
                'link': link,
                'errors': errors.join(', ')
            };
            """
            result = self.driver.execute_script(js_verify_status)
            
            current_url = result.get('url', '')
            status      = result.get('status', '')
            post_link   = result.get('link', '')
            error_str   = result.get('errors', '')
            
            if "post-new.php" in current_url and not edit_url:
                msg = f"Lỗi kẹt tại ô: [{error_str}]" if error_str else "Kẹt không rõ nguyên nhân (chưa điền đủ Required Form)"
                return "", msg
                
            if "nháp" in status or "draft" in status:
                msg = f"Web ép lưu Bản nháp vì thiếu ô: [{error_str}]" if error_str else "Bài bị ép thành Bản nháp (Có thể lỗi plugin chặn)"
                return "", msg
                
            return post_link, ""
        except Exception as e:
            return "", f"Lỗi trình duyệt ngắt kết nối: {e}"

# ══════════════════════════════════════════════════════════════════════════════
# 6. MAIN CONTROLLER
# ══════════════════════════════════════════════════════════════════════════════

class MainController:
    def run(self):
        records = read_pending()
        if not records:
            print(f'\n  {Style.BBLACK}Không có văn bản nào cần đăng.{Style.RESET}')
            return

        total   = len(records)
        ok = skip = fail = 0
        t_start = datetime.now()
        
        failed_list = []

        print(f'\n{"=" * 72}\n  {Style.BOLD}ĐĂNG VĂN BẢN QUA GIAO DIỆN CHROME  —  {total} bài{Style.RESET}\n{"=" * 72}')

        bot = WordPressUIBot()
        try:
            for n, record in enumerate(records, 1):
                rec_id, url_id = record['id'], str(record.get('url_id') or '')
                title = str(record.get('title') or '').strip()
                
                wp_url = get_wp_target(url_id)
                if not wp_url: 
                    fail += 1
                    failed_list.append((title, f"Không tìm thấy cấu hình web đích (target) cho url_id={url_id}", ""))
                    continue

                log_row(n, total, title, wp_url)
                t_row, base_url, post_type = datetime.now(), _get_base_url(wp_url), _get_post_type_from_url(wp_url)

                if not bot.login_if_needed(base_url): 
                    fail += 1
                    failed_list.append((title, "Lỗi đăng nhập vào Admin WordPress", base_url))
                    continue
                
                # 💡 FIX QUAN TRỌNG 5: Nhận lại trạng thái (SKIP/UPDATE/NEW) và URL Edit nếu có
                dup_status, edit_url = bot.check_duplicate(base_url, post_type, record)
                if dup_status == "SKIP": 
                    mark_uploaded(rec_id)
                    skip += 1
                    continue

                max_retries = 2
                success = False
                last_error_msg = ""
                
                for attempt in range(1, max_retries + 1):
                    if attempt > 1:
                        print(f"  {Style.YELLOW}↻ Đang thử lại lần {attempt} cho bài này...{Style.RESET}")

                    try: pdf_list = json.loads(record.get('pdf_urls') or '[]')
                    except Exception: pdf_list = []
                    
                    final_pdfs = []
                    file_upload_error = False
                    
                    for pdf_item in pdf_list:
                        original_url = pdf_item.get('url')
                        if not original_url: continue
                        
                        filename_from_url = unquote(os.path.basename(urlparse(original_url).path))
                        pdf_name = filename_from_url if filename_from_url.lower().endswith(('.pdf', '.doc', '.docx', '.zip', '.rar', '.xls', '.xlsx')) else pdf_item.get('name', 'vanban.pdf')
                        
                        res_url, res_id = bot.download_and_upload_pdf(base_url, original_url, pdf_name)
                        if res_url and res_id:
                            final_pdfs.append({
                                'url': res_url,
                                'id': res_id
                            })
                        else:
                            print(f"  {Style.YELLOW}⚠ Không tải được file {pdf_name}. Dừng xử lý ở lần thử này!{Style.RESET}")
                            file_upload_error = True
                            last_error_msg = "Lỗi tải file đính kèm"
                            break 
                    
                    if file_upload_error:
                        continue 
                    
                    # Truyền thêm biến edit_url vào hàm create_post
                    post_link, error_msg = bot.create_post(base_url, post_type, record, final_pdfs, edit_url)
                    
                    if post_link:
                        mark_uploaded(rec_id) 
                        ok += 1
                        elapsed = (datetime.now() - t_row).total_seconds()
                        log_ok(post_link, elapsed)
                        _write_log_txt(f'[{_now()}]  ✓  SUCCESS  {title}\n         {post_link}')
                        success = True
                        break 
                    else:
                        last_error_msg = error_msg
                        continue
                
                if not success:
                    fail += 1
                    failed_list.append((title, f"Đã thử {max_retries} lần vẫn lỗi: {last_error_msg}", base_url)) 
                    log_err(f'Đăng thất bại: {last_error_msg} — Bỏ qua bài này, giữ upload=0 để xem xét sau')
                    _write_log_txt(f'[{_now()}]  ✗  FAIL  {title} | Lỗi: {last_error_msg}')
                
                time.sleep(2) 
        finally:
            bot.close()
            m, s = divmod(int((datetime.now() - t_start).total_seconds()), 60)
            h, m = divmod(m, 60)
            print(f'\n{"=" * 72}\n  {Style.BOLD}KẾT QUẢ ĐĂNG BÀI{Style.RESET}\n  Thành công : {Style.B_GREEN}{ok}{Style.RESET}')
            if skip: print(f'  Bỏ qua     : {Style.YELLOW}{skip}{Style.RESET}')
            if fail: print(f'  Thất bại   : {Style.RED}{fail}{Style.RESET}')
            print(f'  Tổng       : {total}\n  Thời gian  : {f"{h:02d}:{m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"}\n{"=" * 72}')

            if failed_list:
                print(f'\n{Style.BBLACK}{"─" * 72}{Style.RESET}')
                print(f'  {Style.BOLD}{Style.RED}DANH SÁCH BÀI VIẾT LỖI ({len(failed_list)} bài){Style.RESET}')
                
                grouped_fails = {}
                for f_title, f_err, f_url in failed_list:
                    safe_url = f_url if f_url else "Trường không xác định"
                    if safe_url not in grouped_fails:
                        grouped_fails[safe_url] = []
                    grouped_fails[safe_url].append((f_title, f_err))
                
                for school_url, items in grouped_fails.items():
                    print(f'\n  {Style.BBLACK}▶ Trường: {Style.WHITE}{school_url}{Style.RESET}')
                    for i, (f_title, f_err) in enumerate(items, 1):
                        print(f'    {i}. {Style.CYAN}{_tr(f_title, 60)}{Style.RESET}')
                        print(f'       {Style.RED}↳ Lỗi: {f_err}{Style.RESET}')
                        
                print(f'\n{Style.BBLACK}{"─" * 72}{Style.RESET}\n')

if __name__ == '__main__':
    try: 
        app = MainController()
        app.run()
    except KeyboardInterrupt: 
        print(f'\n{Style.YELLOW}⚠ Dừng chương trình{Style.RESET}')
    except Exception as e: 
        print(f'\n{Style.B_RED}🔥 Lỗi: {e}{Style.RESET}\n{traceback.format_exc()}')