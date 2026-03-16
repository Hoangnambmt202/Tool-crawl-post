import requests
import Objectlink as obl
import MenuLink as menulink
from bs4 import BeautifulSoup
import CameraObject as camob
# Function to fetch and parse the webpage
import json
import mysql.connector
import time
import helpers as hp
from difflib import SequenceMatcher

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from datetime import date
from openpyxl import load_workbook
import time
import re
from urllib.parse import urlparse
FROM_DATE = date(2025, 7, 1)   # chỉ lấy bài từ 01/07/2025 trở về đây
DETAIL_DATE_CSS = "span.post-date"         # ngày bài viết
def clean_spaces(s: str) -> str:
    return " ".join((s or "").split()).strip()



def parse_vn_date_from_soup(soup: BeautifulSoup) -> date | None:
    """
    Parse ngày kiểu: "Thứ hai, 24/11/2025 | 07:36"
    """
    el = soup.find('span',class_='post-date left')
    if not el:
        el = soup.find('div',class_='PostDate')
    print('el')
    print(el)
    if not el:
        return None
    text = clean_spaces(el.get_text(" ", strip=True))
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if not m:
        return None
    d, mo, y = map(int, m.groups())
    try:
        return date(y, mo, d)
    except:
        return None

def to_int(text, default=None):
    s = (text or "").strip()
    m = re.search(r"\d+", s)          # lấy cụm số đầu tiên
    if not m:
        return default
    return int(m.group())

def get_base(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    p = urlparse(url)
    if not p.scheme or not p.netloc:
        # nếu Excel có URL thiếu scheme, ví dụ: thtienhiep.ninhbinh.edu.vn/...
        p = urlparse("https://" + url)
    return f"{p.scheme}://{p.netloc}/"

class VspProducts:
    def __init__(self, base, url, cat,target):
        self.base = base
        self.cat = cat
        self.target = target
        self.camlist = camob.Listcam()
        self.url_links = []
        self.menu_links = menulink.MenuLink()

        service = Service(r"F:\soft\chromedriver-win64\chromedriver-win64\chromedriver.exe")
        chrome_options = Options()
        chrome_options.binary_location = r"F:\soft\chrome-win64\chrome-win64\chrome.exe"
        self.driver = webdriver.Chrome(service=service, options=chrome_options)

        # các biến sẽ được set theo từng dòng excel
        self.url = url
        self.cat_id = ""
        self.url_id = ""
        self.page_link = 1
    def reset_for_row(self):
        # reset queue + menu cho mỗi dòng excel
        self.url_links = []
        self.menu_links = menulink.MenuLink()
        # nếu muốn reset danh sách cam cho mỗi dòng:
        self.camlist = camob.Listcam()

    def get_data(self):
        self.url_links.append(self.url)
        while self.url_links:
            self.extract_data(self.url_links[0])
            self.url_links.pop(0)

    def close(self):
        try:
            self.driver.quit()
        except:
            pass

    def extract_data(self, o_url ):
        time.sleep(1)
        print(o_url)
        chitiet = 0
        for cam in self.camlist.camobs:
            if cam.url == o_url:
                print('chuan bi doc thong tin san pham')
                chitiet = 1
        try:
            # soup = hp.fetch_webpage( o_url)
            self.driver.get(o_url)
            # Lấy nội dung HTML sau khi JavaScript đã tải xong
            html = self.driver.page_source
            # Phân tích nội dung với BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
        except Exception as e:
            print("Error:", e)
            return

        if soup:
            print('dang parse' +  o_url)
            
            
            # read module product
            while True:
                time.sleep(2)
                div_pros = soup.find_all('div', class_="post-item row")
                print('doc lai div_pros')
                if(len(div_pros) > 0):
                    for tag_div in div_pros:
                        
                    
                        h3 = tag_div.find('h3' )
                        print(h3)
                        if h3:
                            tag_a = h3.find('a' )
                            price = 0
                            url = ""
                            name =""
                            photo = ""
                            if tag_a:
                                name=tag_a.text.strip()
                                url =self.base + tag_a.get('href')
                        

                            check = hp.check_cam_url(self.url_id,url,name)
                            if url != '' and name != '' and   check :
                                print("nhap cam")
                                if self.camlist.add_cam(camob.CameraObject(0,name,price, url,photo,self.cat_id)):
                                    self.url_links.append(url)
                        
                tag_title_product = soup.find('h1', class_="title-detail")
                print(tag_title_product)
                if tag_title_product  :
                    print('chi tiet san pham')
                    tag_content =  soup.find('div',class_='content-detail font-size-text mb-20')
                    d = parse_vn_date_from_soup(soup)
                    print('****************date')
                    print(d)
                

                    if  d is not None :
                        print(d < FROM_DATE)
                        if d < FROM_DATE:
                            print(f"Bài {d} < {FROM_DATE} => bỏ qua")
                            return

                    photo = ""
                    short = ""
                    if tag_content:
                        for cam in self.camlist.camobs:
                            if cam.url == o_url:
                                cam.date_publish = d
                                cam.short = self.target
                                for tag in tag_content.find_all(True):
                                        if tag.has_attr('style'):
                                            del tag['style']

                                for img_tag in tag_content.find_all('img'):
                                    if 'src' in img_tag.attrs:  # Kiểm tra nếu có thuộc tính 'src'
                                        src_value = img_tag['src']  # Lấy giá trị thuộc tính 'src'
                                        img_tag.attrs = {'src': src_value}  # Chỉ giữ lại thuộc tính 'src'
                                    else:
                                        img_tag.attrs = {}  # Xóa tất cả thuộc tính nếu không có 'src'
                                for tag in tag_content.find_all(True):
                                        # Nếu thẻ là <img>, chỉ giữ lại thuộc tính 'src'
                                        if tag.name == 'img':
                                            src = tag.get('src')  
                                            tag.attrs = {}
                                            if src:
                                                tag['src'] = src
                                            print('tag img')
                                            print(tag)
                                        else:
                                            # Với các thẻ khác, xóa toàn bộ attribute
                                            tag.attrs = {}       
                                tags_to_removes = tag_content.find_all ('script' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                tags_to_removes = tag_content.find_all ('link' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                cam.description = str(tag_content)
                                cam.display_info()
                                hp.save_data_cam(self.url_id,cam)     #dahua 3 #hk 4 #tiandy 5 #ezviz 6 #vsp là 7  #imou 8  #kb 9
                
                tag_ul = soup.find('ul', class_='ArticleList')
               
                if tag_ul:
                    tag_ils= tag_ul.find_all('li',class_='row')
                    for tag_il in tag_ils:
                        h2 = tag_il.find('h2', class_='Title')
                     
                        if h2:
                            
                            price = 0
                            url = ""
                            name =""
                            photo = ""
                            a3 = h2.find('a')
                            if a3:
                                name=a3.text.strip()
                                url =self.base + a3.get('href')
                                print('check')
                                print(url)
                                print(name)
                                print(self.url_id)
                                check = hp.check_cam_url(self.url_id,url,name)
                            
                                if url != '' and name != '' and   check :
                                    print("nhap cam")
                                    if self.camlist.add_cam(camob.CameraObject(0,name,price, url,photo,self.cat_id)):
                                        self.url_links.append(url)
                        
                tag_title_product = soup.find('div', class_="ArticleHeader")
            
                if tag_title_product  :
                    print('chi tiet san pham')
                    tag_content =  soup.find('div',class_='ArticleContent')
                    d = parse_vn_date_from_soup(soup)
                    print('****************date')
                    print(d)
                

                    if  d is not None :
                        print(d < FROM_DATE)
                        if d < FROM_DATE:
                            print(f"Bài {d} < {FROM_DATE} => bỏ qua")
                            return

                    photo = ""
                    short = ""
                    if tag_content:
                        for cam in self.camlist.camobs:
                            if cam.url == o_url:
                                cam.date_publish = d
                                cam.short = self.target
                                for tag in tag_content.find_all(True):
                                        if tag.has_attr('style'):
                                            del tag['style']

                                for img_tag in tag_content.find_all('img'):
                                    if 'src' in img_tag.attrs:  # Kiểm tra nếu có thuộc tính 'src'
                                        src_value = img_tag['src']  # Lấy giá trị thuộc tính 'src'
                                        img_tag.attrs = {'src': src_value}  # Chỉ giữ lại thuộc tính 'src'
                                    else:
                                        img_tag.attrs = {}  # Xóa tất cả thuộc tính nếu không có 'src'
                                for tag in tag_content.find_all(True):
                                        # Nếu thẻ là <img>, chỉ giữ lại thuộc tính 'src'
                                        if tag.name == 'img':
                                            src = tag.get('src')  
                                            tag.attrs = {}
                                            if src:
                                                tag['src'] = src
                                            print('tag img')
                                            print(tag)
                                        else:
                                            # Với các thẻ khác, xóa toàn bộ attribute
                                            tag.attrs = {}       
                                tags_to_removes = tag_content.find_all ('script' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                tags_to_removes = tag_content.find_all ('link' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                cam.description = str(tag_content)
                                cam.display_info()
                                hp.save_data_cam(self.url_id,cam) 
                # truonghop3
               
                tag_divs= soup.find_all('div',class_='post-title')
                for tag_il in tag_divs:
                    h4 = tag_il.find('h4', class_='entry-title' )
                    if h4:
                        tag_a = h4.find("a")
                        price = 0
                        url = ""
                        name =""
                        photo = ""
                        if tag_a:
                            name=tag_a.text.strip()
                            url =self.base + tag_a.get('href')
                    

                        check = hp.check_cam_url(self.url_id,url,name)
                        if url != '' and name != '' and   check :
                            print("nhap cam")
                            if self.camlist.add_cam(camob.CameraObject(0,name,price, url,photo,self.cat_id)):
                                self.url_links.append(url)
                    
                tag_title_product = soup.find('div', class_="ArticleHeader")
            
                if tag_title_product  :
                    print('chi tiet san pham')
                    tag_content =  soup.find('div',class_='ArticleContent')
                    d = parse_vn_date_from_soup(soup)
                    print('****************date')
                    print(d)
                

                    if  d is not None :
                        print(d < FROM_DATE)
                        if d < FROM_DATE:
                            print(f"Bài {d} < {FROM_DATE} => bỏ qua")
                            return

                    photo = ""
                    short = ""
                    if tag_content:
                        for cam in self.camlist.camobs:
                            if cam.url == o_url:
                                cam.date_publish = d
                                cam.short = self.target
                                for tag in tag_content.find_all(True):
                                        if tag.has_attr('style'):
                                            del tag['style']

                                for img_tag in tag_content.find_all('img'):
                                    if 'src' in img_tag.attrs:  # Kiểm tra nếu có thuộc tính 'src'
                                        src_value = img_tag['src']  # Lấy giá trị thuộc tính 'src'
                                        img_tag.attrs = {'src': src_value}  # Chỉ giữ lại thuộc tính 'src'
                                    else:
                                        img_tag.attrs = {}  # Xóa tất cả thuộc tính nếu không có 'src'
                                for tag in tag_content.find_all(True):
                                        # Nếu thẻ là <img>, chỉ giữ lại thuộc tính 'src'
                                        if tag.name == 'img':
                                            src = tag.get('src')  
                                            tag.attrs = {}
                                            if src:
                                                tag['src'] = src
                                            print('tag img')
                                            print(tag)
                                        else:
                                            # Với các thẻ khác, xóa toàn bộ attribute
                                            tag.attrs = {}       
                                tags_to_removes = tag_content.find_all ('script' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                tags_to_removes = tag_content.find_all ('link' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                cam.description = str(tag_content)
                                cam.display_info()
                                hp.save_data_cam(self.url_id,cam) 
            # read menu
                # tag_div = soup.find('div', class_="default-pagination") 
                #     # if tag_menu:
                # tag_menu_a = tag_div.find_all('a')
                # #read menu link
                # for tag_a in tag_menu_a:
                #     # print(tag_a)
                #     # print(len(tag_menu_a))
                #     url =  tag_a.get('href')
                #     print(url)
                #     if(url ):
                #             if self.menu_links.add_link(url) == 1:
                #                 self.url_links.append(url)
                if chitiet == 1:
                    return
                btns = self.driver.find_elements(By.CSS_SELECTOR, 'div.col-center a.next')   
                print(btns)
                # Không còn nút → dừng
                if   btns:
                    b = btns[0]
                    self.driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});", b
                    )
                    time.sleep(0.2)
                    self.driver.execute_script("arguments[0].click();", b)
                    time.sleep(2.5)
                if len (btns) == 0:
                    btns = self.driver.find_elements(By.CSS_SELECTOR, 'a.page-link')
                    print('do dai btns')
                    bam = 0
                   
                    for btn in btns:
                        btn_page  = to_int(btn.text, default=-1)
                        cur_page  = int(self.page_link)  # nếu self.page_link đã là số, bỏ int cũng được
                        print(btn_page)
                        print(cur_page)
                        if btn_page > cur_page:
                            self.driver.execute_script(
                                "arguments[0].scrollIntoView({block:'center'});", btn
                            )
                            time.sleep(0.2)
                            self.driver.execute_script("arguments[0].click();", btn)
                            time.sleep(2.5)
                            print('bấm rồi')
                            bam = 1
                            self.page_link =  self.page_link + 1  
                    if bam == 0 or len (btns) == 0:
                        print('ket thuc')
                        break
                    
                      
                    


def run_from_excel(
    excel_path: str,
    sheet_name: str = None,
    start_row: int = 2,
    sleep_between: float = 0.5,
    save_every: int = 1,          # lưu sau mỗi N dòng xử lý
):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # đọc header để map cột theo tên
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            headers[str(v).strip().lower()] = col

    required = ["idurl", "source", "target", "cat_id"]
    for k in required:
        if k not in headers:
            raise ValueError(f"Thiếu cột '{k}' trong Excel. Hiện có: {list(headers.keys())}")

    # cột done (bắt buộc theo yêu cầu mới)
    if "done" not in headers:
        raise ValueError("Thiếu cột 'done' trong Excel (header phải là 'done').")

    done_col = headers["done"]

    bot = None
    processed_since_save = 0

    try:
        for r in range(start_row, ws.max_row + 1):
            idurl  = ws.cell(r, headers["idurl"]).value
            source = ws.cell(r, headers["source"]).value
            target = ws.cell(r, headers["target"]).value
            cat_id = ws.cell(r, headers["cat_id"]).value
            done_v = ws.cell(r, done_col).value

            # nếu không có source thì bỏ qua
            if not source:
                continue

            # chỉ chạy khi done == 0 hoặc trống
            # (chấp nhận "0", 0, "", None)
            done_str = str(done_v).strip() if done_v is not None else ""
            if done_str == "1":
                # đã làm rồi
                continue

            idurl  = str(idurl).strip() if idurl is not None else ""
            source = str(source).strip()
            target = str(target).strip() if target else ""
            cat_id = str(cat_id).strip() if cat_id else ""

            base = get_base(source)

            if bot is None:
                bot = VspProducts(base=base, url=source, cat=cat_id, target=target)

            bot.reset_for_row()
            bot.url    = source
            bot.url_id = idurl
            bot.cat_id = cat_id
            bot.base   = base
            bot.target   = target
            print(f"\n=== ROW {r} | idurl={idurl} | cat_id={cat_id}")
            print(f"source={source}")
            print(f"target={target}")

            try:
                bot.get_data()

                # nếu chạy xong không lỗi -> đánh done = 1
                ws.cell(r, done_col).value = 1

            except Exception as e:
                # nếu lỗi thì KHÔNG đánh done = 1, để lần sau chạy lại
                print(f"[ERROR] ROW {r} => {e}")

            processed_since_save += 1
            if processed_since_save >= save_every:
                wb.save(excel_path)
                processed_since_save = 0

            time.sleep(sleep_between)

    finally:
        # lưu nốt nếu còn thay đổi chưa save
        try:
            wb.save(excel_path)
        except:
            pass

        if bot:
            bot.close()



# chạy
run_from_excel(r"D:\WORKSPACE_CODE\Projects\Web\folder\websiteninhbinh\danhsachweb.xlsx", sheet_name=None)
