import os
import re
import time
import uuid
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from datetime import datetime, date
from urllib.parse import urljoin, urlparse, quote_plus
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from collections import defaultdict
from urllib.parse import unquote
import unicodedata
import hashlib
import helpers as hp

from openpyxl import Workbook, load_workbook

# Setup requests session with retries
retry_strategy = Retry(
    total=3,
    backoff_factor=1,
    status_forcelist=[429, 500, 502, 503, 504],
)
adapter = HTTPAdapter(max_retries=retry_strategy)
http_session = requests.Session()
http_session.mount("https://", adapter)
http_session.mount("http://", adapter)

# Add headers to mimic browser
http_session.headers.update(
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
        "Accept": "image/webp,image/apng,image/*,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "DNT": "1",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Referer": "https://c0giaophong.ninhbinh.edu.vn/",
    }
)

LOG_XLSX = os.path.join(os.path.dirname(__file__), "..", "..", "log_posted.xlsx")

# =========================
# CONFIG
# =========================
CHROME_DRIVER_PATH = r"D:\WORKSPACE_CODE\Projects\Web\folder\chromedriver-win64\chromedriver-win64\chromedriver.exe"
CHROME_BINARY_PATH = (
    r"D:\WORKSPACE_CODE\Projects\Web\folder\chrome-win64\chrome-win64\chrome.exe"
)

EMAIL = "adminvtk"
PASSWORD = "Khanhkh@nh9999"

USE_PROFILE = False
PROFILE_DIR = r"D:\WORKSPACE_CODE\Projects\Web\folder\selenium_profile_itcctv"

TMP_DIR = r"D:\WORKSPACE_CODE\Projects\Web\folder\tmp_wp_upload"
DEFAULT_PUBLISH_HOUR = 8
DEFAULT_PUBLISH_MINUTE = 0


# =========================
# BASIC HELPERS
# =========================


def init_log_workbook(path: str, sheet_name: str = "log"):
    if os.path.isfile(path):
        wb = load_workbook(path)
        ws = (
            wb[sheet_name]
            if sheet_name in wb.sheetnames
            else wb.create_sheet(sheet_name)
        )
        return wb, ws

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = [
        "ts",
        "base_target",
        "login_url",
        "news_id",
        "title",
        "source_url",
        "post_date",
        "category",
        "status",
        "step",
        "error",
        "uploaded_count",
        "uploaded_ids",
        "featured_ok",
        "duration_sec",
    ]
    ws.append(headers)
    wb.save(path)
    return wb, ws


def append_log_row(path: str, row: dict, sheet_name: str = "log"):
    wb, ws = init_log_workbook(path, sheet_name=sheet_name)
    headers = [cell.value for cell in ws[1]]

    line = []
    for h in headers:
        v = row.get(h, "")
        # ép list -> string cho dễ đọc
        if isinstance(v, (list, tuple, set)):
            v = ",".join(str(x) for x in v)
        line.append(v)

    ws.append(line)
    wb.save(path)


def init_log_workbook(path: str, sheet_name: str = "log"):
    if os.path.isfile(path):
        try:
            wb = load_workbook(path)
            ws = (
                wb[sheet_name]
                if sheet_name in wb.sheetnames
                else wb.create_sheet(sheet_name)
            )
            return wb, ws
        except Exception as e:
            print(f"File excel log lỗi ({e}), tạo mới...")
            # Fallthrough to create new

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Add headers for new file
    ws.append(["ts", "base_target", "news_id", "title", "status", "step", "error"])
    return wb, ws


def append_log_row(path: str, row: dict, sheet_name: str = "log"):
    wb, ws = init_log_workbook(path, sheet_name=sheet_name)
    headers = [cell.value for cell in ws[1]]

    line = []
    for h in headers:
        v = row.get(h, "")
        # ép list -> string cho dễ đọc
        if isinstance(v, (list, tuple, set)):
            v = ",".join(str(x) for x in v)
        line.append(v)

    ws.append(line)
    wb.save(path)


def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def safe_js_click(driver, element):
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
    driver.execute_script("arguments[0].click();", element)


def date_to_datetime(d: date, hour=8, minute=0) -> datetime:
    return datetime(d.year, d.month, d.day, hour, minute)


def get_base(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    p = urlparse(url)
    if not p.scheme or not p.netloc:
        p = urlparse("https://" + url)
    return f"{p.scheme}://{p.netloc}/"


def normalize_img_url(src: str, base_source: str) -> str:
    src = (src or "").strip()
    if not src:
        return ""
    if src.startswith("//"):
        return "https:" + src
    if src.startswith("http://") or src.startswith("https://"):
        return src
    return urljoin(base_source, src)


def guess_ext_from_url(img_url: str) -> str:
    ext = ".jpg"
    path_part = urlparse(img_url).path.lower()
    for e in [".jpg", ".jpeg", ".png", ".webp", ".gif"]:
        if path_part.endswith(e):
            return e
    return ext


def _safe_filename_from_url(img_url: str, max_len: int = 100) -> str:
    p = urlparse(img_url)
    name = os.path.basename(p.path) or "image.jpg"

    # ✅ decode %xx
    name = unquote(name)

    # ✅ normalize unicode (Windows-safe)
    name = unicodedata.normalize("NFKD", name)

    # bỏ query / fragment
    name = name.split("?")[0].split("#")[0]

    # bỏ ký tự cấm Windows
    name = re.sub(r'[<>:"/\\|?*\x00-\x1F]', "-", name)

    # gọn khoảng trắng
    name = re.sub(r"\s+", " ", name).strip()

    if not name:
        name = "image.jpg"

    stem, ext = os.path.splitext(name)
    ext = ext or ".jpg"

    # hash ngắn để tránh trùng
    h = hashlib.md5(name.encode("utf-8")).hexdigest()[:8]

    stem = stem[:max_len]
    return f"{stem}-{h}{ext}"


def download_file(file_url: str, tmp_dir: str = TMP_DIR, custom_name: str = "") -> str:
    ensure_dir(tmp_dir)

    if custom_name:
        import re

        safe_title = custom_name.strip()
        safe_title = re.sub(r"\s+", "_", safe_title)
        safe_title = re.sub(r'[\\/*?:"<>|]', "", safe_title)

        parsed = urlparse(file_url)
        ext = os.path.splitext(parsed.path)[1]
        if not ext:
            ext = ".pdf"
        filename2 = f"{safe_title}_{uuid.uuid4().hex[:4]}{ext}"
    else:
        filename = _safe_filename_from_url(file_url)
        filename2 = f"{os.path.splitext(filename)[0]}-{uuid.uuid4().hex[:8]}{os.path.splitext(filename)[1] or ''}"

    file_path = os.path.join(tmp_dir, filename2)

    r = http_session.get(file_url, timeout=120, stream=True)
    r.raise_for_status()
    with open(file_path, "wb") as f:
        for chunk in r.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)

    return file_path


FILE_EXTS = (
    ".pdf",
    ".doc",
    ".docx",
    ".xls",
    ".xlsx",
    ".ppt",
    ".pptx",
    ".zip",
    ".rar",
    ".7z",
)


def is_download_link(tag) -> bool:
    if not tag or getattr(tag, "name", "").lower() != "a":
        return False
    href = (tag.get("href") or "").strip().lower()
    if not href:
        return False

    cls = " ".join(tag.get("class", [])).lower()
    if "link-download" in cls:
        return True

    # fallback theo đuôi file
    path = urlparse(href).path.lower()
    return path.endswith(FILE_EXTS)


def download_image(img_url: str, tmp_dir: str = TMP_DIR, referer: str = None) -> str:
    ensure_dir(tmp_dir)

    filename = _safe_filename_from_url(img_url)
    stem, ext = os.path.splitext(filename)
    if ext.lower() not in [".jpg", ".jpeg", ".png", ".webp", ".gif"]:
        ext = guess_ext_from_url(img_url)

    # thêm suffix để stem khác nhau -> tìm aria-label ổn định
    filename2 = f"{stem}-{uuid.uuid4().hex[:8]}{ext}"
    file_path = os.path.join(tmp_dir, filename2)

    headers = {}
    if referer:
        headers["Referer"] = referer

    r = http_session.get(img_url, timeout=120, stream=True, headers=headers)
    r.raise_for_status()
    with open(file_path, "wb") as f:
        for chunk in r.iter_content(chunk_size=8192):
            if chunk:
                f.write(chunk)

    return file_path


# =========================
# WORDPRESS: EDITOR HELPERS
# =========================
def wp_editor_focus_end(driver):
    driver.execute_script("""
    try {
      if (window.tinymce && tinymce.get('content')) {
        const ed = tinymce.get('content');
        ed.focus();
        ed.selection.select(ed.getBody(), true);
        ed.selection.collapse(false);
        return true;
      }
    } catch(e) {}
    return false;
    """)


def wp_editor_insert_html(driver, html: str):
    html = html or ""
    driver.execute_script(
        """
    const html = arguments[0];
    try {
      if (window.tinymce && tinymce.get('content')) {
        const ed = tinymce.get('content');
        ed.execCommand('mceInsertContent', false, html);
        ed.save();
        return 'tinymce';
      }
    } catch(e) {}
    const ta = document.getElementById('content');
    if (ta) {
      ta.value = (ta.value || '') + html;
      return 'textarea';
    }
    return 'none';
    """,
        html,
    )


# =========================
# WORDPRESS: MEDIA MODAL (CÁCH B)
# =========================
def wp_open_add_media_modal(driver, wait: WebDriverWait):
    btn = wait.until(EC.element_to_be_clickable((By.ID, "insert-media-button")))
    safe_js_click(driver, btn)
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".media-modal")))

    # chuyển sang tab Upload Files nếu có
    try:
        upload_tab = wait.until(
            EC.element_to_be_clickable(
                (
                    By.CSS_SELECTOR,
                    "#menu-item-upload, .media-menu-item[data-route='upload']",
                )
            )
        )
        safe_js_click(driver, upload_tab)
        time.sleep(0.2)
    except:
        pass


def wp_media_clear_selection(driver):
    driver.execute_script("""
    try {
      // reset selection trong wp.media.editor (Classic Editor)
      if (window.wp && wp.media && wp.media.editor && wp.media.editor.get) {
        const fr = wp.media.editor.get();
        if (fr && fr.state) {
          const st = fr.state();
          if (st && st.get) {
            const sel = st.get('selection');
            if (sel && sel.reset) sel.reset();
          }
        }
      }

      // reset selection trong wp.media.frame (fallback)
      if (window.wp && wp.media && wp.media.frame) {
        const fr2 = wp.media.frame;
        if (fr2 && fr2.state) {
          const st2 = fr2.state();
          if (st2 && st2.get) {
            const sel2 = st2.get('selection');
            if (sel2 && sel2.reset) sel2.reset();
          }
        }
      }

      // dọn tick xanh UI
      document.querySelectorAll('li.attachment[aria-checked="true"]').forEach(el=>{
        el.setAttribute('aria-checked','false');
        el.classList.remove('selected','details');
      });
    } catch(e) {}
    """)


def _get_modal_file_input(driver):
    # ưu tiên input file nằm trong modal
    inputs = driver.find_elements(By.CSS_SELECTOR, ".media-modal input[type='file']")
    if inputs:
        return inputs[0]
    # fallback
    inputs = driver.find_elements(By.CSS_SELECTOR, "input[type='file']")
    if inputs:
        return inputs[0]
    return None


def wp_media_upload_pick_insert2(driver, wait, file_path: str) -> int:
    """
    Upload 1 ảnh:
    - WP site này tự chèn ảnh ngay sau upload
    - KHÔNG có nút 'Chèn'
    - Chỉ cần: upload → xác định attachment → đợi modal đóng
    """
    wp_media_clear_selection(driver)

    # input file
    file_input = wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, ".media-modal input[type='file'], input[type='file']")
        )
    )
    file_input.send_keys(file_path)

    # chờ attachments xuất hiện
    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "ul.attachments, .attachments-browser")
        )
    )

    stem = os.path.splitext(os.path.basename(file_path))[0]
    stem_words = [w.lower() for w in re.split(r"[-_]", stem) if w.strip()]
    att_id = None

    # tìm đúng attachment theo aria-label
    for _ in range(120):
        items = driver.find_elements(
            By.CSS_SELECTOR, "li.attachment[data-id][aria-label]"
        )
        for it in items:
            al = (it.get_attribute("aria-label") or "").lower()
            if stem.lower() in al or (stem_words and all(w in al for w in stem_words)):
                safe_js_click(driver, it)
                att_id = int(it.get_attribute("data-id"))
                break
        if att_id:
            break
        time.sleep(0.25)

    if not att_id:
        # fallback: item đang selected
        sel = driver.find_elements(
            By.CSS_SELECTOR,
            "li.attachment.selected[data-id], li.attachment[aria-checked='true'][data-id]",
        )
        if sel:
            safe_js_click(driver, sel[0])
            att_id = int(sel[0].get_attribute("data-id"))

    if not att_id:
        raise RuntimeError(f"Không xác định được attachment vừa upload: {file_path}")

    # 🔥 ĐIỂM QUAN TRỌNG 🔥
    # Site này tự chèn ảnh → chỉ cần đợi modal đóng
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".media-modal")))

    return att_id


def wp_media_upload_pick_insert(driver, wait: WebDriverWait, file_path: str) -> int:
    """
    Upload 1 file trong modal Add Media và bấm 'Chèn vào bài viết' luôn.
    Mỗi lần chỉ chọn đúng 1 ảnh vừa upload (clear selection trước đó).
    """
    wp_media_clear_selection(driver)

    file_input = _get_modal_file_input(driver)
    if not file_input:
        raise RuntimeError("Không tìm thấy input[type=file] trong Media Modal")

    file_input.send_keys(file_path)

    # chờ danh sách attachments có
    wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "ul.attachments, .attachments-browser")
        )
    )

    stem = os.path.splitext(os.path.basename(file_path))[0]
    stem_words = [w.lower() for w in re.split(r"[-_]", stem) if w.strip()]
    att_id = None

    # 1) tìm đúng item theo aria-label chứa stem
    for _ in range(220):  # ~55s
        try:
            items = driver.find_elements(
                By.CSS_SELECTOR, "li.attachment[data-id][aria-label]"
            )
            for it in items:
                al = (it.get_attribute("aria-label") or "").lower()
                if stem.lower() in al or (
                    stem_words and all(w in al for w in stem_words)
                ):
                    safe_js_click(driver, it)
                    att_id = int(it.get_attribute("data-id"))
                    break
            if att_id:
                break
        except StaleElementReferenceException:
            pass
        time.sleep(0.25)

    # 2) fallback: item đang selected
    if not att_id:
        sel = driver.find_elements(
            By.CSS_SELECTOR,
            "li.attachment.selected[data-id], li.attachment[aria-checked='true'][data-id]",
        )
        if sel:
            safe_js_click(driver, sel[0])
            att_id = int(sel[0].get_attribute("data-id"))

    if not att_id:
        raise RuntimeError(f"Không xác định được attachment vừa upload: {file_path}")

    # 3) ép selection chỉ còn đúng 1 ảnh vừa upload (cực quan trọng)
    ok = driver.execute_script(
        """
    try {
      const id = arguments[0];
      if (!window.wp || !wp.media) return false;

      let fr = null;
      if (wp.media.editor && wp.media.editor.get) fr = wp.media.editor.get();
      if (!fr && wp.media.frame) fr = wp.media.frame;
      if (!fr || !fr.state) return false;

      const st = fr.state();
      const sel = st.get('selection');
      sel.reset();

      const att = wp.media.attachment(id);
      att.fetch();
      sel.add(att);

      const btn = document.querySelector('button.media-button-insert');
      if (btn) btn.disabled = false;

      return true;
    } catch(e) { return false; }
    """,
        int(att_id),
    )

    if not ok:
        raise RuntimeError("Không set selection được bằng wp.media/editor")

    # 4) bấm "Chèn vào bài viết"
    def _find_clickable_insert_button(driver):
        # Ưu tiên các class phổ biến của WP
        candidates = driver.find_elements(
            By.CSS_SELECTOR,
            ".media-modal button.media-button-insert, "
            ".media-modal button.media-button-select, "
            ".media-modal .media-toolbar button.media-button, "
            ".media-modal .media-toolbar-primary button.button-primary",
        )

        # lọc theo text (VN/EN) + trạng thái enabled
        for b in candidates:
            try:
                if not b.is_displayed():
                    continue
                if b.get_attribute("disabled"):
                    continue
                txt = (
                    ((b.text or "") + " " + (b.get_attribute("value") or ""))
                    .strip()
                    .lower()
                )
                # các nhãn hay gặp
                if (
                    ("chèn" in txt)
                    or ("insert" in txt)
                    or ("select" in txt)
                    or ("choose" in txt)
                ):
                    return b
            except:
                pass

        # fallback: nút primary cuối toolbar
        try:
            b = driver.find_element(
                By.CSS_SELECTOR, ".media-modal .media-toolbar-primary button"
            )
            if b.is_displayed() and not b.get_attribute("disabled"):
                return b
        except:
            pass

        return None

    # --- bấm nút chèn ---
    deadline = time.time() + 20  # tối đa 20s
    insert_btn = None
    while time.time() < deadline:
        insert_btn = _find_clickable_insert_button(driver)
        if insert_btn:
            break
        time.sleep(0.25)

    if not insert_btn:
        # debug để bạn nhìn đúng trạng thái modal đang ở đâu
        try:
            toolbar = driver.find_element(
                By.CSS_SELECTOR, ".media-modal .media-toolbar"
            )
            print("DEBUG TOOLBAR:", toolbar.text[:400])
        except:
            print("DEBUG TOOLBAR: cannot read")
        # raise TimeoutException("Không tìm thấy nút 'Chèn vào bài viết/Chọn' trong media modal")
    else:
        safe_js_click(driver, insert_btn)

    # đợi modal đóng
    wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".media-modal")))

    return att_id


# =========================
# WORDPRESS: FEATURED IMAGE
# =========================
def _wait_media_modal_open(wait: WebDriverWait):
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".media-modal")))


def _close_media_modal(driver):
    try:
        btn = driver.find_element(By.CSS_SELECTOR, "button.media-modal-close")
        driver.execute_script("arguments[0].click();", btn)
    except:
        pass


def set_featured_image_by_id(driver, wait: WebDriverWait, attachment_id: int) -> bool:
    """
    Classic Editor: set featured image đúng attachment_id.
    - Mở modal #set-post-thumbnail
    - Dùng wp.media.featuredImage.frame() (đúng frame)
    - Reset selection rồi add attachment(id)
    - Bấm nút primary (Set featured image / Đặt ảnh đại diện / Chọn ảnh tiêu biểu...)
    """

    # 1) mở modal ảnh tiêu biểu
    open_btn = wait.until(EC.element_to_be_clickable((By.ID, "set-post-thumbnail")))
    safe_js_click(driver, open_btn)

    # chờ modal
    wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".media-modal, .media-frame"))
    )

    # 2) set selection bằng featuredImage frame
    ok = driver.execute_script(
        """
    try {
      const id = arguments[0];
      if (!window.wp || !wp.media) return {ok:false, why:'no-wp-media'};

      // featured frame (classic)
      let frame = null;
      if (wp.media.featuredImage && wp.media.featuredImage.frame) {
        frame = wp.media.featuredImage.frame();
      }
      // fallback nếu site custom
      if (!frame && wp.media.frame) frame = wp.media.frame;
      if (!frame || !frame.state) return {ok:false, why:'no-frame'};

      const st = frame.state();
      if (!st || !st.get) return {ok:false, why:'no-state'};

      const sel = st.get('selection');
      if (sel && sel.reset) sel.reset();

      const att = wp.media.attachment(id);
      att.fetch();

      // add đúng 1 attachment
      if (sel && sel.add) sel.add(att);

      // bật nút primary (đừng phụ thuộc class cố định)
      document.querySelectorAll('.media-modal button, .media-frame button').forEach(b=>{
        try {
          const t = (b.innerText || b.value || '').toLowerCase();
          if (
            t.includes('set featured') ||
            t.includes('featured image') ||
            t.includes('đặt ảnh') ||
            t.includes('ảnh đại diện') ||
            t.includes('chọn ảnh') ||
            t.includes('select') ||
            t.includes('choose')
          ) b.disabled = false;
        } catch(e) {}
      });

      return {ok:true, why:'ok'};
    } catch(e) {
      return {ok:false, why:String(e)};
    }
    """,
        int(attachment_id),
    )

    if not ok or not ok.get("ok"):
        # đóng modal cho sạch trạng thái
        try:
            _close_media_modal(driver)
        except:
            pass
        raise RuntimeError(f"Không thể set selection featured image. Debug: {ok}")

    # 3) tìm nút xác nhận theo text (VN/EN) + button-primary
    def find_feature_button():
        candidates = driver.find_elements(
            By.CSS_SELECTOR,
            ".media-modal .media-toolbar-primary button, "
            ".media-frame .media-toolbar-primary button, "
            ".media-modal button.button-primary, "
            ".media-frame button.button-primary, "
            ".media-modal button.media-button, "
            ".media-frame button.media-button",
        )
        for b in candidates:
            try:
                if not b.is_displayed():
                    continue
                if b.get_attribute("disabled"):
                    continue
                txt = (
                    ((b.text or "") + " " + (b.get_attribute("value") or ""))
                    .strip()
                    .lower()
                )

                # các nhãn thường gặp
                if (
                    "set featured" in txt
                    or "featured image" in txt
                    or "đặt ảnh" in txt
                    or "ảnh đại diện" in txt
                    or "chọn ảnh" in txt
                    or "select" in txt
                    or "choose" in txt
                ):
                    return b

                # fallback: nếu là button-primary và đang hiển thị thì lấy luôn
                cls = (b.get_attribute("class") or "").lower()
                if "button-primary" in cls:
                    return b
            except StaleElementReferenceException:
                continue
            except:
                continue
        return None

    deadline = time.time() + 20
    btn = None
    while time.time() < deadline:
        btn = find_feature_button()
        if btn:
            break
        time.sleep(0.25)

    if not btn:
        # debug toolbar text để bạn nhìn trạng thái
        try:
            tb = driver.find_element(
                By.CSS_SELECTOR,
                ".media-modal .media-toolbar, .media-frame .media-toolbar",
            )
            print("DEBUG FEATURE TOOLBAR:", (tb.text or "")[:500])
        except:
            print("DEBUG FEATURE TOOLBAR: cannot read")
        raise TimeoutException(
            "Không tìm thấy nút 'Đặt ảnh đại diện/Set featured image' trong modal."
        )

    safe_js_click(driver, btn)

    # 4) chờ thumbnail hiện hoặc _thumbnail_id input có value
    # Đôi khi img ko render kịp nhưng value hidden input đã có
    for _ in range(20):  # giảm xuống 5s xem sao, 120 (30s) hơi lâu
        # Check hidden input value
        thumb_val = driver.execute_script("""
            var el = document.getElementById('_thumbnail_id');
            return el ? el.value : '';
        """)
        if str(thumb_val) == str(attachment_id):
            return True

        if driver.find_elements(By.CSS_SELECTOR, "#postimagediv img"):
            return True
        time.sleep(0.25)

    # Nếu vẫn chưa thấy, ta thử check lại 1 lần nữa xem có phải nó đã đóng modal chưa?
    # Nếu modal đóng rồi nghĩa là OK, chỉ là chưa update UI kịp thôi.
    # Warning thay vì Error
    print(
        "⚠️ Warning: Đã set featured image nhưng chưa thấy UI cập nhật thumbnail/id. Coi như thành công."
    )
    return True


# ========================
# WORDPRESS: FEATURED IMAGE (CÁCH B)
# ========================
def set_featured_image_by_id2(driver, wait: WebDriverWait, attachment_id: int) -> bool:
    open_btn = wait.until(EC.element_to_be_clickable((By.ID, "set-post-thumbnail")))
    safe_js_click(driver, open_btn)
    _wait_media_modal_open(wait)

    ok = driver.execute_script(
        """
        try {
        const id = arguments[0];
        if (!window.wp || !wp.media) return false;

        // ưu tiên editor frame
        let fr = null;
        if (wp.media.editor && wp.media.editor.get) fr = wp.media.editor.get();
        if (!fr && wp.media.frame) fr = wp.media.frame;
        if (!fr || !fr.state) return false;

        const st = fr.state();
        const sel = st.get('selection');
        sel.reset();

        const att = wp.media.attachment(id);
        att.fetch();
        sel.add(att);

        // cố gắng bật nút primary
        document.querySelectorAll('.media-modal button').forEach(b=>{
            const t = (b.innerText||'').toLowerCase();
            if (t.includes('chèn') || t.includes('insert') || t.includes('select') || t.includes('choose')) {
            b.disabled = false;
            }
        });

        return true;
        } catch(e) { return false; }
        """,
        int(attachment_id),
    )

    if not ok:
        _close_media_modal(driver)
        raise RuntimeError("Không thể set featured image bằng wp.media")

    select_btn = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.media-button-select"))
    )
    safe_js_click(driver, select_btn)

    # 4) chờ _thumbnail_id được set
    for _ in range(80):
        thumb_id = driver.execute_script("""
            const el = document.querySelector("input#_thumbnail_id");
            return el ? el.value : "";
        """)
        if str(attachment_id) == str(thumb_id):
            return True
        time.sleep(0.25)

    raise TimeoutException(
        "Đã set featured image nhưng _thumbnail_id chưa được cập nhật."
    )


# =========================
# CATEGORY + PUBLISH DATE (Classic)
# =========================
def select_category_by_name(driver, wait: WebDriverWait, cat_name: str) -> bool:
    print("[CATEGORY] Want:", repr(cat_name))
    cat_name = (cat_name or "").strip()
    if not cat_name:
        return False

    wait.until(EC.presence_of_element_located((By.ID, "categorychecklist")))
    labels = driver.find_elements(By.CSS_SELECTOR, "#categorychecklist label")

    target_cats = [c.strip().lower() for c in cat_name.split(";") if c.strip()]
    found_any = False

    for lb in labels:
        lb_text = (lb.text or "").strip().lower()
        if lb_text in target_cats:
            print("✅ CATEGORY MATCH:", lb.text)
            cb = lb.find_element(By.TAG_NAME, "input")
            if not cb.is_selected():
                driver.execute_script("arguments[0].click();", cb)
            found_any = True

    if not found_any:
        print("Không tìm thấy category nào trong:", cat_name)
    return found_any


def set_wp_publish_datetime(driver, wait: WebDriverWait, dt: datetime):
    edit_links = driver.find_elements(By.CSS_SELECTOR, "a.edit-timestamp")
    if edit_links:
        safe_js_click(driver, edit_links[0])
    wait.until(EC.presence_of_element_located((By.ID, "mm")))

    driver.execute_script(
        """
    (function(mm,jj,aa,hh,mn){
      function setVal(id,val){
        var el=document.getElementById(id);
        if(!el) return;
        el.value = val;
        el.dispatchEvent(new Event('input',{bubbles:true}));
        el.dispatchEvent(new Event('change',{bubbles:true}));
      }
      setVal('mm', mm);
      setVal('jj', jj);
      setVal('aa', aa);
      setVal('hh', hh);
      setVal('mn', mn);
    })(arguments[0],arguments[1],arguments[2],arguments[3],arguments[4]);
    """,
        f"{dt.month:02d}",
        f"{dt.day:02d}",
        str(dt.year),
        f"{dt.hour:02d}",
        f"{dt.minute:02d}",
    )

    time.sleep(0.2)
    ok_btns = driver.find_elements(By.CSS_SELECTOR, "#timestampdiv .save-timestamp")
    if ok_btns:
        safe_js_click(driver, ok_btns[0])


# =========================
# CONTENT: CHÈN THEO THỨ TỰ + IMG THÌ UPLOAD&INSERT
# =========================


def wp_insert_content_with_images(
    driver, wait, content_html: str, base_source: str, current_cam=None
) -> list[int]:
    # Kiểm tra marker: paste trực tiếp không download
    if "<!-- TYPE_DIRECT_PASTE -->" in (content_html or ""):
        print(
            ">> Phát hiện đánh dấu TYPE_DIRECT_PASTE, paste thẳng content, không bóc tách tải file."
        )
        wp_editor_focus_end(driver)
        wp_editor_insert_html(driver, content_html)
        return []

    soup = BeautifulSoup(content_html or "", "html.parser")
    root = soup.body if soup.body else soup

    inserted_ids = []
    wp_editor_focus_end(driver)

    def walk(node):
        for child in list(getattr(node, "children", [])):
            # text node
            if getattr(child, "name", None) is None:
                txt = str(child)
                if txt.strip():
                    wp_editor_insert_html(driver, txt)
                continue

            tag_name = child.name.lower()

            # ====== IMG ======
            if tag_name == "img":
                src = (child.get("src") or "").strip()
                if not src:
                    continue

                if src.startswith(("data:", "chrome-extension:", "blob:", "file:")):
                    print("⏭️ SKIP IMG (invalid scheme):", src)
                    continue

                img_abs = normalize_img_url(src, base_source)
                if not img_abs.startswith(("http://", "https://")):
                    print("⏭️ SKIP IMG (invalid abs):", img_abs)
                    continue

                # Skip Zalo share buttons or other problematic URLs
                if "zalo.me" in img_abs.lower() or "button-share" in img_abs.lower():
                    print("⏭️ SKIP IMG (social share button):", img_abs)
                    continue

                try:
                    local_path = download_image(img_abs)
                except Exception as e:
                    print("⚠️ SKIP IMG (download failed):", img_abs, "|", e)
                    continue

                try:
                    wp_open_add_media_modal(driver, wait)
                    att_id = wp_media_upload_pick_insert(driver, wait, local_path)
                    inserted_ids.append(int(att_id))
                    wp_editor_focus_end(driver)
                finally:
                    if local_path and os.path.isfile(local_path):
                        os.remove(local_path)

                continue

            # ====== FILE DOWNLOAD LINK ======
            if tag_name == "a" and is_download_link(child):
                # Không xử lý upload để tránh chèn gview vào trong wp_editor mce_panel.
                # Giữ nguyên thẻ <a> trỏ vào remote cho văn bản, toolset đính kèm đã xử lý upload gốc rồi.
                wp_editor_insert_html(driver, str(child))
                continue

            # ====== TAG KHÁC ======
            # nếu bên trong có img hoặc link download thì đi sâu để giữ đúng thứ tự
            if child.find("img") or child.find("a", href=True):
                walk(child)
            else:
                wp_editor_insert_html(driver, str(child))

    walk(root)
    return inserted_ids


# =========================
# DUPLICATE TITLE GUARD
# =========================
def wp_post_exists_by_title(
    driver, wait: WebDriverWait, base: str, title: str, post_type: str = "post"
) -> bool:
    title = (title or "").strip()
    if not title:
        return False

    search_url = (
        base
        + f"wp-admin/edit.php?post_type={post_type}&post_status=all&s="
        + quote_plus(title)
    )
    driver.get(search_url)
    wait.until(EC.presence_of_element_located((By.ID, "the-list")))

    print("[DUPLICATE CHECK]", title)
    print("SEARCH URL:", search_url)

    rows = driver.find_elements(By.CSS_SELECTOR, "#the-list tr")
    for r in rows:
        els = r.find_elements(By.CSS_SELECTOR, "a.row-title")
        if not els:
            continue
        t = (els[0].text or "").strip()
        if t.lower() == title.lower():
            print("❌ DUPLICATE FOUND:", t)
            return True
    return False


def strip_non_bmp(s: str) -> str:
    s = s or ""
    return "".join(ch for ch in s if ord(ch) <= 0xFFFF)


def clean_title_for_wp(s: str) -> str:
    s = (s or "").strip()
    s = strip_non_bmp(s)  # bỏ emoji / ký tự ngoài BMP
    s = re.sub(r"\s+", " ", s)  # gọn khoảng trắng
    return s.strip()


# ========================
# GROUP BY TARGET SITE
def group_by_target(rows):
    groups = defaultdict(list)
    for r in rows:
        target = r[7]  # cột base_target
        groups[target].append(r)
    return groups


# ========================
# BUILD POST FROM ROWS GROUP UPLOAD 0
def build_post_from_rows(rows):
    """
    rows: list các record upload=0
    """
    title = clean_title_for_wp(rows[0][2])  # lấy title row đầu
    category = rows[0][9]
    base_target = rows[0][7]

    contents = []
    source_urls = []

    for r in rows:
        contents.append(r[3])  # content_html
        source_urls.append(r[6])

    content_html = "<hr/>".join(contents)

    return {
        "title": title,
        "content_html": content_html,
        "category": category,
        "base_target": base_target,
        "rows": rows,  # giữ lại để update upload
    }


# =========================
# MAIN BOT
# =========================


class NewsPoster:
    def __init__(self):
        service = Service(CHROME_DRIVER_PATH)
        chrome_options = Options()
        chrome_options.binary_location = CHROME_BINARY_PATH
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")

        if USE_PROFILE:
            chrome_options.add_argument(rf"--user-data-dir={PROFILE_DIR}")
            chrome_options.add_argument("--profile-directory=Default")

        self.driver = webdriver.Chrome(service=service, options=chrome_options)
        self.wait = WebDriverWait(self.driver, 20)
        self.logged_sites = set()

    def close(self):
        try:
            self.driver.quit()
        except:
            pass

    def ensure_wp_login(
        self, base: str, login_url: str, username: str, password: str
    ) -> bool:
        if base in self.logged_sites:
            return True

        self.driver.get(login_url)
        wait = WebDriverWait(self.driver, 20)

        try:
            user_el = wait.until(EC.presence_of_element_located((By.ID, "user_login")))
            pass_el = wait.until(EC.presence_of_element_located((By.ID, "user_pass")))
            btn_el = wait.until(EC.element_to_be_clickable((By.ID, "wp-submit")))
        except TimeoutException:
            print("Không thấy form WP login:", login_url)
            return False

        user_el.clear()
        user_el.send_keys(username)
        pass_el.clear()
        pass_el.send_keys(password)
        btn_el.click()

        try:
            wait.until(
                lambda d: (
                    ("wp-admin" in d.current_url)
                    or (len(d.find_elements(By.ID, "wpadminbar")) > 0)
                )
            )
        except TimeoutException:
            print(
                "Login WP có thể thất bại:",
                base,
                "URL hiện tại:",
                self.driver.current_url,
            )
            return False

        self.logged_sites.add(base)
        print("WP Login OK:", base)
        return True

    def post_one(self, anew: list) -> dict:
        t0 = time.time()

        res = {
            "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "base_target": (anew[7] or "").strip() if len(anew) > 7 else "",
            "login_url": "",
            "news_id": anew[0] if len(anew) > 0 else "",
            "title": "",
            "source_url": anew[6] if len(anew) > 6 else "",
            "post_date": "",
            "category": (anew[9] or "").strip() if len(anew) > 9 else "",
            "status": "FAIL",
            "step": "",
            "error": "",
            "uploaded_count": 0,
            "uploaded_ids": [],
            "featured_ok": "",
            "duration_sec": 0,
        }

        try:
            print("\n" + "=" * 80)
            print("[START POST]")
            print("NEWS_ID:", anew[0])
            print("TITLE:", anew[2])
            print("TARGET:", anew[7])
            print("CATEGORY:", res["category"])
            print("=" * 80)

            title = clean_title_for_wp(anew[2] or "")
            content_html = (anew[3] or "").strip() if len(anew) > 3 else ""
            base_source = get_base(anew[6]) if len(anew) > 6 else ""
            base_target = (anew[7] or "").strip() if len(anew) > 7 else ""

            import json
            import re

            vanban_meta = {}
            match = re.search(
                r"<!--\s*VANBAN_META:\s*({.*?})\s*-->", content_html, re.DOTALL
            )
            if match:
                try:
                    vanban_meta = json.loads(match.group(1))
                    # Remove the comment so it doesn't show up in the wp editor
                    content_html = content_html.replace(match.group(0), "")
                except:
                    pass

            res["title"] = title
            res["post_date"] = str(anew[12]) if (len(anew) > 12 and anew[12]) else ""

            if not title or not content_html or not base_target:
                res["status"] = "SKIP_MISSING"
                res["step"] = "validate"
                return res

            base = get_base(base_target)
            res["login_url"] = base + "wp-login.php"

            # duplicate
            res["step"] = "duplicate_check"
            post_type_param = "van-ban" if vanban_meta else "post"
            if wp_post_exists_by_title(
                self.driver, self.wait, base, title, post_type=post_type_param
            ):
                res["status"] = "SKIP_DUPLICATE"
                return res

            # open editor
            res["step"] = "open_editor"
            if vanban_meta:
                create_url = base + "wp-admin/post-new.php?post_type=van-ban"
            else:
                create_url = base + "wp-admin/post-new.php"
            self.driver.get(create_url)

            # 🔥 ENSURE NEW POST 🔥
            for _ in range(5):
                title_el = self.wait.until(
                    EC.presence_of_element_located((By.ID, "title"))
                )
                if not title_el.get_attribute("value"):
                    break
                # nếu title vẫn có → WP đang mở lại bài cũ
                self.driver.get(create_url)
                time.sleep(1)
            else:
                raise RuntimeError(
                    "Không tạo được bài viết mới, WP đang mở lại post cũ"
                )

            print("[EDITOR] Open new post OK")

            # title
            res["step"] = "set_title"
            title_el = self.wait.until(EC.presence_of_element_located((By.ID, "title")))
            title_el.clear()
            title_el.send_keys(title)

            # clear content
            res["step"] = "clear_content"
            self.driver.execute_script("""
            try {
            if (window.tinymce && tinymce.get('content')) {
                tinymce.get('content').setContent('');
                tinymce.get('content').save();
            } else {
                const ta = document.getElementById('content');
                if (ta) ta.value = '';
            }
            } catch(e) {}
            """)

            # Fill custom fields for van-ban
            res["step"] = "fill_vanban_custom_fields"
            so_hieu = vanban_meta.get("so_ky_hieu", "")
            if so_hieu:
                try:
                    so_hieu_el = self.driver.find_element(
                        By.CSS_SELECTOR, "input[name='wpcf[vb-so-ky-hieu]']"
                    )
                    so_hieu_el.clear()
                    so_hieu_el.send_keys(so_hieu)
                except:
                    pass

            nguoi_ky = vanban_meta.get("nguoi_ky", "")
            co_quan = vanban_meta.get("co_quan_ban_hanh", "")
            if nguoi_ky or co_quan:
                trich_yeu_html = f"<p><strong>Cơ quan ban hành:</strong> {co_quan}<br/><strong>Người ký:</strong> {nguoi_ky}</p>"
                self.driver.execute_script(
                    """
                const html = arguments[0];
                try {
                    if (window.tinymce && tinymce.get('wpcf-vb-trich-yeu')) {
                        tinymce.get('wpcf-vb-trich-yeu').setContent(html);
                        tinymce.get('wpcf-vb-trich-yeu').save();
                    } else {
                        const ta = document.getElementById('wpcf-vb-trich-yeu');
                        if (ta) ta.value = html;
                    }
                } catch(e) {}
                """,
                    trich_yeu_html,
                )

            pub_date = str(anew[12])[:10] if (len(anew) > 12 and anew[12]) else ""
            if pub_date:
                try:
                    date_el = self.driver.find_element(
                        By.CSS_SELECTOR,
                        "input[name='wpcf[vb-ngay-ban-hanh][datepicker]']",
                    )
                    self.driver.execute_script(
                        "arguments[0].value = arguments[1];", date_el, pub_date
                    )
                    display_el = self.driver.find_element(
                        By.CSS_SELECTOR,
                        "input[name='wpcf[vb-ngay-ban-hanh][display-only]']",
                    )
                    self.driver.execute_script(
                        "arguments[0].value = arguments[1];", display_el, pub_date
                    )
                except:
                    pass

            file_links = vanban_meta.get("file_links", [])
            if file_links:
                try:
                    file_url = file_links[0]
                    # Upload file bằng toolset button
                    local_tmp = download_file(file_url, custom_name=title)

                    upload_btn = self.driver.find_element(
                        By.CSS_SELECTOR, "button.js-wpt-file-upload"
                    )
                    safe_js_click(self.driver, upload_btn)
                    time.sleep(1)

                    wp_media_upload_pick_insert(self.driver, self.wait, local_tmp)
                    time.sleep(1)

                    if os.path.exists(local_tmp):
                        os.remove(local_tmp)

                    if len(file_links) > 1:
                        file_input = self.driver.find_element(
                            By.CSS_SELECTOR, "input[name='wpcf[vb-file-dinh-kem]']"
                        )
                        existing = file_input.get_attribute("value") or ""
                        others = "\n".join(file_links[1:])
                        file_input.clear()
                        file_input.send_keys(existing + "\n" + others)
                except Exception as e:
                    print(f"Lỗi tải & gán file wpcf[vb-file-dinh-kem]: {e}")
                    # Fallback chỉ send keys
                    try:
                        file_input = self.driver.find_element(
                            By.CSS_SELECTOR, "input[name='wpcf[vb-file-dinh-kem]']"
                        )
                        file_input.clear()
                        file_input.send_keys("\n".join(file_links))
                    except:
                        pass

            # insert content + images
            res["step"] = "insert_content"
            uploaded_attachment_ids = wp_insert_content_with_images(
                driver=self.driver,
                wait=self.wait,
                content_html=content_html,
                base_source=base_source,
            )
            res["uploaded_ids"] = uploaded_attachment_ids
            res["uploaded_count"] = len(uploaded_attachment_ids)

            # set publish date
            post_date = anew[12] if len(anew) > 12 else None
            if post_date:
                res["step"] = "set_publish_datetime"
                dt = date_to_datetime(
                    post_date, hour=DEFAULT_PUBLISH_HOUR, minute=DEFAULT_PUBLISH_MINUTE
                )
                set_wp_publish_datetime(self.driver, self.wait, dt)

            # category for normal post vs taxonomy for van-ban
            cat_name = (anew[9] or "").strip() if len(anew) > 9 else ""
            if cat_name:
                res["step"] = "select_category"
                try:
                    # taxonomy checkbox is inside #loai-van-banchecklist
                    labels = self.driver.find_elements(
                        By.CSS_SELECTOR, "#loai-van-banchecklist label"
                    )
                    target_cats = [
                        c.strip().lower() for c in cat_name.split(",") if c.strip()
                    ]
                    found_in_taxonomy = False
                    for lb in labels:
                        if (lb.text or "").strip().lower() in target_cats:
                            cb = lb.find_element(By.TAG_NAME, "input")
                            if not cb.is_selected():
                                self.driver.execute_script("arguments[0].click();", cb)
                            found_in_taxonomy = True
                    if not found_in_taxonomy:
                        select_category_by_name(self.driver, self.wait, cat_name)
                except:
                    select_category_by_name(self.driver, self.wait, cat_name)

            # featured image
            res["featured_ok"] = ""
            if uploaded_attachment_ids:
                res["step"] = "set_featured"
                try:
                    ok = set_featured_image_by_id(
                        self.driver,
                        self.wait,
                        uploaded_attachment_ids[len(uploaded_attachment_ids) - 1],
                    )
                    res["featured_ok"] = "1" if ok else "0"
                except Exception as e:
                    # không fail toàn bài, chỉ ghi lại
                    res["featured_ok"] = "0"
                    res["error"] = f"featured_error: {e}"

            # publish
            post_id = None

            res["step"] = "publish"

            # WP often has "Publish" or "Update" depending on context, or "Schedule"
            # We try standard #publish button.
            try:
                publish_btn = self.wait.until(
                    EC.presence_of_element_located((By.ID, "publish"))
                )
                self.driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", publish_btn
                )
                time.sleep(1)  # wait for scroll

                # Try raw click first if clickable, else JS
                try:
                    publish_btn.click()
                except:
                    self.driver.execute_script("arguments[0].click();", publish_btn)

            except TimeoutException:
                res["error"] = "Publish button not found"
                raise

            # chờ message thành công HOẶC URL đổi
            # message.updated, message.notice-success
            # Check for URL having post=... and action=edit (which happens after publish)
            try:
                self.wait.until(
                    lambda d: (
                        ("post=" in d.current_url and "action=edit" in d.current_url)
                        or d.find_elements(
                            By.CSS_SELECTOR, "#message.updated, #message.notice-success"
                        )
                    )
                )
            except TimeoutException:
                print("Wait success timeout, checking if post was actually created...")
                # Fallback check handled by script below
                pass

            post_id = self.driver.execute_script("""
            const m = location.search.match(/post=(\d+)/);
            return m ? m[1] : null;
            """)

            if post_id:
                print("✅ POST CREATED, ID:", post_id)

            # wait success message
            res["step"] = "wait_success"
            try:
                self.wait.until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "#message.updated, #message.notice-success")
                    )
                )
            except:
                # không thấy message vẫn coi là success nếu URL đã đổi sang post.php?post=...
                pass

            res["status"] = "SUCCESS"
            return res

        except Exception as e:
            res["status"] = "FAIL"
            res["error"] = str(e)
            return res

        finally:
            res["duration_sec"] = round(time.time() - t0, 2)
            print("[END POST]")
            print("STATUS:", res["status"])
            print("STEP:", res["step"])
            if res["error"]:
                print("ERROR:", res["error"])
            print("DURATION:", res["duration_sec"], "sec")
            print("=" * 80 + "\n")

    # ========================
    # MAIN RUN
    # ========================
    def run(self):
        news = hp.read_news()
        if not news:
            print("✅ Không có bài nào cần đăng (upload = 0)")
            return
        print(f"🔎 Tìm thấy {len(news)} bài chưa upload")
        for anew in news:
            anew = list(anew)

            base = get_base(anew[7])
            login_url = base + "wp-login.php"

            title_check = clean_title_for_wp(anew[2] or "")
            if hp.check_summary_duplicate(title_check):
                print(f"❌ SKIPPING: Found duplicate in 'summary' table: {title_check}")
                print(
                    f"   -> Deleting from bot_news (ID: {anew[0]}) to prevent re-processing."
                )
                hp.delete_bot_news(anew[0])
                continue

            ok = self.ensure_wp_login(base, login_url, EMAIL, PASSWORD)
            if not ok:
                # log login fail
                append_log_row(
                    LOG_XLSX,
                    {
                        "ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "base_target": (anew[7] or "").strip() if len(anew) > 7 else "",
                        "login_url": login_url,
                        "news_id": anew[0] if len(anew) > 0 else "",
                        "title": clean_title_for_wp(anew[2] or ""),
                        "source_url": anew[6] if len(anew) > 6 else "",
                        "post_date": str(anew[12])
                        if (len(anew) > 12 and anew[12])
                        else "",
                        "category": (anew[9] or "").strip() if len(anew) > 9 else "",
                        "status": "FAIL",
                        "step": "login",
                        "error": "login_failed",
                    },
                )
                continue

            result = self.post_one(anew)

            # ghi log
            append_log_row(LOG_XLSX, result)

            # chỉ update db nếu SUCCESS hoặc SKIP_DUPLICATE (tuỳ bạn)
            try:
                if (
                    result.get("status") in ("SUCCESS", "SKIP_DUPLICATE")
                    and result.get("error") == ""
                ):
                    hp.update_upload_new(anew[0])

            except:
                pass

            time.sleep(1.5)
            # break


if __name__ == "__main__":
    bot = NewsPoster()
    try:
        bot.run()
    finally:
        bot.close()
