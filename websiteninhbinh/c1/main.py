import os
import sys
import time
from openpyxl import load_workbook

# Add parent directory to sys.path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


from .utils import get_base
from .scraper import VspProducts


def run_from_excel(
    excel_path: str,
    sheet_name: str = None,
    start_row: int = 2,
    sleep_between: float = 0.5,
    save_every: int = 1,
):
    print(f"Loading Excel: {excel_path}")
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            headers[str(v).strip().lower()] = col

    required = ["url_name", "source", "target", "cat_id"]
    for k in required:
        if k not in headers:
            raise ValueError(
                f"Thiếu cột '{k}' trong Excel. Hiện có: {list(headers.keys())}"
            )

    if "done" not in headers:
        raise ValueError("Thiếu cột 'done'.")

    done_col = headers["done"]

    # Check if url_name exists, if not it's fine we just pass empty string if needed
    url_name_col = headers.get("url_name")

    bot = None
    processed_since_save = 0

    try:
        for r in range(start_row, ws.max_row + 1):
            source = ws.cell(r, headers["source"]).value
            target = ws.cell(r, headers["target"]).value
            cat_id = ws.cell(r, headers["cat_id"]).value
            done_v = ws.cell(r, done_col).value
            url_name = ws.cell(r, url_name_col).value if url_name_col else ""

            if not source:
                continue

            done_str = str(done_v).strip() if done_v is not None else ""
            if done_str == "1":
                continue

            source = str(source).strip()
            target = str(target).strip() if target else ""
            cat_id = str(cat_id).strip() if cat_id else ""

            base = get_base(source)

            if bot is None:
                bot = VspProducts(base=base, url=source, cat=cat_id, target=target)

            bot.reset_for_row()
            bot.url = source
            bot.url_name = url_name
            bot.cat_id = cat_id
            bot.base = base
            bot.target = target
            print(f"\n=== ROW {r} | url_name={url_name} | cat_id={cat_id}")
            print(f"source={source}")

            try:
                bot.get_data()
                ws.cell(r, done_col).value = 1
            except Exception as e:
                print(f"[ERROR] ROW {r} => {e}")

            processed_since_save += 1
            if processed_since_save >= save_every:
                wb.save(excel_path)
                processed_since_save = 0

            time.sleep(sleep_between)

    finally:
        try:
            wb.save(excel_path)
        except:
            pass
        if bot:
            bot.close()


if __name__ == "__main__":
    # Example usage
    excel_file = (
        r"D:\WORKSPACE_CODE\Projects\Web\folder\websiteninhbinh\danhsachweb.xlsx"
    )
    if os.path.exists(excel_file):
        run_from_excel(excel_file, sheet_name=None)
    else:
        print(f"File Excel not found: {excel_file}")
