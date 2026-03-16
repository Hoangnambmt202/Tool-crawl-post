from openai import base_url
import requests
import Objectlink as obl
import MenuLink as menulink
from bs4 import BeautifulSoup
import CameraObject as camob

# Function to fetch and parse the webpage
import json
import mysql.connector
import time
import helpers as hp
from difflib import SequenceMatcher
from urllib.parse import urljoin
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from datetime import date
from openpyxl import load_workbook
import time
import re
from urllib.parse import urljoin, urlparse, quote
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

FROM_DATE = date(2023, 1, 1)  # chỉ lấy bài từ 01/07/2025 trở về đây
DETAIL_DATE_CSS = "span.post-date"  # ngày bài viết


def clean_spaces(s: str) -> str:
    return " ".join((s or "").split()).strip()


CHROME_DRIVER_PATH = r"D:\WORKSPACE_CODE\Projects\Web\folder\chromedriver-win64\chromedriver-win64\chromedriver.exe"
CHROME_BINARY_PATH = (
    r"D:\WORKSPACE_CODE\Projects\Web\folder\chrome-win64\chrome-win64\chrome.exe"
)


def parse_date_from_meta(soup: BeautifulSoup) -> date | None:
    """
    Fallback: lấy ngày từ meta.
    Ưu tiên itemprop=dateCreated (ví dụ: 2025-10-26T06:34:32+0700).
    Hỗ trợ thêm một số meta phổ biến khác.
    """
    if not soup:
        return None

    candidates = []

    # 1) itemprop
    for itemprop in ["dateCreated", "datePublished", "dateModified"]:
        tag = soup.find("meta", attrs={"itemprop": itemprop})
        if tag and tag.get("content"):
            candidates.append(tag["content"])

    # 2) Open Graph / Article
    for prop in ["article:published_time", "article:modified_time"]:
        tag = soup.find("meta", attrs={"property": prop})
        if tag and tag.get("content"):
            candidates.append(tag["content"])

    # 3) name=
    for name in [
        "pubdate",
        "publishdate",
        "date",
        "dc.date",
        "dc.date.issued",
        "dcterms.created",
        "dcterms.modified",
    ]:
        tag = soup.find("meta", attrs={"name": name})
        if tag and tag.get("content"):
            candidates.append(tag["content"])

    for s in candidates:
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
        if m:
            y, mo, d = map(int, m.groups())
            try:
                return date(y, mo, d)
            except:
                pass
        # Thử parse dd/mm/yyyy
        d_obj = parse_vn_date_any(s)
        if d_obj:
            return d_obj

    return None


def parse_public_date_from_uicongkhai(soup: BeautifulSoup) -> date | None:
    """
    Tìm trong table/tr:
      <td>Ngày công bố</td><td>13/10/2025</td>
    """
    root = soup.find("div", class_="UICongKhaiNganSach_Default")
    if not root:
        return None

    for tr in root.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) >= 2:
            label = clean_spaces(tds[0].get_text(" ", strip=True)).lower()
            if "ngày công bố" in label:
                return parse_vn_date_any(tds[1].get_text(" ", strip=True))

    # fallback: quét toàn root nếu cần
    return None


# 1. CẬP NHẬT HÀM PARSE NGÀY (Để bắt được ngày trong text nếu thiếu thẻ date)
# ==============================================================================
def parse_vn_date_from_soup(soup: BeautifulSoup) -> date | None:
    # 1. Tìm trong thẻ date hiển thị
    d = parse_vn_date_from_soup1(soup)

    # 2. Nếu không thấy, tìm trong div.title_news (Tiêu đề bài viết thường chứa ngày ở các trang giáo dục)
    if d is None:
        title_tag = soup.find("div", class_="title_news")
        if title_tag:
            d = parse_vn_date_any(title_tag.get_text())

    # 3. Nếu vẫn không thấy, tìm dòng text đầu tiên trong nội dung bài viết
    if d is None:
        content_tag = soup.find("div", class_="media news")
        if content_tag:
            # Lấy 100 ký tự đầu tiên để tìm ngày
            first_text = content_tag.get_text(strip=True)[:200]
            d = parse_vn_date_any(first_text)

    # 4. Fallback các meta tag cũ
    if d is None:
        d = parse_issue_date_from_module34(soup)
    if d is None:
        d = parse_public_date_from_uicongkhai(soup)
    if d is None:
        d = parse_date_from_meta(soup)

    return d


def parse_vn_date_from_soup1(soup: BeautifulSoup) -> date | None:
    """
    Tìm ngày ở các class phổ biến hoặc thẻ span post-date
    """
    candidates = [
        soup.find("span", class_="post-date"),
        soup.find("div", class_="PostDate"),
        soup.find("span", class_="date"),
        soup.find("p", class_="date"),
        soup.find("div", class_="creat_date"),
    ]

    for el in candidates:
        if el:
            text = clean_spaces(el.get_text(" ", strip=True))
            # Regex bắt dd/mm/yyyy
            m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", text)
            if m:
                d, mo, y = map(int, m.groups())
                try:
                    return date(y, mo, d)
                except:
                    pass
    return None


def to_int(text, default=None):
    s = (text or "").strip()
    m = re.search(r"\d+", s)  # lấy cụm số đầu tiên
    if not m:
        return default
    return int(m.group())


def get_base(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    p = urlparse(url)
    if not p.scheme or not p.netloc:
        # nếu Excel có URL thiếu scheme, ví dụ: thtienhiep.ninhbinh.edu.vn/...
        p = urlparse("https://" + url)
    return f"{p.scheme}://{p.netloc}/"


def clean_spaces(s: str) -> str:
    return " ".join((s or "").split()).strip()


def parse_vn_date_any(text: str) -> date | None:
    """
    Bắt ngày kiểu dd/mm/yyyy trong một chuỗi bất kỳ.
    """
    if not text:
        return None
    text = clean_spaces(text)
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", text)
    if not m:
        return None
    d, mo, y = map(int, m.groups())
    try:
        return date(y, mo, d)
    except:
        return None


def normalize_download_links_in_content(tag_content):
    """
    Chuyển:
      <a onclick="downloadFile('id','/upload/...whatever')">Text</a>
    thành:
      <a href="/upload/...whatever" class="link-download" target="_blank">Text</a>
    Không phụ thuộc đuôi file.
    """
    if not tag_content:
        return

    for a in tag_content.find_all("a"):
        onclick = (a.get("onclick") or "").strip()
        if not onclick:
            continue

        # Bắt tham số thứ 2 của downloadFile: '/upload/...'
        # - hỗ trợ ' hoặc "
        # - không quan tâm đuôi file là gì
        m = re.search(
            r"downloadFile\s*\(\s*(['\"]).*?\1\s*,\s*(['\"])\s*(/upload/[^'\"\)\s]+)\s*\2\s*\)",
            onclick,
            flags=re.IGNORECASE,
        )
        if not m:
            continue

        file_path = m.group(3).strip()

        # set về chuẩn a href
        a.attrs.pop("onclick", None)
        a["href"] = file_path
        a["target"] = "_blank"

        # class: giữ class cũ + thêm link-download
        cls = a.get("class", [])
        if isinstance(cls, str):
            cls = cls.split()
        if "link-download" not in cls:
            cls.append("link-download")
        a["class"] = cls

        # tuỳ chọn: bỏ style nếu có
        a.attrs.pop("style", None)


def pick_detail_links(soup):
    links = []

    # UI cũ: td.tg-yw4l a
    links += soup.select("td.tg-yw4l a[href]")

    # UI mới: td a[title*='Xem chi tiết công khai']
    links += soup.select("td a[title*='Xem chi tiết công khai'][href]")
    print("links")
    # print(links)
    # (tuỳ chọn) nếu title có thể khác hoa/thường
    # BeautifulSoup CSS selector không hỗ trợ case-insensitive chuẩn,
    # nên bạn có thể lọc thêm bằng Python bên dưới.

    return links


def parse_issue_date_from_module34(soup: BeautifulSoup) -> date | None:
    """
    Tìm 'Ngày ban hành' trong bảng thuộc #module34.
    Mẫu hay gặp:
      <th>Ngày ban hành</th><td>27/11/2025</td>
    hoặc có thể nằm trong <td class="td-title">Ngày ban hành</td>
    """
    module = soup.find("div", id="module34")
    if not module:
        return None

    # ưu tiên th (hoặc td) có text "Ngày ban hành"
    label_cells = module.find_all(["th", "td"], string=True)
    for cell in label_cells:
        label = clean_spaces(cell.get_text(" ", strip=True)).lower()
        print(label)
        if "ngày ban hành" in label:
            # lấy ô kế tiếp: thường là <td>...</td>
            nxt = cell.find_next(["td", "th"])
            if nxt:
                d = parse_vn_date_any(nxt.get_text(" ", strip=True))
                if d:
                    return d

            # fallback: nếu cùng hàng <tr> có nhiều ô, tìm ô cuối
            tr = cell.find_parent("tr")
            if tr:
                d = parse_vn_date_any(tr.get_text(" ", strip=True))
                if d:
                    return d

    # fallback cuối: quét toàn module nếu có dd/mm/yyyy
    return parse_vn_date_any(module.get_text(" ", strip=True))


def parse_date_module34(soup):
    span = soup.select_one("article.news-detail-layout-type-2 span.post-date")
    if not span:
        return None

    text = clean_spaces(span.get_text())
    m = re.search(r"(\d{2}/\d{2}/\d{4})", text)
    if m:
        d, mth, y = map(int, m.group(1).split("/"))
        return date(y, mth, d)
    return None


def parse_public_media_date(soup):
    em = soup.select_one("em.date-time")
    if em:
        return parse_vn_date_any(em.get_text())
    return None


class VspProducts:
    def __init__(self, base, url, cat, target):
        self.base = base
        self.cat = cat
        self.target = target
        self.camlist = camob.Listcam()
        self.url_links = []
        self.menu_links = menulink.MenuLink()

        service = Service(CHROME_DRIVER_PATH)
        chrome_options = Options()
        chrome_options.binary_location = CHROME_BINARY_PATH
        # Thêm headless để chạy nhanh hơn nếu muốn
        # chrome_options.add_argument("--headless")
        self.driver = webdriver.Chrome(service=service, options=chrome_options)

        self.url = url
        self.cat_id = ""
        self.url_id = ""
        self.page_link = 1

    def reset_for_row(self):
        self.url_links = []
        self.menu_links = menulink.MenuLink()
        self.camlist = camob.Listcam()

    def get_data(self):
        self.url_links.append(self.url)
        while self.url_links:
            self.extract_data(self.url_links[0])
            self.url_links.pop(0)

    def close(self):
        try:
            self.driver.quit()
        except:
            pass

    def extract_data(self, o_url):
        time.sleep(1)

        # Kiểm tra xem URL này có phải là đang đọc chi tiết không
        chitiet = 0
        current_cam = None
        for cam in self.camlist.camobs:
            if cam.url == o_url:
                chitiet = 1
                current_cam = cam
                break

        try:
            self.driver.get(o_url)
            try:
                WebDriverWait(self.driver, 10).until(
                    lambda d: d.find_elements(By.CSS_SELECTOR, "div.content_news")
                    or d.find_elements(By.CSS_SELECTOR, "div.bancanbiet")
                    or d.find_elements(By.CSS_SELECTOR, "div.media.news")
                )
            except:
                pass  # Timeout thì cứ thử parse tiếp, có thể là page trống

            html = self.driver.page_source
            soup = BeautifulSoup(html, "html.parser")
        except Exception as e:
            print("Error loading page:", e)
            return

        if soup:
            # -----------------------------------------------------------
            # PHẦN A: XỬ LÝ CHI TIẾT BÀI VIẾT
            # -----------------------------------------------------------
            if chitiet == 1 and current_cam:
                print(f">> Đang xử lý chi tiết: {current_cam.name}")

                tag_content = None

                # =====================================================
                # CASE 1: Tin tức thường
                # =====================================================
                tag_content = soup.select_one(
                    "div.media.news#gioithieu_noidung"
                ) or soup.select_one("div.content_news")

                if tag_content:
                    print("   -> Cấu trúc: Tin tức thường")

                # =====================================================
                # CASE 2: Thông báo / Công khai
                # =====================================================
                if not tag_content:
                    title_tag = soup.select_one("p.title")
                    public_content = soup.select_one("div.media.news")

                    if title_tag and public_content:
                        print("   -> Cấu trúc: Thông báo / Công khai")
                        current_cam.name = clean_spaces(title_tag.get_text())
                        tag_content = public_content

                # bỏ link tải file nằm ngoài content (nếu có)
                if tag_content:
                    for a in soup.select(
                        "a[href*='.pdf'], a[href*='.doc'], a[href*='.xls']"
                    ):
                        if not tag_content.find(a):
                            a.decompose()

                # =====================================================
                # CASE 3: TIN TỨC HOẠT ĐỘNG – HANAM (action-news + view_img)
                # =====================================================
                if not tag_content:
                    root = soup.select_one("#left-content-modules div.action-news")
                    if root:
                        print("   -> Cấu trúc: Tin tức hoạt động (Hanam)")

                        # title
                        title_tag = root.select_one("a > p.title")
                        if title_tag:
                            current_cam.name = clean_spaces(title_tag.get_text())

                        # nội dung chính
                        view_img = root.select_one("div.view_img")
                        if view_img:
                            tag_content = view_img
                # =====================================================
                # CASE 2.5: TIN TỨC HANAM (_content/tintuc/detailnews)
                # =====================================================
                if not tag_content:
                    detail = soup.select_one("div.news-detail")
                    if detail:
                        print("   -> Cấu trúc: Tin tức Hanam (news-detail)")

                        # TITLE
                        h4 = detail.find("h4")
                        if h4:
                            current_cam.name = clean_spaces(h4.get_text())

                        tag_content = detail
                # =====================================================
                # CASE: news-detail-layout-type-2 (module34 – chuẩn VHV)
                # =====================================================
                if not tag_content:
                    article = soup.select_one("article.news-detail-layout-type-2")
                    if article:
                        print("   -> Cấu trúc: news-detail-layout-type-2")

                        # TITLE
                        h1 = article.select_one("h1.title-detail")
                        if h1:
                            current_cam.name = clean_spaces(h1.get_text())

                        # CONTENT (QUAN TRỌNG)
                        # Ưu tiên tìm div.content-detail bên trong block-core-a5
                        # Hoặc div.content-detail trực tiếp
                        content = article.select_one(
                            "div.block-core-a5 div.content-detail"
                        ) or article.select_one("div.content-detail")

                        if content:
                            tag_content = content
                        else:
                            # Fallback: lấy toàn bộ article-content nếu không tách nhỏ
                            tag_content = article.select_one("div.article-content")

                    d = parse_date_module34(soup) or parse_vn_date_from_soup(soup)
                    # remove module34 trash
                    if article:
                        for sel in [
                            "div.block_share",
                            "div.rating",
                            "div.network-share",
                            "div.author",
                            "div#audio34",
                            "div#audio20",  # New audio id
                            "div.social-connect",  # New social
                            "style",
                        ]:
                            for t in article.select(sel):
                                t.decompose()
                # =====================================================
                # CASE: CÔNG KHAI – content_news + gioithieu_noidung
                # =====================================================
                if not tag_content:
                    container = soup.select_one("div.content_news")
                    content = soup.select_one("div#gioithieu_noidung")

                    if container and content:
                        print("   -> Cấu trúc: Công khai (content_news)")

                        # TITLE
                        title_tag = container.select_one("div.title_news")
                        if title_tag:
                            current_cam.name = clean_spaces(title_tag.get_text())

                        tag_content = content
                    d = parse_public_media_date(soup) or parse_vn_date_from_soup(soup)
                # =====================================================
                # CASE MỚI: project-realty-detail (Article-News)
                # =====================================================
                if not tag_content:
                    project_detail = soup.select_one("article.project-realty-detail")
                    if project_detail:
                        print("   -> Cấu trúc: project-realty-detail")

                        # Title
                        h1 = project_detail.select_one("h1.post-title")
                        if h1:
                            current_cam.name = clean_spaces(h1.get_text())

                        # Date
                        time_tag = project_detail.select_one("div.post-image time")
                        if time_tag:
                            d = parse_vn_date_any(time_tag.get_text())

                        # Content
                        # Tìm div.content-detail nằm trong div.post-content
                        content_div = project_detail.select_one(
                            "div.post-content div.content-detail"
                        )
                        if content_div:
                            tag_content = content_div
                        else:
                            # Fallback lấy cả post-content nếu không thấy detail
                            tag_content = project_detail.select_one("div.post-content")

                        # Remove trash specific to this layout
                        if tag_content:
                            # Sử dụng select an toàn
                            trash_items = tag_content.select_one(
                                "div.content-label, div.social"
                            )
                            for bad in trash_items:
                                bad.decompose()

                # =====================================================
                # KHÔNG TÌM THẤY
                # =====================================================
                if not tag_content:
                    print("❌ Không tìm thấy nội dung bài viết (mọi case)")
                    return
                # 2. Date
                d = parse_vn_date_from_soup(soup)
                print("Date found:", d)

                if d and d < FROM_DATE:
                    print(f"Bài cũ ({d}) < {FROM_DATE}")

                current_cam.date_publish = d if d else None
                # =====================================================
                # CLEAN CONTENT
                # =====================================================
                # remove tác giả
                for t in tag_content.select("div.tac_gia_news"):
                    t.decompose()

                # 4. Normalize download links
                normalize_download_links_in_content(tag_content)

                # 5. Fix images
                for img in tag_content.find_all("img"):
                    src = img.get("src")
                    if not src:
                        continue

                    # Ghép domain nếu thiếu
                    full_src = urljoin(self.base, src)

                    # Encode path (giữ nguyên scheme + netloc)
                    parsed = urlparse(full_src)
                    encoded_path = quote(parsed.path)

                    # giữ query nếu có
                    encoded_src = parsed._replace(path=encoded_path).geturl()

                    img["src"] = encoded_src
                    img["style"] = "max-width:100%;height:auto;"

                # 6. Fix links
                for a in tag_content.find_all("a"):
                    href = a.get("href")
                    if href:
                        a["href"] = urljoin(self.base, href)
                        a["target"] = "_blank"

                # 7. Remove trash
                for t in tag_content.find_all(["script", "style"]):
                    t.decompose()

                # iframe: chỉ xoá iframe KHÔNG phải youtube
                for iframe in tag_content.find_all("iframe"):
                    src = iframe.get("src", "")
                    if "youtube.com" not in src and "youtu.be" not in src:
                        iframe.decompose()

                # 8. Save
                current_cam.description = str(tag_content)
                current_cam.short = self.target

                current_cam.display_info()
                hp.save_data_cam(self.url_id, current_cam)

                print("✅ Đã lưu bài viết")
                return

            # -----------------------------------------------------------
            # PHẦN B: XỬ LÝ DANH SÁCH (LIST VIEW) – HANAM.EDU.VN
            # -----------------------------------------------------------
            print(">> Đang quét danh sách bài viết...")
            count_new = 0
            # --- KIỂM TRA LOẠI 1: Cấu trúc Công khai (div.bancanbiet) ---
            bancanbiet_root = soup.select_one("div.bancanbiet")
            if bancanbiet_root:
                print("--> Phát hiện cấu trúc: Công khai / Bạn cần biết")
                items = bancanbiet_root.select("div.bancanbiet-item")
                print(f"--> Tìm thấy {len(items)} mục.")

                for item in items:
                    # Lấy link trong thẻ a bên trong col-xs-8 (phần text)
                    a_tag = item.select_one("div.col-xs-8 p.text-left a")
                    if not a_tag:
                        # Fallback: tìm a bất kỳ trong item
                        a_tag = item.select_one("a[href]")

                    if not a_tag:
                        continue

                    href = a_tag.get("href")
                    full_url = urljoin(self.base, href)

                    # Lấy tiêu đề
                    title = clean_spaces(a_tag.get("title"))
                    if not title:
                        title = clean_spaces(a_tag.get_text())

                    # Kiểm tra trùng và thêm vào queue
                    if not hp.check_cam_url(self.url_id, full_url, title):
                        continue

                    cam = camob.CameraObject(0, title, 0, full_url, "", self.cat_id)
                    if self.camlist.add_cam(cam):
                        self.url_links.append(full_url)
                        count_new += 1
            # --- KIỂM TRA LOẠI 2: Cấu trúc Tin tức thường (logic cũ) : content_news ---

            content_root = soup.select_one("div.content_news ul.media-list")
            if content_root:
                print("--> Phát hiện cấu trúc: Tin tức thường")
                for a in content_root.select("li.media > a.pull-left[href]"):
                    href = a["href"]
                    url = urljoin(self.base, href)

                    li = a.find_parent("li", class_="media")
                    title_tag = li.select_one("h4.title-content-new") if li else None

                    name = (
                        clean_spaces(title_tag.get_text())
                        if title_tag
                        else url.split("/")[-1]
                    )

                    if not hp.check_cam_url(self.url_id, url, name):
                        continue

                    cam = camob.CameraObject(0, name, 0, url, "", self.cat_id)
                    if self.camlist.add_cam(cam):
                        self.url_links.append(url)
                        count_new += 1
            else:
                print(
                    "❌ Không tìm thấy danh sách bài viết (bancanbiet hoặc content_news)"
                )

            # --- KIỂM TRA LOẠI 2.5: Cấu trúc listType10 (div.news-listType10) ---
            type10_items = soup.select(
                "section.Article-Detail-listType10 div.news-listType10"
            )
            if not type10_items:
                # Fallback: tìm trực tiếp div class base
                type10_items = soup.select("div.news-listType10")

            if type10_items:
                print(
                    f"--> Phát hiện cấu trúc: news-listType10 | {len(type10_items)} bài"
                )
                count_new = 0

                for item in type10_items:
                    # 1. Link + Title
                    # h2 > a
                    h2_a = item.select_one("div.title-news-listType10 h2 a")
                    if not h2_a:
                        continue

                    href = h2_a.get("href")
                    full_url = urljoin(self.base, href)
                    title = clean_spaces(h2_a.get("title") or h2_a.get_text())

                    if not title:
                        continue

                    # 2. Date: 19/01/26 (dd/mm/yy)
                    pub_date = None
                    time_span = item.select_one("span.time-news")
                    if time_span:
                        txt_time = clean_spaces(time_span.get_text())
                        # Chờ regex dd/mm/yy
                        # 19/01/26
                        m = re.search(
                            r"(\d{1,2})[/-](\d{1,2})[/-](\d{2})", txt_time
                        )  # 2 digit year
                        if m:
                            d, mo, y_short = map(int, m.groups())
                            # Giả sử 20xx
                            y = 2000 + y_short
                            try:
                                pub_date = date(y, mo, d)
                            except:
                                pass
                        else:
                            # Fallback standard
                            pub_date = parse_vn_date_any(txt_time)

                    # 3. Image
                    thumb = ""
                    img_tag = item.select_one("div.images-news img")
                    if img_tag:
                        src = img_tag.get("data-original") or img_tag.get("src")
                        thumb = urljoin(self.base, src)

                    # 4. Summary
                    excerpt = ""
                    brief = item.select_one("div.brief-news")
                    if brief:
                        excerpt = clean_spaces(brief.get_text())

                    # Check duplicate
                    if not hp.check_cam_url(self.url_id, full_url, title):
                        continue

                    cam = camob.CameraObject(0, title, 0, full_url, "", self.cat_id)
                    cam.date_publish = pub_date
                    cam.thumb = thumb
                    cam.short = excerpt

                    if self.camlist.add_cam(cam):
                        self.url_links.append(full_url)
                        count_new += 1

                print(f"--> Bài mới thêm vào queue (Type10): {count_new}")

            # --- KIỂM TRA LOẠI 3: Cấu trúc form 1 c1 Tin tức: DANH SÁCH KIỂU HANAM (div.list-item)

            list_root = soup.select("div#left-content-modules div.list-item")

            if list_root:
                print(
                    f"--> Phát hiện cấu trúc: list-item (Hanam) | {len(list_root)} bài"
                )
                count_new = 0

                for item in list_root:
                    a_tag = item.select_one(".news-item-name a[href]")
                    if not a_tag:
                        continue

                    href = a_tag.get("href")
                    full_url = urljoin(self.base, href)

                    title = clean_spaces(a_tag.get_text())
                    if not title:
                        continue

                    # ngày đăng
                    date_text = ""
                    date_tag = item.select_one("span.text-color")
                    if date_tag:
                        date_text = clean_spaces(date_tag.get_text())

                    pub_date = parse_vn_date_any(date_text)

                    # ảnh đại diện
                    img_url = ""
                    img_tag = item.select_one("div.col-xs-4 img[src]")
                    if img_tag:
                        img_url = urljoin(self.base, img_tag["src"])

                    # mô tả ngắn
                    excerpt = ""
                    desc_tag = item.select_one("div.col-xs-8 div p span")
                    if desc_tag:
                        excerpt = clean_spaces(desc_tag.get_text())

                    # tác giả (nếu cần)
                    author = ""
                    author_tag = item.select_one("p.text-right")
                    if author_tag:
                        author = clean_spaces(author_tag.get_text())

                    # check duplicate
                    if not hp.check_cam_url(self.url_id, full_url, title):
                        continue

                    cam = camob.CameraObject(0, title, 0, full_url, "", self.cat_id)

                    cam.date_publish = pub_date
                    cam.short = excerpt
                    cam.thumb = img_url
                    cam.author = author

                    if self.camlist.add_cam(cam):
                        self.url_links.append(full_url)
                        count_new += 1

                print(f"--> Bài mới thêm vào queue: {count_new}")

            # --- KIỂM TRA LOẠI 4: Cấu trúc form 1 c0 Tin tức: DANH SÁCH KIỂU HANAM (div.action-news)
            action_items = soup.select("div#left-content-modules div.action-news")
            if action_items:
                print(
                    f"--> Phát hiện cấu trúc: action-news (Hanam) | {len(action_items)} bài"
                )
                count_new = 0

                for item in action_items:
                    # link + title
                    a_tag = item.select_one("a[href] p.title")
                    if not a_tag:
                        continue
                    a = a_tag.find_parent("a")
                    href = a.get("href")
                    full_url = urljoin(self.base, href)

                    title = clean_spaces(a_tag.get_text())
                    if not title:
                        continue
                    # ngày đăng
                    pub_date = None
                    time_tag = item.select_one("p.time")
                    if time_tag:
                        pub_date = parse_vn_date_any(time_tag.get_text())

                    # ảnh đại diện
                    thumb = ""
                    img_tag = item.select_one("img[src]")
                    if img_tag:
                        thumb = urljoin(self.base, img_tag["src"])

                    # mô tả ngắn
                    excerpt = ""
                    desc_tag = item.select_one("p.text-content")
                    if desc_tag:
                        excerpt = clean_spaces(desc_tag.get_text())

                    # check duplicate
                    if not hp.check_cam_url(self.url_id, full_url, title):
                        continue
                    cam = camob.CameraObject(0, title, 0, full_url, "", self.cat_id)
                    cam.date_publish = pub_date
                    cam.short = excerpt
                    cam.thumb = thumb
                    if self.camlist.add_cam(cam):
                        self.url_links.append(full_url)
                        count_new += 1

                print(f"--> Bài mới thêm vào queue: {count_new}")
            # --- CASE 5 (QUAN TRỌNG): LIST DẠNG UL/LI (Bao gồm cả cungchuyenmuc) ---
            # Cấu trúc:
            # <div id="left-content-modules">
            #    ...
            #    <div class="cungchuyenmuc">...</div>
            #    <ul>
            #       <li> <a href="...">Title</a> 23/10/2025 </li>
            #    </ul>
            # </div>

            main_container = soup.select_one("div#left-content-modules")
            if main_container:
                # Lấy tất cả các thẻ UL trực tiếp hoặc nằm sau cungchuyenmuc
                all_uls = main_container.find_all("ul")

                for ul in all_uls:
                    # Bỏ qua UL phân trang (pagination)
                    if ul.find("li", class_="actived") or ul.find(
                        "a", text=re.compile(r"Trang|Next|Sau", re.I)
                    ):
                        continue
                    # Bỏ qua nếu UL nằm trong div.page
                    if ul.find_parent("div", class_="page"):
                        continue

                    lis = ul.find_all("li", recursive=False)
                    if not lis:
                        continue

                    print(f"--> Tìm thấy UL list ({len(lis)} item). Đang quét...")
                    for li in lis:
                        a = li.find("a", href=True)
                        if not a:
                            continue

                        full_url = urljoin(self.base, a["href"])
                        title = clean_spaces(a.get_text())

                        # --- LOGIC BẮT NGÀY BÊN CẠNH TIÊU ĐỀ ---
                        # Lấy toàn bộ text của LI
                        li_text = li.get_text(" ", strip=True)
                        # Loại bỏ text của tiêu đề để tránh nhiễu
                        remaining_text = li_text.replace(title, "")

                        # Regex tìm ngày ở phần còn lại
                        pub_date = parse_vn_date_any(remaining_text)

                        # Nếu tìm thấy ngày và cũ hơn mốc => Bỏ qua
                        if pub_date and pub_date < FROM_DATE:
                            continue

                        if hp.check_cam_url(self.url_id, full_url, title):
                            cam = camob.CameraObject(
                                0, title, 0, full_url, "", self.cat_id
                            )
                            cam.date_publish = pub_date  # Lưu ngày ngay từ list view

                            if self.camlist.add_cam(cam):
                                self.url_links.append(full_url)
                                count_new += 1

            print(f"--> Tổng cộng thêm {count_new} bài vào hàng đợi.")
            # =====================================================
            # CASE LIST: HANAM list-news-content (_content/tintuc)
            # =====================================================
            list_root = soup.select_one("div.row.list-news-content")
            if list_root:
                items = list_root.select("div.new-content")
                print(f"--> Phát hiện cấu trúc: list-news-content | {len(items)} bài")

                for item in items:
                    a = item.select_one("a.title[href]")
                    if not a:
                        continue

                    href = urljoin(self.base, a["href"])
                    title = clean_spaces(a.get_text())

                    desc = ""
                    p_desc = item.find("p")
                    if p_desc:
                        desc = clean_spaces(p_desc.get_text())

                    if not hp.check_cam_url(self.url_id, href, title):
                        continue

                    cam = camob.CameraObject(0, title, 0, href, "", self.cat_id)
                    cam.short = desc

                    if self.camlist.add_cam(cam):
                        self.url_links.append(href)
                        count_new += 1

            # =====================================================
            # CASE LIST: Cấu trúc form c1 khác (section.section-list)
            # =====================================================
            section = soup.select_one("section.section-list")
            if section:
                print("--> Phát hiện cấu trúc: section-list")
                items = section.select("div.item-article")
                for item in items:
                    a = item.select_one("a[href]")
                    if not a:
                        continue

                    href = urljoin(self.base, a["href"])
                    title = clean_spaces(a.get_text())

                    if not hp.check_cam_url(self.url_id, href, title):
                        continue

                    cam = camob.CameraObject(0, title, 0, href, "", self.cat_id)

                    if self.camlist.add_cam(cam):
                        self.url_links.append(href)
                        count_new += 1
            print(f"--> Tổng cộng thêm {count_new} bài vào hàng đợi.")

            # =====================================================
            # CASE LIST: CÔNG KHAI – aside.content-new (media-list)
            # =====================================================
            aside = soup.select_one("aside.content-new ul.media-list")
            if aside:
                items = aside.select("li.media")
                print(
                    f"--> Phát hiện cấu trúc: Công khai media-list | {len(items)} bài"
                )

                for li in items:
                    a = li.select_one("h4 a[href]")
                    if not a:
                        continue

                    href = urljoin(self.base, a["href"])
                    title = clean_spaces(a.get_text())

                    # ngày đăng
                    pub_date = None
                    date_tag = li.select_one("em.date-time")
                    if date_tag:
                        pub_date = parse_vn_date_any(date_tag.get_text())

                    if pub_date and pub_date < FROM_DATE:
                        continue

                    # ✅ LOGIC ĐÚNG
                    if hp.check_cam_url(self.url_id, href, title):
                        cam = camob.CameraObject(0, title, 0, href, "", self.cat_id)
                        cam.date_publish = pub_date

                        if self.camlist.add_cam(cam):
                            self.url_links.append(href)
                            count_new += 1

                print(f"--> Bài mới thêm vào queue: {count_new}")

            # =====================================================
            # CASE LIST MỚI: Article-Detail-listmeberpost (section#section33)
            # =====================================================
            section_member = soup.select_one(
                "section.section-list.Article-Detail-listmeberpost"
            )
            if section_member:
                items = section_member.select("article.item-block")
                print(
                    f"--> Phát hiện cấu trúc: Article-Detail-listmeberpost | {len(items)} bài"
                )

                for item in items:
                    # 1. Link & Title
                    a_tag = item.select_one("h2.entry-title a")
                    if not a_tag:
                        continue

                    href = urljoin(self.base, a_tag.get("href"))
                    title = clean_spaces(a_tag.get_text())

                    # 2. Date (Time tag hoặc span.date)
                    pub_date = None
                    time_tag = item.select_one("time.post-date")
                    if time_tag:
                        # Ưu tiên text hiển thị (26/12/2025)
                        pub_date = parse_vn_date_any(time_tag.get_text())
                        # Nếu không được, thử lấy từ attribute datetime
                        if not pub_date and time_tag.get("datetime"):
                            pub_date = parse_vn_date_any(time_tag.get("datetime"))

                    # Fallback nếu không có thẻ time
                    if not pub_date:
                        date_span = item.select_one("span.date")
                        if date_span:
                            pub_date = parse_vn_date_any(date_span.get_text())

                    # Check Date Limit
                    if pub_date and pub_date < FROM_DATE:
                        continue

                    # 3. Image
                    thumb = ""
                    img = item.select_one("figure.post-image img")
                    if img:
                        thumb = urljoin(self.base, img.get("src"))

                    # 4. Summary
                    summary = ""
                    content_div = item.select_one("div.post-content")
                    if content_div:
                        summary = clean_spaces(content_div.get_text())

                    # 5. Check DB & Add to Queue
                    if hp.check_cam_url(self.url_id, href, title):
                        cam = camob.CameraObject(0, title, 0, href, "", self.cat_id)
                        cam.date_publish = pub_date
                        cam.thumb = thumb
                        cam.short = summary

                        if self.camlist.add_cam(cam):
                            self.url_links.append(href)
                            count_new += 1

                print(f"--> Bài mới thêm vào queue: {count_new}")

            # -----------------------------------------------------------
            # PHẦN C: PHÂN TRANG (Pagination)
            # -----------------------------------------------------------
            # Chỉ next trang nếu tìm thấy bài mới, hoặc tùy logic của bạn
            # Cấu trúc pagination thường gặp: ul.pagination li a[rel="next"]

            next_btn = None
            try:
                # Cách 1: Tìm theo text "Trang sau" hoặc icon
                # next_btn = self.driver.find_element(By.XPATH, "//a[contains(text(),'Trang sau') or contains(text(),'Next')]")

                # Cách 2: Tìm theo class phổ biến trong Bootstrap (trang Hanam dùng Bootstrap)
                # ul.pagination > li > a[rel='next'] hoặc nút cuối cùng

                # CẬP NHẬT: Tìm nút next ở cả ul.pagination và div.pagination
                # Tìm tất cả thẻ a trong class pagination
                pagination_links = self.driver.find_elements(
                    By.CSS_SELECTOR, ".pagination a"
                )
                if pagination_links:
                    last_link = pagination_links[-1]
                    link_text = last_link.text.strip()
                    link_rel = last_link.get_attribute("rel") or ""

                    # Logic xác định nút Next:
                    # 1. Text chứa ">>" hoặc "Sau" hoặc "Next"
                    # 2. Attribute rel="next"
                    # 3. Hoặc đơn giản là nút cuối cùng nếu trang hiện tại < 5 (để test)

                    is_next = False
                    if ">>" in link_text or "Next" in link_rel or "Sau" in link_text:
                        is_next = True
                    # Logic cũ của bạn: bấm nút cuối nếu page < 5
                    elif last_link.get_attribute("href") and self.page_link < 5:
                        is_next = True

                    if is_next:
                        next_btn = last_link

                if next_btn:
                    print(">> Chuyển sang trang tiếp theo...")
                    self.page_link += 1
                    self.driver.execute_script("arguments[0].click();", next_btn)
                    time.sleep(3)
                else:
                    print(">> Không còn trang tiếp theo.")
            except Exception as ex:
                print("Lỗi phân trang hoặc hết trang:", ex)


def run_from_excel(
    excel_path: str,
    sheet_name: str = None,
    start_row: int = 2,
    sleep_between: float = 0.5,
    save_every: int = 1,  # lưu sau mỗi N dòng xử lý
):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # đọc header để map cột theo tên
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            headers[str(v).strip().lower()] = col

    required = ["idurl", "source", "target", "cat_id"]
    for k in required:
        if k not in headers:
            raise ValueError(
                f"Thiếu cột '{k}' trong Excel. Hiện có: {list(headers.keys())}"
            )

    # cột done (bắt buộc theo yêu cầu mới)
    if "done" not in headers:
        raise ValueError("Thiếu cột 'done' trong Excel (header phải là 'done').")

    done_col = headers["done"]

    bot = None
    processed_since_save = 0

    try:
        for r in range(start_row, ws.max_row + 1):
            idurl = ws.cell(r, headers["idurl"]).value
            source = ws.cell(r, headers["source"]).value
            target = ws.cell(r, headers["target"]).value
            cat_id = ws.cell(r, headers["cat_id"]).value
            done_v = ws.cell(r, done_col).value

            # nếu không có source thì bỏ qua
            if not source:
                continue

            # chỉ chạy khi done == 0 hoặc trống
            # (chấp nhận "0", 0, "", None)
            done_str = str(done_v).strip() if done_v is not None else ""
            if done_str == "1":
                # đã làm rồi
                continue

            idurl = str(idurl).strip() if idurl is not None else ""
            source = str(source).strip()
            target = str(target).strip() if target else ""
            cat_id = str(cat_id).strip() if cat_id else ""

            base = get_base(source)

            if bot is None:
                bot = VspProducts(base=base, url=source, cat=cat_id, target=target)

            bot.reset_for_row()
            bot.url = source
            bot.url_id = idurl
            bot.cat_id = cat_id
            bot.base = base
            bot.target = target
            print(f"\n=== ROW {r} | idurl={idurl} | cat_id={cat_id}")
            print(f"source={source}")
            print(f"target={target}")

            try:
                bot.get_data()

                # nếu chạy xong không lỗi -> đánh done = 1
                ws.cell(r, done_col).value = 1

            except Exception as e:
                # nếu lỗi thì KHÔNG đánh done = 1, để lần sau chạy lại
                print(f"[ERROR] ROW {r} => {e}")

            processed_since_save += 1
            if processed_since_save >= save_every:
                wb.save(excel_path)
                processed_since_save = 0

            time.sleep(sleep_between)
            # break

    finally:
        # lưu nốt nếu còn thay đổi chưa save
        try:
            wb.save(excel_path)
        except:
            pass

        if bot:
            bot.close()


# chạy
run_from_excel(
    r"D:\WORKSPACE_CODE\Projects\Web\folder\websiteninhbinh\danhsachweb - Copy.xlsx",
    sheet_name=None,
)
