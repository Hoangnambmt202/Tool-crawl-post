"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  WEB SCRAPER + DEBUG LOGGER  —  File tích hợp hoàn chỉnh                   ║
║                                                                              ║
║  PHẦN 1 · DebugLogger       — Logger màu sắc VERBOSE / COMPACT              ║
║  PHẦN 2 · SiteStructure     — Định nghĩa & tự detect cấu trúc website      ║
║  PHẦN 3 · SessionStats      — Thống kê toàn phiên scraping                 ║
║  PHẦN 4 · Utility functions — clean_spaces, to_int, get_base               ║
║  PHẦN 5 · Date parsers      — Pipeline parse ngày từ nhiều nguồn           ║
║  PHẦN 6 · HTML processors   — clean_html, normalize_links, attachments     ║
║  PHẦN 7 · VspProducts       — Scraper chính (Selenium)                     ║
║  PHẦN 8 · Excel runner      — Đọc Excel và xử lý từng dòng                ║
╚══════════════════════════════════════════════════════════════════════════════╝

  Chuyển chế độ log bất kỳ lúc nào:
      log.set_mode(verbose=False)   # COMPACT — 1 dòng / URL
      log.set_mode(verbose=True)    # VERBOSE — đầy đủ chi tiết (mặc định)
"""

# ══════════════════════════════════════════════════════════════════════════════
# IMPORTS
# ══════════════════════════════════════════════════════════════════════════════

import sys
import re
import copy
import time
import traceback
from datetime import datetime, date
from enum import Enum
from contextlib import contextmanager
from collections import defaultdict
from typing import Optional

# ── Suppress stdout từ helpers (hp) — tránh in "luu data", "save_data"... ──
import io as _io
@contextmanager
def _suppress_stdout():
    """Suppress bất kỳ print nào phát sinh trong block (dùng cho lời gọi hp.*)."""
    _old = sys.stdout
    sys.stdout = _io.StringIO()
    try:
        yield
    finally:
        sys.stdout = _old

from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from difflib import SequenceMatcher
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, WebDriverException

# ── Module nội bộ (giữ nguyên) ───────────────────────────────────────────────
import Objectlink as obl
import MenuLink as menulink
import CameraObject as camob
import helpers as hp
import json
import mysql.connector


# ══════════════════════════════════════════════════════════════════════════════
# PHẦN 1 · DEBUG LOGGER
# ══════════════════════════════════════════════════════════════════════════════

class Color:
    """ANSI color codes"""
    RESET          = '\033[0m'
    BOLD           = '\033[1m'
    DIM            = '\033[2m'

    BLACK          = '\033[30m'
    RED            = '\033[31m'
    GREEN          = '\033[32m'
    YELLOW         = '\033[33m'
    BLUE           = '\033[34m'
    MAGENTA        = '\033[35m'
    CYAN           = '\033[36m'
    WHITE          = '\033[37m'

    BRIGHT_BLACK   = '\033[90m'
    BRIGHT_RED     = '\033[91m'
    BRIGHT_GREEN   = '\033[92m'
    BRIGHT_YELLOW  = '\033[93m'
    BRIGHT_BLUE    = '\033[94m'
    BRIGHT_MAGENTA = '\033[95m'
    BRIGHT_CYAN    = '\033[96m'
    BRIGHT_WHITE   = '\033[97m'

    BG_BLACK   = '\033[40m'
    BG_RED     = '\033[41m'
    BG_GREEN   = '\033[42m'
    BG_YELLOW  = '\033[43m'
    BG_BLUE    = '\033[44m'
    BG_MAGENTA = '\033[45m'
    BG_CYAN    = '\033[46m'
    BG_WHITE   = '\033[47m'


class Icon:
    """Unicode icons cho các loại log"""
    SUCCESS      = '✓'
    ERROR        = '✗'
    WARNING      = '⚠'
    INFO         = 'ℹ'
    DEBUG        = '⚙'
    ARROW        = '→'
    DOUBLE_ARROW = '⇒'
    BULLET       = '•'
    CLOCK        = '⏱'
    DOWNLOAD     = '⬇'
    UPLOAD       = '⬆'
    LINK         = '🔗'
    PAGE         = '📄'
    DATABASE     = '💾'
    SEARCH       = '🔍'
    PARSE        = '📝'
    CLEAN        = '🧹'
    CHECK        = '☑'
    CROSS        = '☒'
    HOURGLASS    = '⌛'
    ROCKET       = '🚀'
    FIRE         = '🔥'
    STAR         = '⭐'
    FOLDER       = '📁'
    FILE         = '📋'
    DETECT       = '🧬'
    STATS        = '📊'
    UNKNOWN      = '❓'


class LogLevel(Enum):
    DEBUG    = (Color.BRIGHT_BLACK,            Icon.DEBUG,   'DEBUG')
    INFO     = (Color.BLUE,                    Icon.INFO,    'INFO ')
    SUCCESS  = (Color.GREEN,                   Icon.SUCCESS, 'SUCCS')
    WARNING  = (Color.YELLOW,                  Icon.WARNING, 'WARN ')
    ERROR    = (Color.RED,                     Icon.ERROR,   'ERROR')
    CRITICAL = (Color.BRIGHT_RED + Color.BOLD, Icon.FIRE,    'CRIT ')


class DebugLogger:
    """
    Logger cấu trúc hoá cho scraper — thiết kế lại toàn bộ.

    Output format:
      ─── DÒNG N/M  [cat]  src: ...  dst: ...
      ◆ LIST   url
        ⏳ Tải... ✓ 52KB → ArticleList(12) → trang 1/?
         1  [+] 🖼  Tiêu đề bài
                    https://...
        >> Kiểm tra độ phủ ...
        ✓ LIST xong  18 added  1 skipped  2 trang  3.4s

      ◆ DETAIL  Tiêu đề bài
        ⏳ Tải... ✓ 38KB → ArticleHeader → 📅 15/01/2025 → 🖼 2 ảnh  📎 1 file
        ✓ Lưu DB OK  1.2s

      ✓ DÒNG N XONG  18 bài đã lưu  1 lỗi  48.3s

    Các method cũ (debug/info/section/...) được giữ lại nhưng không in ra —
    để không phải sửa hàng trăm call-site nhỏ trong code.
    Warning/error/critical vẫn in để không bỏ sót lỗi thật.
    """

    def __init__(self,
                 show_timestamp: bool = False,
                 show_level: bool = False,
                 indent_level: int = 0,
                 verbose: bool = False):   # verbose=False → dùng format mới

        self.show_timestamp = show_timestamp
        self.show_level     = show_level
        self.indent_level   = indent_level
        self.verbose        = verbose

        self._section_stack: list[int] = []
        self._timer_start: Optional[datetime] = None
        self._on_progress_line: bool = False

        # ── State cho LIST đang xử lý ──────────────────────────────────────
        self._list_url          = ""
        self._list_pages: list  = []   # [(page, html_c, added, skipped, has_err)]
        self._list_cur_page     = 1
        self._list_page_html    = 0    # số item tìm được trong HTML trang này
        self._list_page_added   = 0
        self._list_page_skipped = 0
        self._list_total_added  = 0
        self._list_total_skip   = 0
        self._list_total_indb   = 0   # bài đã có trong DB (subset của skip)
        self._list_has_pg_err   = False
        # Tracking URL sets để phát hiện CMS loop pagination
        self._cur_page_urls:  set = set()   # URLs trang hiện tại
        self._prev_page_urls: set = set()   # URLs trang trước

    # ══════════════════════════════════════════════════════════════════════════
    # CÁC METHOD CŨ — giữ signature, phần lớn im lặng
    # ══════════════════════════════════════════════════════════════════════════

    def set_mode(self, verbose: bool):
        self.verbose = verbose

    def _flush_progress(self):
        if self._on_progress_line:
            print()
            self._on_progress_line = False

    def _p(self, msg: str):
        """Print thẳng, không buffer."""
        self._flush_progress()
        print(msg)

    # ── Suppressed (quá nhiều noise) ──────────────────────────────────────────
    def debug(self, message: str):   pass
    def info(self,  message: str):   pass
    def success(self, message: str): pass

    def section(self, title: str, color=None):
        self._section_stack.append(self.indent_level)
    def end_section(self):
        if self._section_stack:
            self.indent_level = self._section_stack.pop()
    def subsection(self, title: str):    pass
    def step(self, n: int, desc: str):   pass
    def url(self, label: str, u: str):   pass
    def key_value(self, k, v, c=None):   pass
    def html_tag(self, t, a="found"):    pass
    def parse_info(self, e, f, v=""):    pass
    def html_preview(self, tag, l=""):   pass
    def separator(self, *a, **kw):       pass
    def blank_line(self):                pass
    def highlight(self, msg, color=None):pass
    def begin_url(self, *a, **kw):       pass
    def end_url(self, *a, **kw):         pass
    def database(self, op, d=""):        pass
    def clean_operation(self, w, c):     pass
    def timer_start(self, l):            pass
    def timer_end(self, l): return 0.0
    def date_found(self, d, s=""):       pass
    def progress(self, cur, tot, item=""): pass

    def file_operation(self, operation: str, filepath: str, status="processing"):
        """Vẫn in vì liên quan đến save Excel."""
        self._flush_progress()
        icons  = {"processing": "⏳", "success": "✓", "error": "✗"}
        colors = {"processing": Color.YELLOW, "success": Color.GREEN, "error": Color.RED}
        c = colors.get(status, Color.WHITE)
        print(f"  {c}{icons.get(status,'')} {operation}:{Color.RESET} {filepath}")

    # ── Warning / Error / Critical vẫn in nổi bật ────────────────────────────

    def warning(self, message: str):
        self._p(f"  {Color.YELLOW}⚠ {message}{Color.RESET}")

    def error(self, message: str):
        self._p(f"  {Color.RED}✗ {message}{Color.RESET}")

    def critical(self, message: str):
        self._p(f"  {Color.BRIGHT_RED}{Color.BOLD}🔥 {message}{Color.RESET}")

    # ── detect_and_log — vẫn cần, trả về structure list ─────────────────────

    def detect_and_log(self, soup, url: str = "") -> "list[SiteStructureType]":
        results = _detector.detect(soup, url)
        if results[0] == SiteStructureType.UNKNOWN:
            self._p(f"  {Color.YELLOW}⚠ Structure UNKNOWN — không nhận diện được HTML{Color.RESET}")
        return results

    @contextmanager
    def task(self, label: str, color=None):
        yield

    # ══════════════════════════════════════════════════════════════════════════
    # FORMAT HELPER
    # ══════════════════════════════════════════════════════════════════════════

    @staticmethod
    def _trunc(s: str, n: int) -> str:
        return s[:n] + ("…" if len(s) > n else "")

    @staticmethod
    def _fmt_elapsed(secs: float) -> str:
        if secs < 60: return f"{secs:.1f}s"
        m, s = divmod(int(secs), 60)
        return f"{m}m {s}s"

    # ══════════════════════════════════════════════════════════════════════════
    # EXCEL ROW HEADER / FOOTER
    # ══════════════════════════════════════════════════════════════════════════

    def row_start(self, row: int, total: int,
                  source: str = "", cat: str = "", target: str = ""):
        """Header mỗi dòng Excel — 1 dòng kẻ trên, không có kẻ dưới."""
        self._flush_progress()
        self.indent_level   = 0
        self._section_stack = []
        sep = f"{Color.BRIGHT_BLACK}{'─' * 72}{Color.RESET}"
        print(f"\n{sep}")
        cat_s = f"  {Color.YELLOW}[{cat}]{Color.RESET}" if cat else ""
        print(f"  {Color.BOLD}{Color.BRIGHT_WHITE}DÒNG {row}/{total}{Color.RESET}{cat_s}")
        if source:
            print(f"  {Color.BRIGHT_BLACK}src:{Color.RESET} {Color.CYAN}{self._trunc(source, 66)}{Color.RESET}")
        if target:
            print(f"  {Color.BRIGHT_BLACK}dst:{Color.RESET} {Color.GREEN}{self._trunc(target, 66)}{Color.RESET}")

    def row_done(self, row: int, saved: int, errors: int, elapsed: float):
        """Footer mỗi dòng Excel."""
        self._flush_progress()
        st = f"{Color.YELLOW}⚠{Color.RESET}" if errors else f"{Color.GREEN}✓{Color.RESET}"
        err_s = f"  {Color.RED}{errors} lỗi{Color.RESET}" if errors else ""
        print(f"\n{st} DÒNG {row} XONG"
              f"  {Color.GREEN}{saved} bài đã lưu{Color.RESET}"
              f"{err_s}"
              f"  {Color.BRIGHT_BLACK}{self._fmt_elapsed(elapsed)}{Color.RESET}")

    # ══════════════════════════════════════════════════════════════════════════
    # LIST URL — tracking theo từng trang
    # ══════════════════════════════════════════════════════════════════════════

    def list_start(self, url: str):
        """Bắt đầu LIST URL — in header + ⏳ Tải... (chưa xuống dòng)."""
        self._flush_progress()
        self._list_url               = url
        self._list_pages             = []
        self._list_cur_page          = 1
        self._list_page_html         = 0
        self._list_page_added        = 0
        self._list_page_skipped      = 0
        self._list_total_added       = 0
        self._list_total_skip        = 0
        self._list_total_indb        = 0
        self._list_has_pg_err        = False
        self._page_already_recorded  = False   # tránh double-append khi trang lỗi
        self._cur_page_urls          = set()
        self._prev_page_urls         = set()
        print(f"\n◆ {Color.BLUE}LIST{Color.RESET}   {Color.BRIGHT_BLACK}{self._trunc(url, 66)}{Color.RESET}")
        print(f"  {Color.YELLOW}⏳ Tải...{Color.RESET}", end="", flush=True)

    def list_page_loaded(self, kb: float, structure_name: str,
                         item_count: int, page: int, more_pages: bool = True,
                         total_pages: int = 0):
        """Complete inline: ✓ 52KB → ArticleList(12) → trang 1/10"""
        self._list_cur_page     = page
        self._list_page_html    = item_count
        self._list_page_added   = 0
        self._list_page_skipped = 0
        self._cur_page_urls     = set()   # reset tập URL cho trang mới
        # Hiển thị tổng trang cố định: nếu biết total → luôn hiện total
        # (không dùng str(page) nữa để tránh "trang 5/5, 6/6..." khi CMS loop)
        if total_pages > 0:
            more_s = str(total_pages)
        else:
            more_s = "?"
        print(f" {Color.GREEN}✓ {kb:.0f}KB{Color.RESET}"
              f" → {Color.CYAN}{structure_name}({item_count}){Color.RESET}"
              f" → trang {page}/{more_s}", flush=True)

    def list_page_error(self, page: int, error: str, kb: float = 0):
        """Trang load thất bại."""
        self._list_has_pg_err       = True
        self._page_already_recorded = True   # ngăn list_coverage_check append lần nữa
        self._list_pages.append((page, 0, 0, 0, True, error[:60]))
        kb_s = f" {kb:.0f}KB" if kb else ""
        print(f" {Color.RED}✗{kb_s} Trang {page}: {error[:60]}{Color.RESET}", flush=True)

    def list_item(self, idx: int, total: int, title: str, url: str,
                  result: str, has_img: bool = False,
                  skip_reason: str = ""):
        """1 item tìm được trong listing."""
        _icons = {
            "added":            f"{Color.GREEN}[+]{Color.RESET}",
            "replaced":         f"{Color.YELLOW}[↺]{Color.RESET}",
            "kept_old":         f"{Color.BRIGHT_BLACK}[=]{Color.RESET}",
            "skipped_url":      f"{Color.BRIGHT_BLACK}[=]{Color.RESET}",
            "filtered_section": f"{Color.BRIGHT_BLACK}[~]{Color.RESET}",
            "in_db":            f"{Color.GREEN}[✓]{Color.RESET}",
        }
        _default_reasons = {
            "kept_old":         "giữ bài cũ (URL mới kém hơn)",
            "skipped_url":      "trùng URL",
            "filtered_section": "khác chuyên mục",
            "in_db":            "đã có trong DB",
        }
        st      = _icons.get(result, f"{Color.BRIGHT_BLACK}[?]{Color.RESET}")
        img_s   = f" {Color.CYAN}🖼{Color.RESET}" if has_img else "  "
        num_s   = f"{Color.BRIGHT_BLACK}{idx:>3}{Color.RESET}"
        reason  = skip_reason or _default_reasons.get(result, "")
        reason_s = (f"  {Color.BRIGHT_BLACK}← {reason}{Color.RESET}"
                    if reason else "")

        print(f"  {num_s}  {st}{img_s}  {self._trunc(title, 56)}{reason_s}")
        if result != "filtered_section":
            print(f"              {Color.BRIGHT_BLACK}{self._trunc(url, 64)}{Color.RESET}")

        # Cập nhật counter + URL tracking
        if result in ("added", "replaced"):
            self._list_page_added   += 1
            self._list_total_added  += 1
            self._cur_page_urls.add(url)
        elif result == "in_db":
            # Đã có trong DB → tính vào skip (để coverage diff = 0) nhưng track riêng
            self._list_page_skipped += 1
            self._list_total_skip   += 1
            self._list_total_indb   += 1
            self._cur_page_urls.add(url)
        elif result == "filtered_section":
            pass   # khác chuyên mục → không tính, không làm coverage diff tăng
        else:
            # skipped_url, kept_old
            self._list_page_skipped += 1
            self._list_total_skip   += 1
            self._cur_page_urls.add(url)

    def list_next_page(self, next_page: int):
        """Lưu data trang hiện tại, in ⏳ tải trang tiếp (chưa xuống dòng)."""
        self._list_pages.append((
            self._list_cur_page,
            self._list_page_html,
            self._list_page_added,
            self._list_page_skipped,
            False, ""
        ))
        self._prev_page_urls = self._cur_page_urls   # lưu để guard loop detection
        print(f"  {Color.YELLOW}⏳ Tải trang {next_page}...{Color.RESET}", end="", flush=True)

    def list_coverage_check(self, list_url: str = "", base: str = ""):
        """In block >> Kiểm tra độ phủ sau khi hết pagination."""
        # Lưu trang cuối — chỉ khi chưa được record bởi list_page_error
        if not self._page_already_recorded:
            self._list_pages.append((
                self._list_cur_page,
                self._list_page_html,
                self._list_page_added,
                self._list_page_skipped,
                self._list_has_pg_err, ""
            ))
        self._page_already_recorded = False   # reset cho trang tiếp
        total_html = sum(p[1] for p in self._list_pages)

        print(f"\n  {Color.BRIGHT_BLACK}>> Kiểm tra độ phủ{Color.RESET}")
        all_ok = True
        for pg, html_c, add_c, skip_c, err, err_msg in self._list_pages:
            if err:
                all_ok = False
                print(f"     Trang {pg} : {Color.RED}LỖI — {err_msg or 'không tải được'}{Color.RESET}")
            else:
                skip_s = (f"  {Color.BRIGHT_BLACK}{skip_c} bỏ qua{Color.RESET}"
                          if skip_c else "")
                print(f"     Trang {pg} : {Color.BRIGHT_BLACK}{html_c} item HTML{Color.RESET}"
                      f" → {Color.GREEN}+{add_c} added{Color.RESET}{skip_s}")

        # Tổng item được xử lý (không tính filtered_section vì đúng là khác mục)
        processed = self._list_total_added + self._list_total_skip
        diff = total_html - processed
        diff_s = (f"{Color.YELLOW}chênh {diff} — kiểm tra lại{Color.RESET}"
                  if diff > 0 else f"{Color.BRIGHT_BLACK}chênh 0{Color.RESET}")
        print(f"     Tổng HTML {total_html}"
              f"  →  +{self._list_total_added} added"
              f"  →  {diff_s}")

        if not all_ok:
            verdict = f"{Color.YELLOW}⚠ CHƯA CHẮC — có trang lỗi, có thể bỏ sót{Color.RESET}"
            is_warn = True
        elif diff > 0:
            verdict = f"{Color.YELLOW}⚠ CÓ THỂ BỎ SÓT — HTML > queue{Color.RESET}"
            is_warn = True
        else:
            verdict = f"{Color.GREEN}ĐẦY ĐỦ — không bỏ sót{Color.RESET}"
            is_warn = False
        print(f"     Đánh giá : {verdict}")

        # Ghi vào stats để hiển thị trong báo cáo
        if is_warn and list_url:
            stats.record_coverage_warning(
                url=list_url, base=base,
                html_total=total_html,
                added=self._list_total_added,
                diff=diff,
            )

    def list_done(self, elapsed: float):
        """Tóm tắt cuối LIST."""
        pg_count = len(self._list_pages)
        err_pg   = sum(1 for p in self._list_pages if p[4])
        st   = f"{Color.YELLOW}⚠{Color.RESET}" if err_pg else f"{Color.GREEN}✓{Color.RESET}"
        err_s = f"  {Color.RED}{err_pg} trang lỗi{Color.RESET}" if err_pg else ""
        # Phần skipped thuần (không tính in_db)
        pure_skip = self._list_total_skip - self._list_total_indb
        indb_s  = (f"  {Color.GREEN}{self._list_total_indb} in_db{Color.RESET}"
                   if self._list_total_indb else "")
        skip_s  = (f"  {Color.BRIGHT_BLACK}{pure_skip} skipped{Color.RESET}"
                   if pure_skip else "")
        print(f"\n  {st} LIST xong"
              f"  {Color.GREEN}{self._list_total_added} added{Color.RESET}"
              f"{indb_s}{skip_s}"
              f"  {pg_count} trang{err_s}"
              f"  {Color.BRIGHT_BLACK}{self._fmt_elapsed(elapsed)}{Color.RESET}")

    # ══════════════════════════════════════════════════════════════════════════
    # DETAIL URL
    # ══════════════════════════════════════════════════════════════════════════

    def detail_start(self, title: str, url: str,
                     idx: int = 0, total: int = 0):
        """Bắt đầu DETAIL — in header + ⏳ Tải..."""
        self._flush_progress()
        counter_s = ""
        if idx and total:
            counter_s = f"  {Color.BRIGHT_BLACK}[{idx}/{total}]{Color.RESET}"
        elif idx:
            counter_s = f"  {Color.BRIGHT_BLACK}[{idx}]{Color.RESET}"
        print(f"\n◆ {Color.MAGENTA}DETAIL{Color.RESET}  {self._trunc(title, 56)}{counter_s}")
        print(f"  {Color.YELLOW}⏳ Tải...{Color.RESET}", end="", flush=True)

    def detail_loaded(self, kb: float, structure_name: str,
                      date_str: str = "", img_count: int = 0, file_count: int = 0):
        """Complete inline sau khi đọc xong content."""
        parts = [f"{Color.GREEN}✓ {kb:.0f}KB{Color.RESET}",
                 f"{Color.CYAN}{structure_name}{Color.RESET}"]
        if date_str:
            parts.append(f"{Color.YELLOW}📅 {date_str}{Color.RESET}")
        if img_count:
            parts.append(f"🖼 {img_count} ảnh")
        if file_count:
            parts.append(f"📎 {file_count} file")
        if not img_count and not file_count and not date_str:
            parts.append(f"{Color.BRIGHT_BLACK}(không đính kèm){Color.RESET}")
        print(f" {' → '.join(parts)}", flush=True)

    def detail_load_error(self, error: str, elapsed: float = 0):
        """Detail load/parse thất bại."""
        print(f" {Color.RED}✗ {error[:70]}{Color.RESET}", flush=True)
        el_s = f"  {Color.BRIGHT_BLACK}{self._fmt_elapsed(elapsed)}{Color.RESET}" if elapsed else ""
        print(f"  {Color.RED}✗ Bỏ qua{Color.RESET}{el_s}")

    def detail_saved(self, elapsed: float):
        """Lưu DB thành công."""
        print(f"  {Color.GREEN}✓ Lưu DB OK{Color.RESET}"
              f"  {Color.BRIGHT_BLACK}{self._fmt_elapsed(elapsed)}{Color.RESET}")

    def detail_error(self, error: str, elapsed: float):
        """Lưu DB thất bại."""
        print(f"  {Color.RED}✗ Lỗi lưu: {error[:60]}{Color.RESET}"
              f"  {Color.BRIGHT_BLACK}{self._fmt_elapsed(elapsed)}{Color.RESET}")

    # ══════════════════════════════════════════════════════════════════════════
    # SITE SUMMARY
    # ══════════════════════════════════════════════════════════════════════════

    def site_summary(self, base: str, cat_stats: dict, failed: list):
        """Tổng kết 1 trường sau khi xử lý hết các dòng Excel của trường đó."""
        self._flush_progress()
        base_s = base.replace("https://","").replace("http://","").rstrip("/")
        print(f"\n== KẾT THÚC TRƯỜNG: {base_s} {'=' * max(0, 50-len(base_s))}")

        total_saved = sum(cat_stats.values())
        site_fails  = [f for f in failed if f.get("base") == base]

        # Bảng chuyên mục
        # Bài đã có trong DB của trường này
        indb_site = stats.in_db_count.get(base, 0)
        indb_note = (f"  {Color.GREEN}+{indb_site} đã có trong DB{Color.RESET}"
                     if indb_site else "")
        print(f"\n  {'Chuyên mục':<38} {'Lấy được':>8}   {'Lỗi':>4}")
        print(f"  {'─'*38} {'─'*8}   {'─'*4}")
        for cat, cnt in sorted(cat_stats.items(), key=lambda x: -x[1]):
            cat_err = sum(1 for f in site_fails if f.get("cat") == cat)
            err_s = f"{Color.RED}{cat_err:>4}{Color.RESET}" if cat_err else f"{'0':>4}"
            print(f"  {cat:<38} {Color.GREEN}{cnt:>8}{Color.RESET}   {err_s}")
        print(f"  {'─'*38} {'─'*8}   {'─'*4}")
        total_err = len(site_fails)
        err_tot_s = f"{Color.RED}{total_err:>4}{Color.RESET}" if total_err else f"{'0':>4}"
        print(f"  {'TỔNG':<38} {Color.BOLD}{total_saved:>8}{Color.RESET}   {err_tot_s}{indb_note}")

        # Danh sách lỗi của trường — tách list / article
        list_fails    = [f for f in site_fails if f.get("type") == "list"]
        article_fails = [f for f in site_fails if f.get("type") != "list"]

        if list_fails:
            print(f"\n  List không lấy được ({len(list_fails)}):")
            for f in list_fails:
                url_s    = f.get("url") or ""
                reason_s = self._trunc(f.get("reason") or "", 68)
                row_s    = f"  [Dòng {f['row']}]" if f.get("row") else ""
                print(f"  {Color.YELLOW}⊘{Color.RESET}{row_s}  {Color.BRIGHT_BLACK}{url_s}{Color.RESET}")
                print(f"     lý do : {Color.YELLOW}{reason_s}{Color.RESET}")

        if article_fails:
            print(f"\n  Bài viết không lấy được ({len(article_fails)}):")
            for f in article_fails:
                title_s  = self._trunc(f.get("title") or "[không rõ tiêu đề]", 60)
                url_s    = f.get("url") or ""
                reason_s = self._trunc(f.get("reason") or "", 68)
                row_s    = f"  [Dòng {f['row']}]" if f.get("row") else ""
                print(f"  {Color.RED}✗{Color.RESET}{row_s}  {title_s}")
                print(f"     {Color.BRIGHT_BLACK}{url_s}{Color.RESET}")
                print(f"     lý do : {Color.YELLOW}{reason_s}{Color.RESET}")

    # ══════════════════════════════════════════════════════════════════════════
    # SESSION SUMMARY (thay thế print_session_summary cũ)
    # ══════════════════════════════════════════════════════════════════════════

    def print_session_summary(self, stats: "SessionStats"):
        """Báo cáo toàn phiên — được gọi ở cuối run_from_excel()."""
        self._flush_progress()
        print(f"\n\n{'=' * 72}")
        print(f"  BÁO CÁO TOÀN PHIÊN")
        print(f"{'=' * 72}")

        # ── Bảng tổng hợp theo trường ─────────────────────────────────────────
        if stats.cat_saved:
            # Header: Trường | Tổng chuyên mục | Bài viết | Lấy được | Cảnh báo | Lỗi
            print(f"\n  {'Trường':<36} {'CM':>4} {'Bài viết':>9} {'Lấy được':>9} {'Cảnh báo':>9} {'Lỗi':>5}")
            print(f"  {'─'*36} {'─'*4} {'─'*9} {'─'*9} {'─'*9} {'─'*5}")
            grand_saved = grand_err = grand_cat = grand_warn = grand_seen = 0
            for base, cats in sorted(stats.cat_saved.items()):
                saved  = sum(cats.values())
                err    = sum(1 for f in stats.failed_articles if f.get("base") == base)
                # Bài viết = số bài tìm thấy trong listing (saved + bị lọc/skip)
                seen   = stats.total_seen.get(base, saved)
                # Cảnh báo = số URL listing bị ⚠ CÓ THỂ BỎ SÓT
                warn   = sum(1 for w in stats.coverage_warnings if w.get("base") == base)
                grand_saved += saved; grand_err += err; grand_cat += len(cats)
                grand_warn  += warn;  grand_seen += seen
                base_s = base.replace("https://","").replace("http://","").rstrip("/")[:35]
                err_s  = f"{Color.RED}{err:>5}{Color.RESET}" if err else f"{'0':>5}"
                warn_s = (f"{Color.YELLOW}{warn:>9}{Color.RESET}" if warn
                          else f"{'0':>9}")
                print(f"  {base_s:<36} {len(cats):>4} {seen:>9} "
                      f"{Color.GREEN}{saved:>9}{Color.RESET} "
                      f"{warn_s} {err_s}")
            print(f"  {'─'*36} {'─'*4} {'─'*9} {'─'*9} {'─'*9} {'─'*5}")
            warn_gs = (f"{Color.YELLOW}{grand_warn}{Color.RESET}" if grand_warn
                       else "0")
            err_gs  = f"{Color.RED}{grand_err}{Color.RESET}" if grand_err else "0"
            print(f"  {'TỔNG':<36} {grand_cat:>4} {grand_seen:>9} "
                  f"{Color.BOLD}{grand_saved:>9}{Color.RESET} "
                  f"{warn_gs:>9}     {err_gs}")

        # ── Thời gian ─────────────────────────────────────────────────────────
        total_saved = sum(sum(c.values()) for c in stats.cat_saved.values()) if stats.cat_saved else 0
        avg_s = f"  TB: {stats.total_time/total_saved:.1f}s/bài" if total_saved and hasattr(stats,"total_time") and stats.total_time else ""
        print(f"\n  Thời gian: {stats.elapsed()}{avg_s}")

        # ── Danh sách cảnh báo lấy bài không đầy đủ ──────────────────────────
        if stats.coverage_warnings:
            print(f"\n  {'─'*68}")
            print(f"  {Color.YELLOW}⚠ CẢNH BÁO — CÓ THỂ LẤY THIẾU BÀI  ({len(stats.coverage_warnings)} URL){Color.RESET}")
            print(f"  {'─'*68}")
            from collections import defaultdict as _dd2
            by_base_w: dict = _dd2(list)
            for w in stats.coverage_warnings:
                by_base_w[w.get("base","")].append(w)
            for base, items in sorted(by_base_w.items()):
                base_s = base.replace("https://","").replace("http://","").rstrip("/")
                print(f"\n    {Color.BOLD}{base_s}{Color.RESET}  ({len(items)} URL)")
                for w in items:
                    url_s = w.get("url","")
                    diff  = w.get("diff", 0)
                    html_t = w.get("html_total", 0)
                    added  = w.get("added", 0)
                    print(f"    {Color.YELLOW}⚠{Color.RESET}  {Color.CYAN}{url_s}{Color.RESET}")
                    print(f"       HTML phát hiện: {html_t}  →  Lưu được: {added}"
                          f"  {Color.YELLOW}(chênh {diff}){Color.RESET}")
            print(f"  {Color.YELLOW}→ Kiểm tra lại các URL trên, có thể cần chạy lại hoặc tăng MAX_LIST_PAGES{Color.RESET}")

        # ── Chi tiết lỗi ─────────────────────────────────────────────────────
        if stats.failed_articles:
            from collections import defaultdict as _dd

            # Tách list-level vs article-level
            list_fails    = [f for f in stats.failed_articles if f.get("type") == "list"]
            article_fails = [f for f in stats.failed_articles if f.get("type") != "list"]

            total_label = len(stats.failed_articles)
            print(f"\n  {'─'*68}")
            print(f"  CHI TIẾT LỖI  ({len(list_fails)} trang list + {len(article_fails)} bài viết = {total_label} mục)")
            print(f"  {'─'*68}")

            # ── Phần 1: List không tải được ──────────────────────────────────
            if list_fails:
                by_base_l: dict = _dd(list)
                for f in list_fails:
                    by_base_l[f.get("base", "")].append(f)
                print(f"\n  {Color.YELLOW}▸ TRANG DANH SÁCH KHÔNG TẢI ĐƯỢC"
                      f"  ({len(list_fails)} URL){Color.RESET}")
                for base, items in sorted(by_base_l.items()):
                    base_s = base.replace("https://","").replace("http://","").rstrip("/")
                    print(f"\n    {Color.BOLD}{base_s}{Color.RESET}  ({len(items)} URL)")
                    for item in items:
                        row_n    = item.get("row", 0)
                        row_lbl  = f"Dòng {row_n}" if row_n else "Dòng ?"
                        cat_lbl  = item.get("cat", "") or "?"
                        url_s    = item.get("url") or ""
                        reason_s = item.get("reason") or "lỗi không xác định"
                        print(f"    {Color.YELLOW}[{row_lbl}]{Color.RESET}"
                              f"  {Color.BRIGHT_BLACK}{cat_lbl}{Color.RESET}")
                        print(f"      {Color.RED}✗ List không tải được:{Color.RESET}"
                              f" {Color.CYAN}{url_s}{Color.RESET}")
                        print(f"        {Color.YELLOW}↳ {reason_s}{Color.RESET}")

            # ── Phần 2: Bài viết không lấy được ──────────────────────────────────
            if article_fails:
                by_base_a: dict = _dd(list)
                for f in article_fails:
                    by_base_a[f.get("base", "")].append(f)
                print(f"\n  {Color.RED}▸ BÀI VIẾT KHÔNG LẤY ĐƯỢC"
                      f"  ({len(article_fails)} bài){Color.RESET}")
                for base, items in sorted(by_base_a.items()):
                    base_s = base.replace("https://","").replace("http://","").rstrip("/")
                    print(f"\n    {Color.BOLD}{Color.BRIGHT_WHITE}▶ {base_s}{Color.RESET}"
                          f"  ({len(items)} bài)")
                    by_row_a: dict = _dd(list)
                    for item in items:
                        by_row_a[item.get("row", 0)].append(item)
                    for row_n, row_items in sorted(by_row_a.items()):
                        row_lbl = f"Dòng {row_n}" if row_n else "Dòng ?"
                        cat_lbl = row_items[0].get("cat", "") or ""
                        print(f"\n      {Color.YELLOW}[{row_lbl}]{Color.RESET}"
                              f"  {Color.BRIGHT_BLACK}{cat_lbl}{Color.RESET}")
                        for item in row_items:
                            title_s  = item.get("title") or "[không có tiêu đề]"
                            url_s    = item.get("url") or ""
                            reason_s = item.get("reason") or "lỗi không xác định"
                            print(f"      {Color.RED}✗{Color.RESET}  {title_s}")
                            print(f"         {Color.BRIGHT_BLACK}{url_s}{Color.RESET}")
                            print(f"         {Color.YELLOW}↳ {reason_s}{Color.RESET}")

        else:
            print(f"\n  {Color.GREEN}✓ Không có lỗi nào trong phiên này.{Color.RESET}")

        print(f"\n{'=' * 72}\n")


# ══════════════════════════════════════════════════════════════════════════════
# PHẦN 2 · SITE STRUCTURE — định nghĩa & tự detect
# ══════════════════════════════════════════════════════════════════════════════

class SiteStructureType(Enum):
    """
    Toàn bộ cấu trúc website đã biết.
    Thêm type mới vào đây + vào STRUCTURE_SELECTOR_MAP khi gặp website lạ.
    """
    # ── Trang CHI TIẾT ────────────────────────────────────────────────────────
    TYPE_TITLE_DETAIL       = "TYPE-1  · title-detail"           # h1.title-detail + div.content-detail
    TYPE_ARTICLE_HEADER     = "TYPE-2  · ArticleHeader"          # div.ArticleHeader + div.ArticleContent
    TYPE_ARTICLE_NEWS       = "TYPE-3  · Article-News"           # article.Article-News
    TYPE_NEWS_DETAIL        = "TYPE-4  · news-detail-default"    # section.news-detail-default
    TYPE_DOWNLOAD_DETAIL    = "TYPE-5  · download-detail"        # article.download-detail-layout-default
    TYPE_VIEW_DETAIL        = "TYPE-6  · view-detail"            # div.view-detail
    TYPE_ARTICLE_CONTENT    = "TYPE-7  · article-content"        # div.article-content
    TYPE_MODULE34           = "TYPE-8  · module34"               # div#module34 (văn bản pháp quy)
    TYPE_MODULE16           = "TYPE-9  · module16"               # div#module16
    TYPE_CONGKHAI_NGAN_SACH = "TYPE-10 · CongKhaiNganSach"       # div.UICongKhaiNganSach_Default
    TYPE_CTL_DIVCONTENT     = "TYPE-11 · ctl01_divContent"       # div#ctl01_divContent (cổng CP)
    TYPE_TABLE_LAYOUT       = "TYPE-12 · table.table"            # Bảng biểu
    TYPE_ENTRY_CONTENT      = "TYPE-13 · entry-content"          # div.entry-content (WordPress/CMS VN)
    TYPE_MAIN_CONTENT       = "TYPE-14 · main-content"           # div.main-content / div.content-main
    TYPE_BOX_CONTENT        = "TYPE-15 · box-content"            # div.box-content / div.page-content
    TYPE_CONTENT_INNER      = "TYPE-16 · content-inner"          # div.content-inner / div.detail-inner
    TYPE_DIV_CONTENT        = "TYPE-17 · div.content"            # div.content (generic CMS VN)
    TYPE_THUC_DON           = "TYPE-18 · thuc-don"               # Bảng thực đơn tuần (edu.vn)
    TYPE_IFRAME_DOC         = "TYPE-19 · iframe-doc"             # Tài liệu nhúng iframe / PDF viewer
    TYPE_ZONE_CONTENT       = "TYPE-20 · zone-content"           # div.zone-content (portal CP VN)
    TYPE_INLINE_DOWNLOAD    = "TYPE-21 · inline-download"        # div.inline-download (tải file nội tuyến)
    TYPE_VIDEO_EMBED        = "TYPE-22 · video-embed"            # Trang chứa video YouTube / video embed
    TYPE_ALBUM_GALLERY      = "TYPE-23 · container-album"        # Trang thư viện ảnh / album (hanam.edu.vn)
    TYPE_MEDIA_NEWS         = "TYPE-24 · media.news"             # div.media.news / div#gioithieu_noidung (hanam.edu.vn detail)

    # ── Trang DANH SÁCH ───────────────────────────────────────────────────────
    LIST_POST_ITEM          = "LIST-A  · post-item"              # div.post-item.row
    LIST_ARTICLE_LIST       = "LIST-B  · ArticleList"            # ul.ArticleList / ul.down-list
    LIST_POST_TITLE         = "LIST-C  · post-title"             # div.post-title
    LIST_TITLE_DOC          = "LIST-D  · title-documment"        # a.title-documment
    LIST_CONGKHAI_TABLE     = "LIST-E  · td.tg-yw4l"             # Bảng công khai ngân sách
    LIST_PHAL               = "LIST-F  · phal-list"              # ul.phal-list.row
    LIST_NEWS_CONTENT       = "LIST-G  · news-content"           # div.list-news > div.news-content (hanam.edu.vn)

    # ── Không nhận diện được ─────────────────────────────────────────────────
    UNKNOWN                 = "UNKNOWN · không nhận diện được"


# CSS selector tương ứng với từng loại cấu trúc.
# Thứ tự trong dict = thứ tự ưu tiên detect (detail trước, list sau).
STRUCTURE_SELECTOR_MAP: dict[SiteStructureType, str] = {
    SiteStructureType.TYPE_TITLE_DETAIL:       "h1.title-detail",
    SiteStructureType.TYPE_ARTICLE_HEADER:     "div.ArticleHeader",
    SiteStructureType.TYPE_ARTICLE_NEWS:       "article.Article-News",
    SiteStructureType.TYPE_NEWS_DETAIL:        "section.news-detail-default",
    SiteStructureType.TYPE_DOWNLOAD_DETAIL:    "article.download-detail-layout-default",
    SiteStructureType.TYPE_VIEW_DETAIL:        "div.view-detail",
    SiteStructureType.TYPE_ARTICLE_CONTENT:    "div.article-content",
    SiteStructureType.TYPE_MODULE34:           "div#module34",
    SiteStructureType.TYPE_MODULE16:           "div#module16",
    SiteStructureType.TYPE_CONGKHAI_NGAN_SACH: "div.UICongKhaiNganSach_Default",
    SiteStructureType.TYPE_CTL_DIVCONTENT:     "div#ctl01_divContent",
    SiteStructureType.TYPE_TABLE_LAYOUT:       "table.table",
    # ── Cấu trúc bổ sung — CMS Việt Nam / edu.vn ────────────────────────────
    SiteStructureType.TYPE_ENTRY_CONTENT:      "div.entry-content",
    SiteStructureType.TYPE_MAIN_CONTENT:       "div.main-content, div.content-main",
    SiteStructureType.TYPE_BOX_CONTENT:        "div.box-content, div.page-content",
    SiteStructureType.TYPE_CONTENT_INNER:      "div.content-inner, div.detail-inner",
    SiteStructureType.TYPE_DIV_CONTENT:        "div.content",
    SiteStructureType.TYPE_THUC_DON:           "div.thuc-don, table.thuc-don, div.menu-week",
    SiteStructureType.TYPE_IFRAME_DOC:         "iframe[src], div.iframe-wrapper iframe",
    SiteStructureType.TYPE_ZONE_CONTENT:       "div.zone-content, div.region-content, div.portlet-content",
    SiteStructureType.TYPE_INLINE_DOWNLOAD:    "div.inline-download, div.download-type4, ul.list-download.type4, div.listfile",
    SiteStructureType.TYPE_VIDEO_EMBED:        "iframe[src*='youtube.com'], iframe[src*='youtube-nocookie.com'], iframe[src*='youtu.be'], div.videoWrapper, div.video-embed, div.post-video",
    SiteStructureType.TYPE_ALBUM_GALLERY:      "div.container_album, ul.libs-images",
    SiteStructureType.TYPE_MEDIA_NEWS:         "div.media.news, div#gioithieu_noidung",
    SiteStructureType.LIST_POST_ITEM:          "div.post-item",
    SiteStructureType.LIST_ARTICLE_LIST:       "ul.ArticleList, ul.down-list, ul.phal-list",
    SiteStructureType.LIST_POST_TITLE:         "div.post-title, div.title-news-listType10, div.item-info",
    SiteStructureType.LIST_TITLE_DOC:          "a.title-documment",
    SiteStructureType.LIST_CONGKHAI_TABLE:     "td.tg-yw4l",
    SiteStructureType.LIST_PHAL:               "ul.phal-list",
    SiteStructureType.LIST_NEWS_CONTENT:       "div.list-news",
}


class SiteStructureDetector:
    """
    Nhận BeautifulSoup, trả về danh sách SiteStructureType tìm thấy.
    Dùng qua instance toàn cục `_detector` hoặc qua log.detect_and_log().
    """

    def detect(self, soup, url: str = "") -> list[SiteStructureType]:
        """
        Chạy qua tất cả selector.
        Trả về [SiteStructureType.UNKNOWN] nếu không match gì.
        """
        found: list[SiteStructureType] = []
        for struct_type, selector in STRUCTURE_SELECTOR_MAP.items():
            try:
                el = soup.select_one(selector)
                if el and len(el.get_text(strip=True)) > 10:
                    found.append(struct_type)
            except Exception:
                pass
        return found if found else [SiteStructureType.UNKNOWN]

    def primary(self, soup, url: str = "") -> SiteStructureType:
        return self.detect(soup, url)[0]


# Instance toàn cục — dùng nội bộ
_detector = SiteStructureDetector()


# ══════════════════════════════════════════════════════════════════════════════
# PHẦN 3 · SESSION STATISTICS
# ══════════════════════════════════════════════════════════════════════════════

class SessionStats:
    """Thu thập thống kê toàn phiên scraping."""

    def __init__(self):
        self.reset()

    def reset(self):
        self.total_urls: int                    = 0
        self.success_urls: int                  = 0
        self.error_urls: int                    = 0
        self.skipped_urls: int                  = 0
        self.unknown_count: int                 = 0
        self.structure_counts: dict[str, int]   = defaultdict(int)
        self.errors: list[tuple[str, str]]      = []   # (url, error_msg)
        self.unknown_urls: list[str]            = []
        self._start: datetime                   = datetime.now()
        # Mở rộng — theo trường & chuyên mục
        self.cat_saved: dict                    = defaultdict(lambda: defaultdict(int))
        # {base → {cat → saved_count}}
        self.failed_articles: list              = []
        # [{title, url, base, cat, reason}]
        self.total_time: float                  = 0.0  # tổng giây xử lý bài
        self.total_seen: dict                   = defaultdict(int)
        # {base → total articles seen (saved + skipped)}
        self.coverage_warnings: list            = []
        # [{url, base, html_total, added, diff}] — listing có thể bỏ sót bài
        self.in_db_count: dict                  = defaultdict(int)
        # {base: số bài đã có trong DB (in_db) trong phiên này}

    def record_coverage_warning(self, url: str, base: str,
                                html_total: int, added: int, diff: int):
        self.coverage_warnings.append({
            "url": url, "base": base,
            "html_total": html_total, "added": added, "diff": diff,
        })

    def record_saved(self, base: str, cat: str):
        """Ghi nhận 1 bài lưu thành công."""
        self.cat_saved[base][cat or "Chưa phân loại"] += 1
        self.total_seen[base] += 1

    def record_in_db(self, base: str):
        """Ghi nhận 1 bài đã tồn tại trong DB (không lưu lại, nhưng coi là xử lý xong)."""
        self.in_db_count[base] += 1

    def record_failed(self, title: str, url: str, base: str,
                      cat: str, reason: str, row: int = 0,
                      fail_type: str = "article"):
        """
        Ghi nhận 1 lỗi.
        fail_type: "list"    — list page không tải được (chưa biết bài nào)
                   "article" — bài viết cụ thể không lấy được
        """
        self.failed_articles.append({
            "title": title, "url": url,
            "base": base, "cat": cat or "Chưa phân loại",
            "reason": reason,
            "row": row,
            "type": fail_type,
        })

    def record_url(self, url: str,
                   structure: SiteStructureType,
                   success: bool = True,
                   error_msg: str = ""):
        self.total_urls += 1
        if success:
            self.success_urls += 1
        else:
            self.error_urls += 1
            if error_msg:
                self.errors.append((url, error_msg))

        if structure == SiteStructureType.UNKNOWN:
            self.unknown_count += 1
            self.unknown_urls.append(url)
        else:
            self.structure_counts[structure.value] += 1

    def record_skip(self):
        self.skipped_urls += 1

    def elapsed(self) -> str:
        secs   = (datetime.now() - self._start).total_seconds()
        m, s   = divmod(int(secs), 60)
        h, m   = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"


# ══════════════════════════════════════════════════════════════════════════════
# KHỞI TẠO GLOBAL  (dùng ngay trong toàn file)
# ══════════════════════════════════════════════════════════════════════════════

log   = DebugLogger(verbose=True)   # ← đổi thành False để bật COMPACT
stats = SessionStats()


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION  ← Xem và chỉnh sửa tại config.py
# ══════════════════════════════════════════════════════════════════════════════
from config import (
    CHROMEDRIVER_PATH, CHROME_BINARY,
    EXCEL_PATH, ERROR_LOG_FILE,
    FROM_DATE, DETAIL_DATE_CSS, MAX_LIST_PAGES,
    SHOW_CHROME_WINDOW,
)


# ── Ghi log lỗi ra file ──────────────────────────────────────────────────────

def _write_error_log(row: int, source: str, idurl: str, error: str):
    """
    Ghi 1 dòng lỗi vào ERROR_LOG_FILE.
    Định dạng: [timestamp]  ROW:<n>  IDURL:<id>  <url>  ERROR: <msg>
    Tạo file nếu chưa có. Bỏ qua nếu ERROR_LOG_FILE rỗng/None.
    """
    if not ERROR_LOG_FILE:
        return
    try:
        import os as _os
        _os.makedirs(_os.path.dirname(ERROR_LOG_FILE), exist_ok=True)
        ts  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # Cắt error message cho gọn — traceback dài có thể nhiễu
        err_short = error.replace("\n", " ").replace("\r", "")[:300]
        line = (f"[{ts}]  ROW:{row:<5}  IDURL:{str(idurl):<8}"
                f"  {source}\n"
                f"         ERROR: {err_short}\n")
        with open(ERROR_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line)
    except Exception as e:
        print(f"[FailLogger] Không thể ghi file log: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# PHẦN 4 · UTILITY FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def clean_spaces(s: str) -> str:
    return " ".join((s or "").split()).strip()


def to_int(text, default=None):
    s = (text or "").strip()
    m = re.search(r"\d+", s)
    if not m:
        return default
    return int(m.group())


def get_base(url: str) -> str:
    url = (url or "").strip()
    if not url:
        log.warning("get_base: URL rỗng")
        return ""
    p = urlparse(url)
    if not p.scheme or not p.netloc:
        p = urlparse("https://" + url)
    result = f"{p.scheme}://{p.netloc}/"
    log.debug(f"get_base → {result}")
    return result


# ── Danh sách từ khóa báo lỗi HTTP trong tiêu đề / nội dung trang ────────────
_HTTP_ERROR_PATTERNS = [
    # Tiêu đề / code HTTP
    (r"<title>[^<]*\b(403|404|500|502|503)\b[^<]*</title>",          "HTTP {code} (từ <title>)"),
    (r"<title>[^<]*(forbidden|access denied|not found|error)[^<]*</title>",
                                                                       "HTTP error (từ <title>)"),
    # Thông báo lỗi phổ biến trong body
    (r"\b403\s*[-–]?\s*forbidden\b",                                  "403 Forbidden"),
    (r"\baccess\s+denied\b",                                          "Access Denied"),
    (r"\bforbidden\b",                                                 "Forbidden"),
    (r"\b404\s*[-–]?\s*not\s+found\b",                               "404 Not Found"),
    (r"trang\s+không\s+tồn\s+tại",                                   "404 (VN: trang không tồn tại)"),
    (r"không\s+tìm\s+thấy\s+trang",                                  "404 (VN: không tìm thấy trang)"),
    (r"\b500\s*[-–]?\s*internal\s+server\s+error\b",                 "500 Internal Server Error"),
]
_MIN_REAL_PAGE_BYTES = 500   # trang thật luôn > 500 bytes


def _detect_silent_http_error(html: str, url: str = "") -> str | None:
    """
    Phát hiện các lỗi HTTP mà Selenium không throw exception:
      - Trang quá nhỏ (< _MIN_REAL_PAGE_BYTES bytes)  →  khả năng trang lỗi/trắng
      - <title> hoặc body chứa từ khóa 403/404/Forbidden/…

    Trả về chuỗi mô tả lỗi nếu phát hiện, None nếu bình thường.
    """
    if not html:
        return "Trang rỗng (0 bytes)"

    size = len(html)

    # Kiểm tra kích thước trước (nhanh)
    if size < _MIN_REAL_PAGE_BYTES:
        # Vẫn kiểm tra pattern để có message cụ thể hơn
        lower = html.lower()
        for pattern, label in _HTTP_ERROR_PATTERNS:
            if re.search(pattern, lower, re.IGNORECASE):
                return f"{label} — trang chỉ {size} bytes"
        return f"Trang quá nhỏ ({size} bytes) — nghi ngờ 403/trang lỗi"

    # Trang đủ lớn nhưng vẫn kiểm tra pattern trong 2 KB đầu (nhanh, tránh scan toàn bộ)
    head = html[:2048].lower()
    for pattern, label in _HTTP_ERROR_PATTERNS:
        if re.search(pattern, head, re.IGNORECASE):
            return f"{label} — phát hiện trong head trang"

    return None   # bình thường


# ══════════════════════════════════════════════════════════════════════════════
# PHẦN 5 · DATE PARSERS
# ══════════════════════════════════════════════════════════════════════════════

def parse_vn_date_any(text: str) -> date | None:
    """Bắt ngày kiểu dd/mm/yyyy HOẶC 'ngày dd tháng mm năm yyyy' trong chuỗi bất kỳ."""
    log.debug(f"parse_vn_date_any: '{text}'")
    text = clean_spaces(text)

    # Pattern 1: dd/mm/yyyy
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if m:
        d, mo, y = map(int, m.groups())
        try:
            result = date(y, mo, d)
            log.date_found(result.strftime("%d/%m/%Y"), "parse_vn_date_any (dd/mm/yyyy)")
            return result
        except Exception:
            pass

    # Pattern 2: ngày dd tháng mm năm yyyy
    m = re.search(r"ngày\s+(\d{1,2})\s+tháng\s+(\d{1,2})\s+năm\s+(\d{4})", text, re.IGNORECASE)
    if m:
        d, mo, y = map(int, m.groups())
        try:
            result = date(y, mo, d)
            log.date_found(result.strftime("%d/%m/%Y"), "parse_vn_date_any (text VN)")
            return result
        except Exception:
            pass

    # Pattern 3: yyyy-mm-dd (ISO)
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", text)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            result = date(y, mo, d)
            log.date_found(result.strftime("%d/%m/%Y"), "parse_vn_date_any (ISO)")
            return result
        except Exception:
            pass

    return None


def parse_vn_date_from_soup1(soup: BeautifulSoup) -> date | None:
    """Parse ngày từ DOM: DETAIL_DATE_CSS từ config, rồi các selector hardcode phổ biến."""
    log.subsection("Parsing date từ DOM element")

    # Ưu tiên: time[datetime] — chuẩn HTML5
    el = soup.find('time', attrs={"datetime": True})
    if el:
        log.parse_info("time[datetime]", True, el["datetime"])
        d = parse_vn_date_any(el["datetime"])
        if d:
            return d

    # Ưu tiên tiếp: DETAIL_DATE_CSS từ config.py (mặc định "span.post-date")
    # → dùng select_one để hỗ trợ CSS selector đầy đủ thay vì chỉ find()
    if DETAIL_DATE_CSS:
        try:
            el = soup.select_one(DETAIL_DATE_CSS)
            if el:
                text = clean_spaces(el.get_text(" ", strip=True))
                log.parse_info(f"config:{DETAIL_DATE_CSS}", True, text[:40])
                d = parse_vn_date_any(text)
                if d:
                    log.date_found(d.strftime("%d/%m/%Y"), f"config:{DETAIL_DATE_CSS}")
                    return d
        except Exception:
            pass   # CSS selector không hợp lệ → tiếp tục

    for selector, desc in [
        (lambda s: s.find('span', class_='post-date left'),   "span.post-date.left"),
        (lambda s: s.find('div',  class_='PostDate'),          "div.PostDate"),
        (lambda s: s.find('span', class_='post-date'),         "span.post-date"),
        (lambda s: s.find('span', class_='date'),              "span.date"),
        (lambda s: s.find('p',    class_='date'),              "p.date"),
        (lambda s: s.find('div',  class_='date'),              "div.date"),
        (lambda s: s.find('span', class_='news-date'),         "span.news-date"),
        (lambda s: s.find('div',  class_='publish-date'),      "div.publish-date"),
    ]:
        el = selector(soup)
        log.parse_info(desc, bool(el))
        if el:
            text = clean_spaces(el.get_text(" ", strip=True))
            log.debug(f"Date element text: '{text}'")
            d = parse_vn_date_any(text)
            if d:
                log.date_found(d.strftime("%d/%m/%Y"), desc)
                return d

    log.warning("parse_vn_date_from_soup1: không tìm thấy element ngày")
    return None


def parse_issue_date_from_module34(soup: BeautifulSoup) -> date | None:
    """Tìm 'Ngày ban hành' trong bảng thuộc div#module34."""
    log.subsection("Parsing 'Ngày ban hành' từ module34")

    module = soup.find("div", id="module34")
    log.parse_info("div#module34", bool(module))
    if not module:
        return None

    for cell in module.find_all(["th", "td"], string=True):
        label = clean_spaces(cell.get_text(" ", strip=True)).lower()
        if "ngày ban hành" in label:
            log.parse_info("Ngày ban hành cell", True)
            nxt = cell.find_next(["td", "th"])
            if nxt:
                d = parse_vn_date_any(nxt.get_text(" ", strip=True))
                if d:
                    return d
            tr = cell.find_parent("tr")
            if tr:
                d = parse_vn_date_any(tr.get_text(" ", strip=True))
                if d:
                    return d

    log.warning("parse_issue_date_from_module34: không tìm thấy 'Ngày ban hành'")
    return None


def parse_public_date_from_uicongkhai(soup: BeautifulSoup) -> date | None:
    """Tìm 'Ngày công bố' trong UICongKhaiNganSach_Default."""
    log.subsection("Parsing 'Ngày công bố' từ UICongKhaiNganSach")

    root = soup.find("div", class_="UICongKhaiNganSach_Default")
    log.parse_info("UICongKhaiNganSach_Default", bool(root))
    if not root:
        return None

    for tr in root.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) >= 2:
            label = clean_spaces(tds[0].get_text(" ", strip=True)).lower()
            if "ngày công bố" in label:
                log.parse_info("Ngày công bố", True)
                return parse_vn_date_any(tds[1].get_text(" ", strip=True))

    log.warning("parse_public_date_from_uicongkhai: không tìm thấy 'Ngày công bố'")
    return None


def parse_date_from_meta(soup: BeautifulSoup) -> date | None:
    """Fallback: lấy ngày từ meta tags."""
    log.subsection("Parsing date từ meta tags")
    if not soup:
        return None

    candidates = []

    for itemprop in ["dateCreated", "datePublished", "dateModified"]:
        tag = soup.find("meta", attrs={"itemprop": itemprop})
        found = bool(tag and tag.get("content"))
        log.parse_info(f"meta[itemprop={itemprop}]", found, tag["content"] if found else "")
        if found:
            candidates.append(tag["content"])

    for prop in ["article:published_time", "article:modified_time"]:
        tag = soup.find("meta", attrs={"property": prop})
        found = bool(tag and tag.get("content"))
        log.parse_info(f"meta[property={prop}]", found, tag["content"] if found else "")
        if found:
            candidates.append(tag["content"])

    for name in ["pubdate", "publishdate", "date", "dc.date"]:
        tag = soup.find("meta", attrs={"name": name})
        found = bool(tag and tag.get("content"))
        log.parse_info(f"meta[name={name}]", found, tag["content"] if found else "")
        if found:
            candidates.append(tag["content"])

    def _to_date(s: str) -> date | None:
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", (s or "").strip())
        if not m:
            return None
        y, mo, d = map(int, m.groups())
        try:
            return date(y, mo, d)
        except Exception:
            return None

    for s in candidates:
        d = _to_date(s)
        if d:
            log.date_found(d.strftime("%Y-%m-%d"), "meta tag")
            return d

    log.warning("parse_date_from_meta: không tìm thấy ngày hợp lệ trong meta tags")
    return None


def parse_vn_date_from_soup(soup: BeautifulSoup) -> date | None:
    """
    Pipeline parse ngày theo thứ tự ưu tiên:
      1. DOM element  →  2. module34  →  3. UICongKhai  →  4. Meta tags
    """
    log.section("DATE PARSING PIPELINE", Color.YELLOW)

    for fn, label in [
        (parse_vn_date_from_soup1,          "DOM element"),
        (parse_issue_date_from_module34,    "module34"),
        (parse_public_date_from_uicongkhai, "UICongKhai"),
        (parse_date_from_meta,              "meta tags"),
    ]:
        d = fn(soup)
        if d is not None:
            log.success(f"✓ Ngày tìm thấy từ {label}: {d}")
            log.end_section()
            return d

    log.error("✗ Không tìm thấy ngày từ bất kỳ nguồn nào")
    log.end_section()
    return None


# ══════════════════════════════════════════════════════════════════════════════
# PHẦN 6 · HTML PROCESSORS
# ══════════════════════════════════════════════════════════════════════════════

def normalize_download_links_in_content(tag_content, base_url: str = ""):
    """
    Chuẩn hoá link tải trong nội dung.
    base_url: URL tuyệt đối gốc của trang (để convert relative href → absolute).
    """
    log.subsection("Normalizing download links")
    if not tag_content:
        log.warning("normalize_download_links: không có content")
        return

    link_count = processed_count = 0

    for a in tag_content.find_all("a"):
        link_count += 1
        href    = (a.get("href") or "").strip()
        onclick = (a.get("onclick") or "").strip()
        file_path = None

        if onclick:
            # Lazy match cả 2 args (hỗ trợ Unicode + space trong tên file/label)
            # rồi chọn arg nào trông giống path/URL file
            _m2 = re.search(
                r"downloadFile\s*\(\s*(['\"])(.*?)\1\s*,\s*(['\"])(.*?)\3",
                onclick, flags=re.IGNORECASE | re.DOTALL
            )
            _m1 = re.search(
                r"downloadFile\s*\(\s*(['\"])(.*?)\1\s*\)",
                onclick, flags=re.IGNORECASE | re.DOTALL
            )
            if _m2:
                _a1 = _m2.group(2).strip()
                _a2 = _m2.group(4).strip()
                def _is_path(s):
                    sl = s.lower()
                    return (sl.startswith(('/', 'http', '../')) or
                            any(sl.endswith(e) for e in (
                                '.pdf','.doc','.docx','.xls','.xlsx',
                                '.ppt','.pptx','.zip','.rar','.7z',
                                '.txt','.odt','.ods','.mp4','.mp3')) or
                            '?' in sl)
                file_path = _a1 if _is_path(_a1) else (_a2 if _is_path(_a2) else _a1)
                log.success(f"  ✓ onclick arg1={_a1!r} arg2={_a2!r} → {file_path!r}")
            elif _m1:
                file_path = _m1.group(2).strip()
                log.success(f"  ✓ onclick 1-arg: {file_path!r}")

        if not file_path and href:
            if href.lower().startswith(("javascript:", "#", "mailto:", "tel:")):
                continue
            file_path = href

        if not file_path:
            continue

        # Convert relative → absolute
        if file_path and not file_path.startswith(('http://', 'https://', 'data:', 'javascript:')):
            if base_url:
                file_path = urljoin(base_url, file_path)

        a.attrs.pop("onclick", None)
        a["href"]   = file_path
        a["target"] = "_blank"
        cls = a.get("class", [])
        if isinstance(cls, str):
            cls = cls.split()
        if "link-download" not in cls:
            cls.append("link-download")
        a["class"] = cls
        a.attrs.pop("style", None)

        processed_count += 1

    log.info(f"Links: {processed_count}/{link_count} đã chuẩn hoá")


def normalize_img_srcs_in_content(tag_content, base_url: str = ""):
    """
    Convert tất cả img[src] relative → absolute URL trong content.
    Xử lý cả lazy-load (data-src, data-original) và inline style background.
    Gọi SAU normalize_download_links_in_content, TRƯỚC clean_html_content.
    """
    if not tag_content or not base_url:
        return
    for img in tag_content.find_all('img'):
        for attr in ('src', 'data-src', 'data-original', 'data-lazy', 'data-url'):
            val = (img.get(attr) or '').strip()
            if val and not val.startswith(('http://', 'https://', 'data:', '//')):
                img[attr] = urljoin(base_url, val)
            elif val and val.startswith('//'):
                img[attr] = 'https:' + val


def pick_detail_links(soup):
    """Tìm các link chi tiết công khai (2 kiểu UI cũ và mới)."""
    log.subsection("Picking detail links")
    links = []

    old_ui = soup.select("td.tg-yw4l a[href]")
    log.parse_info("td.tg-yw4l a[href]", bool(old_ui),
                   f"{len(old_ui)} found" if old_ui else "")
    links += old_ui

    new_ui = soup.select("td a[title*='Xem chi tiết công khai'][href]")
    log.parse_info("td a[title*='Xem chi tiết']", bool(new_ui),
                   f"{len(new_ui)} found" if new_ui else "")
    links += new_ui

    log.info(f"Tổng detail links: {len(links)}")
    return links



def _extract_doc_url_from_embed(iframe_src):
    """Trích URL tài liệu từ Google Drive / Docs / Scribd iframe embed."""
    if not iframe_src:
        return ""
    s = iframe_src
    # Google Docs Viewer: docs.google.com/viewer?url=ENCODED_URL
    m = re.search(r"docs[.]google[.]com/viewer[?].*?url=([^&\s\"']+)", s, re.I)
    if m:
        try:
            from urllib.parse import unquote
            return unquote(m.group(1))
        except Exception:
            return m.group(1)
    # Google Drive: drive.google.com/file/d/ID/preview
    m = re.search(r"drive[.]google[.]com/file/d/([A-Za-z0-9_-]+)", s, re.I)
    if m:
        return "https://drive.google.com/uc?id={}&export=download".format(m.group(1))
    # Google Docs/Sheets/Slides
    m = re.search(r"docs[.]google[.]com/(?:document|spreadsheets|presentation)/d/([A-Za-z0-9_-]+)", s, re.I)
    if m:
        return "https://docs.google.com/uc?id={}&export=download".format(m.group(1))
    # Scribd embed
    m = re.search(r"scribd[.]com/embeds/(\d+)", s, re.I)
    if m:
        return "https://www.scribd.com/doc/{}".format(m.group(1))
    return ""


def clean_html_content(tag_content):
    """
    Làm sạch HTML: xóa script/style/rác, giữ YouTube iframe,
    strip inline style, giữ href cho <a>, src cho <img>.
    """
    log.section("CLEANING HTML CONTENT", Color.MAGENTA)
    if not tag_content:
        return

    log.step(1, "Xóa các element rác")

    # ── Bước 0: Xóa script/style; noscript → unwrap (giữ img lazy-load) ───
    for _stag in list(tag_content.find_all(['script', 'style'])):
        _stag.decompose()
    for _ns in list(tag_content.find_all('noscript')):
        if _ns.find('img'):
            _ns.unwrap()   # lazy-load: giữ lại <img> bên trong
        else:
            _ns.decompose()

    # ── Sidebar/nav dư: xóa trước khi làm gì khác ──────────────────────────
    JUNK_SELECTORS = [
        'div.sidebar', 'aside', 'nav', 'div[id*="sidebar"]',
        'div[class*="sidebar"]', 'div[class*="side-bar"]',
        'div[class*="widget"]', 'div.related-news', 'div.related-posts',
        'div.tags-list', 'div.share-social', 'div.social-share',
        'div.bread-crumb', 'div.breadcrumb', 'ol.breadcrumb',
        'div.author-box', 'div.comment-area', 'div#comments',
        'div.print-button', 'div.rating-box',
    ]
    for _jsel in JUNK_SELECTORS:
        for _jtag in tag_content.select(_jsel):
            _jtag.decompose()

    specific_garbage = [
        '.network-share', '.fb-share-button',
        '.button-bookmark', 'div[id^="audio"]',
        # module34 / news-detail-layout-type-2 — rác không liên quan bài viết
        'h1.title-detail',          # tiêu đề đã có WP title riêng, không lặp trong body
        'div.social-connect',       # Zalo / FB share
        'div.block-core-a3',        # wrapper header: tiêu đề + social (xóa hết nếu còn sót)
        'div.rating',               # star-rating widget
        'div#star-rating',
        'div[id^="stringrating"]',  # "Tổng số điểm..."
        'div.clearfix.mt-10',       # lượt xem + tác giả (chân bài)
        'div.author',               # "Lượt xem: N", "Tác giả: ..."
        'div.block_share',          # nút in, font-size, email
        'span.post-date',           # ngày đăng — đã lưu riêng vào cam.date_publish
        'span.drash',
        # hanam.edu.vn media.news — rác cuối bài
        'div.tac_gia_news',         # "Tác giả: thathanhson"
    ]
    for selector in specific_garbage:
        for tag in tag_content.select(selector):
            tag.decompose()
            log.debug(f"  Đã xóa: {selector}")

    log.step(2, "Strip inline styles & xử lý iframe")
    for tag in tag_content.find_all(True):
        # Guard: một số BS4 tag có attrs=None (Comment, ProcessingInstruction...)
        if not isinstance(getattr(tag, 'attrs', None), dict):
            continue

        if tag.name == 'iframe':
            src = tag.get('src', '')
            # Giữ lại mọi dạng YouTube iframe kể cả youtube-nocookie.com
            is_youtube = any(d in src for d in (
                'youtube.com', 'youtube-nocookie.com', 'youtu.be'
            ))
            if is_youtube:
                tag.attrs = {
                    'src': src, 'width': '100%', 'height': '450',
                    'frameborder': '0', 'allowfullscreen': 'true',
                    'style': 'display:block; margin:10px 0;',
                }
                log.info(f"  ✓ Giữ YouTube iframe: {src[:60]}")
            else:
                # Thử chuyển embed document (GDrive/GDocs Viewer/Scribd) → link tải
                _dl = _extract_doc_url_from_embed(src)
                if _dl:
                    from bs4 import Tag
                    _a = tag.find_parent().__class__.__new__(tag.find_parent().__class__)
                    _atag = tag.parent.find_parent() if False else None
                    # Dùng NavigableString wrapper
                    import copy as _cp
                    _new_soup_str = ('<a href="{}" target="_blank" class="link-download">'
                                     '&#128196; Xem / Tải tài liệu đính kèm</a>').format(_dl)
                    from bs4 import BeautifulSoup as _BSX
                    _a_node = _BSX(_new_soup_str, 'html.parser').find('a')
                    tag.replace_with(_a_node)
                    log.info("  ✓ Iframe embed → link: {}".format(_dl[:80]))
                else:
                    tag.decompose()
            continue

        if tag.has_attr('style') and tag.name != 'iframe':
            del tag['style']

        if tag.name == 'a':
            href  = tag.get('href')
            cls   = tag.get('class', [])
            # Giữ class link-download để 2_Dang_bai nhận diện được file link
            keep_cls = [c for c in cls if c == 'link-download'] if cls else []
            tag.attrs = {}
            if href:
                tag['href'] = href
            tag['target'] = '_blank'
            if keep_cls:
                tag['class'] = keep_cls

        elif tag.name == 'img':
            src = tag.get('src')
            tag.attrs = {}
            if src:
                tag['src'] = src
            tag['style'] = "max-width:100%; height:auto;"

    log.success("✓ HTML cleaning hoàn tất")



def convert_rg_gallery_to_imgs(tag_content, base_url: str = ""):
    """
    Chuyển #rg-gallery (thư viện ảnh CMS edu.vn) → danh sách <img> thường.
    Mỗi thumbnail <img data-large="..."> → 1 <img src="URL_ảnh_gốc">.
    Xóa toàn bộ widget gallery, thay bằng ảnh thật để có thể upload.
    """
    gallery = tag_content.find('div', id='rg-gallery')
    if not gallery:
        gallery = tag_content.select_one('div.rg-gallery')
    if not gallery:
        return 0

    imgs_out = []
    for li in gallery.select('div.rg-thumbs li'):
        img = li.find('img')
        if not img:
            continue
        large = img.get('data-large', '').strip()
        if not large:
            # Fallback: dùng href của <a> chứa img
            a_parent = img.find_parent('a')
            large = (a_parent.get('href', '') if a_parent else '')
        if not large or 'javascript' in large:
            continue
        # Chuẩn hoá URL
        if not large.startswith('http'):
            large = urljoin(base_url, large) if base_url else large
        if 'no-image' in large.lower():
            continue
        from bs4 import BeautifulSoup as _BSG
        new_img = _BSG(
            f'<img src="{large}" style="max-width:100%;height:auto;margin:4px 2px;">',
            'html.parser'
        ).find('img')
        if new_img:
            imgs_out.append(new_img)

    if not imgs_out:
        gallery.decompose()
        return 0

    # Tạo wrapper thay thế gallery
    from bs4 import BeautifulSoup as _BS
    wrapper_html = '<div class="gallery-imgs" style="display:flex;flex-wrap:wrap;gap:4px;margin:10px 0;">' +         ''.join(str(i) for i in imgs_out) + '</div>'
    wrapper = _BS(wrapper_html, 'html.parser').find('div')
    gallery.replace_with(wrapper)
    log.info(f"  🖼 rg-gallery → {len(imgs_out)} ảnh")
    return len(imgs_out)

def find_and_merge_attachments(soup, main_content_tag):
    """
    Tìm file đính kèm và link YouTube nằm ngoài nội dung chính,
    chèn vào đầu main_content_tag.
    """
    log.subsection("Scanning for external attachments & YouTube links")

    keywords       = ['tập tin đính kèm', 'tải về', 'download', 'file đính kèm',
                      'văn bản đính kèm', 'video', 'clip', 'youtube']
    file_extensions = ['.pdf', '.doc', '.docx', '.xls', '.xlsx', '.zip', '.rar']
    found_links    = []
    found_youtube  = []

    # Cache set identity của link/iframe trong content để tránh O(n²)
    _content_link_ids   = {id(x) for x in main_content_tag.find_all('a')}
    _content_iframe_ids = {id(x) for x in main_content_tag.find_all('iframe')}

    # Scan iframe embed ngoài content → extract doc URL (GDrive/GDocs/Scribd)
    for _ifr in soup.find_all('iframe', src=True):
        if id(_ifr) in _content_iframe_ids:
            continue
        _idoc = _extract_doc_url_from_embed(_ifr.get('src', ''))
        if _idoc:
            from bs4 import BeautifulSoup as _BSF
            _ia = _BSF('<a href="{}" target="_blank" class="link-download">📄 Tài liệu đính kèm</a>'.format(_idoc), 'html.parser').find('a')
            found_links.append(_ia)
            log.info("  Found iframe doc: {}".format(_idoc[:80]))

    for a in soup.find_all('a', href=True):
        if id(a) in _content_link_ids:
            continue
        href    = a['href']
        href_l  = href.lower()
        text    = a.get_text(" ", strip=True).lower()
        title   = (a.get('title') or "").lower()
        onclick = (a.get('onclick') or "").lower()

        # ── YouTube link ngoài content ─────────────────────────────────────
        vid = _yt_video_id(href)
        if vid:
            found_youtube.append(vid)
            log.info(f"  Found external YouTube: {vid}")
            continue

        # ── File đính kèm thông thường ─────────────────────────────────────
        is_file = (any(ext in href_l for ext in file_extensions)
                   or any(kw in text or kw in title for kw in keywords)
                   or 'downloadfile' in onclick)

        if is_file:
            new_a = copy.copy(a)
            new_a.string = f" [Tải về: {a.get_text(strip=True) or 'Tập tin'}] "
            new_a['class'] = ['link-download']   # để 2_Dang_bai nhận diện là file
            new_a.attrs.pop('style', None)
            found_links.append(new_a)
            log.info(f"  Found external attachment: {href_l}")

    # ── Chèn phần YouTube embed ────────────────────────────────────────────
    if found_youtube:
        yt_div = soup.new_tag('div', style="margin-bottom:20px;")
        yt_hdr = soup.new_tag('strong')
        yt_hdr.string = "🎬 Video liên quan:"
        yt_div.append(yt_hdr)
        yt_div.append(soup.new_tag('br'))
        for vid in found_youtube:
            yt_div.append(_make_yt_iframe(vid, soup))
        main_content_tag.insert(0, yt_div)
        log.success(f"✓ Merged {len(found_youtube)} YouTube iframe(s) vào content")

    # ── Chèn phần file đính kèm ───────────────────────────────────────────
    if found_links:
        div    = soup.new_tag('div')
        header = soup.new_tag('strong')
        header.string = "Tài liệu đính kèm: "
        div.append(header)
        div.append(soup.new_tag('br'))
        for link in found_links:
            div.append(link)
            div.append(soup.new_tag('br'))
        div.append(soup.new_tag('hr'))
        main_content_tag.insert(0, div)
        log.success(f"✓ Merged {len(found_links)} external attachments vào content")

    if not found_links and not found_youtube:
        log.info("Không tìm thấy file đính kèm/YouTube ngoài content")


# ══════════════════════════════════════════════════════════════════════════════
# YOUTUBE HELPERS
# ══════════════════════════════════════════════════════════════════════════════

_YT_URL_RE = re.compile(
    r'(?:https?://)?(?:www\.)?'
    r'(?:youtube\.com|youtube-nocookie\.com|youtu\.be)'
    r'(?:'
        r'/watch\?(?:[^#&\s]*&)*v='          # youtube.com/watch?v=ID
        r'|/embed/'                           # youtube.com/embed/ID  (và nocookie)
        r'|/'                                 # youtu.be/ID
    r')'
    r'([\w\-]{11})',
    re.IGNORECASE
)


def _yt_video_id(url: str) -> str | None:
    """Trích video-id (11 ký tự) từ URL YouTube. Trả None nếu không hợp lệ."""
    m = _YT_URL_RE.search(url or "")
    return m.group(1) if m else None


def _make_yt_iframe(video_id: str, soup) -> object:
    """Tạo div bao + iframe embed YouTube responsive 16:9.
    Gán data-yt-wrapped='1' để embed_youtube_links_in_content không bọc lại.
    """
    wrapper = soup.new_tag(
        "div",
        style="position:relative;padding-bottom:56.25%;height:0;overflow:hidden;margin:14px 0;",
        **{"data-yt-wrapped": "1"}
    )
    iframe = soup.new_tag(
        "iframe",
        src=f"https://www.youtube.com/embed/{video_id}",
        width="100%",
        height="450",
        frameborder="0",
        allow="accelerometer; autoplay; clipboard-write; encrypted-media; gyroscope; picture-in-picture",
        allowfullscreen="true",
        style="position:absolute;top:0;left:0;width:100%;height:100%;border:0;",
    )
    wrapper.append(iframe)
    return wrapper


def embed_youtube_links_in_content(tag_content, soup=None) -> int:
    """
    Chuyển đổi link YouTube dạng <a href> thành <iframe> embed trong nội dung.
    Xử lý 3 trường hợp:
      1. <a href="youtube...">bất kỳ nội dung</a>  → thay bằng wrapper+iframe
      2. <a href="youtube..."><img ...></a>          → thay bằng wrapper+iframe
      3. Text node thuần chứa URL youtube (hiếm)    → bọc thành iframe

    LƯU Ý: Gọi hàm này TRƯỚC clean_html_content() để iframe được giữ lại.

    Trả về số lượng embed đã chuyển đổi.
    """
    if not tag_content:
        return 0

    # Dùng soup của tag_content nếu không truyền vào
    if soup is None:
        soup = tag_content

    converted = 0

    # ── Giữ nguyên iframe YouTube đã có trong content (kể cả nocookie) ────────
    # Chỉ đảm bảo responsive, không convert lại
    for iframe_tag in list(tag_content.find_all('iframe')):
        src = iframe_tag.get('src', '')
        is_yt = any(d in src for d in ('youtube.com', 'youtube-nocookie.com', 'youtu.be'))
        if is_yt:
            # Đảm bảo có wrapper responsive nếu chưa có
            parent = iframe_tag.parent
            # Kiểm tra parent đã là wrapper (class videoWrapper HOẶC data-yt-wrapped)
            _already_wrapped = (
                (parent and 'videoWrapper' in ' '.join(parent.get('class', [])))
                or (parent and parent.get('data-yt-wrapped'))
            )
            if _already_wrapped:
                pass  # đã có wrapper — giữ nguyên
            else:
                vid = _yt_video_id(src)
                if vid:
                    wrapper = _make_yt_iframe(vid, soup)
                    iframe_tag.replace_with(wrapper)
                    log.info(f"  🎬 Bọc YouTube iframe: {vid}  ({src[:50]})")
                    converted += 1

    # ── Case 1 & 2: thẻ <a> trỏ tới YouTube ──────────────────────────────────
    for a_tag in list(tag_content.find_all('a', href=True)):
        href = a_tag.get('href', '')
        vid  = _yt_video_id(href)
        if not vid:
            continue
        iframe_wrap = _make_yt_iframe(vid, soup)
        a_tag.replace_with(iframe_wrap)
        log.info(f"  🎬 YouTube <a> → iframe: {vid}  ({href[:60]})")
        converted += 1

    # ── Case 3: text node chứa URL youtube chưa được bọc thẻ <a> ─────────────
    from bs4 import NavigableString
    for node in list(tag_content.descendants):
        if not isinstance(node, NavigableString):
            continue
        if node.parent and node.parent.name == 'a':
            continue   # đã xử lý ở Case 1-2
        txt = str(node)
        vid = _yt_video_id(txt)
        if vid:
            iframe_wrap = _make_yt_iframe(vid, soup)
            node.replace_with(iframe_wrap)
            log.info(f"  🎬 YouTube text-node → iframe: {vid}")
            converted += 1

    if converted:
        log.success(f"✓ embed_youtube_links: chuyển {converted} link → iframe")
    return converted


# ══════════════════════════════════════════════════════════════════════════════
# PHẦN 7 · SCRAPER CHÍNH
# ══════════════════════════════════════════════════════════════════════════════

class VspProducts:
    def __init__(self, base, url, cat, target):
        log.section("KHỞI TẠO SCRAPER", Color.CYAN)

        self.base       = base
        self.cat        = cat
        self.target     = target
        self.camlist    = camob.Listcam()
        self.url_links: list[tuple[str, str]] = []   # (url, cat_id)
        self.menu_links = menulink.MenuLink()

        log.key_value("Base URL",  base,   Color.CYAN)
        log.key_value("Category",  cat,    Color.YELLOW)
        log.key_value("Target",    target, Color.GREEN)

        log.info(f"{Icon.ROCKET} Khởi động Chrome driver...")
        service        = Service(CHROMEDRIVER_PATH)
        chrome_options = Options()
        chrome_options.binary_location = CHROME_BINARY
        chrome_options.add_argument("--window-size=1024,768")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")

        try:
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            # Tôn trọng config SHOW_CHROME_WINDOW giống 2_Dang_bai.py
            if not SHOW_CHROME_WINDOW:
                try:
                    wid = self.driver.execute_cdp_cmd('Browser.getWindowForTarget', {})
                    self.driver.execute_cdp_cmd('Browser.setWindowBounds', {
                        'windowId': wid['windowId'],
                        'bounds':   {'windowState': 'minimized'},
                    })
                except Exception:
                    try:
                        self.driver.minimize_window()
                    except Exception:
                        pass
            log.success("✓ Chrome driver sẵn sàng")
        except Exception as e:
            log.critical(f"✗ Không thể khởi động Chrome driver: {e}")
            raise

        self.url      = url
        self.cat_id   = ""
        self.post_type = "posts"   # WP REST endpoint slug
        self.url_id   = ""
        self.page_link = 1
        self._excel_row = 0   # số dòng Excel hiện tại (set bởi run_from_excel)

        log.end_section()

    def reset_for_row(self):
        """Reset queue + menu cho mỗi dòng Excel."""
        self.url_links: list[tuple[str, str]] = []
        self.menu_links = menulink.MenuLink()
        self.camlist    = camob.Listcam()
        self.page_link  = 1   # BUG FIX: reset pagination cho mỗi URL mới

    @staticmethod
    def _norm_url(url: str) -> str:
        """Chuẩn hóa URL: xóa double-slash, strip categoryId noise param, strip trailing slash."""
        if not url:
            return url
        p = urlparse(url)
        clean = re.sub(r'/{2,}', '/', p.path)
        # Strip trailing slash (trừ root path "/")
        if clean != '/' and clean.endswith('/'):
            clean = clean.rstrip('/')
        # Strip categoryId — CMS portal dùng param này để tracking menu,
        # không phải phần định danh bài viết
        qs = p.query
        if qs:
            params = [kv for kv in qs.split('&')
                      if not re.match(r'^categoryid=', kv, re.IGNORECASE)]
            qs = '&'.join(params)
        return p._replace(path=clean, query=qs).geturl()

    def _in_same_section(self, link_url: str, page_url: str,
                          excel_cat: str = "") -> bool:
        """
        Lọc sidebar: chỉ chấp nhận link hợp lệ.

        Nếu cat_id ĐÃ được điền trong Excel → chỉ kiểm tra cùng domain.
        (CMS portal hay gán ?categoryId cho bài từ thư mục khác nhưng cùng chuyên mục)

        Nếu cat_id CHƯA điền → kiểm tra cùng path prefix để lọc sidebar.
        Ngoại lệ: nếu section là root ('') → chấp nhận tất cả cùng domain
        (CMS dùng URL flat như /bai-viet.html trực tiếp dưới domain)
        """
        link_parsed = urlparse(link_url)
        page_parsed = urlparse(page_url)

        # Khác domain → luôn từ chối
        if link_parsed.netloc != page_parsed.netloc:
            return False

        # cat_id từ Excel có giá trị → tin tưởng listing, chỉ cần cùng domain
        if excel_cat and excel_cat.strip():
            return True

        # cat_id chưa có → lọc theo path prefix (tránh sidebar)
        page_path = page_parsed.path.rstrip('/')
        link_path = link_parsed.path.rstrip('/')
        if '.' in page_path.split('/')[-1]:
            section = '/'.join(page_path.split('/')[:-1])
        else:
            section = page_path

        # Nếu section = '' (page ở root domain, kiểu /listing.html)
        # → không thể lọc path-based, chấp nhận tất cả cùng domain
        if section == '':
            return True

        return link_path.startswith(section + '/') or link_path == section

    @staticmethod
    def _extract_id_from_url(url: str) -> int:
        """
        Trích số ID từ URL slug.
        Ví dụ: .../bai-viet-abc-12345.html → 12345
                .../p=12345               → 12345
        Trả về 0 nếu không tìm thấy.
        """
        # Query param ?p=ID hoặc ?id=ID
        m = re.search(r'[?&](?:p|id|post_id)=(\d+)', url)
        if m:
            return int(m.group(1))
        # Số cuối trong slug trước .html
        m = re.search(r'-(\d{3,})(?:\.html?)?(?:[?#]|$)', url)
        if m:
            return int(m.group(1))
        return 0

    @staticmethod
    def _extract_thumbnail_url(tag, base: str = "") -> str:
        """
        Lấy URL ảnh thumbnail từ listing item HTML.
        Bỏ qua ảnh no-image.jpg (placeholder).
        Trả về URL tuyệt đối hoặc "" nếu không có.
        """
        if tag is None:
            return ""
        # Ưu tiên figure.post-image img (CauTruc1)
        for selector in (
            'figure.post-image img',
            'div.col-xs-4 img',          # CauTruc2 / hanam.edu.vn
            'img.img-responsive',
            'img',
        ):
            img = tag.select_one(selector) if hasattr(tag, 'select_one') else tag.find('img')
            if img is None:
                continue
            src = img.get('src') or ''
            _sl = src.lower()
            # Lazy-load placeholder? → thử data-src / data-original / data-lazy
            _is_placeholder = (
                not src
                or src.startswith('data:')
                or any(_p in _sl for _p in ('blank.gif', 'transparent.gif',
                                             '1x1.gif', 'loading.gif',
                                             'spacer.gif', 'placeholder'))
            )
            if _is_placeholder:
                src = (img.get('data-src') or img.get('data-original')
                       or img.get('data-lazy') or img.get('data-url') or '')
            if not src:
                continue
            # Bỏ qua ảnh no-image
            _sl = src.lower()
            if 'no-image' in _sl or _sl.endswith('/no-image.jpg'):
                continue
            # Chuyển relative → absolute
            if src.startswith('/') and base:
                from urllib.parse import urljoin
                src = urljoin(base, src)
            return src
        return ""

    @staticmethod
    def _score_listing_item(tag=None) -> dict:
        """
        Chấm điểm 1 item trong listing HTML (không fetch thêm trang).
        Trả về dict gồm: has_img, has_avatar, text_len
        """
        if tag is None:
            return {"has_img": False, "has_avatar": False, "text_len": 0}
        has_img    = bool(tag.find('img'))
        has_avatar = bool(tag.find('img', class_=lambda c: c and
                          any(k in c for k in ('avatar', 'thumb', 'feature', 'thumbnail'))))
        text_len   = len(tag.get_text(strip=True))
        return {"has_img": has_img, "has_avatar": has_avatar, "text_len": text_len}

    def _smart_add(self, name: str, url: str, cat_id: str,
                   hp, position: int = 0, tag=None) -> str:
        """
        Thêm bài vào hàng đợi với logic chọn lọc thông minh.

        Tiêu chí so sánh khi trùng title (theo thứ tự ưu tiên):
          1. Có ảnh/media trong listing item (has_img)
          2. Có avatar/thumbnail (has_avatar)
          3. Text snippet dài hơn (text_len — thường = mô tả dài hơn)
          4. ID trong URL cao hơn (mới hơn)
          → Nếu bài mới tốt hơn ở bất kỳ tiêu chí nào → thay thế

        Trả về: "added" | "skipped_url" | "replaced" | "kept_old"
        """
        # ── Case 1: URL trong url_links ─────────────────────────────────────
        if url in [u for u, _ in self.url_links]:
            # Nếu chưa có trong camlist (chỉ từ fallback prefix scan menu)
            # → thêm vào camlist để chitiet nhận ra đúng, không add url_links lần nữa
            in_camlist = any(self._norm_url(c.url) == url
                             for c in self.camlist.camobs)
            if not in_camlist and name:
                new_cam = camob.CameraObject(0, name, 0, url, "", cat_id)
                new_cam._listing_score = self._score_listing_item(tag)
                new_cam._thumbnail_url = self._extract_thumbnail_url(tag, self.base)
                self.camlist.add_cam(new_cam)
                return "added"
            return "skipped_url"

        # ── Case 2: trùng title, khác URL ────────────────────────────────────
        existing_cam = next(
            (c for c in self.camlist.camobs if c.name.strip() == name.strip()), None)

        if existing_cam:
            old_url   = self._norm_url(existing_cam.url)
            old_score = getattr(existing_cam, '_listing_score', None) or \
                        {"has_img": False, "has_avatar": False, "text_len": 0, "id": 0}
            new_score = self._score_listing_item(tag)
            new_score["id"] = self._extract_id_from_url(url)
            old_score.setdefault("id", self._extract_id_from_url(old_url))

            # So sánh theo thứ tự ưu tiên
            def is_better(n, o) -> bool:
                if n["has_img"]    != o["has_img"]:    return n["has_img"]
                if n["has_avatar"] != o["has_avatar"]:  return n["has_avatar"]
                if n["text_len"]   != o["text_len"]:   return n["text_len"] > o["text_len"]
                return n["id"] > o["id"]   # mới hơn = tốt hơn

            if is_better(new_score, old_score):
                self.url_links = [(u, c) for u, c in self.url_links if u != old_url]
                self.camlist.camobs = [
                    c for c in self.camlist.camobs if c.name.strip() != name.strip()]
                new_cam = camob.CameraObject(0, name, 0, url, "", cat_id)
                new_cam._listing_score = new_score
                new_cam._thumbnail_url = self._extract_thumbnail_url(tag, self.base)
                self.camlist.add_cam(new_cam)
                self.url_links.append((url, cat_id))
                return "replaced"
            else:
                return "kept_old"

        # ── Case 3: kiểm tra DB rồi thêm ────────────────────────────────────
        check = hp.check_cam_url(self.url_id, url, name)
        if not check:
            # hp.check_cam_url trả False = đã tồn tại trong DB → bỏ qua, hiện [✓]
            stats.record_in_db(self.base)
            return "in_db"
        if url and name:
            new_cam = camob.CameraObject(0, name, 0, url, "", cat_id)
            new_cam._listing_score = self._score_listing_item(tag)
            new_cam._listing_score["id"] = self._extract_id_from_url(url)
            new_cam._thumbnail_url = self._extract_thumbnail_url(tag, self.base)
            if self.camlist.add_cam(new_cam):
                self.url_links.append((url, cat_id))
                return "added"
        return "skipped_url"

    @staticmethod
    def _scrape_cat_from_page(soup) -> str:
        """
        Lấy tên chuyên mục từ trang web khi cat_id chưa có trong Excel.
        Thứ tự thử: breadcrumb → meta category → h1 page title.
        """
        # 1. Breadcrumb — thường là mục cuối cùng trước trang hiện tại
        for sel in ('nav.breadcrumb a', 'div.breadcrumb a', 'ul.breadcrumb li a',
                    'div.bread-crumb a', 'ol.breadcrumb li a'):
            crumbs = soup.select(sel)
            if len(crumbs) >= 2:
                # Lấy mục áp cuối (cuối = trang hiện tại, áp cuối = chuyên mục)
                cat = crumbs[-2].get_text(strip=True)
                if cat and len(cat) < 80:
                    return cat
        # 2. Meta / schema category
        for sel in ('span.cat-name', 'a.category-name', 'div.post-category a',
                    'span.category', 'div.category a', 'p.category a'):
            el = soup.select_one(sel)
            if el:
                cat = el.get_text(strip=True)
                if cat and len(cat) < 80:
                    return cat
        # 3. Tiêu đề trang section (h2 đầu tiên trong main)
        for sel in ('main h2:first-of-type', 'div.main-content h2',
                    'div.content h2', 'section h2'):
            el = soup.select_one(sel)
            if el:
                cat = el.get_text(strip=True)
                if cat and len(cat) < 80:
                    return cat
        return ""

    def _try_save_cam(self, cam, o_url: str, t_start, struct_type,
                      soup=None) -> bool:
        """
        Kiểm tra trùng lặp rồi lưu DB.

        Nếu đã có bài với CÙNG tiêu đề VÀ nội dung trong DB → in log
        '-> đã tồn tại trước đó' và trả về True mà không lưu lại.

        Trả về True nếu đã xử lý (dù lưu hay bỏ qua),
        để caller có thể `return` ngay sau khi gọi hàm này.

        soup: BeautifulSoup của trang — dùng để lấy title khi cam.name rỗng
              (trường hợp URL trực tiếp từ Excel, không qua listing).
        """
        # Nếu cam là synthetic (name rỗng), lấy title từ page
        if not cam.name and soup is not None:
            for _sel in ('h1.title-detail', 'h1.entry-title', 'h1',
                         'title', '.post-title h1', 'article h1'):
                _el = soup.select_one(_sel)
                if _el:
                    _t = clean_spaces(_el.get_text(" ", strip=True))
                    if _t and len(_t) > 3:
                        cam.name = _t
                        log.info(f"  ℹ Title từ trang: {_t[:60]}")
                        break

        title   = cam.name or ""
        content = cam.description or ""

        # Kiểm tra trùng tiêu đề + nội dung trong database
        already_exists = False
        try:
            already_exists = hp.check_exists_title_content(
                self.url_id, title, content)
        except AttributeError:
            # Nếu helpers chưa có hàm này thì bỏ qua kiểm tra
            pass
        except Exception as e:
            log.warning(f"Lỗi kiểm tra trùng lặp: {e}")

        if already_exists:
            print(f"  -> đã tồn tại trong DB")
            elapsed = (datetime.now() - t_start).total_seconds()
            stats.record_url(o_url, struct_type)
            stats.record_in_db(self.base)   # track riêng — không tính vào "Lấy được"
            log.detail_saved(elapsed)
            return True

        # Chưa có → lưu bình thường
        cam.display_info()
        # Gán post_type vào cam trước khi lưu (để hp.save_data_cam biết loại bài)
        if not getattr(cam, 'post_type', ''):
            cam.post_type = getattr(self, 'post_type', 'posts')
        with _suppress_stdout():
            hp.save_data_cam(self.url_id, cam)
        elapsed = (datetime.now() - t_start).total_seconds()
        stats.record_url(o_url, struct_type)
        stats.record_saved(self.base, self.cat_id)
        log.detail_saved(elapsed)
        return True

    def close(self):
        """Đóng Chrome driver khi kết thúc."""
        try:
            self.driver.quit()
            log.success("✓ Đã đóng driver")
        except Exception as e:
            log.error(f"Lỗi khi đóng driver: {e}")

    def restart_driver(self):
        """Khởi động lại Chrome driver sau khi crash."""
        log.warning("♻ Đang khởi động lại Chrome driver...")
        try:
            self.driver.quit()
        except Exception:
            pass
        service        = Service(CHROMEDRIVER_PATH)
        chrome_options = Options()
        chrome_options.binary_location = CHROME_BINARY
        chrome_options.add_argument("--window-size=1024,768")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-gpu")
        try:
            self.driver = webdriver.Chrome(service=service, options=chrome_options)
            if not SHOW_CHROME_WINDOW:
                try:
                    wid = self.driver.execute_cdp_cmd('Browser.getWindowForTarget', {})
                    self.driver.execute_cdp_cmd('Browser.setWindowBounds', {
                        'windowId': wid['windowId'],
                        'bounds':   {'windowState': 'minimized'},
                    })
                except Exception:
                    try: self.driver.minimize_window()
                    except Exception: pass
            log.success("✓ Driver đã khởi động lại")
        except Exception as e:
            log.critical(f"✗ Không thể restart driver: {e}")
            raise

    # ── get_data ──────────────────────────────────────────────────────────────

    def get_data(self):
        """Entry point — khám phá sub-menu rồi xử lý hàng đợi URL."""
        log.section("BẮT ĐẦU TRÍCH XUẤT DỮ LIỆU", Color.GREEN)

        try:
            log.info(f"Tải URL gốc để tìm sub-menu: {self.url}")
            self.driver.get(self.url)
            time.sleep(2)
            soup      = BeautifulSoup(self.driver.page_source, 'html.parser')
            sub_menus = self.discover_sub_menus(soup)  # dict: url → cat_id

            # Thêm link mẹ trước (giữ cat_id từ Excel)
            norm_self = self._norm_url(self.url)
            if norm_self not in [u for u, _ in self.url_links]:
                self.url_links.append((norm_self, self.cat_id))

            for url, cat in sub_menus.items():
                norm_u = self._norm_url(url)
                if norm_u not in [u for u, _ in self.url_links]:
                    # Nếu Excel đã có cat_id → dùng cho tất cả sub-URL
                    # (không override bằng tên menu con)
                    effective_cat = self.cat_id if self.cat_id else cat
                    self.url_links.append((norm_u, effective_cat))

        except Exception as e:
            log.error(f"Lỗi khi tìm sub-menu: {e}")
            self.url_links.append((self.url, self.cat_id))

        processed      = 0
        processed_urls = set()

        while self.url_links:
            current_url, current_cat = self.url_links.pop(0)
            if current_url in processed_urls:
                continue

            # Cập nhật cat_id hiện tại cho extract_data dùng
            self.cat_id = current_cat
            self._url_index = processed + 1

            self.extract_data(current_url)
            processed_urls.add(current_url)
            processed += 1

        log.success(f"✓ Đã xử lý {processed} URL (gồm cả sub-menu)")
        log.end_section()

    # ── extract_data ──────────────────────────────────────────────────────────

    def extract_data(self, o_url: str):
        """Trích xuất dữ liệu từ 1 URL (listing hoặc detail)."""
        t_start = datetime.now()

        # Reset page_link cho mỗi URL mới — tránh pagination bị lệch giữa các URL trong queue
        self.page_link = 1

        # ── Chuẩn hóa URL: xóa double-slash sau domain ───────────────────────
        parsed   = urlparse(o_url)
        clean_path = re.sub(r'/{2,}', '/', parsed.path)   # //foo → /foo
        o_url    = parsed._replace(path=clean_path).geturl()
        # Áp dụng _norm_url đầy đủ (strip trailing slash, strip categoryId)
        # để so sánh nhất quán với cam.url (đã qua _norm_url trong _smart_add)
        o_url    = self._norm_url(o_url)
        log.debug(f"URL sau chuẩn hóa: {o_url}")

        # ── Xác định detail / listing ────────────────────────────────────────
        chitiet = any(self._norm_url(cam.url) == o_url for cam in self.camlist.camobs)

        # ── In header URL và ⏳ Tải... ────────────────────────────────────────
        if chitiet:
            _detail_title = ""
            _detail_idx   = 0
            _detail_total = len(self.camlist.camobs)
            for _i, _cam in enumerate(self.camlist.camobs, 1):
                if self._norm_url(_cam.url) == o_url:
                    _detail_title = _cam.name
                    _detail_idx   = _i
                    break
            if not _detail_title:
                _detail_title = o_url
            log.detail_start(_detail_title, o_url,
                             idx=_detail_idx, total=_detail_total)
        else:
            log.list_start(o_url)
        time.sleep(1)

        # ── Tải trang ────────────────────────────────────────────────────────
        _load_err_msg = ""
        try:
            self.driver.get(o_url)
            time.sleep(1)
            html = self.driver.page_source
        except TimeoutException:
            _load_err_msg = "Timeout"
        except WebDriverException as e:
            _load_err_msg = f"WebDriver: {str(e)[:60]}"
        except Exception as e:
            _load_err_msg = f"Lỗi tải: {str(e)[:60]}"

        if _load_err_msg:
            log.detail_load_error(_load_err_msg, (datetime.now()-t_start).total_seconds())
            stats.record_url(o_url, SiteStructureType.UNKNOWN,
                             success=False, error_msg=_load_err_msg)
            if not chitiet:
                log.list_page_error(1, _load_err_msg)
                log.list_coverage_check(list_url=o_url, base=get_base(o_url))
                log.list_done((datetime.now()-t_start).total_seconds())
                base_k = get_base(o_url)
                stats.record_failed("", o_url, base_k, self.cat_id, _load_err_msg,
                                    row=self._excel_row, fail_type="list")
            return

        # ── Kiểm tra lỗi HTTP im lặng (403 / 404 / trang trắng) ─────────────
        http_err = _detect_silent_http_error(html, o_url)
        if http_err:
            log.detail_load_error(http_err, (datetime.now()-t_start).total_seconds())
            stats.record_url(o_url, SiteStructureType.UNKNOWN,
                             success=False, error_msg=http_err)
            if not chitiet:
                log.list_page_error(1, http_err)
                log.list_coverage_check(list_url=o_url, base=get_base(o_url))
                log.list_done((datetime.now()-t_start).total_seconds())
            base_k = get_base(o_url)
            # Nếu là trang bài viết → lấy tiêu đề từ camlist
            _fail_title = ""
            if chitiet:
                _found_cam = next(
                    (c for c in self.camlist.camobs
                     if self._norm_url(c.url) == o_url), None)
                if _found_cam:
                    _fail_title = _found_cam.name or ""
            stats.record_failed(_fail_title, o_url, base_k, self.cat_id, http_err,
                                row=self._excel_row,
                                fail_type="article" if chitiet else "list")
            return

        # ── MAIN LOOP ─────────────────────────────────────────────────────────
        detected_structure = SiteStructureType.UNKNOWN
        saved_date_str     = ""

        _page_num    = 1   # số trang hiện tại trong pagination
        _total_pages = 0   # tổng số trang (0 = chưa biết)
        _page1_html  = ""  # HTML trang 1 để detect pagination trước khi render items

        while True:
            soup = BeautifulSoup(html, 'html.parser')
            time.sleep(1)

            # ── Detect cấu trúc ──────────────────────────────────────────────
            structures = log.detect_and_log(soup, o_url)
            detected_structure = structures[0]
            _struct_name = (detected_structure.value.split("·")[1].strip()
                            if "·" in detected_structure.value
                            else detected_structure.value)

            # ── cat_id: dùng từ Excel nếu có, nếu không scrape từ trang ─────
            if not self.cat_id:
                scraped_cat = self._scrape_cat_from_page(soup)
                if scraped_cat:
                    self.cat_id = scraped_cat

            # ── Detect tổng số trang — chạy MỖI trang, lấy giá trị lớn nhất ────────
            # Sliding-window pagination: trang 1 chỉ hiển 1-5,
            # sang trang 6 mới thấy số 6-9 → phải cập nhật mỗi lần.
            if not chitiet:
                _pag_soup = soup
                # 1. BeautifulSoup — a.last / btn-end
                _last_a = (_pag_soup.select_one('a.last')
                           or _pag_soup.select_one('.default-pagination a.last')
                           or _pag_soup.select_one('div.page ul li.btn-end a'))
                if _last_a:
                    _href = _last_a.get('href', '')
                    _m = re.search(r'page=(\d+)', _href)
                    if _m:
                        _total_pages = max(_total_pages, int(_m.group(1)))
                    else:
                        _nums = [int(a.text.strip()) for a in
                                 _pag_soup.select('.default-pagination a, div.page ul li a')
                                 if a.text.strip().isdigit()]
                        if _nums: _total_pages = max(_total_pages, max(_nums))
                # 2. Lấy max số trong window phân trang hiện tại (bắt sliding window)
                _pag_nums = [int(a.text.strip()) for a in
                             _pag_soup.select(
                                 'div.col-center a, a.page-link, '
                                 'div.page ul li a, div.page-news ul li a')
                             if a.text.strip().isdigit()]
                if _pag_nums:
                    _total_pages = max(_total_pages, max(_pag_nums))
                # 3. Selenium live DOM — theo href nút >> để biết trang cuối thật sự
                try:
                    for _lsel in ['div.col-center a.last', 'a.last',
                                  'li.btn-end a', 'a[title="Cuối"]', 'a[title="Last"]']:
                        _lels = self.driver.find_elements(By.CSS_SELECTOR, _lsel)
                        if _lels:
                            _lh = _lels[0].get_attribute('href') or ''
                            _lm = re.search(r'page=(\d+)', _lh)
                            if _lm:
                                _total_pages = max(_total_pages, int(_lm.group(1)))
                                break
                    # Lấy max số từ tất cả nút phân trang live (kể cả khi không có a.last)
                    _live_nums = []
                    for _lsel2 in ['div.col-center a', 'a.page-link',
                                   'div.page ul li a', 'div.page-news ul li a']:
                        for _lel in self.driver.find_elements(By.CSS_SELECTOR, _lsel2):
                            _lt = _lel.text.strip()
                            if _lt.isdigit(): _live_nums.append(int(_lt))
                    if _live_nums:
                        _total_pages = max(_total_pages, max(_live_nums))
                except Exception:
                    pass
                if _page_num == 1:
                    log.debug(f"Tổng số trang detect: {_total_pages or '?'}")

            # ── Hoàn thành dòng ⏳ Tải... (lần đầu vào loop) ─────────────────
            if not chitiet and _page_num == 1:
                # Đếm item HTML trên trang
                # Đếm items theo từng selector — lấy giá trị lớn nhất có selector
                _cnt_post   = len(soup.select('div.post-item.row') or
                                  soup.find_all('div', class_='post-item'))
                _ul         = (soup.find('ul', class_='ArticleList')
                               or soup.find('ul', class_='down-list')
                               or soup.find('ul', class_='phal-list row'))
                _cnt_ul     = len(_ul.find_all('li')) if _ul else 0
                # Article-News: đây là selector chính của CauTruc1
                _cnt_art    = len(soup.select('article.Article-News, article.list-memberpost-new'))
                _cnt_ptitle = len(soup.find_all('div', class_='post-title') or
                                  soup.find_all('div', class_='title-news-listType10') or
                                  soup.find_all('div', class_='item-info'))
                _cnt_docs   = len(soup.find_all('a', class_='title-documment'))
                _cnt_ck     = len(soup.select('td.tg-yw4l a'))
                _cnt_hanam  = len(soup.find_all('div', class_='list-item'))
                _cnt_news_c  = len(soup.select('div.list-news div.news-content'))
                _cnt_media   = len(soup.select('ul.media-list li.media'))
                # Lấy selector nào có nhiều item nhất (tránh đếm trùng)
                _html_count = max(_cnt_post, _cnt_ul, _cnt_art,
                                  _cnt_ptitle, _cnt_docs, _cnt_ck,
                                  _cnt_hanam, _cnt_news_c, _cnt_media)
                log.list_page_loaded(len(html)/1024, _struct_name,
                                     _html_count, _page_num,
                                     total_pages=_total_pages)
                log._list_page_html = _html_count
            elif not chitiet:
                # Trang 2+ — hoàn thành dòng ⏳ Tải trang N...
                _cnt_post  = len(soup.select('div.post-item.row') or
                                  soup.find_all('div', class_='post-item'))
                _ul2       = (soup.find('ul', class_='ArticleList')
                              or soup.find('ul', class_='down-list')
                              or soup.find('ul', class_='phal-list row'))
                _cnt_ul2   = len(_ul2.find_all('li')) if _ul2 else 0
                _cnt_art2  = len(soup.select('article.Article-News, article.list-memberpost-new'))
                _cnt_pt2   = len(soup.find_all('div', class_='post-title') or
                                  soup.find_all('div', class_='item-info') or [])
                _cnt_d2    = len(soup.find_all('a', class_='title-documment'))
                _cnt_h2    = len(soup.find_all('div', class_='list-item'))
                _cnt_nc2   = len(soup.select('div.list-news div.news-content'))
                _cnt_med2  = len(soup.select('ul.media-list li.media'))
                _html_cnt2 = max(_cnt_post, _cnt_ul2, _cnt_art2,
                                 _cnt_pt2, _cnt_d2, _cnt_h2, _cnt_nc2, _cnt_med2)
                log.list_page_loaded(len(html)/1024, _struct_name,
                                     _html_cnt2, _page_num,
                                     total_pages=_total_pages)
                log._list_page_html = _html_cnt2

            # ════════════════════════════════════════════════════════════════
            # 1. LISTING — post-item
            # ════════════════════════════════════════════════════════════════
            div_pros = (soup.select('div.post-item.row')
                        or soup.find_all('div', class_='post-item'))

            if div_pros and not chitiet:
                _global_idx = log._list_total_added + log._list_total_skip + 1
                for idx, tag_div in enumerate(div_pros, 1):
                    h3 = tag_div.find('h3') or tag_div.select_one('.entry-title')
                    if not h3: continue
                    tag_a = h3.find('a')
                    if not tag_a: continue
                    raw_href = tag_a.get('href') or ''
                    name = tag_a.text.strip()
                    url  = self._norm_url(urljoin(self.base, raw_href))
                    has_img = bool(tag_div.find('img'))
                    if not self._in_same_section(url, o_url, self.cat_id):
                        log.list_item(_global_idx, 0, name, url,
                                      "filtered_section", has_img)
                        _global_idx += 1
                        continue
                    result = self._smart_add(name, url, self.cat_id, hp, position=idx, tag=tag_div)
                    log.list_item(_global_idx, 0, name, url, result, has_img)
                    _global_idx += 1

            # ════════════════════════════════════════════════════════════════
            # 2. DETAIL — TYPE 1: h1.title-detail
            # ════════════════════════════════════════════════════════════════
            tag_title_product = soup.find('h1', class_="title-detail")

            if chitiet and tag_title_product:

                log.section("📄 CHI TIẾT — TYPE-1 (title-detail)", Color.GREEN)
                log.parse_info("h1.title-detail", True, tag_title_product.text.strip()[:80])

                tag_content        = (soup.find('div', class_='content-detail')
                                      or soup.find('div', class_='content-detail font-size-text mb-20'))
                # Tìm container file đính kèm — thử nhiều selector CMS
                tag_files_container = (
                    soup.find('div', class_='pull-left mt-5')
                    or soup.find('div', class_='listfile')
                    or soup.find('div', class_='attach-list')
                    or soup.find('ul',  class_='list-download')
                    or soup.find('div', class_='inline-download')
                    or soup.find('div', class_='download-type4')
                    or soup.select_one('div.file-attach, div.files-attach,'
                                       'div.attachments, div[class*="attachment"]')
                )
                # Fallback: tìm block chứa "Tài liệu/Hình ảnh đính kèm"
                if not tag_files_container:
                    for _blk in soup.find_all(['div', 'ul', 'section']):
                        _blk_txt = _blk.get_text(" ", strip=True).lower()
                        if any(kw in _blk_txt for kw in ('tài liệu đính kèm',
                               'hình ảnh đính kèm', 'tài liệu/hình ảnh')):
                            if _blk.find('a', href=True):
                                tag_files_container = _blk
                                break

                d = parse_vn_date_from_soup(soup)
                if d:
                    saved_date_str = d.strftime("%d/%m/%Y")
                    log.key_value("Date", saved_date_str, Color.YELLOW)
                    # Lọc bài cũ hơn FROM_DATE
                    if d < FROM_DATE:
                        log.warning(f"Bỏ qua — ngày {saved_date_str} trước FROM_DATE {FROM_DATE}")
                        stats.record_url(o_url, SiteStructureType.TYPE_TITLE_DETAIL)
                        return

                # Tìm nút "Đọc bài viết" / "Xem tài liệu" — CMS dùng JS để ẩn/hiện content
                _doc_btn_html = ""
                _doc_btns = soup.select('a.btn, a.button, a[class*="btn"]')
                for _db in _doc_btns:
                    _db_txt = _db.get_text(strip=True).lower()
                    if any(kw in _db_txt for kw in ('đọc bài', 'xem tài liệu',
                                                    'tải về', 'download', 'xem file',
                                                    'xem văn bản', 'mở tài liệu')):
                        _db_href = _db.get('href', '')
                        if _db_href and 'javascript' not in _db_href.lower():
                            if not _db_href.startswith('http'):
                                _db_href = urljoin(self.base, _db_href)
                            _db_label = _db.get_text(strip=True) or 'Đọc bài viết'
                            _doc_btn_html += (f'<p>➤ <a href="{_db_href}" target="_blank"'
                                              f' class="link-download">{_db_label}</a></p>')

                # Tính is_empty: tag_content rỗng hay chỉ có whitespace/br
                _tc_text = tag_content.get_text(strip=True) if tag_content else ""
                _tc_has_content = bool(_tc_text or (tag_content and tag_content.find('img')))

                if tag_content or _doc_btn_html or tag_files_container:
                    for cam in self.camlist.camobs:
                        if self._norm_url(cam.url) == o_url:
                            cam.short = self.target
                            if d:
                                cam.date_publish = d
                            if 'Công khai' in cam.name:
                                cam.cat_id = 'Công khai'

                            attachment_html = ""
                            _file_cnt = 0
                            if tag_files_container:
                                all_links = tag_files_container.find_all('a')
                                if all_links:
                                    attachment_html = ("<div class='attachments-list' style='margin-bottom:20px;"
                                                       "padding:10px;background:#f9f9f9;border-left:4px solid #007bff;'>"
                                                       "<strong>📂 Tài liệu/Hình ảnh đính kèm:</strong><br/>")
                                    def _pick_link(a_tag, base_url):
                                        """Ưu tiên onclick.downloadFile() → href → bỏ qua."""
                                        onclick_str = (a_tag.get('onclick') or '').strip()
                                        _m2 = re.search(
                                            r"downloadFile\s*\(\s*(['\"])(.*?)\1\s*,\s*(['\"])(.*?)\3",
                                            onclick_str, re.IGNORECASE | re.DOTALL)
                                        _m1 = re.search(
                                            r"downloadFile\s*\(\s*(['\"])(.*?)\1\s*\)",
                                            onclick_str, re.IGNORECASE | re.DOTALL)
                                        if _m2:
                                            _a1 = _m2.group(2).strip()
                                            _a2 = _m2.group(4).strip()
                                            def _isp(s):
                                                sl = s.lower()
                                                return (sl.startswith(('/', 'http', '../')) or
                                                        any(sl.endswith(e) for e in (
                                                            '.pdf','.doc','.docx','.xls','.xlsx',
                                                            '.ppt','.pptx','.zip','.rar','.7z',
                                                            '.txt','.odt','.ods')) or '?' in sl)
                                            fp = _a1 if _isp(_a1) else (_a2 if _isp(_a2) else _a1)
                                        elif _m1:
                                            fp = _m1.group(2).strip()
                                        else:
                                            fp = (a_tag.get('href') or '').strip()
                                        if not fp or fp in ('#',) or fp.startswith('javascript'):
                                            return None
                                        if not fp.startswith('http'):
                                            fp = urljoin(base_url, fp)
                                        return fp
                                    for a in all_links:
                                        link = _pick_link(a, self.base)
                                        text = a.get_text(strip=True) or 'Tài liệu'
                                        if not link:
                                            continue
                                        attachment_html += (f'➤ <a href="{link}" target="_blank"'
                                                            f' class="link-download">{text}</a><br/>')
                                        _file_cnt += 1
                                    attachment_html += "</div><hr/>"

                            _img_cnt = len(tag_content.find_all('img')) if tag_content else 0
                            _tc_size = len(str(tag_content))/1024 if tag_content else 0
                            log.detail_loaded(_tc_size, "title-detail",
                                              saved_date_str, _img_cnt, _file_cnt)

                            # ── Pipeline chuẩn (thống nhất với SMART DETECTION) ──
                            if tag_content:
                                # normalize TRƯỚC clean (clean sẽ xóa onclick nhưng giữ class link-download)
                                normalize_download_links_in_content(tag_content, base_url=o_url)
                                normalize_img_srcs_in_content(tag_content, base_url=o_url)
                                find_and_merge_attachments(soup, tag_content)
                                embed_youtube_links_in_content(tag_content, soup)
                                convert_rg_gallery_to_imgs(tag_content, self.base)
                                clean_html_content(tag_content)
                                _tc_str = str(tag_content) if _tc_has_content else ""
                            else:
                                _tc_str = ""
                            cam.description = attachment_html + _doc_btn_html + _tc_str

                            # Ảnh đại diện: ưu tiên thumbnail từ listing, fallback ảnh đầu trong content
                            _thumb = getattr(cam, '_thumbnail_url', '')
                            if not _thumb:
                                for _fi in (tag_content.find_all('img') if tag_content else []):
                                    _src = _fi.get('src', '') or ''
                                    _sl2 = _src.lower()
                                    # Lazy placeholder → thử data-src
                                    if (not _src or _src.startswith('data:') or
                                            any(_p in _sl2 for _p in ('blank.gif','loading.gif',
                                                                       'spacer.gif','transparent.gif'))):
                                        _src = (_fi.get('data-src') or _fi.get('data-original') or '')
                                    if not _src:
                                        continue
                                    if 'no-image' in _src.lower():
                                        continue
                                    _thumb = urljoin(self.base, _src) if _src.startswith('/') else _src
                                    break   # ảnh đầu tiên hợp lệ
                            # Fallback: YouTube thumbnail nếu không có ảnh nào trong content
                            if not _thumb:
                                for _ifr in (tag_content.find_all('iframe') if tag_content else []):
                                    _yt_id = _yt_video_id(_ifr.get('src', ''))
                                    if _yt_id:
                                        _thumb = f"https://img.youtube.com/vi/{_yt_id}/maxresdefault.jpg"
                                        log.info(f"  🎬 Thumbnail từ YouTube: {_yt_id}")
                                        break
                            if _thumb:
                                cam.Photo = [_thumb]

                            # ── Kiểm tra trùng tiêu đề+nội dung trước khi lưu ──
                            # (display_info gọi bên trong _try_save_cam, không gọi lại ở đây)
                            self._try_save_cam(cam, o_url, t_start,
                                               SiteStructureType.TYPE_TITLE_DETAIL, soup=soup)
                            return

            # ════════════════════════════════════════════════════════════════
            # 3. LISTING — ArticleList / down-list / phal-list
            # Dùng cờ _listing_handled: chỉ chạy 1 handler listing / trang
            # (tránh double-processing khi nhiều handler cùng match 1 trang)
            # Thứ tự ưu tiên: handler CÓ thumbnail trước, handler không có thumbnail sau
            # ════════════════════════════════════════════════════════════════
            if not chitiet:
                _listing_handled = (log._list_page_added + log._list_page_skipped) > 0
                _gidx = log._list_total_added + log._list_total_skip + 1

                # ── ul.media-list li.media (hanam.edu.vn CMS mới) ───────────
                # Có thumbnail từ img.media-object — ưu tiên cao nhất
                if not _listing_handled:
                    _media_lis = soup.select('ul.media-list li.media')
                    for _ml in _media_lis:
                        _h4 = (_ml.find('h4', class_='media-heading')
                               or _ml.find('h4', class_='title-content-new')
                               or _ml.find('h4') or _ml.find('h3'))
                        if not _h4: continue
                        _a_img  = _ml.select_one('a.pull-left[href]')
                        _a_body = _ml.select_one('div.media-body a[href]')
                        tag_a   = _a_img or _a_body
                        if not tag_a: continue
                        href_raw = tag_a.get('href') or ''
                        if 'javascript' in href_raw or not href_raw: continue
                        name    = _h4.get_text(strip=True)
                        url     = self._norm_url(urljoin(self.base, href_raw))
                        has_img = bool(_ml.find('img', class_='media-object')
                                       or _ml.find('img'))
                        if not self._in_same_section(url, o_url, self.cat_id):
                            log.list_item(_gidx, 0, name, url, 'filtered_section', has_img)
                            _gidx += 1; continue
                        result = self._smart_add(name, url, self.cat_id, hp, tag=_ml)
                        log.list_item(_gidx, 0, name, url, result, has_img)
                        _gidx += 1
                    if _media_lis:
                        _listing_handled = True

                # ── video-grid-type8 trong section.video-type2 (video listing) ─
                # Cấu trúc: section.section-grid.video-type2 > div.row
                #            > div.item-video > article.video-grid-type8
                #            > div.avatar > a[href] + img
                #            > div.prim-wrap > h3.prim-titl > a (tiêu đề)
                if not _listing_handled:
                    _vg_items = soup.select('div.item-video article.video-grid-type8')
                    for _vi in _vg_items:
                        _vh = (_vi.find('h3', class_='prim-titl') or
                               _vi.find('h2', class_='prim-titl') or
                               _vi.find('h3') or _vi.find('h2'))
                        if not _vh:
                            continue
                        _va = _vh.find('a', href=True)
                        if not _va:
                            _va = _vi.select_one('div.avatar a[href]')
                        if not _va:
                            continue
                        href_r = _va.get('href') or ''
                        if not href_r or 'javascript' in href_r:
                            continue
                        name    = _va.get('title') or _va.get_text(strip=True)
                        url     = self._norm_url(urljoin(self.base, href_r))
                        _img    = _vi.select_one('div.avatar img')
                        has_img = bool(_img)
                        if not self._in_same_section(url, o_url, self.cat_id):
                            log.list_item(_gidx, 0, name, url, 'filtered_section', has_img)
                            _gidx += 1
                            continue
                        result = self._smart_add(name, url, self.cat_id, hp, tag=_vi)
                        log.list_item(_gidx, 0, name, url, result, has_img)
                        _gidx += 1
                    if _vg_items:
                        _listing_handled = True

                # ── phal-list row (CMS edu.vn gallery/photo listing) ──────
                # Cấu trúc cũ: ul.phal-list.row > li > h2.Title a
                # Cấu trúc mới: ul.phal-list.row > li > article.video-grid-type8
                #               > div.prim-wrap > h3.prim-titl > a
                # Có thumbnail từ div.avatar > img trong li
                if not _listing_handled:
                    _phal_ul = soup.find('ul', class_=lambda c: c and 'phal-list' in c and 'row' in c)
                    if _phal_ul:
                        _phal_items = _phal_ul.find_all('li')
                        for _pli in _phal_items:
                            _ph = (_pli.find('h3', class_='prim-titl') or  # video-grid-type8 style
                                   _pli.find('h2', class_='prim-titl') or
                                   _pli.find('h2', class_='Title') or
                                   _pli.find('h3', class_='Title') or
                                   _pli.find('h4', class_='Title') or
                                   _pli.find('h2') or _pli.find('h3') or _pli.find('h4'))
                            if not _ph:
                                continue
                            _pa = _ph.find('a', href=True)
                            if not _pa:
                                _pa = _pli.find('a', href=True)
                            if not _pa:
                                continue
                            name    = _pa.get('title') or _pa.get_text(strip=True)
                            href_r  = _pa.get('href') or ''
                            if not href_r or 'javascript' in href_r:
                                continue
                            url     = self._norm_url(urljoin(self.base, href_r))
                            _img    = (_pli.select_one('div.avatar img') or
                                       _pli.find('img'))
                            has_img = bool(_img)
                            if not self._in_same_section(url, o_url, self.cat_id):
                                log.list_item(_gidx, 0, name, url, 'filtered_section', has_img)
                                _gidx += 1
                                continue
                            result = self._smart_add(name, url, self.cat_id, hp, tag=_pli)
                            log.list_item(_gidx, 0, name, url, result, has_img)
                            _gidx += 1
                        if _phal_items:
                            _listing_handled = True

                # ── ArticleList / down-list / phal-list ───────────────────
                tag_ul = (soup.find('ul', class_='ArticleList')
                          or soup.find('ul', class_='down-list')
                          or soup.find('ul', class_='phal-list row'))
                if tag_ul and not _listing_handled:
                    tag_ils = tag_ul.find_all('li', class_='row') or tag_ul.find_all('li')
                    for tag_il in tag_ils:
                        h2 = (tag_il.find('h2', class_='Title') or tag_il.find('h5')
                              or tag_il.find('h4') or tag_il.find('div', class_='avatar'))
                        if not h2: continue
                        a3 = h2.find('a')
                        if not a3: continue
                        name = a3.text.strip()
                        url  = self._norm_url(urljoin(self.base, a3.get('href') or ''))
                        has_img = bool(tag_il.find('img'))
                        if not self._in_same_section(url, o_url, self.cat_id):
                            log.list_item(_gidx, 0, name, url,
                                          "filtered_section", has_img)
                            _gidx += 1
                            continue
                        result = self._smart_add(name, url, self.cat_id, hp, tag=tag_il)
                        log.list_item(_gidx, 0, name, url, result, has_img)
                        _gidx += 1
                    if tag_ils:
                        _listing_handled = True

                # ── div.list-item (hanam.edu.vn / edu.vn CMS) ────────────
                # Đặt TRƯỚC div.post-title vì handler này có thumbnail (img-responsive)
                if not _listing_handled:
                    list_item_divs = soup.find_all('div', class_='list-item')
                    _valid_li = [t for t in list_item_divs
                                 if t.find('div', class_='news-item-name')]
                    for tag_li in _valid_li:
                        name_div = tag_li.find('div', class_='news-item-name')
                        tag_a = name_div.find('a', href=True)
                        if not tag_a: continue
                        name = tag_a.get_text(strip=True)
                        url  = self._norm_url(urljoin(self.base, tag_a.get('href') or ''))
                        has_img = bool(tag_li.find('img', class_='img-responsive'))
                        if not self._in_same_section(url, o_url, self.cat_id):
                            log.list_item(_gidx, 0, name, url, "filtered_section", has_img)
                            _gidx += 1
                            continue
                        result = self._smart_add(name, url, self.cat_id, hp, tag=tag_li)
                        log.list_item(_gidx, 0, name, url, result, has_img)
                        _gidx += 1
                    if _valid_li:
                        _listing_handled = True

                # ── LIST-G: div.list-news > div.news-content ─────────────
                # Đặt TRƯỚC div.post-title vì handler này có thumbnail
                if not _listing_handled:
                    list_news = soup.find('div', class_='list-news')
                    if list_news:
                        news_items = list_news.find_all('div', class_='news-content')
                        for tag_ni in news_items:
                            col_right = tag_ni.find('div', class_=lambda c: c and 'col-md-8' in c)
                            tag_a = (col_right.find('a', href=True) if col_right
                                     else tag_ni.find('a', href=True))
                            if not tag_a: continue
                            name = tag_a.get_text(strip=True)
                            href = tag_a.get('href') or ''
                            if not href or 'javascript' in href: continue
                            url  = self._norm_url(urljoin(self.base, href))
                            has_img = bool(tag_ni.find('img'))
                            if not self._in_same_section(url, o_url, self.cat_id):
                                log.list_item(_gidx, 0, name, url, "filtered_section", has_img)
                                _gidx += 1
                                continue
                            result = self._smart_add(name, url, self.cat_id, hp, tag=tag_ni)
                            log.list_item(_gidx, 0, name, url, result, has_img)
                            _gidx += 1
                        if news_items:
                            _listing_handled = True

                # ── post-title / item-info (fallback — không có thumbnail) ──
                if not _listing_handled:
                    tag_divs = (soup.find_all('div', class_='post-title')
                                or soup.find_all('div', class_='title-news-listType10')
                                or soup.find_all('div', class_='item-info'))
                    for tag_il in tag_divs:
                        h4 = (tag_il.find('h4', class_='entry-title')
                              or tag_il.find('h2') or tag_il.find('h5'))
                        if not h4: continue
                        tag_a = h4.find("a")
                        if not tag_a: continue
                        name = tag_a.text.strip()
                        url  = self._norm_url(urljoin(self.base, tag_a.get('href') or ''))
                        # Thử lấy img từ article cha (nếu có) để có thumbnail
                        parent_art = tag_il.find_parent('article')
                        has_img    = bool(parent_art.find('img') if parent_art else tag_il.find('img'))
                        if not self._in_same_section(url, o_url, self.cat_id):
                            log.list_item(_gidx, 0, name, url, "filtered_section", has_img)
                            _gidx += 1
                            continue
                        result = self._smart_add(name, url, self.cat_id, hp,
                                                 tag=parent_art or tag_il)
                        log.list_item(_gidx, 0, name, url, result, has_img)
                        _gidx += 1
                    if tag_divs:
                        _listing_handled = True

                # ── a.title-documment ─────────────────────────────────────
                if not _listing_handled:
                    docs = [a for a in soup.find_all('a', class_='title-documment')
                            if self._in_same_section(
                                self._norm_url(urljoin(self.base, a.get('href') or '')), o_url, self.cat_id)]
                    for tag_a in docs:
                        name = tag_a.text.strip()
                        url  = self._norm_url(urljoin(self.base, tag_a.get('href') or ''))
                        result = self._smart_add(name, url, self.cat_id, hp)
                        log.list_item(_gidx, 0, name, url, result, has_img=False)
                        _gidx += 1
                    if docs:
                        _listing_handled = True



            # ════════════════════════════════════════════════════════════════
            # 4. DETAIL — TYPE 2: ArticleHeader
            # ════════════════════════════════════════════════════════════════
            tag_title_product = soup.find('div', class_="ArticleHeader")

            if chitiet and tag_title_product:

                log.section("📄 CHI TIẾT — TYPE-2 (ArticleHeader)", Color.GREEN)
                tag_content = soup.find('div', class_='ArticleContent')
                log.parse_info("ArticleContent", bool(tag_content))

                d = parse_vn_date_from_soup(soup)
                if d:
                    saved_date_str = d.strftime("%d/%m/%Y")
                    log.key_value("Date", saved_date_str, Color.YELLOW)
                    if d < FROM_DATE:
                        log.warning(f"Bỏ qua — ngày {saved_date_str} trước FROM_DATE {FROM_DATE}")
                        stats.record_url(o_url, SiteStructureType.TYPE_ARTICLE_HEADER)
                        return

                if tag_content:
                    for cam in self.camlist.camobs:
                        if self._norm_url(cam.url) == o_url:
                            cam.date_publish = d
                            cam.short        = self.target
                            if 'Công khai' in cam.name:
                                cam.cat_id = 'Công khai'

                            normalize_download_links_in_content(tag_content, base_url=o_url)
                            normalize_img_srcs_in_content(tag_content, base_url=o_url)
                            find_and_merge_attachments(soup, tag_content)
                            embed_youtube_links_in_content(tag_content, soup)
                            convert_rg_gallery_to_imgs(tag_content, self.base)
                            clean_html_content(tag_content)
                            cam.description = str(tag_content)

                            # Ảnh đại diện: ưu tiên ảnh trong content, fallback YouTube thumbnail
                            _thumb = getattr(cam, '_thumbnail_url', '')
                            if not _thumb:
                                for _fi in (tag_content.find_all('img') if tag_content else []):
                                    _src = _fi.get('src', '') or ''
                                    _sl2 = _src.lower()
                                    if (not _src or _src.startswith('data:') or
                                            any(_p in _sl2 for _p in ('blank.gif','loading.gif',
                                                                       'spacer.gif','transparent.gif'))):
                                        _src = (_fi.get('data-src') or _fi.get('data-original') or '')
                                    if not _src:
                                        continue
                                    if 'no-image' in _src.lower():
                                        continue
                                    _thumb = urljoin(self.base, _src) if _src.startswith('/') else _src
                                    break
                            # Fallback: YouTube thumbnail
                            if not _thumb:
                                for _ifr in (tag_content.find_all('iframe') if tag_content else []):
                                    _yt_id = _yt_video_id(_ifr.get('src', ''))
                                    if _yt_id:
                                        _thumb = f"https://img.youtube.com/vi/{_yt_id}/maxresdefault.jpg"
                                        break
                            if _thumb:
                                cam.Photo = [_thumb]

                            log.html_preview(tag_content)
                            # ── Kiểm tra trùng tiêu đề+nội dung trước khi lưu ──
                            self._try_save_cam(cam, o_url, t_start,
                                               SiteStructureType.TYPE_ARTICLE_HEADER, soup=soup)
                            # log.success("✓ Đã lưu (TYPE-2)") ← bỏ: _try_save_cam đã log
                            log.end_url(SiteStructureType.TYPE_ARTICLE_HEADER, "ok",
                                        saved_date_str,
                                        (datetime.now() - t_start).total_seconds())
                            log.end_section()
                            return

            # ════════════════════════════════════════════════════════════════
            # 5. LISTING — Công khai ngân sách (td.tg-yw4l)
            # ════════════════════════════════════════════════════════════════
            if not chitiet:
                detail_links = pick_detail_links(soup)
                _ck_links = [(tag_a.get_text(" ", strip=True),
                              self._norm_url(urljoin(self.base, (tag_a.get("href") or "").strip())))
                             for tag_a in detail_links
                             if (tag_a.get("href") or "").strip()
                             and self._in_same_section(
                                 self._norm_url(urljoin(self.base, (tag_a.get("href") or "").strip())), o_url, self.cat_id)]
                if _ck_links:
                    log.section(f"📝 DANH SÁCH — Công khai ngân sách ({len(_ck_links)} link)", Color.CYAN)
                    for name, url in _ck_links:
                        log.key_value("Detail", name[:80], Color.CYAN)
                        result = self._smart_add(name, url, self.cat_id, hp)
                        if result == "added":
                            log.success("  ✓ Đã thêm")
                        elif result == "replaced":
                            log.warning(f"  ↺ Thay thế bài cũ (ID mới hơn)")
                        elif result == "kept_old":
                            log.debug(f"  → Giữ bài cũ")
                        elif result == "skipped_url":
                            log.debug(f"  → Trùng URL")
                    log.end_section()

            # ════════════════════════════════════════════════════════════════
            # 6. SMART CONTENT DETECTION — fallback cho detail page
            # ════════════════════════════════════════════════════════════════
            if chitiet:

                log.section("📄 SMART DETECTION — fallback detail", Color.YELLOW)

                # Thứ tự ưu tiên — thêm selector mới vào đây nếu gặp website lạ
                # QUAN TRỌNG: div.content-detail phải trước div.article-content
                # vì article-content là container cha bao gồm cả tiêu đề + rác
                content_selectors = [
                    ('div#ctl01_divContent',                  SiteStructureType.TYPE_CTL_DIVCONTENT),
                    ('div.content-detail',                    SiteStructureType.TYPE_TITLE_DETAIL),   # ← ưu tiên cao hơn article-content
                    ('div.article-content div.block-core-a5', SiteStructureType.TYPE_ARTICLE_CONTENT), # chỉ lấy phần nội dung thuần
                    ('div.article-content',                   SiteStructureType.TYPE_ARTICLE_CONTENT),
                    ('div#news-content',                      SiteStructureType.TYPE_VIEW_DETAIL),
                    ('div.view-detail',                       SiteStructureType.TYPE_VIEW_DETAIL),
                    ('div.detail-content',                    SiteStructureType.TYPE_VIEW_DETAIL),
                    ('div.news-detail-content',               SiteStructureType.TYPE_NEWS_DETAIL),
                    ('div.v_content',                         SiteStructureType.TYPE_VIEW_DETAIL),
                    ('section.news-detail-default',           SiteStructureType.TYPE_NEWS_DETAIL),
                    ('article.Article-News',                  SiteStructureType.TYPE_ARTICLE_NEWS),
                    ('article.download-detail-layout-default',SiteStructureType.TYPE_DOWNLOAD_DETAIL),
                    ('div.UICongKhaiNganSach_Default',        SiteStructureType.TYPE_CONGKHAI_NGAN_SACH),
                    ('div#module34',                          SiteStructureType.TYPE_MODULE34),
                    ('div#module16',                          SiteStructureType.TYPE_MODULE16),
                    # ── Bổ sung: CMS edu.vn / portal VN ─────────────────────
                    ('div.entry-content',                     SiteStructureType.TYPE_ENTRY_CONTENT),
                    ('div.main-content',                      SiteStructureType.TYPE_MAIN_CONTENT),
                    ('div.content-main',                      SiteStructureType.TYPE_MAIN_CONTENT),
                    ('div.box-content',                       SiteStructureType.TYPE_BOX_CONTENT),
                    ('div.page-content',                      SiteStructureType.TYPE_BOX_CONTENT),
                    ('div.content-inner',                     SiteStructureType.TYPE_CONTENT_INNER),
                    ('div.detail-inner',                      SiteStructureType.TYPE_CONTENT_INNER),
                    ('div.post-content',                      SiteStructureType.TYPE_ENTRY_CONTENT),
                    ('div.single-content',                    SiteStructureType.TYPE_ENTRY_CONTENT),
                    ('div.thuc-don',                          SiteStructureType.TYPE_THUC_DON),
                    ('table.thuc-don',                        SiteStructureType.TYPE_THUC_DON),
                    ('div.zone-content',                      SiteStructureType.TYPE_ZONE_CONTENT),
                    ('div.region-content',                    SiteStructureType.TYPE_ZONE_CONTENT),
                    ('div.portlet-content',                   SiteStructureType.TYPE_ZONE_CONTENT),
                    ('div#ctl00_ContentPlaceHolder1_divContent', SiteStructureType.TYPE_CTL_DIVCONTENT),
                    ('div#ContentPlaceHolder1_divContent',    SiteStructureType.TYPE_CTL_DIVCONTENT),
                    # ── inline-download-Type4 ─────────────────────────────────
                    ('div.inline-download',                   SiteStructureType.TYPE_INLINE_DOWNLOAD),
                    ('div.download-type4',                    SiteStructureType.TYPE_INLINE_DOWNLOAD),
                    ('ul.list-download.type4',                SiteStructureType.TYPE_INLINE_DOWNLOAD),
                    ('div.listfile',                          SiteStructureType.TYPE_INLINE_DOWNLOAD),
                    # ── hanam.edu.vn — div.media.news / div#gioithieu_noidung ──
                    ('div.media.news',                        SiteStructureType.TYPE_MEDIA_NEWS),
                    ('div#gioithieu_noidung',                 SiteStructureType.TYPE_MEDIA_NEWS),
                    # ── video-embed (YouTube / youtube-nocookie / video nhúng) ──
                    ("iframe[src*='youtube.com']",            SiteStructureType.TYPE_VIDEO_EMBED),
                    ("iframe[src*='youtube-nocookie.com']",   SiteStructureType.TYPE_VIDEO_EMBED),
                    ("iframe[src*='youtu.be']",               SiteStructureType.TYPE_VIDEO_EMBED),
                    ('div.video-embed',                       SiteStructureType.TYPE_VIDEO_EMBED),
                    ('div.post-video',                        SiteStructureType.TYPE_VIDEO_EMBED),
                    ('div.videoWrapper',                      SiteStructureType.TYPE_VIDEO_EMBED),
                    # ── Fallback rộng: tìm thẻ article bất kỳ ───────────────
                    ('table.table',                           SiteStructureType.TYPE_TABLE_LAYOUT),
                    ('article',                               SiteStructureType.TYPE_ARTICLE_NEWS),
                    # ── Fallback cuối: div.content (generic) ─────────────────
                    ('div.content',                           SiteStructureType.TYPE_DIV_CONTENT),
                ]

                tag_content        = None
                detected_structure = SiteStructureType.UNKNOWN

                for selector, struct_type in content_selectors:
                    el = soup.select_one(selector)
                    if el is None:
                        continue
                    # Chấp nhận element nếu:
                    #   - có text > 5 ký tự, HOẶC
                    #   - chứa iframe (video embed), HOẶC
                    #   - chứa link tải file (href có extension)
                    has_content = (
                        len(el.get_text(strip=True)) > 5
                        or bool(el.find('iframe'))
                        or bool(el.find('a', href=True))
                    )
                    if has_content:
                        tag_content        = el
                        detected_structure = struct_type
                        log.success(f"✓ Content tìm thấy: {selector} → {struct_type.value}")
                        break

                if not tag_content:
                    # Thật sự không tìm được — cảnh báo rõ
                    elapsed = (datetime.now()-t_start).total_seconds()
                    log.detail_error("Không detect được cấu trúc HTML", elapsed)
                    stats.record_url(o_url, SiteStructureType.UNKNOWN,
                                     success=False, error_msg="Không detect được cấu trúc")
                    # Lấy tiêu đề từ camlist (URL đã được thêm vào queue từ listing)
                    _found_cam = next(
                        (c for c in self.camlist.camobs
                         if self._norm_url(c.url) == o_url), None)
                    _fail_title = _found_cam.name if _found_cam else ""
                    stats.record_failed(_fail_title, o_url, self.base, self.cat_id,
                                        "Không detect được cấu trúc HTML",
                                        row=self._excel_row, fail_type="article")
                    return

                # ── Xử lý đặc biệt TYPE-22: trang chỉ có video YouTube ───────
                # Nếu detect ra video-embed nhưng tag_content là iframe trực tiếp
                # → bọc lại thành div để pipeline xử lý bình thường
                if detected_structure == SiteStructureType.TYPE_VIDEO_EMBED:
                    if tag_content.name == 'iframe':
                        wrapper = soup.new_tag('div', **{'class': 'video-embed-wrapper'})
                        iframe_copy = copy.copy(tag_content)
                        wrapper.append(iframe_copy)
                        tag_content = wrapper
                    # Nếu trong trang có thêm text content → fallback sang body
                    body_text = soup.body.get_text(strip=True) if soup.body else ""
                    if len(body_text) > 200:
                        # Trang có text thực — giữ nguyên tag_content đã tìm được
                        pass
                    log.info(f"🎬 TYPE-22 video-embed: {tag_content.name}")

                d = parse_vn_date_from_soup(soup)
                if d:
                    saved_date_str = d.strftime("%d/%m/%Y")
                    if d < FROM_DATE:
                        log.warning(f"Bỏ qua — ngày {saved_date_str} trước FROM_DATE {FROM_DATE}")
                        stats.record_url(o_url, detected_structure)
                        return

                # Tìm cam tương ứng với URL này trong camlist (đến từ listing)
                _target_cam = next(
                    (c for c in self.camlist.camobs if self._norm_url(c.url) == o_url), None)

                # Nếu không có cam nào khớp (URL trực tiếp từ Excel, không qua listing)
                # → tạo synthetic cam với title lấy từ h1 của trang
                if _target_cam is None and not chitiet:
                    _title_from_page = ""
                    for _sel in ('h1.title-detail', 'h1.entry-title', 'h1',
                                 '.post-title h1', 'article h1'):
                        _el = soup.select_one(_sel)
                        if _el:
                            _t = clean_spaces(_el.get_text(" ", strip=True))
                            if _t and len(_t) > 3:
                                _title_from_page = _t
                                log.info(f"  ℹ URL trực tiếp — title từ trang: {_t[:60]}")
                                break
                    _target_cam = camob.CameraObject(0, _title_from_page, 0, o_url, "", self.cat_id)
                    self.camlist.add_cam(_target_cam)
                    log.info("  ℹ Tạo cam tổng hợp cho URL trực tiếp từ Excel")

                for cam in self.camlist.camobs:
                    if self._norm_url(cam.url) == o_url:
                        cam.date_publish = d
                        cam.short        = self.target
                        if 'Công khai' in cam.name:
                            cam.cat_id = 'Công khai'

                        normalize_download_links_in_content(tag_content, base_url=o_url)
                        normalize_img_srcs_in_content(tag_content, base_url=o_url)
                        find_and_merge_attachments(soup, tag_content)
                        embed_youtube_links_in_content(tag_content, soup)
                        convert_rg_gallery_to_imgs(tag_content, self.base)
                        _img_cnt2 = len(tag_content.find_all('img'))
                        _file_cnt2 = len(tag_content.find_all('a', href=True))
                        log.detail_loaded(len(str(tag_content))/1024,
                                          _struct_name, saved_date_str,
                                          _img_cnt2, _file_cnt2)
                        clean_html_content(tag_content)
                        cam.description = str(tag_content)

                        # Ảnh đại diện: ưu tiên ảnh trong content, fallback YouTube thumbnail
                        _thumb = getattr(cam, '_thumbnail_url', '')
                        if not _thumb:
                            _first_img = tag_content.find('img')
                            if _first_img:
                                _src = _first_img.get('src', '')
                                if _src and 'no-image' not in _src.lower():
                                    _thumb = urljoin(self.base, _src) if _src.startswith('/') else _src
                        # Fallback: lấy thumbnail từ YouTube embed nếu không có ảnh
                        if not _thumb:
                            for _ifr in tag_content.find_all('iframe'):
                                _yt_id = _yt_video_id(_ifr.get('src', ''))
                                if _yt_id:
                                    _thumb = f"https://img.youtube.com/vi/{_yt_id}/maxresdefault.jpg"
                                    log.info(f"  🎬 Thumbnail từ YouTube: {_yt_id}")
                                    break
                        if _thumb:
                            cam.Photo = [_thumb]

                        # ── Kiểm tra trùng tiêu đề+nội dung trước khi lưu ──
                        self._try_save_cam(cam, o_url, t_start,
                                           detected_structure, soup=soup)
                        return

                log.end_section()

            # ════════════════════════════════════════════════════════════════
            # 7. PAGINATION
            # ════════════════════════════════════════════════════════════════
            if chitiet:
                time.sleep(2)
                elapsed = (datetime.now() - t_start).total_seconds()
                stats.record_url(o_url, detected_structure)
                log.end_url(detected_structure, "ok", saved_date_str, elapsed)
                return

            # ── Giới hạn cứng để tránh loop vô hạn ──────────────────────────
            if _page_num >= MAX_LIST_PAGES:
                log.warning(f"Đã tải {MAX_LIST_PAGES} trang — dừng (giới hạn an toàn, xem MAX_LIST_PAGES)")
                elapsed = (datetime.now() - t_start).total_seconds()
                stats.record_url(o_url, detected_structure)
                log.list_coverage_check(list_url=o_url, base=get_base(o_url))
                log.list_done(elapsed)
                break

            # ── Dừng nếu đã qua hết tổng số trang đã biết ───────────────────
            if _total_pages > 0 and _page_num >= _total_pages:
                elapsed = (datetime.now() - t_start).total_seconds()
                stats.record_url(o_url, detected_structure)
                log.list_coverage_check(list_url=o_url, base=get_base(o_url))
                log.list_done(elapsed)
                break

            # ── CMS loop detection: trang mới trùng URL trang trước → dừng ──
            # log._cur_page_urls được populate bởi list_item()
            # log._prev_page_urls được gán bởi list_next_page()
            if (_page_num > 1
                    and log._cur_page_urls
                    and log._prev_page_urls
                    and log._cur_page_urls == log._prev_page_urls):
                log.warning(f"⚠ CMS loop phát hiện — trang {_page_num} trùng trang trước, dừng")
                elapsed = (datetime.now() - t_start).total_seconds()
                stats.record_url(o_url, detected_structure)
                log.list_coverage_check(list_url=o_url, base=get_base(o_url))
                log.list_done(elapsed)
                break

            log.section("PHÂN TRANG", Color.MAGENTA)

            btns = self.driver.find_elements(By.CSS_SELECTOR, 'div.col-center a.next')
            if btns:
                log.parse_info("a.next button", True)
                b = btns[0]
                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", b)
                time.sleep(0.2)
                self.driver.execute_script("arguments[0].click();", b)
                time.sleep(2.5)
                html = self.driver.page_source
                _page_num += 1
                log.list_next_page(_page_num)
                continue

            # Fallback: hanam.edu.vn — ?dm=&page=N  (div.page)
            btns_hanam = self.driver.find_elements(
                By.CSS_SELECTOR, 'div.page ul li a')
            # Fallback thêm: div.page-news (CMS mới hanam.edu.vn — LIST-G)
            if not btns_hanam:
                btns_hanam = self.driver.find_elements(
                    By.CSS_SELECTOR, 'div.page-news ul li a')
            moved_hanam = False
            for btn in btns_hanam:
                btn_page = to_int(btn.text, default=-1)
                if btn_page > int(self.page_link):
                    self.driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});", btn)
                    time.sleep(0.2)
                    self.driver.execute_script("arguments[0].click();", btn)
                    time.sleep(2.0)
                    self.page_link += 1
                    html = self.driver.page_source
                    _page_num += 1
                    log.list_next_page(_page_num)
                    moved_hanam = True
                    break
            if moved_hanam:
                log.end_section()
                continue

            # Fallback 3: a.page-link (Bootstrap CMS)
            log.parse_info("a.next button", False)
            btns  = self.driver.find_elements(By.CSS_SELECTOR, 'a.page-link')
            moved = False
            for btn in btns:
                btn_page = to_int(btn.text, default=-1)
                if btn_page > int(self.page_link):
                    self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                    time.sleep(0.2)
                    self.driver.execute_script("arguments[0].click();", btn)
                    time.sleep(1.5)
                    self.page_link += 1
                    html = self.driver.page_source
                    _page_num += 1
                    log.list_next_page(_page_num)
                    moved = True
                    break

            # Fallback 4: div.col-center a (số trang — CMS SSDH edu.vn / mnthuyson)
            # Selector này được dùng trong detect _total_pages nên phải có cùng ở đây
            if not moved:
                for _sel4 in ['div.col-center a', 'div.default-pagination a',
                              'ul.pagination a', 'nav.pagination a']:
                    _btns4 = self.driver.find_elements(By.CSS_SELECTOR, _sel4)
                    for btn in _btns4:
                        btn_page = to_int(btn.text.strip(), default=-1)
                        if btn_page == _page_num + 1:
                            self.driver.execute_script(
                                "arguments[0].scrollIntoView({block:'center'});", btn)
                            time.sleep(0.2)
                            self.driver.execute_script("arguments[0].click();", btn)
                            time.sleep(2.0)
                            self.page_link = btn_page
                            html = self.driver.page_source
                            _page_num += 1
                            log.list_next_page(_page_num)
                            moved = True
                            break
                    if moved:
                        break

            if not moved:
                elapsed = (datetime.now() - t_start).total_seconds()
                stats.record_url(o_url, detected_structure)
                log.list_coverage_check(list_url=o_url, base=get_base(o_url))
                log.list_done(elapsed)
                break

            log.end_section()

    # ── discover_sub_menus ────────────────────────────────────────────────────

    def _is_valid_nav_href(self, href: str) -> bool:
        """Lọc bỏ href rác: javascript, #, mailto, tel, link ngoài domain."""
        if not href:
            return False
        low = href.lower().strip()
        if low.startswith(("javascript:", "#", "mailto:", "tel:")):
            return False
        full = urljoin(self.base, href)
        return full.startswith(self.base.rstrip('/'))

    def _collect_links_from_li(self, li_tag, collected: dict,
                               depth: int = 0, parent_cat: str = ""):
        """
        Đệ quy lấy tất cả href trong <li> này và mọi <li> con.
        collected: dict  url → cat_id  (cat_id = text của <a> đó trong menu)
        parent_cat: cat_id của cấp cha (dùng làm fallback nếu <a> không có text)
        """
        indent = "  " * depth

        direct_a = li_tag.find('a', href=True, recursive=False)
        this_cat = parent_cat   # mặc định kế thừa cha

        if direct_a and self._is_valid_nav_href(direct_a['href']):
            full     = urljoin(self.base, direct_a['href']).rstrip('/')
            this_cat = direct_a.get_text(strip=True) or parent_cat
            if full not in collected:
                collected[full] = this_cat
                log.debug(f"{indent}[cấp {depth}] «{this_cat}» → {full}")

        for sub_ul in li_tag.find_all('ul', recursive=False):
            for sub_li in sub_ul.find_all('li', recursive=False):
                self._collect_links_from_li(sub_li, collected,
                                            depth + 1, parent_cat=this_cat)

    def discover_sub_menus(self, soup) -> dict[str, str]:
        """
        Tìm TẤT CẢ link con trong menu, đệ quy mọi cấp.
        Trả về dict: { url → cat_id }
        cat_id = text của thẻ <a> tương ứng trong menu.

        Chiến lược:
          1. Tìm <a> trong nav/header/div.menu trỏ đúng self.url (nhiều selector)
          2. Leo lên <li> cha → đệ quy toàn bộ <ul>/<li> con
          3. Fallback: quét toàn bộ zone lấy mọi link có prefix = self.url
        """
        log.section("KHÁM PHÁ SUB-MENU (đệ quy toàn cấp)", Color.CYAN)

        target_url    = self.url.rstrip('/')
        collected: dict[str, str] = {}   # url → cat_id
        found_main_li = None
        root_cat      = ""

        # ── Bước 1: Tìm zone chứa nav ────────────────────────────────────────
        # CMS edu.vn thường dùng <div class="main-nav"> hoặc <div class="menu-…">
        # thay vì thẻ HTML5 <nav> → cần thử nhiều selector
        def _find_search_zone(soup):
            # Ưu tiên thẻ <nav> chuẩn
            z = soup.find('nav')
            if z:
                return z
            # Thẻ <header> (thường bọc nav)
            z = soup.find('header')
            if z:
                return z
            # CMS edu.vn: div.main-nav, div.nav-main, div.menu-main...
            for cls_pat in ('main-nav', 'nav-main', 'menu-main', 'main-menu',
                            'navigation', 'site-navigation', 'primary-nav',
                            'top-nav', 'topnav', 'mainmenu', 'main_nav'):
                z = soup.find(attrs={'class': re.compile(r'\b' + cls_pat + r'\b', re.I)})
                if z:
                    return z
            # id-based
            for id_pat in ('nav', 'menu', 'navigation', 'mainmenu', 'main-nav'):
                z = soup.find(id=re.compile(r'^' + id_pat + r'$', re.I))
                if z:
                    return z
            # Fallback: toàn bộ body
            return soup

        search_zone = _find_search_zone(soup)
        zone_name   = getattr(search_zone, 'name', '?')
        zone_class  = ' '.join((search_zone.get('class') or []))[:40] if hasattr(search_zone, 'get') else ''
        log.debug(f"Search zone: <{zone_name} class='{zone_class}'>")

        # ── Bước 2: Tìm <li> chứa link gốc ──────────────────────────────────
        for a in search_zone.find_all('a', href=True):
            # Normalize cả 2 đầu để so sánh chính xác (bỏ trailing slash + fragment)
            href_full = urljoin(self.base, a['href']).split('#')[0].rstrip('/')
            if href_full == target_url:
                parent_li = a.find_parent('li')
                if parent_li:
                    found_main_li = parent_li
                    root_cat = a.get_text(strip=True)
                    log.info(f"✓ Menu gốc tìm thấy: «{root_cat}»")
                    break

        if not found_main_li:
            log.debug(f"Không tìm được <li> cho URL: {target_url}")

        # ── Bước 3: Đệ quy từ <li> gốc ──────────────────────────────────────
        if found_main_li:
            self._collect_links_from_li(found_main_li, collected,
                                        depth=0, parent_cat=root_cat)
            collected.pop(target_url, None)   # bỏ link mẹ khỏi kết quả
            if not collected:
                # Không có link con → URL lá, hoàn toàn bình thường
                log.debug("Trang leaf (không có link con trong menu) — xử lý 1 URL")

        # ── Bước 4: Fallback prefix scan ─────────────────────────────────────
        # Chỉ chạy khi Bước 2 thất bại (menu item không tồn tại trong zone)
        if not collected and not found_main_li:
            for a in search_zone.find_all('a', href=True):
                if not self._is_valid_nav_href(a['href']):
                    continue
                full = urljoin(self.base, a['href']).rstrip('/')
                if full.startswith(target_url + '/') and full != target_url:
                    cat = a.get_text(strip=True) or self.cat_id
                    collected.setdefault(full, cat)
                    log.debug(f"  [prefix-scan] «{cat}» → {full}")

            if collected:
                log.info(f"Sub-menu qua prefix scan: {len(collected)} link")
            else:
                # URL đơn lẻ hoặc nav render bằng JS — xử lý 1 trang, không cảnh báo
                log.debug("Không tìm được sub-link — xử lý 1 URL đơn (có thể nav JS-render)")

        log.success(f"✓ Tìm được {len(collected)} sub-link (mọi cấp)")
        for url, cat in sorted(collected.items()):
            log.info(f"  [{cat}] → {url}")
        log.end_section()
        return collected


def expand_excel_with_submenus(
    excel_path: str,
    sheet_name: str = None,
    start_row: int = 2,
):
    """
    BƯỚC TIỀN XỬ LÝ — chạy TRƯỚC run_from_excel().

    Đọc từng dòng Excel có source URL, tải trang, tìm toàn bộ link con
    trong menu (đệ quy mọi cấp), rồi ghi bổ sung xuống Excel.

    Mỗi link con được ghi thành 1 dòng mới với:
      - idurl  : tự tăng (max idurl hiện có + 1)
      - source : url con
      - target : kế thừa từ dòng mẹ
      - cat_id : text của <a> trong menu (đúng tên mục)
      - done   : để trống (chưa scrape)

    Sau khi hàm này chạy xong, mở Excel kiểm tra rồi mới gọi run_from_excel().
    """
    log.section("MỞ RỘNG EXCEL VỚI SUB-MENU LINKS", Color.BRIGHT_MAGENTA)
    log.file_operation("Reading", excel_path, "processing")

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # ── Đọc header ────────────────────────────────────────────────────────────
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            headers[str(v).strip().lower()] = col

    required = ["idurl", "source", "target", "cat_id", "done"]
    # post_type là cột tùy chọn — không bắt buộc (mặc định = "post")
    for k in required:
        if k not in headers:
            raise ValueError(f"Thiếu cột '{k}' trong Excel")

    col_idurl  = headers["idurl"]
    col_source = headers["source"]
    col_target = headers["target"]
    col_cat    = headers["cat_id"]
    col_done   = headers["done"]

    # ── Thu thập idurl lớn nhất hiện có ──────────────────────────────────────
    max_idurl = 0
    existing_sources: set[str] = set()
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(r, col_idurl).value
        try:
            max_idurl = max(max_idurl, int(v or 0))
        except (ValueError, TypeError):
            pass
        src = ws.cell(r, col_source).value
        if src:
            existing_sources.add(str(src).strip().rstrip('/'))

    log.info(f"max idurl hiện tại: {max_idurl} | {len(existing_sources)} source đã có")

    # ── Khởi tạo Chrome (dùng chung 1 driver) ────────────────────────────────
    service        = Service(CHROMEDRIVER_PATH)
    chrome_options = Options()
    chrome_options.binary_location = CHROME_BINARY
    driver = webdriver.Chrome(service=service, options=chrome_options)

    new_rows: list[dict] = []   # tích lũy rows mới, tránh ghi giữa chừng

    try:
        for r in range(start_row, ws.max_row + 1):
            source = ws.cell(r, col_source).value
            target = ws.cell(r, col_target).value
            if not source:
                continue

            source = str(source).strip()
            target = str(target).strip() if target else ""
            base   = get_base(source)

            log.info(f"\n  Dòng {r}: {source}")

            # Tạo bot tạm để gọi discover_sub_menus
            # (dùng lại driver, không tạo driver mới)
            class _TmpBot:
                pass
            tmp            = _TmpBot()
            tmp.base       = base
            tmp.url        = source
            tmp.cat_id     = str(ws.cell(r, col_cat).value or "")
            tmp.driver     = driver
            tmp._is_valid_nav_href  = VspProducts._is_valid_nav_href.__get__(tmp, type(tmp))
            tmp._collect_links_from_li = VspProducts._collect_links_from_li.__get__(tmp, type(tmp))
            tmp.discover_sub_menus  = VspProducts.discover_sub_menus.__get__(tmp, type(tmp))

            try:
                driver.get(source)
                time.sleep(2)
                soup      = BeautifulSoup(driver.page_source, 'html.parser')
                sub_menus = tmp.discover_sub_menus(soup)   # dict url→cat_id
            except Exception as e:
                log.error(f"  Lỗi tải trang {source}: {e}")
                continue

            added = 0
            for url, cat in sub_menus.items():
                url_clean = url.rstrip('/')
                if url_clean in existing_sources:
                    log.debug(f"  Bỏ qua (đã có): {url_clean}")
                    continue
                max_idurl += 1
                new_rows.append({
                    "idurl":  max_idurl,
                    "source": url_clean,
                    "target": target,
                    "cat_id": cat,
                    "done":   "",
                })
                existing_sources.add(url_clean)
                added += 1

            log.success(f"  ✓ Thêm {added} link con mới từ dòng {r}")

    finally:
        driver.quit()

    # ── Ghi các dòng mới xuống cuối sheet ────────────────────────────────────
    if new_rows:
        next_row = ws.max_row + 1
        for row_data in new_rows:
            ws.cell(next_row, col_idurl).value  = row_data["idurl"]
            ws.cell(next_row, col_source).value = row_data["source"]
            ws.cell(next_row, col_target).value = row_data["target"]
            ws.cell(next_row, col_cat).value    = row_data["cat_id"]
            ws.cell(next_row, col_done).value   = row_data["done"]
            next_row += 1

        wb.save(excel_path)
        log.success(f"✓ Đã ghi {len(new_rows)} dòng mới vào Excel: {excel_path}")
        log.highlight("→ Mở Excel kiểm tra cat_id, xoá/sửa nếu cần, rồi chạy run_from_excel()")
    else:
        log.info("Không có link con mới nào cần thêm.")

    log.end_section()


# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════

def run_from_excel(
    excel_path: str,
    sheet_name: str = None,
    start_row: int = 2,
    sleep_between: float = 0.5,
    save_every: int = 1,
):
    """Đọc Excel và xử lý từng dòng."""
    stats.reset()

    log.section("BẮT ĐẦU XỬ LÝ EXCEL", Color.BRIGHT_MAGENTA)
    log.file_operation("Reading", excel_path, "processing")

    # ── Ghi header phiên vào error log ───────────────────────────────────────
    if ERROR_LOG_FILE:
        try:
            import os as _os
            _os.makedirs(_os.path.dirname(ERROR_LOG_FILE), exist_ok=True)
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with open(ERROR_LOG_FILE, "a", encoding="utf-8") as f:
                f.write(f"\n{'='*80}\n[{ts}]  PHIÊN MỚI  —  {excel_path}\n{'='*80}\n")
        except Exception:
            pass

    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    log.success(f"✓ Đã mở: {excel_path}  (sheet: {ws.title})")

    # ── Header ───────────────────────────────────────────────────────────────
    log.subsection("Đọc header row")
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            headers[str(v).strip().lower()] = col
            log.debug(f"  Col {col}: {v}")

    required = ["idurl", "source", "target", "cat_id", "done"]
    log.subsection("Kiểm tra cột bắt buộc")
    for k in required:
        log.parse_info(k, k in headers, f"Col {headers[k]}" if k in headers else "")
        if k not in headers:
            raise ValueError(f"Thiếu cột '{k}' trong Excel")
    log.success("✓ Tất cả cột bắt buộc đã có")

    done_col             = headers["done"]
    post_type_col        = headers.get("post_type", None)   # tùy chọn
    bot                  = None
    processed_since_save = 0

    # ── Tính số dòng thực (tránh bug max_row của openpyxl) ───────────────────
    #    openpyxl.max_row đếm cả dòng có format/style nhưng không có dữ liệu.
    #    → Duyệt ngược từ cuối, tìm dòng cuối cùng thực sự có giá trị.
    real_last_row = start_row - 1
    for r in range(ws.max_row, start_row - 1, -1):
        if any(ws.cell(r, headers[k]).value not in (None, "") for k in required):
            real_last_row = r
            break

    if real_last_row < start_row:
        log.warning("Không có dòng dữ liệu nào trong sheet — kết thúc sớm.")
        return

    total_rows = real_last_row - start_row + 1
    log.info(f"max_row (openpyxl) = {ws.max_row}  →  dòng dữ liệu thực: {start_row}–{real_last_row} ({total_rows} dòng)")

    log.separator()
    log.highlight(f"Bắt đầu xử lý {total_rows} dòng...")
    log.separator()

    EMPTY_ROW_STOP = 5   # dừng hẳn nếu gặp liên tiếp N dòng trống
    consecutive_empty = 0

    _row_t_start = datetime.now()
    _prev_base   = None

    try:
        for r in range(start_row, real_last_row + 1):
            idurl     = ws.cell(r, headers["idurl"]).value
            source    = ws.cell(r, headers["source"]).value
            target    = ws.cell(r, headers["target"]).value
            cat_id    = ws.cell(r, headers["cat_id"]).value
            done_v    = ws.cell(r, done_col).value
            post_type = (str(ws.cell(r, post_type_col).value or "").strip()
                         if post_type_col else "")
            post_type = post_type or "posts"   # mặc định

            cur_base = get_base(str(source)) if source else ""

            # Tổng kết trường trước khi chuyển sang trường mới
            if _prev_base and cur_base != _prev_base and _prev_base in stats.cat_saved:
                log.site_summary(_prev_base, dict(stats.cat_saved[_prev_base]),
                                 stats.failed_articles)

            log.row_start(r - start_row + 1, total_rows,
                          source=str(source)[:70] if source else "",
                          cat=str(cat_id) if cat_id else "",
                          target=str(target)[:70] if target else "")
            _row_t_start = datetime.now()

            if not source:
                log.warning(f"Không có source URL — bỏ qua")
                stats.record_skip()
                consecutive_empty += 1
                if consecutive_empty >= EMPTY_ROW_STOP:
                    log.warning(f"⛔ {consecutive_empty} dòng trống liên tiếp — dừng sớm")
                    break
                continue

            consecutive_empty = 0

            if str(done_v).strip() == "1":
                _prev_base = cur_base   # BUG FIX: cập nhật để tránh site_summary sai
                continue

            idurl  = str(idurl).strip()  if idurl  else ""
            source = str(source).strip()
            target = str(target).strip() if target else ""
            cat_id = str(cat_id).strip() if cat_id else ""
            base   = get_base(source)

            if bot is None:
                bot = VspProducts(base=base, url=source, cat=cat_id, target=target)

            bot.reset_for_row()
            bot.url        = source
            bot.url_id     = idurl
            bot.cat_id     = cat_id
            bot.base       = base
            bot.target     = target
            bot.post_type  = post_type   # "posts", "pages", "van-ban", v.v.
            bot._excel_row = r - start_row + 1

            _saved_before = 0
            _saved_after  = 0
            _in_db_before = 0
            _fail_snapshot = len(stats.failed_articles)   # snapshot lỗi trước khi chạy dòng này
            try:
                # Snapshot số bài đã lưu / đã có DB trước khi chạy dòng này
                _saved_before = sum(
                    v for cat_d in stats.cat_saved.values()
                    for v in cat_d.values()
                )
                _in_db_before = sum(stats.in_db_count.values())

                bot.get_data()

                _saved_after = sum(
                    v for cat_d in stats.cat_saved.values()
                    for v in cat_d.values()
                )
                _in_db_after = sum(stats.in_db_count.values())

                newly_saved = _saved_after - _saved_before
                newly_in_db = _in_db_after - _in_db_before

                if newly_saved > 0:
                    # Có bài mới thực sự được lưu → done=1
                    ws.cell(r, done_col).value = 1
                elif newly_in_db > 0:
                    # Tất cả bài đều đã có trong DB → cũng done=1 (không cần lấy lại)
                    ws.cell(r, done_col).value = 1
                    log.info(f"  Tất cả {newly_in_db} bài đã có trong DB — đánh dấu done=1")
                else:
                    # Không lấy được bài nào, không có bài nào trong DB
                    # → Giữ done=0 để chạy lại (có thể là lỗi mạng/structure)
                    log.warning("Không lưu được bài nào và không có in_db — giữ done=0 để chạy lại")

            except Exception as e:
                err_msg   = f"{type(e).__name__}: {e}"
                tb_msg    = traceback.format_exc()
                log.critical(f"LỖI dòng {r}: {err_msg}")
                log.error(tb_msg)
                _write_error_log(row=r, source=source, idurl=idurl,
                                 error=f"{err_msg}\n{tb_msg}")
                # Nếu driver crash → restart để các dòng tiếp theo không die theo
                _is_driver_crash = any(kw in err_msg for kw in (
                    'WebDriverException', 'MaxRetryError', 'NewConnectionError',
                    'WinError 10061', 'HTTPConnectionPool', 'BROWSER_DIED',
                    'disconnected', 'chrome not reachable',
                ))
                if bot and _is_driver_crash:
                    try:
                        bot.restart_driver()
                    except Exception as _restart_err:
                        log.critical(f"Restart driver thất bại: {_restart_err} — dừng phiên")
                        break

            processed_since_save += 1
            if processed_since_save >= save_every:
                wb.save(excel_path)
                processed_since_save = 0

            _prev_base   = cur_base
            _row_elapsed = (datetime.now() - _row_t_start).total_seconds()
            _row_saved   = _saved_after - _saved_before   # số bài lưu được trong dòng này
            # Đếm lỗi phát sinh trong dòng này (dựa vào snapshot trước/sau)
            _row_err     = len(stats.failed_articles) - _fail_snapshot
            log.row_done(r - start_row + 1, _row_saved, _row_err, _row_elapsed)
            time.sleep(sleep_between)

    finally:
        try:
            wb.save(excel_path)
            log.file_operation("Final save", excel_path, "success")
        except Exception as e:
            log.error(f"Lỗi lưu file cuối: {e}")

        if bot:
            bot.close()

        # ── Tổng kết trường cuối + toàn phiên ───────────────────────────────
        if _prev_base and _prev_base in stats.cat_saved:
            log.site_summary(_prev_base, dict(stats.cat_saved[_prev_base]),
                             stats.failed_articles)
        log.print_session_summary(stats)


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # ── BƯỚC 1: Tìm & ghi link con xuống Excel (chạy 1 lần, kiểm tra rồi comment lại)
    # expand_excel_with_submenus(EXCEL_PATH)

    # ── BƯỚC 2: Scrape toàn bộ sau khi đã kiểm tra Excel
    try:
        run_from_excel(EXCEL_PATH, sheet_name=None)
    except KeyboardInterrupt:
        log.warning(f"\n{Icon.WARNING} Người dùng dừng chương trình")
    except Exception as e:
        log.critical(f"\n{Icon.FIRE} Lỗi nghiêm trọng: {e}")
        log.error(traceback.format_exc())