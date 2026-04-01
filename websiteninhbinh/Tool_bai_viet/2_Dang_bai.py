# -*- coding: utf-8 -*-
"""
TOOL ĐĂNG BÀI WORDPRESS ĐA LUỒNG - PHIÊN BẢN HỢP NHẤT TOÀN DIỆN (FINAL INTEGRITY)
-----------------------------------------------------------------------------
Phiên bản này KHÔNG CẮT BỚT LOGIC. Nó hợp nhất:
1. Logic thao tác UI bền bỉ từ 'dangbai_a.py' (Tìm nút theo text, cuộn chuột, JS ép chọn).
2. Hệ thống đa luồng, chống crash, chống bot từ phiên bản mới.
3. Các bản vá lỗi: 403, 400, WinError 10061, High-Res Image.

CẤU TRÚC:
1. Config & Import
2. Logging
3. Smart Wait & Browser Utils (Định nghĩa trước để tránh lỗi)
4. File & Network Utils
5. Editor & Media Core (Logic dangbai_a.py tích hợp)
6. Post Logic (Chuyên mục, Featured Img)
7. Content Processing
8. Bot Controller
9. Main Execution
"""

import os
import re
import sys
import unicodedata
import time
import random
import shutil
import requests
import urllib3
import hashlib
import mimetypes
import traceback
import threading
import io as _io
from contextlib import contextmanager
from concurrent.futures import ThreadPoolExecutor

# ── Suppress stdout từ helpers (hp) — tránh in "save_data", "luu data"... ──
@contextmanager
def _suppress_stdout():
    """Suppress bất kỳ print nào phát sinh trong block (dùng cho lời gọi hp.*)."""
    _old = sys.stdout
    sys.stdout = _io.StringIO()
    try:
        yield
    finally:
        sys.stdout = _old
from datetime import datetime, date
from urllib.parse import urljoin, urlparse, unquote, quote_plus
from bs4 import BeautifulSoup, Tag
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException, 
    StaleElementReferenceException, 
    WebDriverException, 
    ElementClickInterceptedException,
    NoSuchElementException
)
# Import lỗi mạng để xử lý crash
from urllib3.exceptions import MaxRetryError, NewConnectionError

# Giả định file helpers.py nằm cùng thư mục
import helpers as hp
from openpyxl import Workbook, load_workbook

from user_agents import USER_AGENTS

# Tắt cảnh báo SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ==============================================================================
# 1. CẤU HÌNH  ← Xem và chỉnh sửa tại config.py
# ==============================================================================
import queue as _queue
from config import (
    CHROMEDRIVER_PATH, CHROME_BINARY, USE_PROFILE, PROFILE_DIR,
    TMP_DIR, LOG_XLSX,
    WP_EMAIL as EMAIL, WP_PASSWORD as PASSWORD,
    DEFAULT_PUBLISH_HOUR, DEFAULT_PUBLISH_MINUTE,
    PAUSE_TIME, STABILITY_PAUSE, UPLOAD_TIMEOUT,
    DUPLICATE_MODE,
    USE_REST_API, REST_DOMAIN_CONCURRENCY,
    REST_DELAY_MIN, REST_DELAY_MAX, REST_UPLOAD_WORKERS,
    MAX_RETRIES_PER_POST, MAX_THREADS_PER_SITE,
    MAX_CONCURRENT_WORKERS,
    XOA_FILE_SAU_KHI_DANG as Xoa_file_sau_khi_dang,
    SHOW_CHROME_WINDOW, DASHBOARD_REFRESH_MS, TITLE_MAX_LEN,
    TITLE_HEAD_LEN, TITLE_TAIL_LEN,
    LOG_TXT,
)
# Dynamic refill — import mềm để không crash nếu chưa thêm vào config.py
try:
    from config import DYNAMIC_REFILL
except ImportError:
    DYNAMIC_REFILL = False
try:
    from config import REFILL_INTERVAL
except ImportError:
    REFILL_INTERVAL = 60
try:
    from config import REFILL_EMPTY_STOP
except ImportError:
    REFILL_EMPTY_STOP = 3

# ==============================================================================
# 2. LOGGING & STYLE
# ==============================================================================

class Style:
    RESET   = '\033[0m'
    BOLD    = '\033[1m'
    DIM     = '\033[2m'
    RED     = '\033[91m'
    GREEN   = '\033[92m'
    YELLOW  = '\033[93m'
    CYAN    = '\033[96m'
    WHITE   = '\033[97m'
    BBLACK   = '\033[90m'    # bright black / dark gray
    B_RED    = '\033[1;91m'
    B_GREEN  = '\033[1;92m'
    B_YELLOW = '\033[1;93m'
    B_CYAN   = '\033[1;96m'
    B_WHITE  = '\033[1;97m'
    MAGENTA  = '\033[95m'

# ── Bật ANSI color trên Windows (PowerShell / cmd) ──────────────────────────
def _enable_ansi_windows():
    """
    Windows không bật VT100 ANSI codes mặc định.
    Gọi SetConsoleMode để enable — không làm gì nếu không phải Windows.
    """
    import sys
    if sys.platform != 'win32':
        return
    try:
        import ctypes, ctypes.wintypes
        kernel32 = ctypes.windll.kernel32
        # stdout handle
        handle = kernel32.GetStdHandle(-11)   # STD_OUTPUT_HANDLE
        old_mode = ctypes.wintypes.DWORD(0)
        kernel32.GetConsoleMode(handle, ctypes.byref(old_mode))
        ENABLE_VT = 0x0004
        kernel32.SetConsoleMode(handle, old_mode.value | ENABLE_VT)
    except Exception:
        pass

_enable_ansi_windows()

log_lock       = threading.Lock()   # giữ lại cho append_log_row (Excel)
db_lock        = threading.Lock()
_pending_titles: dict = {}          # base_url → set[norm_title] đang được xử lý
_pending_lock   = threading.Lock()      # bảo vệ _pending_titles
_dl_cache      = {}
_dl_cache_lock = threading.Lock()

# ── Dynamic refill — shared state giữa MainController, workers và RefillThread ──
_site_queues  : dict            = {}               # base_url → queue.Queue (1 queue/site)
_assigned_ids : set             = set()            # news_id đã được nhận vào queue/đang xử lý
_assigned_lock: threading.Lock  = threading.Lock()
_refill_stop  : threading.Event = threading.Event()# set khi RefillThread kết thúc


def _release_assigned(news_id) -> None:
    """
    Bỏ news_id khỏi tập đã chiếm sau khi worker xử lý xong (bất kể kết quả).
    Chỉ hoạt động khi DYNAMIC_REFILL = True.
    Nếu bài bị lỗi và chưa được update_upload_new → lần quét tiếp theo sẽ re-add.
    """
    if DYNAMIC_REFILL:
        with _assigned_lock:
            _assigned_ids.discard(str(news_id))

def init_log_workbook(path: str, sheet_name: str = "log"):
    import zipfile
    if os.path.isfile(path):
        try:
            wb = load_workbook(path)
            ws = (wb[sheet_name] if sheet_name in wb.sheetnames
                  else wb.create_sheet(sheet_name))
            return wb, ws
        except (zipfile.BadZipFile, EOFError, Exception):
            try: os.remove(path)
            except: pass
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = ["ts", "worker_name", "base_target", "login_url",
               "news_id", "title", "source_url", "post_date",
               "category", "status", "step", "error",
               "uploaded_count", "uploaded_ids", "featured_ok",
               "public_url", "duration_sec"]
    ws.append(headers)
    wb.save(path)
    return wb, ws

def append_log_row(path: str, row: dict, sheet_name: str = "log"):
    with log_lock:
        try:
            wb, ws = init_log_workbook(path, sheet_name=sheet_name)
            headers = [cell.value for cell in ws[1]]
            if "worker_name" not in row:
                row["worker_name"] = threading.current_thread().name
            line = [row.get(h, "") for h in headers]
            line = [",".join(map(str, v)) if isinstance(v, (list, tuple, set))
                    else v for v in line]
            ws.append(line)
            wb.save(path)
        except Exception as e:
            # debug_print định nghĩa muộn hơn — dùng print để tránh forward reference
            print(f"❌ Lỗi ghi log Excel: {e}")


def get_random_ua():
    return random.choice(USER_AGENTS)

_ANSI_RE = re.compile(r'\033\[[0-9;]*[mABCDEFGHJKSTsuhl]')

def _sa(s) -> str:
    """Strip ANSI codes — dùng để tính độ rộng visible."""
    return _ANSI_RE.sub('', str(s or ''))

def _tr(s, n: int) -> str:
    """Truncate text đến n ký tự visible (không tính ANSI). Fallback chung."""
    s = str(s or '').strip()
    plain = _sa(s)
    return (plain[:n - 1] + '\u2026') if len(plain) > n else plain

def _tr_title(s: str) -> str:
    """
    Format tiêu đề bài viết theo cấu trúc HEAD...TAIL:
    - Giữ TITLE_HEAD_LEN ký tự đầu + '...' + TITLE_TAIL_LEN ký tự cuối
    - Ví dụ: 'MỘT SỐ HÌNH ẢNH VỀ LỄ KỶ ... 5 - 2026'
    - Nếu chuỗi đủ ngắn thì giữ nguyên.
    """
    s = _sa(str(s or '')).strip()
    limit = TITLE_HEAD_LEN + TITLE_TAIL_LEN + 3   # 3 = len('...')
    if len(s) <= limit:
        return s
    head = s[:TITLE_HEAD_LEN].rstrip()
    tail = s[-TITLE_TAIL_LEN:].lstrip()
    return f'{head}... {tail}'

def _fmt_t(secs: float) -> str:
    """Định dạng thời gian: 'MMm:SSs' nếu >= 60s, ngược lại 'ss.Xs'."""
    if secs >= 60:
        m, s = divmod(int(secs), 60)
        return f'{m:02d}m:{s:02d}s'
    return f'{secs:.1f}s'


# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# LIVE DASHBOARD — shared-state, dedicated render thread
# ══════════════════════════════════════════════════════════════════════════════
#
# Kiến trúc:
#   worker threads  ─write─►  _slot_data[i]  ◄─read─  _RenderThread ─► stdout
#   (dùng _slot_lock để tránh race condition)
#
# Quy tắc:
#   • _RenderThread là thread DUY NHẤT ghi stdout.
#   • Worker ghi vào _slot_data, KHÔNG ghi stdout trực tiếp.
#   • _RenderThread đọc theo chu kỳ DASHBOARD_REFRESH_MS ms rồi redraw.
#
# Format mỗi slot (2 dòng):
#   [Luồng N] [MM:SS]  domain  [Chuyên mục]  Tiêu đề
#     Quét trùng: X  📝 soạn mới  🖼 media(3)  ⧖ step hiện tại
# Khi xong:
#   [Luồng N] [MM:SS]  domain  [Chuyên mục]  Tiêu đề
#     Quét trùng: X  📝 soạn mới  ✅ Xong  3.2s   📅 2025-11-21
# ══════════════════════════════════════════════════════════════════════════════

# ── Shared state (viết bởi workers, đọc bởi render thread) ───────────────────
_slot_lock : threading.Lock    = threading.Lock()
_slot_data : dict              = {}   # slot_idx → dict trạng thái bài
_slot_pool : list              = []   # danh sách slot index trống
_slot_pool_lock: threading.Lock = threading.Lock()
_n_slots   : int               = 0    # tổng số slot (= min(n_workers, fit_term))
_n_workers_total: int          = 0    # n_workers thực tế (hiển thị "+N chờ")


def _slot_new() -> dict:
    """Tạo dict trạng thái mặc định cho 1 slot."""
    return {
        'label':    '',
        'base':     '',
        'title':    '',
        'category': '',
        'dup_s':    '…',
        'steps':    [],
        'cur':      '',
        'date_s':   '',
        'final_s':  '',
        't0':       0.0,
        'duration': 0.0,
        '_done_at': 0.0,   # thời điểm set final_s (để clear sau N giây)
        '_partial': '',    # buffer cho debug_print end=''
    }


def _claim_slot() -> int:
    """Lấy 1 slot trống. Trả về -1 nếu hết slot (worker tự poll lại)."""
    with _slot_pool_lock:
        if _slot_pool:
            idx = _slot_pool.pop(0)
            with _slot_lock:
                _slot_data[idx] = _slot_new()
            return idx
        return -1


def _free_slot(idx: int):
    """Trả slot về pool. Không xóa _slot_data ngay — render thread dọn sau."""
    if idx >= 0:
        with _slot_pool_lock:
            _slot_pool.append(idx)


# ── Thread-local buffer (mỗi worker có 1 bản riêng) ─────────────────────────
_tl = threading.local()


def _abuf() -> dict:
    """Trả về dict thread-local, tạo mới nếu chưa có."""
    if not hasattr(_tl, 'slot_idx'):
        _tl.slot_idx = -1
    if not hasattr(_tl, 'buf'):
        _tl.buf = {}
    return _tl.buf


def _make_label() -> str:
    name = threading.current_thread().name
    m = re.search(r'_(\d+)$', name)
    return f'Luồng {int(m.group(1)) + 1}' if m else name


# ── Render helpers ────────────────────────────────────────────────────────────

def _rl1(s: dict, now: float) -> str:
    """Dòng 1: [Luồng N] [MM:SS]  domain  [cat]  title"""
    elapsed = now - s['t0'] if s['t0'] else 0.0
    t_str   = _fmt_t(elapsed)
    cat_s   = (f'  {Style.YELLOW}[{s["category"]}]{Style.RESET}'
               if s['category'] else '')
    title_s = f'  {s["title"]}' if s['title'] else ''
    return (
        f'{Style.B_CYAN}[{s["label"]}]{Style.RESET}'
        f' {Style.DIM}[{t_str}]{Style.RESET}'
        f'  {Style.B_WHITE}{s["base"][:42]}{Style.RESET}'
        f'{cat_s}{title_s}'
    )


def _rl2(s: dict, now: float) -> str:
    """Dòng 2: Quét trùng  steps  ⧖cur | ✅final  elapsed  date"""
    steps_s = ('  '.join(s['steps']) + '  ') if s['steps'] else ''
    fs = s['final_s']
    if fs:
        dur    = s['duration'] or (now - s['t0'] if s['t0'] else 0.0)
        dur_s  = f'  {Style.DIM}{_fmt_t(dur)}{Style.RESET}'
        date_s = (f'  📅 {Style.DIM}{s["date_s"]}{Style.RESET}'
                  if s['date_s'] else '')
        if fs.startswith('✅'):
            fs_s = f'{Style.B_GREEN}{fs}{Style.RESET}'
        elif fs.startswith('❌'):
            fs_s = f'{Style.B_RED}{fs}{Style.RESET}'
        elif fs.startswith('⏭'):
            fs_s = f'{Style.YELLOW}{fs}{Style.RESET}'
        elif fs.startswith('⚠'):
            fs_s = f'{Style.B_YELLOW}{fs}{Style.RESET}'
        else:
            fs_s = f'{Style.DIM}{fs}{Style.RESET}'
        return (f'  {Style.BBLACK}Quét trùng:{Style.RESET} {_dup_color(s["dup_s"])}'
                f'  {steps_s}{fs_s}{dur_s}{date_s}')
    # Đang chạy
    cur_s = (f'{Style.DIM}⧖ {Style.RESET}{s["cur"]}'
             if s['cur'] else '')
    return (f'  {Style.BBLACK}Quét trùng:{Style.RESET} {_dup_color(s["dup_s"])}'
            f'  {steps_s}{cur_s}')


def _dup_color(s: str) -> str:
    """Tô màu chuỗi trạng thái quét trùng."""
    if not s:
        return s
    sl = s.lower()
    if 'lỗi'           in sl: return f'{Style.RED}{s}{Style.RESET}'
    if 'bỏ qua'        in sl: return f'{Style.B_RED}{s}{Style.RESET}'
    if 'đã xóa'        in sl: return f'{Style.B_GREEN}{s}{Style.RESET}'
    if 'nháp'          in sl: return f'{Style.B_YELLOW}{s}{Style.RESET}'
    if 'đăng'          in sl and '0 bài trùng' not in sl:
                               return f'{Style.RED}{s}{Style.RESET}'
    if 'không kiểm tra' in sl: return f'{Style.DIM}{s}{Style.RESET}'
    if '0 bài'         in sl: return f'{Style.BBLACK}{s}{Style.RESET}'
    return f'{Style.DIM}{s}{Style.RESET}'


# ── Log file ──────────────────────────────────────────────────────────────────
_log_file_lock = threading.Lock()


def _write_log_file(lines: list):
    try:
        text = '\n'.join(_sa(l).rstrip() for l in lines if _sa(l).strip())
        if not text:
            return
        with _log_file_lock:
            with open(LOG_TXT, 'a', encoding='utf-8') as f:
                ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                f.write(f'[{ts}]\n{text}\n\n')
    except Exception:
        pass


# ── vis helpers ───────────────────────────────────────────────────────────────
def _vis_len(s: str) -> int:
    return len(_ANSI_RE.sub('', s))


def _vis_trunc(s: str, max_w: int) -> str:
    if _vis_len(s) <= max_w:
        return s
    visible, result, i = 0, [], 0
    while i < len(s) and visible < max_w:
        m = _ANSI_RE.match(s, i)
        if m:
            result.append(m.group()); i = m.end()
        else:
            result.append(s[i]); visible += 1; i += 1
    result.append(Style.RESET)
    return ''.join(result)


# ── Render thread ─────────────────────────────────────────────────────────────
_DONE_LINGER = 2.5   # giây giữ kết quả cuối trên slot trước khi clear

class _RenderThread(threading.Thread):
    """
    Thread duy nhất ghi stdout.
    Mỗi DASHBOARD_REFRESH_MS ms: đọc _slot_data → vẽ lại block dashboard.

    Cursor management:
      Windows → SetConsoleCursorPosition (ctypes, tuyệt đối, luôn đúng)
      Other   → \033[{H}A (ANSI cursor-up)
    """
    SEP = '─'

    def __init__(self, n_slots: int, n_total: int):
        super().__init__(name='RenderThread', daemon=True)
        self._n        = n_slots
        self._n_total  = n_total
        self._H        = n_slots * 2 + 1   # 1 separator + 2 dòng × n_slots
        self._ready    = False
        self._interval = max(0.1, DASHBOARD_REFRESH_MS / 1000.0)
        self._top_row  = -1   # dòng đầu block (Windows: tọa độ tuyệt đối)
        # Windows Console API
        self._con_ok   = False
        self._h_stdout = None
        if sys.platform == 'win32':
            try:
                import ctypes, ctypes.wintypes as _wt
                self._ctypes = ctypes
                self._wt     = _wt
                k32 = ctypes.windll.kernel32
                h   = k32.GetStdHandle(-11)   # STD_OUTPUT_HANDLE
                if h and h != -1:
                    self._h_stdout = h
                    self._k32      = k32
                    self._con_ok   = True
            except Exception:
                pass

    def _get_cursor_row(self) -> int:
        """Lấy row hiện tại của cursor (Windows Console API)."""
        if not self._con_ok:
            return -1
        try:
            import ctypes
            class _COORD(ctypes.Structure):
                _fields_ = [('X', self._wt.SHORT), ('Y', self._wt.SHORT)]
            class _SMALL_RECT(ctypes.Structure):
                _fields_ = [('Left',self._wt.SHORT),('Top',self._wt.SHORT),
                             ('Right',self._wt.SHORT),('Bottom',self._wt.SHORT)]
            class _CSBI(ctypes.Structure):
                _fields_ = [('dwSize',_COORD),('dwCursorPosition',_COORD),
                             ('wAttributes',self._wt.WORD),
                             ('srWindow',_SMALL_RECT),
                             ('dwMaximumWindowSize',_COORD)]
            csbi = _CSBI()
            if self._k32.GetConsoleScreenBufferInfo(self._h_stdout,
                                                     ctypes.byref(csbi)):
                return csbi.dwCursorPosition.Y
        except Exception:
            pass
        return -1

    def _set_cursor_row(self, row: int):
        """Di chuyển cursor đến đầu dòng row (Windows Console API)."""
        if not self._con_ok or row < 0:
            return False
        try:
            import ctypes
            class _COORD(ctypes.Structure):
                _fields_ = [('X', self._wt.SHORT), ('Y', self._wt.SHORT)]
            coord = _COORD(0, row)
            return bool(self._k32.SetConsoleCursorPosition(self._h_stdout,
                                                            coord))
        except Exception:
            return False

    def _reserve(self):
        """In H dòng trắng để đặt chỗ cho dashboard block."""
        if not self._ready:
            sys.stdout.flush()
            sys.stdout.write('\n' * self._H)
            sys.stdout.flush()
            self._ready = True

    def _goto_top(self):
        """Về đầu block bằng ANSI relative cursor-up — hoạt động đúng dù terminal scroll."""
        sys.stdout.write(f'\033[{self._H}F')
        sys.stdout.flush()

    def _draw(self):
        now = time.time()
        w   = shutil.get_terminal_size((120, 40)).columns

        # Lấy snapshot (không giữ lock lâu)
        with _slot_lock:
            snap = {k: dict(v) if isinstance(v, dict) else v
                    for k, v in _slot_data.items()}

        waiting = max(0, self._n_total - self._n)
        wait_s  = (f' {Style.DIM}(+{waiting} chờ){Style.RESET}'
                   if waiting > 0 else '')
        sep_len = max(10, min(78, w) - _vis_len(wait_s))

        lines_out = []
        lines_out.append(
            f'\r\033[2K{Style.BBLACK}{self.SEP * sep_len}{Style.RESET}{wait_s}'
        )

        for idx in range(self._n):
            s = snap.get(idx)
            if s and s.get('label'):
                l1 = _vis_trunc(_rl1(s, now), w)
                l2 = _vis_trunc(_rl2(s, now), w)
            else:
                l1 = Style.DIM + '  [chờ]' + Style.RESET
                l2 = ''
            lines_out.append(f'\r\033[2K{l1}')
            lines_out.append(f'\r\033[2K  {l2}' if l2 else '\r\033[2K')

        # Gộp cursor-up + erase + draw thành 1 write duy nhất (atomic)
        # \033[?25l / \033[?25h = ẩn/hiện cursor khi đang vẽ
        # \033[nF   = lên n dòng + về cột 0 (relative → không bị stale khi scroll)
        payload = (f'\033[?25l\033[{self._H}F'
                   + '\n'.join(lines_out) + '\n\033[?25h')
        sys.stdout.write(payload)
        sys.stdout.flush()

    def _cleanup_done_slots(self):
        """Dọn slot đã xong quá _DONE_LINGER giây → xóa data."""
        now = time.time()
        with _slot_lock:
            for idx, s in list(_slot_data.items()):
                if (s.get('final_s') and s.get('_done_at')
                        and now - s['_done_at'] >= _DONE_LINGER):
                    _slot_data[idx] = _slot_new()

    def run(self):
        self._reserve()
        while True:
            time.sleep(self._interval)
            self._cleanup_done_slots()
            self._draw()   # goto_top đã được nhúng vào trong _draw (atomic)


# ── Singleton render thread ───────────────────────────────────────────────────
_render_thread: '_RenderThread | None' = None
_render_lock   = threading.Lock()


def _ensure_render() -> '_RenderThread':
    global _render_thread
    if _render_thread is None:
        with _render_lock:
            if _render_thread is None:
                rt = _RenderThread(_n_slots or 4, _n_workers_total or 4)
                rt.start()
                _render_thread = rt
    return _render_thread


def init_dashboard(n_workers: int):
    """
    Gọi trước khi spawn workers.
    n_slots = min(n_workers, vừa terminal) để cursor-up hoạt động.
    """
    global _n_slots, _n_workers_total, _slot_pool
    term_h  = shutil.get_terminal_size((40, 40)).lines
    n_slots = max(2, min(n_workers, max(2, (term_h - 3) // 2)))
    _n_slots         = n_slots
    _n_workers_total = n_workers
    _slot_pool       = list(range(n_slots))
    # Khởi động render thread ngay để reserve màn hình
    _ensure_render()


# ── Public API (gọi từ worker threads) ───────────────────────────────────────

def post_log_start(base: str, title: str):
    """Bắt đầu 1 bài: claim slot, điền thông tin cơ bản."""
    _ensure_render()
    # Poll cho đến khi có slot trống
    while True:
        idx = _claim_slot()
        if idx >= 0:
            break
        time.sleep(0.05)
    _tl.slot_idx = idx
    with _slot_lock:
        s = _slot_data[idx]
        s['label']  = _make_label()
        s['base']   = (base.rstrip('/')
                        .replace('https://', '').replace('http://', ''))
        s['title']  = _tr_title(title)
        s['t0']     = time.time()


def post_log_set(**kw):
    """Cập nhật 1 hoặc nhiều trường của slot hiện tại."""
    idx = getattr(_tl, 'slot_idx', -1)
    if idx < 0:
        return
    with _slot_lock:
        s = _slot_data.get(idx)
        if s:
            for k, v in kw.items():
                if k in s:
                    s[k] = v


def post_log_step(step: str):
    """Thêm bước hoàn thành vào steps, xóa cur."""
    idx = getattr(_tl, 'slot_idx', -1)
    if idx < 0:
        return
    with _slot_lock:
        s = _slot_data.get(idx)
        if s:
            s['steps'].append(step)
            s['cur'] = ''


def post_log_flush():
    """Bài xong: đánh dấu _done_at, ghi log file, trả slot về pool."""
    idx = getattr(_tl, 'slot_idx', -1)
    _tl.slot_idx = -1
    if idx < 0:
        return
    with _slot_lock:
        s = _slot_data.get(idx)
        if s:
            s['_done_at'] = time.time()
            # Ghi log file
            now = time.time()
            _write_log_file([
                _sa(_rl1(s, now)),
                _sa(_rl2(s, now)),
            ])
    _free_slot(idx)


def debug_print(msg, end='\n', flush=False):
    """
    Cập nhật 'cur' của slot hiện tại.
    end=''  → gộp vào _partial buffer (chưa hiển thị)
    end='\\n' → flush partial + msg vào cur
    slot == -1 (ngoài bài) → ghi file log
    """
    idx   = getattr(_tl, 'slot_idx', -1)
    raw   = str(msg)
    plain = _sa(raw).strip()

    if idx >= 0:
        with _slot_lock:
            s = _slot_data.get(idx)
            if s is None:
                return
            if end == '':
                s['_partial'] = (s['_partial'] + raw).rstrip() if s['_partial'] else raw
            else:
                full_raw   = (s['_partial'] + raw).strip() if s['_partial'] else raw
                full_plain = _sa(full_raw).strip()
                s['_partial'] = ''
                if full_plain:
                    s['cur'] = _vis_trunc(full_raw, 68)
    else:
        if plain:
            _write_log_file([plain])


# ==============================================================================
# 3. SMART WAIT UTILS (ĐỊNH NGHĨA TRƯỚC ĐỂ TRÁNH LỖI UNDEFINED)
# ==============================================================================

def wait_for_wp_test_cookie(driver, timeout=10):
    end = time.time() + timeout
    while time.time() < end:
        for c in driver.get_cookies():
            if c.get("name") == "wordpress_test_cookie":
                return True
        time.sleep(0.2)
    return False

def assert_driver_alive(driver):
    try:
        driver.execute_script("return 1")
    except Exception:
        raise Exception("BROWSER_DIED")

def check_browser_alive(driver):
    """Kiểm tra kết nối driver, ném lỗi nếu trình duyệt chết"""
    try:
        _ = driver.current_url
    except (MaxRetryError, NewConnectionError, WebDriverException) as e:
        # Nhận diện các lỗi mất kết nối phổ biến
        if "10061" in str(e) or "refused" in str(e) or "disconnected" in str(e):
            raise Exception("BROWSER_DIED")
        raise e

def wait_for_page_load(driver, timeout=30):
    """Chờ trang tải xong hoàn toàn (document.readyState = complete)"""
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
    except: pass

def smart_click(driver, by_locator, timeout=10):
    """
    Thử click thông minh: 
    1. Chờ xuất hiện -> Click. 
    2. Nếu bị che (Intercepted) -> Click bằng JS.
    """
    try:
        element = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable(by_locator))
        element.click()
    except ElementClickInterceptedException:
        element = driver.find_element(*by_locator)
        driver.execute_script("arguments[0].click();", element)
    except TimeoutException:
        return False
    except Exception as e:
        check_browser_alive(driver)
        return False
    return True

def smart_send_keys(driver, by_locator, text, timeout=10):
    """
    Điền text thông minh: 
    1. Chờ -> Clear -> Send.
    2. Kiểm tra lại giá trị, nếu sai -> JS Injection.
    """
    try:
        element = WebDriverWait(driver, timeout).until(EC.visibility_of_element_located(by_locator))
        element.clear()
        element.send_keys(text)
        # Verify
        if element.get_attribute("value") != text:
            driver.execute_script("arguments[0].value = arguments[1];", element, text)
            driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", element)
        return True
    except Exception as e:
        check_browser_alive(driver)
        return False

def force_cleanup_modals(driver):
    """Dọn dẹp modal rác ngay lập tức (Logic từ fix trước)"""
    try:
        driver.execute_script("""
            var closers = document.querySelectorAll('.media-modal-close');
            for(var i=0; i<closers.length; i++) { closers[i].click(); }
            
            var modals = document.querySelectorAll('.media-modal, .media-modal-backdrop');
            for(var i=0; i<modals.length; i++){
                modals[i].style.display = 'none';
                modals[i].remove();
            }
            document.body.classList.remove('modal-open');
        """)
        time.sleep(0.5)
    except: pass

def safe_js_click(driver, element):
    """Click an toàn với scroll (Logic cũ)"""
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
        time.sleep(0.1)
        driver.execute_script("arguments[0].click();", element)
    except: 
        try: element.click()
        except: pass

def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)

def get_base(url: str) -> str:
    url = (url or "").strip()
    if not url: return ""
    p = urlparse(url)
    if not p.scheme or not p.netloc: p = urlparse("https://" + url)
    return f"{p.scheme}://{p.netloc}/"

def format_duration(seconds):
    if seconds < 60: return f"{seconds:.1f}s"
    m = int(seconds // 60); s = int(seconds % 60)
    return f"{m} phút {s}s"

def normalize_img_url(src: str, base_source: str) -> str:
    src = (src or "").strip()
    if not src: return ""
    if src.startswith("//"): return "https:" + src
    if src.startswith("http://") or src.startswith("https://"): return src
    return urljoin(base_source, src)

def date_to_datetime(d, hour=8, minute=0) -> datetime:
    """Chuyển date / datetime / string "dd/mm/yyyy" → datetime. Giống _rest_format_date."""
    if isinstance(d, datetime):
        return d.replace(hour=hour, minute=minute, second=0, microsecond=0)
    if isinstance(d, date):
        return datetime(d.year, d.month, d.day, hour, minute)
    # Fallback: parse string
    try:
        parsed = datetime.strptime(str(d).strip(), "%d/%m/%Y")
        return parsed.replace(hour=hour, minute=minute)
    except Exception:
        pass
    try:
        parsed = datetime.strptime(str(d).strip()[:10], "%Y-%m-%d")
        return parsed.replace(hour=hour, minute=minute)
    except Exception:
        pass
    raise ValueError(f"date_to_datetime: không parse được '{d}'")

def cleanup_temp_file(file_path):
    """Xóa file tạm và xóa entry trong _dl_cache để tránh stale reference.
    Thread-safe: dùng try/except để bẫy FileNotFoundError khi nhiều luồng
    cùng xoá một file (race condition bình thường trong multi-thread upload).
    """
    try:
        if file_path and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except FileNotFoundError:
                pass   # luồng khác đã xoá trước — bình thường
        # Xóa cache entry để tránh stale path reference
        if file_path:
            with _dl_cache_lock:
                stale = [k for k, v in _dl_cache.items() if v == file_path]
                for k in stale:
                    del _dl_cache[k]
    except Exception:
        pass

_HTTP_ERROR_MAP = {
    "403": ("403", "Forbidden"),
    "forbidden": ("403", "Forbidden"),
    "access denied": ("403", "Access Denied"),
    "404": ("404", "Not Found"),
    "not found": ("404", "Not Found"),
    "trang không tồn tại": ("404", "Không tồn tại (VN)"),
    "không tìm thấy": ("404", "Không tìm thấy (VN)"),
    "502": ("502", "Bad Gateway"),
    "bad gateway": ("502", "Bad Gateway"),
    "503": ("503", "Service Unavailable"),
    "service unavailable": ("503", "Service Unavailable"),
    "500": ("500", "Internal Server Error"),
    "internal server error": ("500", "Internal Server Error"),
}

def detect_http_error(driver) -> tuple[str, str] | None:
    """
    Kiểm tra trang hiện tại có lỗi HTTP không.
    Trả về (code, reason) nếu có lỗi, None nếu bình thường.
    Ví dụ: ("403", "Forbidden"), ("502", "Bad Gateway")
    """
    try:
        title   = (driver.title or "").lower()
        snippet = driver.page_source[:1000].lower()
        check   = title + " " + snippet
        for keyword, (code, reason) in _HTTP_ERROR_MAP.items():
            if keyword in check:
                return (code, reason)
    except Exception:
        pass
    return None

def is_403(driver):
    """Backward-compat wrapper — dùng detect_http_error() thay thế."""
    err = detect_http_error(driver)
    return err is not None and err[0] == "403"

def clean_title_for_wp(s: str) -> str:
    s = (s or "").strip()
    s = "".join(ch for ch in s if ord(ch) <= 0xFFFF)
    return re.sub(r"\s+", " ", s).strip()

# ==============================================================================
# 4. XỬ LÝ FILE (HIGH-RES & SMART NAMING)
# ==============================================================================

FILE_EXTS = (".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".zip", ".rar", ".7z")

def _try_get_high_res_url(url: str) -> str:
    """Đoán link ảnh gốc, bỏ đuôi resize"""
    high_res = re.sub(r'-\d+x\d+(\.[a-zA-Z]{3,4})$', r'\1', url)
    high_res = high_res.replace('_thumb.', '.')
    return high_res

def _safe_filename_from_url(url: str) -> str:
    p = urlparse(url)
    original_name = unquote(os.path.basename(p.path))
    stem, ext = os.path.splitext(original_name)
    if not ext or len(ext) > 5:
        url_lower = url.lower()
        if "png" in url_lower: ext = ".png"
        elif "pdf" in url_lower: ext = ".pdf"
        elif "doc" in url_lower: ext = ".doc"
        elif "docx" in url_lower: ext = ".docx"
        elif "xls" in url_lower: ext = ".xls"
        elif "xlsx" in url_lower: ext = ".xlsx"
        else: ext = ".jpg"
    ext = ext.lower()

    # Tạo hash ngắn từ URL để tránh collision khi 2 URL khác → cùng tên file
    url_hash = hashlib.md5(url.encode('utf-8')).hexdigest()[:8]

    if ext in FILE_EXTS:
        # Tài liệu: tên gốc + hash ngắn để tránh TOCTOU race giữa các luồng
        clean_stem = re.sub(r'[^a-zA-Z0-9\-_]', '-', stem).strip('-')
        if len(clean_stem) < 3:
            clean_stem = f"document-{url_hash}"
        return f"{clean_stem}-{url_hash}{ext}"
    else:
        # Ảnh: MD5 đầy đủ (đã unique theo URL)
        return f"{url_hash}{ext}"

def download_file_resource(url: str, tmp_dir: str = TMP_DIR, retries: int = 3,
                          use_cache: bool = True) -> str:
    # Cache hit — tránh download lại cùng 1 URL trong phiên
    if use_cache:
        with _dl_cache_lock:
            cached = _dl_cache.get(url)
        if cached and os.path.exists(cached) and os.path.getsize(cached) > 0:
            return cached
    ensure_dir(tmp_dir)
    headers = {
        'User-Agent': get_random_ua(),
        'Accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
        'Referer': get_base(url)
    }

    is_image   = not any(url.lower().endswith(e) for e in FILE_EXTS)
    target_url = url

    # Logic High-Res
    if is_image:
        high_res_url = _try_get_high_res_url(url)
        if high_res_url != url:
            try:
                h = requests.head(high_res_url, headers=headers, timeout=3, verify=False)
                if h.status_code == 200:
                    target_url = high_res_url
            except Exception:
                pass

    # Tính filename và path theo target_url hiện tại
    def _make_path(t_url: str) -> str:
        fname = _safe_filename_from_url(t_url)
        return os.path.join(tmp_dir, fname)

    file_path = _make_path(target_url)
    if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
        return file_path

    for attempt in range(retries):
        try:
            r = requests.get(target_url, headers=headers, timeout=(15, 30),
                             stream=True, verify=False)
            r.raise_for_status()
            with open(file_path, "wb") as f:
                for chunk in r.iter_content(chunk_size=16384):
                    if chunk:
                        f.write(chunk)
            if use_cache:
                with _dl_cache_lock:
                    _dl_cache[url] = file_path
            return file_path
        except requests.exceptions.HTTPError as e:
            if 400 <= e.response.status_code < 500:
                if target_url != url:
                    # Fallback về URL gốc — phải recalculate filename để tránh
                    # lưu vào path sai (tên high-res ≠ tên URL gốc)
                    target_url = url
                    file_path  = _make_path(target_url)
                    if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
                        return file_path
                    continue
                return None
            time.sleep(0.5)
        except Exception:
            if attempt < retries - 1:
                time.sleep(0.5)
            else:
                return None
    return None

def is_download_link(tag) -> bool:
    if tag is None or tag.name != 'a': return False
    href = tag.get("href")
    if not href or not isinstance(href, str): return False
    
    hl = href.lower()
    blacklist = [
        'facebook.com/sharer', 'facebook.com/hashtag', 'sharer.php',
        'twitter.com/intent', 'twitter.com/share', 'zalo.me', 'plus.google.com',
        'linkedin.com/share'
    ]
    if any(b in hl for b in blacklist): return False

    # Check class link-download (Logic dangbai_a.py)
    cls = " ".join(tag.get("class", [])).lower()
    if "link-download" in cls: return True

    # TYPE_INLINE_DOWNLOAD: link nằm trong container file đã biết
    parent = tag.parent
    while parent and parent.name not in ('body', '[document]', None):
        p_cls = " ".join(parent.get("class", [])).lower()
        # Thêm 'attachments-list' (1_Lay_bai tạo) + 'attachment' chung
        if any(k in p_cls for k in ('inline-download', 'listfile', 'download-type4',
                                     'list-download', 'attach-list', 'file-list',
                                     'attachments-list', 'attachments', 'file-attach',
                                     'files-attach')):
            # Trong container tải file → chấp nhận link bất kể extension
            if not any(b in hl for b in ('javascript:', 'mailto:', '#')):
                return True
        parent = parent.parent

    if hl.endswith(('.php', '.html', '.aspx', '.htm', '.jsp')): return False
    return urlparse(hl).path.lower().endswith(FILE_EXTS)

# ==============================================================================
# 5. EDITOR & MEDIA CORE (HỢP NHẤT)
# ==============================================================================

def wp_editor_focus_end(driver):
    driver.execute_script("""
    try {
      if (window.tinymce && tinymce.get('content')) {
        const ed = tinymce.get('content');
        ed.focus();
        ed.selection.select(ed.getBody(), true);
        ed.selection.collapse(false);
      }
    } catch(e) {}
    """)

def wp_editor_insert_html(driver, html: str):
    driver.execute_script("""
    const html = arguments[0];
    try {
      if (window.tinymce && tinymce.get('content')) {
        const ed = tinymce.get('content');
        ed.execCommand('mceInsertContent', false, html);
        ed.save(); return;
      }
    } catch(e) {}
    const ta = document.getElementById('content');
    if (ta) ta.value += html;
    """, html or "")


def wp_editor_insert_html_raw(driver, html: str):
    """
    Chèn HTML vào textarea trực tiếp (bypass TinyMCE) — dùng cho iframe/shortcode.
    TinyMCE Visual mode strip <iframe>, Text mode (quicktags textarea) thì không.
    Quy trình:
      1. ed.save() sync TinyMCE → textarea
      2. Append vào textarea
      3. ed.setContent() load lại vào TinyMCE nếu đang ở Visual
    """
    driver.execute_script("""
    const html = arguments[0];
    try {
      const ta = document.getElementById('content');
      if (window.tinymce && tinymce.get('content')) {
        const ed = tinymce.get('content');
        ed.save();                          // TinyMCE → textarea
        if (ta) ta.value += html;           // append vào textarea
        if (!ed.isHidden()) {
          ed.setContent(ta ? ta.value : ed.getContent() + html);
        }
      } else if (ta) {
        ta.value += html;
      }
    } catch(e) {
      try {
        const ta = document.getElementById('content');
        if (ta) ta.value += html;
      } catch(e2) {}
    }
    """, html or "")

def wp_open_add_media_modal(driver, wait: WebDriverWait):
    try:
        # Check nếu modal chưa mở
        if not driver.execute_script("return document.querySelector('.media-modal.wp-core-ui') ? document.querySelector('.media-modal.wp-core-ui').offsetParent !== null : false;"):
            smart_click(driver, (By.ID, "insert-media-button"))
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".media-modal.wp-core-ui")))
        return True
    except: return False

def safe_switch_tab(driver, tab_name_keyword):
    """Dùng JS tìm và click tab, tránh lỗi null"""
    try:
        driver.execute_script(f"""
            var tabs = document.querySelectorAll('.media-menu-item');
            for(var i=0; i<tabs.length; i++){{
                if(tabs[i].textContent.includes('{tab_name_keyword}')){{
                    tabs[i].click(); return;
                }}
            }}
            var idx = '{tab_name_keyword}' === 'Upload' ? 0 : 1;
            if(tabs[idx]) tabs[idx].click();
        """)
        time.sleep(0.2)
    except: pass

def wp_media_clear_selection(driver):
    """Hàm từ dangbai_a.py: Xóa sạch các ảnh đã chọn trước đó"""
    driver.execute_script("""
    try {
      if (window.wp && wp.media) {
        if (wp.media.frame && wp.media.frame.state) {
           var sel = wp.media.frame.state().get('selection');
           if(sel) sel.reset();
        }
        if (wp.media.editor && wp.media.editor.get) {
            var fr = wp.media.editor.get();
            if (fr.state) {
                var sel = fr.state().get('selection');
                if(sel) sel.reset();
            }
        }
        document.querySelectorAll('li.attachment[aria-checked="true"]').forEach(el=>{
            el.setAttribute('aria-checked','false');
            el.classList.remove('selected','details');
        });
      }
    } catch(e) {}
    """)

def wp_media_upload_only(driver, wait: WebDriverWait, file_path: str) -> int:
    """
    Upload file lên WP media library và trả về attachment_id,
    nhưng KHÔNG click Insert — không chèn vào content editor.
    Dùng riêng cho ảnh đại diện (featured image).
    """
    filename = os.path.basename(file_path)
    # Kiểm tra thư viện trước
    att_id = wp_find_existing_media(driver, wait, filename)
    if att_id:
        # Đóng modal, không insert
        driver.execute_script("try{jQuery('.media-modal-close, .media-modal .media-modal-close').first().click();}catch(e){}")
        time.sleep(0.3)
        return att_id

    # Upload mới
    debug_print(f"   {Style.CYAN}📤 Featured img:{Style.RESET} {Style.DIM}{filename[:22]}{Style.RESET}", end="", flush=True)
    try:
        wp_media_clear_selection(driver)
        wp_open_add_media_modal(driver, wait)
        safe_switch_tab(driver, "Upload")
        inp = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']")))
        inp.send_keys(file_path)

        def file_uploaded(d):
            if d.find_elements(By.CSS_SELECTOR, ".upload-error, .error-message"): return "error"
            sel = d.find_elements(By.CSS_SELECTOR, "li.attachment.selected")
            if sel and sel[0].get_attribute("data-id"): return sel[0].get_attribute("data-id")
            return False

        result_id = WebDriverWait(driver, UPLOAD_TIMEOUT, poll_frequency=0.5).until(file_uploaded)
        if result_id and result_id != "error":
            att_id = int(result_id)
            debug_print(f" {Style.GREEN}✓ ID:{att_id}{Style.RESET}", flush=True)
        else:
            debug_print(f" {Style.RED}✗{Style.RESET}", flush=True)
            att_id = None
    except TimeoutException:
        debug_print(f" {Style.RED}✗ timeout{Style.RESET}", flush=True)
        att_id = None
    except Exception as e:
        debug_print(f" {Style.RED}✗ {e}{Style.RESET}", flush=True)
        att_id = None

    # Đóng modal — KHÔNG click Insert
    driver.execute_script("try{jQuery('.media-modal-close, .media-modal .media-modal-close').first().click();}catch(e){}")
    time.sleep(0.3)
    return att_id


def wp_find_existing_media(driver, wait, filename: str) -> int:
    """Tìm file đã có trong WP media library theo tên. Trả về att_id hoặc None.
    Luôn đóng modal trước khi return để không để lại trạng thái lơ lửng.
    """
    stem = os.path.splitext(filename)[0]
    found_id = None
    for attempt in range(2):
        try:
            wp_open_add_media_modal(driver, wait)
            safe_switch_tab(driver, "Thư viện")

            try:
                search_input = wait.until(EC.presence_of_element_located((By.ID, "media-search-input")))
                search_input.clear()
                time.sleep(0.1)
                search_input.send_keys(stem)
            except: pass

            start_wait = time.time()
            while time.time() - start_wait < 6.0:
                found_id = driver.execute_script("""
                    var target = arguments[0].toLowerCase();
                    var items = document.querySelectorAll('li.attachment.save-ready');
                    for (var i = 0; i < items.length; i++) {
                        var label = (items[i].getAttribute('aria-label') || "").toLowerCase();
                        if (label.indexOf(target) !== -1) {
                            items[i].click();
                            return items[i].getAttribute('data-id');
                        }
                    }
                    return null;
                """, stem)
                if found_id: break
                time.sleep(0.1)

            if found_id:
                # Tìm thấy — giữ modal mở, caller sẽ dùng ngay
                return int(found_id)
        except: pass

        try: driver.find_element(By.ID, "media-search-input").clear()
        except: pass

    # Không tìm thấy — đóng modal để trạng thái sạch cho upload tiếp theo
    try:
        driver.execute_script(
            "try{jQuery('.media-modal-close, .media-modal .media-modal-close').first().click();}catch(e){}")
        time.sleep(0.2)
    except: pass
    return None


def wp_media_upload_pick_insert(driver, wait: WebDriverWait, file_path: str) -> int:
    filename = os.path.basename(file_path)
    short    = filename[:22] + "…" if len(filename) > 22 else filename

    # Tìm trong thư viện trước
    att_id = wp_find_existing_media(driver, wait, filename)

    # Upload nếu chưa có
    if not att_id:
        debug_print(f"   {Style.CYAN}📤 {short}{Style.RESET}", end="", flush=True)
        try:
            wp_media_clear_selection(driver)
            wp_open_add_media_modal(driver, wait)
            safe_switch_tab(driver, "Upload")

            inp = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='file']")))
            inp.send_keys(file_path)

            try:
                def file_uploaded(d):
                    if d.find_elements(By.CSS_SELECTOR, ".upload-error, .error-message"): return "error"
                    sel = d.find_elements(By.CSS_SELECTOR, "li.attachment.selected")
                    if sel and sel[0].get_attribute("data-id"): return sel[0].get_attribute("data-id")
                    return False

                result_id = WebDriverWait(driver, UPLOAD_TIMEOUT, poll_frequency=0.5).until(file_uploaded)

                if result_id == "error":
                    debug_print(f" {Style.RED}✗ lỗi WP{Style.RESET}")
                elif result_id:
                    att_id = int(result_id)
                    debug_print(f" {Style.GREEN}✓ ID:{att_id}{Style.RESET}")

            except TimeoutException:
                debug_print(f" {Style.RED}✗ timeout{Style.RESET}")
                force_cleanup_modals(driver)
                return None

        except Exception as e:
            if "BROWSER_DIED" in str(e): raise e
            debug_print(f" {Style.RED}✗ {e}{Style.RESET}")
            return None

    # Chèn vào editor
    if att_id:
        wp_open_add_media_modal(driver, wait)

        driver.execute_script("""
            try {
                var id = arguments[0]; var wp = window.wp;
                if (wp && wp.media) {
                    var frame = null;
                    if (wp.media.editor && wp.media.editor.get) frame = wp.media.editor.get();
                    if (!frame && wp.media.frame) frame = wp.media.frame;
                    if(frame && frame.state) {
                        var sel = frame.state().get('selection');
                        if(sel) {
                            sel.reset();
                            var att = wp.media.attachment(id);
                            att.fetch();
                            sel.add(att);
                        }
                    }
                    var btn = document.querySelector('button.media-button-insert');
                    if (btn) btn.disabled = false;
                }
            } catch(e) {}
        """, att_id)

        inserted = False
        try:
            btn = driver.find_element(By.CSS_SELECTOR, ".media-button-insert:not([disabled])")
            if btn.is_displayed():
                driver.execute_script("arguments[0].click();", btn)
                inserted = True
        except: pass

        if not inserted:
            # 2. Tìm theo text đa ngôn ngữ
            candidates = driver.find_elements(By.CSS_SELECTOR, 
                ".media-modal button.media-button-insert, .media-modal button.media-button-select, "
                ".media-modal .media-toolbar button.media-button, .media-modal .media-toolbar-primary button.button-primary"
            )
            for b in candidates:
                if not b.is_displayed(): continue
                txt = ((b.text or "") + " " + (b.get_attribute("value") or "")).lower()
                if any(k in txt for k in ['chèn', 'insert', 'select', 'choose', 'đặt']):
                    try: 
                        driver.execute_script("arguments[0].click();", b)
                        inserted = True; break
                    except: pass
        
        if inserted:
             try:
                 wait.until(EC.invisibility_of_element_located((By.CSS_SELECTOR, ".media-modal")))
                 debug_print(f" {Style.GREEN}OK{Style.RESET}", flush=True)
             except: pass
        else:
            driver.execute_script("try{jQuery('.media-modal-close').click();}catch(e){}")
            debug_print(f" {Style.YELLOW}(Skip){Style.RESET}", flush=True)
            
    return att_id

# ==============================================================================
# 6. FEATURED IMAGE & CHUYÊN MỤC
# ==============================================================================

def set_featured_image_by_id(driver, wait, attachment_id: int) -> bool:
    if not attachment_id: return False
    debug_print(f"   {Style.CYAN}🖼 Ảnh đại diện ID:{attachment_id}{Style.RESET}", end="", flush=True)
    try:
        # Cách 1: Inject ID
        driver.execute_script("""
            try {
                var id = arguments[0];
                var input = document.getElementById('_thumbnail_id');
                if (input) { input.value = id; if (window.WpSetThumbnailHTML) { window.WpSetThumbnailHTML(id); return "injected"; } }
                if (window.wp && wp.data && wp.data.dispatch) { wp.data.dispatch('core/editor').editPost({ featured_media: id }); return "gutenberg"; }
            } catch(e) {}
        """, attachment_id)

        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#postimagediv img")))
            debug_print(f" {Style.GREEN}✓{Style.RESET}", flush=True); return True
        except TimeoutException: pass

        # Cách 2: UI Automation
        driver.execute_script("try{jQuery('.media-modal-close').click();}catch(e){}") 
        smart_click(driver, (By.ID, "set-post-thumbnail"))
        wp_open_add_media_modal(driver, wait)

        # JS Select
        driver.execute_script("""
            try {
                var id = arguments[0]; var wp = window.wp;
                var frame = (wp.media.featuredImage && wp.media.featuredImage.frame()) || wp.media.frame;
                if (frame) { var att = wp.media.attachment(id); att.fetch(); frame.state().get('selection').reset().add(att); }
            } catch(e) {}
        """, attachment_id)
        time.sleep(0.5)
        
        # Tìm nút bấm theo class
        candidates = driver.find_elements(By.CSS_SELECTOR, ".media-modal button.media-button-select, .media-modal button.button-primary")
        for b in candidates:
            if b.is_displayed(): 
                driver.execute_script("arguments[0].click();", b)
                break
        
        try:
            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "#postimagediv img")))
            debug_print(f" {Style.GREEN}✓{Style.RESET}", flush=True); return True
        except TimeoutException:
            debug_print(f" {Style.YELLOW}⚠ không xác nhận được{Style.RESET}", flush=True)
            return False

    except Exception as e:
        if "BROWSER_DIED" in str(e): raise e
        debug_print(f" {Style.RED}✗ {e}{Style.RESET}", flush=True); return False

def select_or_create_category(driver, wait, cat_path: str, cat_cache: dict = None):
    """
    Chọn hoặc tạo chuyên mục theo path "Cha/Con/Cháu".
    - Nếu không có "/" → hành vi cũ (tick 1 chuyên mục).
    - Nếu có "/" → tìm/tạo từng cấp, tick chuyên mục lá (cấp sâu nhất).
    cat_cache: dict chia sẻ trong cùng 1 worker.
    """
    cat_path = (cat_path or "").strip()
    if not cat_path: return False
    if cat_cache is None: cat_cache = {}

    # Cấp lá = phần cuối của path — đây là chuyên mục cần tick
    parts    = [p.strip() for p in cat_path.split('/') if p.strip()]
    cat_leaf = parts[-1] if parts else cat_path

    # Đã biết từ trước → tick thẳng
    if cat_path.lower() in cat_cache:
        driver.execute_script("""
            var name = arguments[0].toLowerCase();
            var labels = document.querySelectorAll('#categorychecklist label');
            for (var i = 0; i < labels.length; i++) {
                if (labels[i].innerText.trim().toLowerCase() === name) {
                    var cb = labels[i].querySelector('input');
                    if (!cb.checked) cb.click();
                    return true;
                }
            }
        """, cat_leaf)
        return True

    try:
        debug_print(f"   {Style.CYAN}📂 Chuyên mục:{Style.RESET} {Style.DIM}'{cat_path}'{Style.RESET}", end="")
        wait.until(EC.presence_of_element_located((By.ID, "categorychecklist")))
        cat_box = driver.find_element(By.ID, "categorydiv")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cat_box)

        def _find_and_tick(name):
            return driver.execute_script("""
                var name = arguments[0].toLowerCase();
                var labels = document.querySelectorAll('#categorychecklist label');
                for (var i = 0; i < labels.length; i++) {
                    if (labels[i].innerText.trim().toLowerCase() === name) {
                        var cb = labels[i].querySelector('input');
                        if (!cb.checked) cb.click();
                        return true;
                    }
                }
                return false;
            """, name)

        # Thử tick cấp lá trực tiếp trước
        if _find_and_tick(cat_leaf):
            cat_cache[cat_path.lower()] = True
            debug_print(f" {Style.GREEN}✓{Style.RESET}")
            return True

        debug_print(f" {Style.YELLOW}(tạo mới...){Style.RESET}", end="", flush=True)
        try:
            # Tạo từng cấp nếu chưa tồn tại (WP tự nest bằng newcategory_parent)
            parent_id = 0
            for part in parts:
                # Tìm đã có chưa
                exists = driver.execute_script("""
                    var name = arguments[0].toLowerCase();
                    var labels = document.querySelectorAll('#categorychecklist label');
                    for (var i = 0; i < labels.length; i++) {
                        if (labels[i].innerText.trim().toLowerCase() === name)
                            return true;
                    }
                    return false;
                """, part)
                if not exists:
                    smart_click(driver, (By.ID, "category-add-toggle"))
                    smart_send_keys(driver, (By.ID, "newcategory"), part)
                    if parent_id:
                        # Chọn parent trong dropdown newcategory_parent
                        try:
                            driver.execute_script("""
                                var sel = document.getElementById('newcategory_parent');
                                if (!sel) return;
                                for (var i=0; i<sel.options.length; i++) {
                                    if (sel.options[i].value == arguments[0]) {
                                        sel.value = arguments[0]; return;
                                    }
                                }
                            """, str(parent_id))
                        except: pass
                    smart_click(driver, (By.ID, "category-add-submit"))
                    import time as _t; _t.sleep(1.2)

                # Lấy id của cấp vừa tạo/tìm từ DOM
                parent_id = driver.execute_script("""
                    var name = arguments[0].toLowerCase();
                    var inputs = document.querySelectorAll('#categorychecklist input[type=checkbox]');
                    for (var i=0; i<inputs.length; i++) {
                        var lbl = inputs[i].parentElement;
                        if (lbl && lbl.innerText.trim().toLowerCase() === name)
                            return inputs[i].value;
                    }
                    return 0;
                """, part) or 0

            # Tick cấp lá sau khi tạo xong hierarchy
            WebDriverWait(driver, 10, poll_frequency=0.5).until(
                lambda d: _find_and_tick(cat_leaf))
            cat_cache[cat_path.lower()] = True
            debug_print(f" {Style.GREEN}✓ (đã tạo){Style.RESET}")
            return True
        except: pass

        # Fallback tick cái đầu tiên
        cbs = driver.find_elements(By.CSS_SELECTOR, "#categorychecklist input")
        if cbs: driver.execute_script("arguments[0].click();", cbs[0])
    except: pass
    return False

def set_wp_publish_datetime(driver, wait, dt: datetime):
    try:
        driver.execute_script("try{ document.querySelector('a.edit-timestamp').click(); }catch(e){}")
        wait.until(EC.presence_of_element_located((By.ID, "mm")))
        driver.execute_script("""
            function s(id,v){var e=document.getElementById(id);if(e){e.value=v;}}
            s('mm', arguments[0]); s('jj', arguments[1]); s('aa', arguments[2]);
            s('hh', arguments[3]); s('mn', arguments[4]);
        """, f"{dt.month:02d}", f"{dt.day:02d}", str(dt.year), f"{dt.hour:02d}", f"{dt.minute:02d}")
        
        smart_click(driver, (By.CSS_SELECTOR, ".save-timestamp"))
        
        if dt > datetime.now():
            debug_print(f"   {Style.YELLOW}🕒 Lên lịch: {dt.strftime('%d/%m/%Y %H:%M')}{Style.RESET}")
        else:
            debug_print(f"   {Style.CYAN}📅 {dt.strftime('%d/%m/%Y %H:%M')}{Style.RESET}")
    except: pass

def click_publish_ensure_visibility(driver, wait):
    force_cleanup_modals(driver)
    debug_print(f"   {Style.CYAN}📢 Đăng...{Style.RESET}", end="", flush=True)
    driver.execute_script("window.onbeforeunload = null;")

    try:
        driver.execute_script("window.scrollTo(0, 0);")
        pub_btn = wait.until(EC.presence_of_element_located((By.ID, "publish")))

        val = pub_btn.get_attribute("value").lower()
        is_schedule = "lên lịch" in val or "schedule" in val

        driver.execute_script("arguments[0].click();", pub_btn)

        def check_published(d):
            links = d.find_elements(By.PARTIAL_LINK_TEXT, "Xem bài viết") + d.find_elements(By.PARTIAL_LINK_TEXT, "View post")
            if links: return links[0].get_attribute("href")
            msgs = d.find_elements(By.CSS_SELECTOR, "#message.updated, .notice-success")
            if msgs and ("scheduled" in msgs[0].text.lower() or "lên lịch" in msgs[0].text.lower()): return "SCHEDULED"
            return False

        try:
            result = WebDriverWait(driver, 45, poll_frequency=0.5).until(check_published)
            if result == "SCHEDULED":
                debug_print(f" {Style.GREEN}✓ Lên lịch{Style.RESET}"); return "SCHEDULED"
            else:
                debug_print(f" {Style.GREEN}{Style.BOLD}✓ OK{Style.RESET}"); return result
        except TimeoutException:
            debug_print(f" {Style.RED}✗ Timeout{Style.RESET}"); return "TIMEOUT"

    except Exception as e:
        if "BROWSER_DIED" in str(e): raise e
        debug_print(f" {Style.RED}✗ {e}{Style.RESET}"); return "FAIL"

# ==============================================================================
# 8. XỬ LÝ NỘI DUNG (MAIN PROCESSING)
# ==============================================================================

def wp_insert_content_with_images(driver, wait, content_html, base_source) -> list:
    soup = BeautifulSoup(content_html or "", "html.parser")
    root = soup.body if soup.body else soup
    inserted_ids = []
    if not root: return []

    # ── Xóa rác CMS trước khi chèn vào WP ──────────────────────────────────
    # Các element này được scrape do bắt nhầm div cha (article-content thay vì
    # content-detail), không thuộc nội dung bài viết.
    _JUNK_SEL = [
        'h1.title-detail',        # tiêu đề — WP đã có title riêng
        'div.social-connect',     # FB/Zalo share buttons
        'div.block-core-a3',      # wrapper header (title + social)
        'div.rating', 'div#star-rating', 'div[id^="stringrating"]',
        'div.clearfix.mt-10',     # lượt xem + tác giả
        'div.author',
        'div.block_share',        # nút in, font-size
        'span.post-date', 'span.drash',
        'div.tac_gia_news',       # hanam.edu.vn: "Tác giả: thathanhson"
        '.network-share', '.fb-share-button', '.button-bookmark',
        'div[id^="audio"]',
    ]
    for _sel in _JUNK_SEL:
        for _el in root.select(_sel):
            _el.decompose()

    wp_editor_focus_end(driver)

    try:
        imgs  = root.find_all('img')
        files = [a for a in root.find_all('a') if is_download_link(a)]
        total_media = len(imgs) + len(files)
        if total_media > 0:
            debug_print(f"   {Style.CYAN}➜ {total_media} media ({len(imgs)} ảnh + {len(files)} file){Style.RESET}")
    except: return []

    _media_done = [0]   # counter dùng trong closure walk

    def walk(node):
        for child in list(node.children):
            if not isinstance(child, Tag):
                txt = str(child).strip()
                if txt: wp_editor_insert_html(driver, str(child))
                continue
            tn = child.name.lower()

            check_browser_alive(driver)

            if tn == "img":
                src = normalize_img_url(child.get("src"), base_source)
                if src:
                    path = download_file_resource(src)
                    if path:
                        att_id = wp_media_upload_pick_insert(driver, wait, path)
                        if att_id: inserted_ids.append(int(att_id))
                        if Xoa_file_sau_khi_dang:
                            cleanup_temp_file(path)
                    wp_editor_focus_end(driver)
                continue

            if tn == "a" and is_download_link(child):
                href = normalize_img_url(child.get("href"), base_source)
                if href:
                    path = download_file_resource(href)
                    if path:
                        att_id = wp_media_upload_pick_insert(driver, wait, path)
                        if att_id: inserted_ids.append(int(att_id))
                        if Xoa_file_sau_khi_dang:
                            cleanup_temp_file(path)
                    wp_editor_focus_end(driver)
                continue

            # ── YouTube iframe (kể cả youtube-nocookie.com) ──────────────────
            if tn == "iframe":
                src = child.get("src", "")
                is_yt = any(d in src for d in (
                    'youtube.com', 'youtube-nocookie.com', 'youtu.be'
                ))
                if is_yt:
                    safe_html = (
                        f'<div style="position:relative;padding-bottom:56.25%;'
                        f'height:0;overflow:hidden;margin:12px 0;">'
                        f'<iframe src="{src}" width="100%" height="450" '
                        f'frameborder="0" allowfullscreen="true" '
                        f'style="position:absolute;top:0;left:0;width:100%;height:100%;border:0;">'
                        f'</iframe></div>'
                    )
                    wp_editor_insert_html_raw(driver, safe_html)
                    wp_editor_focus_end(driver)
                continue  # bỏ qua iframe không phải YouTube

            if child.find('img') or any(is_download_link(a) for a in child.find_all('a')):
                walk(child)
            elif child.find('iframe'):
                walk(child)
            else:
                wp_editor_insert_html(driver, str(child))

    try: walk(root)
    except Exception as e:
        if "BROWSER_DIED" in str(e) or "WinError 10061" in str(e) or "HTTPConnectionPool" in str(e):
            debug_print(f"      {Style.RED}⚠ Mất kết nối trình duyệt!{Style.RESET}")
            raise Exception("BROWSER_DIED")
        debug_print(f"      {Style.RED}⚠ Lỗi duyệt HTML: {str(e)[:50]}{Style.RESET}")

    return inserted_ids

# ==============================================================================
# 9. BOT CONTROLLER
# ==============================================================================

# ── Helper: chuẩn hoá tiêu đề để so sánh ─────────────────────────────────────
def _norm_title(s: str) -> str:
    """Lowercase + collapse whitespace + strip dấu phụ."""
    if not s:
        return ""
    # NFKC để chuẩn hoá ký tự Unicode (tránh 2 cách encode cùng 1 chữ)
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _dup_api(base: str, title: str, post_type: str = "posts"):
    """
    Tìm bài trùng qua REST API /wp/v2/{post_type}?search=...
    Quét TẤT CẢ trang (pagination qua X-WP-TotalPages header).
    Trả về list[{id, status, page}] | None (REST không dùng được) | "403".
    """
    import html as _html_mod
    norm  = _norm_title(title)
    auth  = (EMAIL, PASSWORD)
    found = []
    _pt   = post_type if post_type else "posts"
    _MAX_PAGES = 50   # hard cap tránh vòng lặp vô hạn khi header sai

    for status_param in ["any", "publish,draft,pending,private"]:
        page        = 1
        total_pages = 1   # cập nhật sau response đầu tiên

        while page <= total_pages and page <= _MAX_PAGES:
            url = (f"{base}wp-json/wp/v2/{_pt}"
                   f"?search={quote_plus(title)}"
                   f"&per_page=100&status={status_param}"
                   f"&page={page}&_fields=id,title,status")
            try:
                r = requests.get(url, auth=auth, timeout=12, verify=False)
            except Exception:
                return None   # network lỗi → fallback Selenium

            if r.status_code == 401:
                return None
            if r.status_code == 403:
                return "403"
            if r.status_code == 400:
                # page vượt tổng (WP trả 400 khi page > total_pages)
                break
            if r.status_code not in (200, 201):
                return None

            # Đọc tổng số trang từ header (chỉ cần lần đầu)
            if page == 1:
                try:
                    _tp = int(r.headers.get("X-WP-TotalPages", 1))
                    # Guard: WP một số version trả 0 khi không có kết quả
                    total_pages = max(_tp, 1)
                except (ValueError, TypeError):
                    total_pages = 1

            items = r.json()
            if not isinstance(items, list):
                return None
            if not items:
                break   # hết bài → dừng sớm

            seen_ids = {x["id"] for x in found}
            for item in items:
                if item["id"] in seen_ids:
                    continue
                raw = item.get("title", {})
                t   = raw.get("rendered", "") if isinstance(raw, dict) else str(raw)
                try:
                    t = _html_mod.unescape(t)
                except Exception:
                    pass
                if _norm_title(t) == norm:
                    found.append({
                        "id":     item["id"],
                        "status": item.get("status", ""),
                        "page":   page,   # thông tin debug
                    })
            page += 1

        # Nếu status=any hoạt động (không cần thử lại với status riêng lẻ)
        if found or status_param == "any":
            break

    # Đánh dấu số trang đã quét vào metadata (để caller hiển thị)
    # page đã tăng lên 1 sau vòng cuối, nên pages_scanned = page - 1
    _ps = max(page - 1, 1)
    if found:
        found[0]['_pages'] = _ps
    elif _ps > 1:
        # Không tìm thấy nhưng đã quét nhiều trang → metadata
        found = [{'_pages': _ps, '_empty': True}]
    return found


def _dup_delete_api(base: str, post_id: int,
                    post_type: str = "posts") -> bool:
    """Xóa bài qua REST API (force=true → xóa thẳng không qua trash)."""
    _pt = post_type if post_type else "posts"
    url = f"{base}wp-json/wp/v2/{_pt}/{post_id}?force=true"
    try:
        r = requests.delete(url, auth=(EMAIL, PASSWORD), timeout=10, verify=False)
        return r.status_code in (200, 201)
    except Exception:
        return False


def _dup_selenium_fallback(driver, wait, base, title, mode,
                           post_type: str = "posts"):
    """
    Fallback khi REST API không dùng được.
    Dùng wp-admin search + Selenium nhưng với title normalization tốt hơn.
    Trả về "403" | True | False (giống check_process_duplicate).
    Guard: trả về None ngay nếu driver=None (gọi từ REST path).
    """
    # REST path gọi với driver=None → không thể dùng Selenium, bỏ qua
    if driver is None:
        post_log_set(dup_s="không kiểm tra (REST-only, API lỗi tạm)")
        return False

    norm = _norm_title(title)
    _pt  = post_type if post_type else "posts"
    try:
        s_url = (f"{base}wp-admin/edit.php"
                 f"?post_type={_pt}&s={quote_plus(title)}&post_status=all")
        driver.get(s_url)
        err = detect_http_error(driver)
        if err:
            code, _ = err
            return "403" if code == "403" else False

        wait.until(EC.presence_of_element_located((By.ID, "the-list")))
        if driver.find_elements(By.CSS_SELECTOR, "tr.no-items"):
            post_log_set(dup_s="0 bài trùng (Selenium)")
            return False

        rows_data = []
        for row in driver.find_elements(By.CSS_SELECTOR, "#the-list tr"):
            try:
                t_el = row.find_elements(By.CSS_SELECTOR, ".row-title")
                if not t_el:
                    continue
                if _norm_title(t_el[0].text) != norm:
                    continue
                row_class = row.get_attribute("class") or ""
                is_draft  = any(k in row_class for k in
                                ("status-draft", "status-pending", "status-auto-draft"))
                trash_a = row.find_elements(By.CSS_SELECTOR,
                                            "span.trash a, a.submitdelete")
                trash_href = trash_a[0].get_attribute("href") if trash_a else None
                rows_data.append({"is_draft": is_draft, "trash_href": trash_href})
            except StaleElementReferenceException:
                continue

        if not rows_data:
            post_log_set(dup_s="0 bài trùng (Selenium)")
            return False

        n_draft = sum(1 for r in rows_data if r["is_draft"])
        n_pub   = len(rows_data) - n_draft

        deleted = 0
        for rd in rows_data:
            is_draft   = rd["is_draft"]
            trash_href = rd["trash_href"]

            if mode == 1:
                parts = []
                if n_draft: parts.append(f"{n_draft} nháp")
                if n_pub:   parts.append(f"{n_pub} đăng")
                post_log_set(dup_s=f"{', '.join(parts)} (Selenium) → bỏ qua")
                return True

            should_delete = (
                mode == 4
                or (mode == 2 and is_draft)
                or (mode == 3 and not is_draft)
            )
            if should_delete and trash_href:
                driver.get(trash_href)
                deleted += 1

        parts = []
        if n_draft: parts.append(f"{n_draft} nháp")
        if n_pub:   parts.append(f"{n_pub} đăng")
        dup_label = ", ".join(parts) if parts else "0 bài trùng"
        post_log_set(dup_s=f"{dup_label} (Selenium){f' → đã xóa {deleted}' if deleted else ''}")
        return False

    except Exception as e:
        post_log_set(dup_s=f"lỗi quét ({str(e)[:30]})")
        return False





def _claim_title(base: str, norm_title: str) -> bool:
    """
    Đánh dấu title đang được xử lý (atomic check-and-set).
    Trả về True nếu claim thành công, False nếu title đã có luồng khác đang làm.
    """
    with _pending_lock:
        s = _pending_titles.setdefault(base, set())
        if norm_title in s:
            return False
        s.add(norm_title)
        return True


def _release_title(base: str, norm_title: str):
    """Trả title về pool sau khi xong (thành công hay thất bại)."""
    with _pending_lock:
        _pending_titles.get(base, set()).discard(norm_title)


def check_process_duplicate(driver, wait, base, title, mode: int,
                            post_type: str = "posts"):
    """
    Kiểm tra và xử lý bài viết trùng theo DUPLICATE_MODE:
      0 → Không làm gì → False (cứ đăng)
      1 → Nếu tồn tại → True (bỏ qua)
      2 → Xóa bản nháp trùng → False (vẫn đăng mới)
      3 → Xóa bài đã đăng trùng → False (vẫn đăng mới)
      4 → Xóa tất cả bài trùng (nháp + đăng) → False (vẫn đăng mới)

    Trả về:
      "403"  — Gặp lỗi 403 (dừng site)
      True   — Bỏ qua bài này (mode 1 + tồn tại)
      False  — Tiến hành đăng bình thường
    """
    if mode == 0:
        post_log_set(dup_s="0 bài trùng (không kiểm tra)")
        return False

    # ── Ưu tiên REST API — nhanh hơn và đáng tin hơn Selenium ────────────────
    api_result = _dup_api(base, title, post_type=post_type)

    if api_result == "403":
        return "403"

    if api_result is None:
        # REST API không dùng được → fallback Selenium (driver=None từ REST path trả False ngay)
        return _dup_selenium_fallback(driver, wait, base, title, mode,
                                      post_type=post_type)

    # api_result là list[{id, status, ...}]
    # Lọc ra metadata items (_empty=True)
    _pages_scanned = api_result[0].get('_pages', 1) if api_result else 1
    real_results   = [x for x in api_result if not x.get('_empty')]
    _pages_s = f' ({_pages_scanned} trang)' if _pages_scanned > 1 else ''

    if not real_results:
        post_log_set(dup_s=f"0 bài trùng{_pages_s}")
        return False

    # Đếm nháp / đã đăng (chỉ dùng real_results, không đếm metadata)
    n_draft = sum(1 for x in real_results
                  if x["status"] in ("draft","pending","auto-draft","private"))
    n_pub   = len(real_results) - n_draft

    deleted = 0
    for item in real_results:
        post_id  = item["id"]
        status   = item["status"]
        is_draft = status in ("draft", "pending", "auto-draft", "private")

        if mode == 1:
            parts = []
            if n_draft: parts.append(f"{n_draft} nháp")
            if n_pub:   parts.append(f"{n_pub} đăng")
            post_log_set(dup_s=f"{', '.join(parts)}{_pages_s} → bỏ qua")
            return True

        should_delete = (
            mode == 4
            or (mode == 2 and is_draft)
            or (mode == 3 and not is_draft)
        )
        if should_delete:
            if _dup_delete_api(base, post_id, post_type=post_type):
                deleted += 1

    parts = []
    if n_draft: parts.append(f"{n_draft} nháp")
    if n_pub:   parts.append(f"{n_pub} đăng")
    dup_label = ", ".join(parts) if parts else "0 bài trùng"
    post_log_set(dup_s=f"{dup_label}{_pages_s}{f' → đã xóa {deleted}' if deleted else ''}")
    return False

def clean_title(s): return "".join(ch for ch in (s or "") if ord(ch) <= 0xFFFF).strip()

def map_category_name(raw_name):
    if not raw_name: return "Tin tức chung"
    rn = raw_name.strip().lower()
    if "tin tức hoạt động" in rn: return "TIN TỨC HOẠT ĐỘNG"
    if "công khai" in rn: return "CÔNG KHAI"
    return raw_name

def force_real_user_gesture(driver, element):
    """
    Tạo user gesture thật để unlock input (chống treo login lâu ngày)
    """
    driver.execute_script("""
        arguments[0].scrollIntoView({block:'center'});
    """, element)
    time.sleep(0.2)

    # Click thật (ActionChains dễ bị detect, dùng JS + mouse event)
    driver.execute_script("""
        var ev = new MouseEvent('mousedown', {bubbles:true});
        arguments[0].dispatchEvent(ev);
        arguments[0].focus();
    """, element)
    time.sleep(0.3)


# ==============================================================================
# 8a. REST API POSTER — không cần Chrome, nhanh ~10x so với Selenium
# ==============================================================================

# Per-domain semaphore: giới hạn N request đồng thời đến cùng 1 domain (anti-ban)
_domain_semaphores: dict = {}
_domain_sem_lock   = threading.Lock()

def _get_domain_sem(base_url: str) -> threading.Semaphore:
    """Trả về semaphore của domain, tạo mới nếu chưa có."""
    domain = urlparse(base_url).netloc
    with _domain_sem_lock:
        if domain not in _domain_semaphores:
            _domain_semaphores[domain] = threading.Semaphore(REST_DOMAIN_CONCURRENCY)
        return _domain_semaphores[domain]


def _rest_upload_media(session: requests.Session, base: str,
                       file_path: str) -> tuple:
    """
    Upload 1 file lên WP media library qua REST API.
    Trả về (attachment_id, source_url) — source_url lấy thẳng từ response
    của WP, không cần gọi GET thứ 2.
    """
    filename = os.path.basename(file_path)
    mime     = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    try:
        with open(file_path, "rb") as fh:
            data = fh.read()
        r = session.post(
            f"{base}wp-json/wp/v2/media",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Type"       : mime,
            },
            data=data, timeout=120,
        )
        if r.status_code in (200, 201):
            j = r.json()
            return j.get("id"), j.get("source_url", "")
        debug_print(f"   {Style.YELLOW}⚠ media upload HTTP {r.status_code}:{Style.RESET} {Style.DIM}{r.text[:80]}{Style.RESET}")
    except Exception as e:
        debug_print(f"   {Style.YELLOW}⚠ media upload lỗi:{Style.RESET} {Style.DIM}{e}{Style.RESET}")
    return None, ""


# Số luồng song song cho upload ảnh — lấy từ config
_UPLOAD_WORKERS = REST_UPLOAD_WORKERS

def _rest_process_content(session: requests.Session,
                           base: str, source_base: str,
                           content_html: str) -> tuple:
    """
    Xử lý HTML content cho REST API:
      0. Strip junk CMS (h1.title-detail, rating, author...) — nhất quán với Selenium path
      1. Download tất cả ảnh + file song song (ThreadPoolExecutor)
      2. Upload lên WP media song song (tối đa _UPLOAD_WORKERS luồng)
         → source_url lấy thẳng từ response upload, không GET thứ 2
      3. Trả về (content_đã_sửa, featured_attachment_id hoặc None)
    """
    if not content_html:
        return content_html, None

    soup = BeautifulSoup(content_html, "html.parser")

    # ── Bước 0: Strip junk CMS — PHẢI làm trước collect media ────────────────
    # Nhất quán với wp_insert_content_with_images (Selenium path)
    _JUNK_SEL = [
        'h1.title-detail',
        'div.social-connect', 'div.block-core-a3',
        'div.rating', 'div#star-rating', 'div[id^="stringrating"]',
        'div.clearfix.mt-10', 'div.author',
        'div.block_share', 'span.post-date', 'span.drash',
        'div.tac_gia_news',
        '.network-share', '.fb-share-button', '.button-bookmark',
        'div[id^="audio"]',
    ]
    for _sel in _JUNK_SEL:
        for _el in soup.select(_sel):
            _el.decompose()

    first_att_id = None

    # ── Bước 1: Collect tất cả media cần upload ──────────────────────────────
    img_jobs  = []   # list of (img_tag, normalized_src)
    file_jobs = []   # list of (a_tag, normalized_href)

    # Các attr lazy-load cần thu thập (theo thứ tự ưu tiên)
    _IMG_ATTRS = ('src', 'data-src', 'data-original', 'data-lazy',
                  'data-lazy-src', 'data-original-src', 'data-url')

    _seen_img_urls: set = set()   # tránh upload trùng cùng URL

    for img in soup.find_all("img"):
        picked_url = ""
        for attr in _IMG_ATTRS:
            raw = (img.get(attr) or "").strip()
            if not raw or raw.startswith("data:"):
                continue
            nsrc = normalize_img_url(raw, source_base)
            if nsrc and nsrc not in _seen_img_urls:
                picked_url = nsrc
                break
        # Fallback: srcset — lấy URL đầu tiên
        if not picked_url:
            srcset = (img.get("srcset") or "").strip()
            if srcset:
                first_entry = srcset.split(",")[0].strip().split()[0]
                if first_entry:
                    candidate = normalize_img_url(first_entry, source_base)
                    if candidate and candidate not in _seen_img_urls:
                        picked_url = candidate
        if picked_url:
            img_jobs.append((img, picked_url))
            _seen_img_urls.add(picked_url)

    for a in soup.find_all("a", href=True):
        if not is_download_link(a): continue
        href = normalize_img_url(a.get("href", ""), source_base)
        if href: file_jobs.append((a, href))

    if not img_jobs and not file_jobs:
        return str(soup), None

    # ── Bước 2: Download + upload song song ──────────────────────────────────
    def _process_one(url: str) -> tuple:
        """Download rồi upload. Trả về (url_gốc, att_id, source_url_mới)."""
        path = download_file_resource(url)
        if not path:
            return url, None, ""
        att_id, new_url = _rest_upload_media(session, base, path)
        if Xoa_file_sau_khi_dang:
            cleanup_temp_file(path)
        return url, att_id, new_url

    all_urls    = [u for _, u in img_jobs] + [u for _, u in file_jobs]
    url_results = {}   # original_url → (att_id, new_url)

    with ThreadPoolExecutor(max_workers=_UPLOAD_WORKERS) as pool:
        for orig_url, att_id, new_url in pool.map(_process_one, all_urls):
            url_results[orig_url] = (att_id, new_url)

    # ── Bước 3: Thay src/href trong soup ─────────────────────────────────────
    for img, nsrc in img_jobs:
        att_id, new_url = url_results.get(nsrc, (None, ""))
        if att_id:
            if new_url:
                img["src"] = new_url
            # Xoá tất cả lazy-load attrs + srcset cũ để trình duyệt/WP
            # không fallback về domain nguồn sau khi đã upload thành WP URL
            for _la in ('data-src', 'data-original', 'data-lazy', 'data-lazy-src',
                        'data-original-src', 'data-url', 'srcset'):
                if _la in img.attrs:
                    del img.attrs[_la]
            if first_att_id is None:
                first_att_id = att_id
        # Upload thất bại → giữ nguyên src gốc, không để ảnh hỏng

    for a, href in file_jobs:
        att_id, new_url = url_results.get(href, (None, ""))
        if att_id and new_url:
            a["href"] = new_url
        # Nếu upload fail: giữ href gốc (link về nguồn) thay vì để link hỏng

    debug_print(f"   {Style.CYAN}📦 Upload media:{Style.RESET} {Style.BOLD}{len(img_jobs)}{Style.RESET} ảnh + {Style.BOLD}{len(file_jobs)}{Style.RESET} file"
                f" ({_UPLOAD_WORKERS} luồng)")
    return str(soup), first_att_id


def _rest_get_or_create_category(session: requests.Session,
                                  base: str, cat_path: str,
                                  _cache: dict) -> int | None:
    """
    Lấy/tạo category theo path dạng "Cha/Con/Cháu".
    - Nếu cat_path không có "/" → 1 cấp như cũ.
    - Nếu có "/" → tạo/tìm từng cấp theo đúng hierarchy.
    _cache: dict được chia sẻ trong cùng 1 worker.
    """
    if not cat_path or not cat_path.strip():
        return None

    parts = [p.strip() for p in cat_path.split('/') if p.strip()]
    if not parts:
        return None

    # Cảnh báo nếu chỉ có 1 cấp nhưng tên gợi ý thiếu cha
    # (ví dụ: "Video clip" thay vì "Tài nguyên/Video clip")
    if len(parts) == 1:
        _KNOWN_CHILDREN = {
            "video clip", "thư viện ảnh", "tài liệu", "văn bản pháp quy",
            "thông tư", "nghị định", "thông báo", "hoạt động chuyên môn",
        }
        if parts[0].lower() in _KNOWN_CHILDREN:
            debug_print(f"{Style.YELLOW}⚠ Chuyên mục '{parts[0]}' có thể thiếu cấp cha "
                        f"— xem xét điền dạng 'Cha/{parts[0]}' vào cột cat_id{Style.RESET}")

    # Lấy toàn bộ categories 1 lần để giảm HTTP calls
    all_cats_key = f"{base}||__all__"
    if all_cats_key not in _cache:
        try:
            cats = []
            page = 1
            while True:
                r = session.get(
                    f"{base}wp-json/wp/v2/categories",
                    params={"per_page": 100, "page": page, "_fields": "id,name,parent"},
                    timeout=10,
                )
                if not r.ok:
                    break
                batch = r.json()
                if not isinstance(batch, list) or not batch:
                    break
                cats.extend(batch)
                if len(batch) < 100:
                    break
                page += 1
            _cache[all_cats_key] = cats
        except Exception:
            _cache[all_cats_key] = []

    all_cats: list = _cache.get(all_cats_key, [])

    def _find_cat(name: str, parent_id: int) -> int | None:
        """Tìm category theo name + parent_id trong list đã cache."""
        name_l = name.strip().lower()
        for c in all_cats:
            if (c.get("name", "").strip().lower() == name_l
                    and c.get("parent", 0) == parent_id):
                return c["id"]
        return None

    def _create_cat(name: str, parent_id: int) -> int | None:
        """Tạo category mới, cập nhật cache all_cats."""
        payload = {"name": name}
        if parent_id:
            payload["parent"] = parent_id
        try:
            r = session.post(
                f"{base}wp-json/wp/v2/categories",
                json=payload, timeout=10,
            )
            if r.status_code in (200, 201):
                c = r.json()
                all_cats.append({"id": c["id"], "name": name, "parent": parent_id})
                return c["id"]
        except Exception:
            pass
        return None

    parent_id = 0
    last_id   = None
    path_built = ""

    for part in parts:
        path_built = f"{path_built}/{part}" if path_built else part
        cache_key  = f"{base}||{path_built.lower()}"

        if cache_key in _cache:
            last_id   = _cache[cache_key]
            parent_id = last_id
            continue

        cat_id = _find_cat(part, parent_id)
        if cat_id is None:
            cat_id = _create_cat(part, parent_id)
        if cat_id is None:
            return last_id      # tạo thất bại → trả cấp đã tạo được

        _cache[cache_key] = cat_id
        last_id   = cat_id
        parent_id = cat_id

    return last_id


def _rest_format_date(date_val) -> str | None:
    """Chuyển date/str → ISO 8601 mà WP REST chấp nhận: 2025-10-04T08:00:00"""
    if date_val is None:
        return None
    if isinstance(date_val, datetime):
        return date_val.strftime("%Y-%m-%dT%H:%M:%S")
    if isinstance(date_val, date):
        return datetime(date_val.year, date_val.month, date_val.day,
                        DEFAULT_PUBLISH_HOUR, DEFAULT_PUBLISH_MINUTE).strftime("%Y-%m-%dT%H:%M:%S")
    try:
        d = datetime.strptime(str(date_val).strip(), "%d/%m/%Y")
        return d.replace(hour=DEFAULT_PUBLISH_HOUR,
                         minute=DEFAULT_PUBLISH_MINUTE).strftime("%Y-%m-%dT%H:%M:%S")
    except Exception:
        return None


def rest_post_one(anew, session: requests.Session,
                  cat_cache: dict) -> dict:
    """
    Đăng 1 bài qua REST API. Không cần Chrome.
    Trả về dict res (cùng schema với NewsPoster.post_one).
    """
    t0    = time.time()
    title = clean_title(anew[2])
    base  = get_base(anew[7])
    res   = {
        "ts"           : datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "base_target"  : base,
        "news_id"      : anew[0],
        "title"        : title,
        "status"       : "FAIL",
        "uploaded_ids" : [],
    }

    post_log_start(base, title)   # khởi buffer cho bài này
    _norm_t = _norm_title(title)   # cần ở cuối hàm để _release_title

    sem = _get_domain_sem(base)
    with sem:
        try:
            # ── 1. Kiểm tra trùng — dùng _pending_titles chống race condition ──
            # Nếu 2 luồng cùng xử lý 1 title: luồng thứ 2 skip ngay lập tức.
            # Khác với lock/mutex: luồng khác title chạy song song bình thường.
            res["step"] = "check_duplicate"
            _pt_dup     = (str(anew[13]).strip() if len(anew) > 13 and anew[13] else "posts")

            if not _claim_title(base, _norm_t):
                # Luồng khác đang đăng bài này → bỏ qua
                res["status"] = "SKIP_DUPLICATE"
                post_log_set(dup_s="đang xử lý ở luồng khác")
                post_log_set(final_s="⏭ Trùng (song song)", duration=round(time.time()-t0,1))
                post_log_flush(); return res

            try:   # đảm bảo luôn release dù có lỗi
                dup = check_process_duplicate(None, None, base, title, DUPLICATE_MODE,
                                              post_type=_pt_dup)
                if dup == "403":
                    res["status"] = "403_FORBIDDEN"
                    post_log_set(final_s="❌ 403 Forbidden", duration=round(time.time()-t0,1))
                    _release_title(base, _norm_t)
                    post_log_flush(); return res
                if dup is True:
                    res["status"] = "SKIP_DUPLICATE"
                    post_log_set(final_s="⏭ Trùng", duration=round(time.time()-t0,1))
                    _release_title(base, _norm_t)
                    post_log_flush(); return res
            except Exception as _e_dup:
                _release_title(base, _norm_t)
                raise _e_dup

            # ── 2. Xử lý content + upload media ─────────────────────────────
            res["step"]  = "content"
            # Dùng URL đầy đủ của bài gốc (không chỉ base domain) để urljoin
            # xử lý đúng path relative kiểu ../images/ hay ./files/
            source_base  = (str(anew[6]).strip() if len(anew) > 6 and anew[6] else "") or base
            content_html = anew[3] or ""
            content_new, first_content_id = _rest_process_content(
                session, base, source_base, content_html)

            # Đếm media đã upload để hiển thị
            _csoup     = BeautifulSoup(content_html, "html.parser") if content_html else None
            _img_count = len(_csoup.find_all("img")) if _csoup else 0
            post_log_step(f"{Style.B_GREEN}📝 soạn mới{Style.RESET}")
            if _img_count:
                post_log_step(f"{Style.CYAN}🖼 media({_img_count}){Style.RESET}")

            # ── 3. Featured image ─────────────────────────────────────────────
            featured_id = None
            thumb_urls  = anew[4] if (len(anew) > 4 and isinstance(anew[4], list)
                                      and anew[4]) else []
            thumb_url   = thumb_urls[0] if thumb_urls else ""
            _thumb_from_content = False   # True nếu thumb đã upload cùng content

            def _norm_url_cmp(u):
                return u.split('?')[0].rstrip('/').lower()

            if thumb_url and "no-image" not in thumb_url.lower():
                # Kiểm tra thumb URL có trùng với ảnh đầu tiên trong content không
                # (xét cả src gốc lẫn lazy-load attrs để tránh upload lại)
                _content_src_urls = []
                if content_html:
                    _cs2 = BeautifulSoup(content_html, "html.parser")
                    _LAZY_A = ('src', 'data-src', 'data-original', 'data-lazy',
                               'data-lazy-src', 'data-original-src', 'data-url')
                    for _img2 in _cs2.find_all("img"):
                        for _a2 in _LAZY_A:
                            _v2 = (_img2.get(_a2) or "").strip()
                            if _v2 and not _v2.startswith("data:"):
                                _content_src_urls.append(
                                    _norm_url_cmp(normalize_img_url(_v2, source_base)))

                _thumb_norm = _norm_url_cmp(thumb_url)
                if _thumb_norm in _content_src_urls and first_content_id:
                    # Thumb chính là ảnh đầu trong content — dùng ID đã upload
                    featured_id = first_content_id
                    _thumb_from_content = True
                else:
                    # Thumb riêng biệt — cần download + upload độc lập
                    thumb_path = download_file_resource(thumb_url)
                    if thumb_path:
                        att_id, _ = _rest_upload_media(session, base, thumb_path)
                        if Xoa_file_sau_khi_dang:
                            cleanup_temp_file(thumb_path)
                        if att_id:
                            featured_id = att_id
                        else:
                            debug_print(f"   {Style.YELLOW}⚠ Upload thumb thất bại — fallback ảnh content{Style.RESET}")
                    else:
                        debug_print(f"   {Style.YELLOW}⚠ Download thumb thất bại: {thumb_url[:60]}{Style.RESET}")

            # Fallback: dùng ảnh đầu tiên trong content CHỈ KHI thumb_url không có
            # (không fallback khi thumb_url có nhưng upload fail — tránh dùng ảnh sai)
            if featured_id is None and not thumb_url and first_content_id:
                featured_id = first_content_id

            # ── 4. Category — chỉ dùng từ Excel (cột cat_id, anew[9]) ─────────
            res["step"] = "meta"
            # anew[9] = db_cat: dạng "Cha" hoặc "Cha/Con" hoặc "Cha/Con/Cháu"
            cat_path = (anew[9] or "").strip() if len(anew) > 9 else ""
            if not cat_path:
                # Fallback: nếu cột 9 trống, thử cột 8 (raw_cat từ nguồn)
                cat_path = (anew[8] or "").strip() if len(anew) > 8 else ""
            res["category"] = cat_path
            post_log_set(category=cat_path.split('/')[-1] if cat_path else "")
            cat_id  = _rest_get_or_create_category(session, base, cat_path, cat_cache) if cat_path else None
            post_log_step(f"{Style.YELLOW}📂 chuyên mục{Style.RESET}")

            # ── 5. Tạo post ──────────────────────────────────────────────────
            res["step"]    = "publish"
            date_str       = anew[12] if len(anew) > 12 else None
            wp_date        = _rest_format_date(date_str)

            _is_future_date = (
                wp_date is not None
                and wp_date > datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
            )
            if wp_date:
                post_log_set(date_s=wp_date.replace("T", " "))

            payload: dict = {
                "title"  : title,
                "content": content_new,
                "status" : "future" if _is_future_date else "publish",
            }
            if cat_id:
                payload["categories"] = [cat_id]
            if featured_id:
                payload["featured_media"] = featured_id
            if wp_date:
                payload["date"] = wp_date

            # Endpoint dựa vào post_type (anew[13] nếu có, mặc định "posts")
            _pt = (str(anew[13]).strip() if len(anew) > 13 and anew[13] else "posts")
            _pt = _pt if _pt else "posts"
            r = session.post(
                f"{base}wp-json/wp/v2/{_pt}",
                json=payload, timeout=45,
            )

            if r.status_code in (200, 201):
                post_data     = r.json()
                post_id       = post_data.get("id")
                public_url    = post_data.get("link", "")
                res["status"] = "SCHEDULED" if _is_future_date else "SUCCESS"
                res["public_url"] = public_url

                # ── Xác nhận bài đã đăng thật sự ────────────────────────────
                try:
                    v_url = (f"{base}wp-json/wp/v2/{_pt}/{post_id}"
                             f"?_fields=id,status,link")
                    rv = session.get(v_url, timeout=10)
                    if rv.ok:
                        v_data   = rv.json()
                        v_status = v_data.get("status", "")
                        expected = "future" if _is_future_date else "publish"
                        if v_status != expected:
                            res["status"] = f"VERIFY_FAIL_{v_status.upper()}"
                            post_log_set(final_s=f"⚠ Đăng xong nhưng status={v_status}")
                        else:
                            icon = "🕒 Lên lịch" if _is_future_date else "✅ Xong"
                            post_log_set(final_s=icon)
                    else:
                        icon = "🕒 Lên lịch" if _is_future_date else "✅ Xong"
                        post_log_set(final_s=icon)
                except Exception:
                    icon = "🕒 Lên lịch" if _is_future_date else "✅ Xong"
                    post_log_set(final_s=icon)
            elif r.status_code == 403:
                res["status"] = "403_FORBIDDEN"
                res["error"]  = "403 POST /wp/v2/posts"
                post_log_set(final_s="❌ 403 Forbidden")
            else:
                res["status"] = f"REST_HTTP_{r.status_code}"
                res["error"]  = r.text[:200]
                post_log_set(final_s=f"❌ HTTP {r.status_code}")

        except Exception as e:
            res["error"]  = str(e)
            res["status"] = "FAIL"
            post_log_set(final_s=f"❌ Lỗi: {str(e)[:40]}")

    # Luôn release title sau khi xong (dù thành công hay thất bại)
    _release_title(base, _norm_t)

    # Anti-ban: jitter delay
    time.sleep(random.uniform(REST_DELAY_MIN, REST_DELAY_MAX))

    res["duration_sec"] = round(time.time() - t0, 2)
    post_log_set(duration=round(time.time() - t0, 1))
    post_log_flush()
    return res


def _rest_api_works(base: str) -> bool:
    """
    Kiểm tra nhanh xem site có hỗ trợ WP REST API với Basic Auth không.
    Dùng endpoint /wp/v2/users/me (yêu cầu auth).
    """
    try:
        r = requests.get(
            f"{base}wp-json/wp/v2/users/me",
            auth=(EMAIL, PASSWORD),
            timeout=8, verify=False,
        )
        return r.status_code == 200
    except Exception:
        return False


class NewsPoster:
    def __init__(self):
        self.options = Options()
        self.options.binary_location = CHROME_BINARY
        # ── Ổn định (fix GetHandleVerifier crash trên Windows) ─────────────
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_argument("--disable-gpu")
        self.options.add_argument("--disable-extensions")
        self.options.add_argument("--disable-notifications")
        self.options.add_argument("--disable-software-rasterizer")
        self.options.add_argument("--disable-background-networking")
        self.options.add_argument("--disable-background-timer-throttling")
        self.options.add_argument("--disable-renderer-backgrounding")
        self.options.add_argument("--disable-backgrounding-occluded-windows")
        # Giới hạn renderer process — nguồn gốc GetHandleVerifier
        self.options.add_argument("--renderer-process-limit=2")
        # Tắt features gây crash handle trên Windows multi-process
        self.options.add_argument("--disable-features=VizDisplayCompositor,TranslateUI")
        self.options.add_argument("--no-first-run")
        self.options.add_argument("--mute-audio")
        # Kích thước cố định; minimize sẽ được gọi sau khi driver khởi tạo xong
        self.options.add_argument("--window-size=1024,768")
        # Tải trang: eager = DOM ready không cần network idle
        self.options.add_argument("--page-load-strategy=eager")
        self.options.add_argument(f"user-agent={get_random_ua()}")
        self.options.add_argument("--blink-settings=imagesEnabled=true")

        # ── Anti-Detect ──────────────────────────────────────────────────────
        self.options.add_argument("--disable-blink-features=AutomationControlled")
        self.options.add_experimental_option("excludeSwitches", ["enable-automation"])
        self.options.add_experimental_option("useAutomationExtension", False)

        if USE_PROFILE: self.options.add_argument(rf"--user-data-dir={PROFILE_DIR}")
        
        self.service = Service(CHROMEDRIVER_PATH)
        self.driver = None
        self.wait = None
        
        # Khởi tạo driver ngay lập tức
        self.start_driver()
        
        self.logged_sites = set()
        self.failed_sites = {}
        self.cat_cache    = {}   # Cache chuyên mục: tránh duyệt lại DOM mỗi bài

    def start_driver(self):
        try:
            if self.driver:
                try: self.driver.quit()
                except: pass
                self.driver = None

            service = Service(CHROMEDRIVER_PATH)
            self.driver = webdriver.Chrome(service=service, options=self.options)

            # [ANTI-DETECT]
            self.driver.execute_cdp_cmd(
                "Page.addScriptToEvaluateOnNewDocument",
                {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"}
            )

            self.wait = WebDriverWait(self.driver, 45)

            # Thu nhỏ cửa sổ Chrome ngay sau khi mở
            if not SHOW_CHROME_WINDOW:
                try:
                    # Cách 1: Chrome DevTools Protocol — không cần sleep, không flicker
                    wid = self.driver.execute_cdp_cmd('Browser.getWindowForTarget', {})
                    self.driver.execute_cdp_cmd('Browser.setWindowBounds', {
                        'windowId': wid['windowId'],
                        'bounds':   {'windowState': 'minimized'},
                    })
                except Exception:
                    try:
                        # Cách 2: Selenium fallback với delay nhỏ
                        time.sleep(0.3)
                        self.driver.minimize_window()
                    except Exception:
                        pass

        except Exception as e:
            debug_print(f"{Style.RED}❌ Lỗi khởi tạo Driver: {e}{Style.RESET}")
            raise e

    def restart_driver(self):
        debug_print(f"{Style.YELLOW}♻️ Đang khởi động lại trình duyệt (fresh driver)...{Style.RESET}")
        self.start_driver()

    def ensure_wp_login(self, base_url, user, pwd):
        assert_driver_alive(self.driver)
        real_base = get_base(base_url)

        if self.failed_sites.get(real_base, 0) >= 3:
            post_log_set(login_s="Thất bại (≥3 lần)")
            return False
        if real_base in self.logged_sites:
            return True   # login_s không đổi — đã set lần trước

        for attempt in range(3):
            try:
                self.driver.get(real_base + "wp-login.php")
                wait_for_page_load(self.driver)

                if not wait_for_wp_test_cookie(self.driver):
                    raise Exception("WP_TEST_COOKIE_NOT_READY")

                if is_403(self.driver):
                    post_log_set(login_s="403 Forbidden")
                    return "403"

                if self.driver.find_elements(By.ID, "wpadminbar"):
                    self.logged_sites.add(real_base)
                    post_log_set(login_s="OK (cookie)")
                    return True

                wait = WebDriverWait(self.driver, 20)
                user_el = wait.until(EC.visibility_of_element_located((By.ID, "user_login")))
                pass_el = wait.until(EC.visibility_of_element_located((By.ID, "user_pass")))

                user_el.clear(); user_el.send_keys(user)
                pass_el.clear(); pass_el.send_keys(pwd)
                smart_click(self.driver, (By.ID, "wp-submit"))

                WebDriverWait(self.driver, 15).until(
                    lambda d: "wp-admin" in d.current_url or d.find_elements(By.ID, "wpadminbar")
                )

                self.logged_sites.add(real_base)
                self.failed_sites[real_base] = 0
                login_label = f"OK (lần {attempt+1})" if attempt > 0 else "OK"
                post_log_set(login_s=login_label)
                return True

            except Exception as e:
                if "BROWSER_DIED" in str(e) or "10061" in str(e):
                    post_log_set(login_s=f"crash lần {attempt+1}")
                    raise e
                post_log_set(login_s=f"Lỗi lần {attempt+1}")
                if attempt < 2:
                    time.sleep(1)
                    continue

        self.failed_sites[real_base] = self.failed_sites.get(real_base, 0) + 1
        post_log_set(login_s=f"Thất bại ({self.failed_sites[real_base]}/3)")
        return False
    
    def verify_public_url(self, url):
        try:
            headers = {'User-Agent': get_random_ua()}
            r = requests.get(url, headers=headers, verify=False, timeout=15)
            return r.status_code == 200
        except: return False

    def post_one(self, anew, idx, total_in_site):
        t0 = time.time(); title = clean_title(anew[2]); base = get_base(anew[7])
        res = {"ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "base_target": base,
               "news_id": anew[0], "title": title, "status": "FAIL", "uploaded_ids": []}

        post_log_start(base, title)   # khởi buffer cho bài này

        try:
            check_browser_alive(self.driver)

            # Đọc post_type từ dữ liệu bài (anew[13]), mặc định "posts"
            _post_type = (str(anew[13]).strip() if len(anew) > 13 and anew[13] else "posts") or "posts"

            res["step"] = "check_duplicate"
            dup = check_process_duplicate(self.driver, self.wait, base, title, DUPLICATE_MODE,
                                          post_type=_post_type)
            if dup == "403":
                res["status"] = "403_FORBIDDEN"
                post_log_set(final_s="❌ 403 Forbidden", duration=round(time.time()-t0,1))
                post_log_flush(); return res
            if dup:
                res["status"] = "SKIP_DUPLICATE"
                post_log_set(final_s="⏭ Trùng", duration=round(time.time()-t0,1))
                post_log_flush(); return res

            post_log_step("📝 soạn mới")
            # Thêm &classic-editor để force Classic Editor nếu WP có plugin
            # (Gutenberg không có id="publish" → Selenium path fail)
            if _post_type and _post_type != "posts":
                self.driver.get(f"{base}wp-admin/post-new.php?post_type={_post_type}&classic-editor")
            else:
                self.driver.get(f"{base}wp-admin/post-new.php?classic-editor")
            wait_for_page_load(self.driver)

            http_err = detect_http_error(self.driver)
            if http_err:
                code, reason = http_err
                res["status"] = f"HTTP_{code}"
                res["error"]  = f"HTTP {code} {reason} khi mở post-new.php"
                post_log_set(final_s=f"❌ HTTP {code}", duration=round(time.time()-t0,1))
                post_log_flush(); return res

            if not smart_send_keys(self.driver, (By.ID, "title"), title):
                res["status"] = "FAIL_AT_TITLE"
                post_log_set(final_s="❌ Lỗi nhập tiêu đề", duration=round(time.time()-t0,1))
                post_log_flush(); return res

            res["step"] = "content"
            # Truyền URL đầy đủ của bài gốc (không chỉ base domain) để urljoin
            # xử lý đúng path relative kiểu ../images/ hay ./files/
            _source_url = (str(anew[6]).strip() if len(anew) > 6 and anew[6] else "") or base
            ids = wp_insert_content_with_images(self.driver, self.wait, anew[3], _source_url)
            res["uploaded_ids"] = ids
            if ids:
                post_log_step(f"🖼 media({len(ids)})")

            res["step"] = "meta"
            # anew[9] = cat_id từ Excel: "Cha" hoặc "Cha/Con" hoặc "Cha/Con/Cháu"
            cat_path = (anew[9] or "").strip()
            if not cat_path:
                cat_path = (anew[8] or "").strip() if len(anew) > 8 else ""
            select_or_create_category(self.driver, self.wait, cat_path,
                                      cat_cache=self.cat_cache)
            res["category"] = cat_path
            post_log_set(category=cat_path.split('/')[-1] if cat_path else "")
            post_log_step("📂 chuyên mục")

            if len(anew) > 12 and anew[12]:
                wp_dt = date_to_datetime(anew[12], DEFAULT_PUBLISH_HOUR, DEFAULT_PUBLISH_MINUTE)
                set_wp_publish_datetime(self.driver, self.wait, wp_dt)
                post_log_set(date_s=str(wp_dt)[:16])

            # ── Ảnh đại diện ─────────────────────────────────────────────────
            _featured_id  = None
            _thumb_urls   = anew[4] if (len(anew) > 4 and isinstance(anew[4], list)
                                        and anew[4]) else []
            _thumb_url    = _thumb_urls[0] if _thumb_urls else ""
            _thumb_upload_attempted = False

            if _thumb_url and 'no-image' not in _thumb_url.lower():
                _thumb_upload_attempted = True
                _thumb_path = download_file_resource(_thumb_url)
                if _thumb_path:
                    _thumb_att = wp_media_upload_only(self.driver, self.wait, _thumb_path)
                    if Xoa_file_sau_khi_dang:
                        cleanup_temp_file(_thumb_path)
                    if _thumb_att:
                        _featured_id = int(_thumb_att)
                    else:
                        debug_print(f"   {Style.YELLOW}⚠ Selenium: upload thumb thất bại — fallback ảnh content{Style.RESET}")
                else:
                    debug_print(f"   {Style.YELLOW}⚠ Selenium: download thumb thất bại: {_thumb_url[:60]}{Style.RESET}")

            # Fallback: dùng ảnh đầu trong content CHỈ KHI không có thumb_url
            # (không fallback khi có thumb_url nhưng upload fail — tránh ảnh sai)
            if _featured_id is None and not _thumb_upload_attempted and ids:
                _featured_id = ids[0]

            if _featured_id:
                wp_editor_focus_end(self.driver)
                set_featured_image_by_id(self.driver, self.wait, _featured_id)
                res["featured_ok"] = True

            res["step"] = "publish"
            pub_res = click_publish_ensure_visibility(self.driver, self.wait)

            post_http_err = detect_http_error(self.driver)
            if post_http_err and pub_res not in ("SCHEDULED",) and not (
                    isinstance(pub_res, str) and pub_res.startswith("http")):
                code, reason = post_http_err
                res["status"] = f"HTTP_{code}_AFTER_PUBLISH"
                res["error"]  = f"HTTP {code} {reason} sau khi bấm Đăng"
                post_log_set(final_s=f"❌ HTTP {code} sau publish", duration=round(time.time()-t0,1))
                post_log_flush(); return res

            if pub_res and str(pub_res).startswith("http"):
                if self.verify_public_url(pub_res):
                    res["status"] = "SUCCESS"; res["public_url"] = pub_res
                    post_log_set(final_s="✅ Xong")
                else:
                    res["status"] = "PUBLISHED_BUT_NOT_ACCESSIBLE"
                    post_log_set(final_s="⚠ Đăng nhưng không truy cập được")
            else:
                res["status"] = str(pub_res)
                post_log_set(final_s=f"⚠ {pub_res}")

        except Exception as e:
            res["error"] = str(e)
            if "BROWSER_DIED" in str(e) or "WinError 10061" in str(e) or "HTTPConnectionPool" in str(e):
                res["status"] = "BROWSER_CRASHED"
            post_log_set(final_s=f"❌ {str(e)[:40]}")

        duration = round(time.time() - t0, 2)
        res["duration_sec"] = duration
        post_log_set(duration=round(duration, 1))
        post_log_flush()
        return res

    def close(self):
        try: self.driver.quit()
        except: pass

def process_website_batch(site_data):
    """
    Worker xử lý bài cho 1 site.

    Chiến lược 2 tầng:
      1. REST API  — thử trước: nhanh, không cần Chrome, ít RAM
      2. Selenium  — fallback nếu REST không khả dụng

    Hỗ trợ 2 chế độ lấy bài:
      • Static  (DYNAMIC_REFILL=False) — rows là list cố định từ lúc khởi động
      • Dynamic (DYNAMIC_REFILL=True)  — rows là queue.Queue; worker kéo bài liên tục
        cho đến khi queue rỗng VÀ _refill_stop được set (tức RefillThread đã dừng)
    """
    base_url, rows = site_data
    _dynamic = isinstance(rows, _queue.Queue)
    site_start = time.time()

    use_rest = USE_REST_API and _rest_api_works(base_url)
    time.sleep(random.uniform(0.2, 1.0) if use_rest else random.uniform(1.0, 4.0))
    mode_str = "REST API" if use_rest else "Selenium"
    debug_print(f"\n{Style.BOLD}🚀 START [{mode_str}]: {base_url}{Style.RESET}")

    # ── Helper: lấy row tiếp theo từ queue (dynamic) ────────────────────────
    def _next_row_dynamic():
        """
        Kéo 1 row từ queue. Block tối đa 3s mỗi lần thử.
        Trả về None khi queue rỗng VÀ RefillThread đã dừng (không còn bài nào nữa).

        Chống race: RefillThread có thể put() 1 item vào queue ngay trước khi
        set _refill_stop. Để không bỏ sót, sau khi phát hiện _refill_stop đã
        set ta thử get() không chặn thêm 1 lần trước khi trả None.
        """
        while True:
            try:
                return rows.get(timeout=3)
            except _queue.Empty:
                if _refill_stop.is_set():
                    # Thử vét nốt phần tử cuối nếu có (không block)
                    try:
                        return rows.get_nowait()
                    except _queue.Empty:
                        return None   # thực sự hết bài
                # RefillThread vẫn chạy → tiếp tục chờ

    # ════════════════════════════════════════════════════════════════════════
    # REST API path
    # ════════════════════════════════════════════════════════════════════════
    if use_rest:
        session = requests.Session()
        session.auth    = (EMAIL, PASSWORD)
        session.verify  = False
        session.headers.update({"User-Agent": get_random_ua()})
        cat_cache: dict = {}
        idx = 0

        # Tạo iterator thống nhất: list hoặc queue
        row_iter = iter(rows) if not _dynamic else None

        while True:
            # ── Lấy row tiếp theo ──────────────────────────────────────────
            if _dynamic:
                row = _next_row_dynamic()
                if row is None:
                    break           # hết bài
            else:
                try:
                    row = next(row_iter)
                except StopIteration:
                    break

            idx += 1
            should_stop = False
            try:
                res = rest_post_one(list(row), session, cat_cache)
                append_log_row(LOG_XLSX, res)

                if res["status"] in ("SUCCESS", "SCHEDULED"):
                    with db_lock:
                        try:
                            with _suppress_stdout():
                                hp.update_upload_new(row[0])
                        except Exception:
                            pass
                elif res["status"] == "SKIP_DUPLICATE":
                    debug_print(f"{Style.YELLOW}⏭ [REST {base_url}] Bỏ qua trùng — không đánh dấu done (bài #{idx}).{Style.RESET}")
                elif res["status"] == "403_FORBIDDEN":
                    debug_print(f"{Style.RED}⚠ [REST {base_url}] 403 — dừng site.{Style.RESET}")
                    should_stop = True
                elif res["status"].startswith("REST_HTTP_"):
                    debug_print(f"{Style.YELLOW}⚠ [REST]{Style.RESET} {Style.DIM}{res['status']} — bỏ qua bài #{idx}.{Style.RESET}")

            except Exception as e:
                debug_print(f"{Style.RED}❌ [REST {base_url}] row #{idx}: {e}{Style.RESET}")
            finally:
                # Luôn giải phóng claim sau mỗi bài (dù thành công hay lỗi)
                _release_assigned(row[0])

            if should_stop:
                break

        session.close()
        # ── Drain queue khi REST worker thoát sớm (403) ─────────────────────
        if _dynamic and DYNAMIC_REFILL:
            drained = 0
            while True:
                try:
                    leftover = rows.get_nowait()
                    _release_assigned(leftover[0])
                    drained += 1
                except _queue.Empty:
                    break
            if drained:
                debug_print(
                    f"{Style.YELLOW}⚠ [REST {base_url}] "
                    f"Đã release {drained} bài còn trong queue — "
                    f"RefillThread sẽ re-add ở lần quét tiếp.{Style.RESET}"
                )
        site_dur = time.time() - site_start
        debug_print(f"\n🏁 {Style.GREEN}REST DONE: {base_url} (⏱️ {format_duration(site_dur)}){Style.RESET}")
        return

    # ════════════════════════════════════════════════════════════════════════
    # Selenium fallback path
    # ════════════════════════════════════════════════════════════════════════
    worker_bot   = None
    post_retry   = 0
    current_row  = None          # row đang xử lý (None = chưa lấy hoặc đã xong)
    row_idx      = 0             # đếm bài (chỉ để log)
    row_iter     = iter(rows) if not _dynamic else None

    def _fetch_next() -> "tuple | None":
        """Lấy row tiếp theo từ list hoặc queue."""
        nonlocal row_idx
        if _dynamic:
            r = _next_row_dynamic()
        else:
            try:
                r = next(row_iter)
            except StopIteration:
                r = None
        if r is not None:
            row_idx += 1
        return r

    try:
        worker_bot = NewsPoster()

        while True:
            # ── Lấy row mới nếu chưa có ────────────────────────────────────
            if current_row is None:
                current_row = _fetch_next()
                if current_row is None:
                    break          # hết bài

            row = current_row

            # ── Login ───────────────────────────────────────────────────────
            if not worker_bot.ensure_wp_login(base_url, EMAIL, PASSWORD):
                worker_bot.restart_driver()
                if not worker_bot.ensure_wp_login(base_url, EMAIL, PASSWORD):
                    debug_print(f"{Style.RED}⚠️ [Selenium {base_url}] Login thất bại. Bỏ site.{Style.RESET}")
                    _release_assigned(row[0])
                    break

            res = worker_bot.post_one(list(row), row_idx, "?")
            append_log_row(LOG_XLSX, res)

            # ── Crash → retry cùng row ──────────────────────────────────────
            if res["status"] == "BROWSER_CRASHED":
                debug_print(f"{Style.RED}🚨 [Selenium {base_url}] Browser crashed.{Style.RESET}")
                post_retry += 1
                if post_retry > MAX_RETRIES_PER_POST:
                    debug_print(f"{Style.RED}⏭ Bỏ qua bài #{row_idx} (crash {post_retry}x).{Style.RESET}")
                    _release_assigned(row[0])
                    current_row = None
                    post_retry  = 0
                    # Driver đã crash nhiều lần → restart trước khi lấy bài tiếp theo
                    try:
                        worker_bot.restart_driver()
                    except Exception as _re:
                        debug_print(f"{Style.RED}✗ Restart thất bại: {_re} — dừng site.{Style.RESET}")
                        break
                else:
                    # Vẫn còn lượt retry → restart rồi thử lại cùng row
                    try:
                        worker_bot.restart_driver()
                    except Exception as _re:
                        debug_print(f"{Style.RED}✗ Restart thất bại: {_re} — dừng site.{Style.RESET}")
                        _release_assigned(row[0])
                        break
                continue          # thử lại cùng row (nếu current_row chưa cleared)

            # ── Các kết quả không retry ─────────────────────────────────────
            _release_assigned(row[0])
            current_row = None
            post_retry  = 0

            if res["status"] == "403_FORBIDDEN":
                debug_print(f"{Style.RED}⚠️ [Selenium {base_url}] 403. Stop.{Style.RESET}")
                break

            if res["status"].startswith("HTTP_"):
                debug_print(f"{Style.YELLOW}⚠️ {res['status']} — bỏ qua #{row_idx}.{Style.RESET}")
                continue

            if res["status"] in ("SUCCESS", "SCHEDULED"):
                with db_lock:
                    try:
                        with _suppress_stdout():
                            hp.update_upload_new(row[0])
                    except Exception:
                        pass
            elif res["status"] == "SKIP_DUPLICATE":
                debug_print(f"{Style.YELLOW}⏭ [Selenium {base_url}] Bỏ qua trùng — không đánh dấu done.{Style.RESET}")

    except Exception as e:
        debug_print(f"{Style.RED}❌ WORKER ERROR ({base_url}): {e}{Style.RESET}")
        if current_row is not None:
            _release_assigned(current_row[0])
    finally:
        if worker_bot:
            worker_bot.close()
        # ── Drain queue khi worker thoát sớm (403, login fail, exception) ──
        # Các bài còn trong queue không có worker nào kéo nữa.
        # Release claim để RefillThread có thể re-add chúng ở lần quét tiếp theo.
        if _dynamic and DYNAMIC_REFILL:
            drained = 0
            while True:
                try:
                    leftover = rows.get_nowait()
                    _release_assigned(leftover[0])
                    drained += 1
                except _queue.Empty:
                    break
            if drained:
                debug_print(
                    f"{Style.YELLOW}⚠ [Selenium {base_url}] "
                    f"Đã release {drained} bài còn trong queue — "
                    f"RefillThread sẽ re-add ở lần quét tiếp.{Style.RESET}"
                )
        site_dur = time.time() - site_start
        debug_print(f"\n🏁 {Style.GREEN}SELENIUM DONE: {base_url} (⏱️ {format_duration(site_dur)}){Style.RESET}")

def _news_sort_key(row) -> "date":
    """
    Key sắp xếp bài: cũ nhất lên trước.
    row[12] = date_publish; None/thiếu → xếp cuối.
    Dùng được cả trong MainController lẫn _RefillThread.
    """
    d = row[12] if len(row) > 12 else None
    if d is None:
        return date(9999, 12, 31)
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    try:
        return datetime.strptime(str(d).strip(), "%d/%m/%Y").date()
    except Exception:
        return date(9999, 12, 31)


# ══════════════════════════════════════════════════════════════════════════════
# 8b. REFILL THREAD — tự động quét DB và phân phối bài mới vào hàng chờ
# ══════════════════════════════════════════════════════════════════════════════

class _RefillThread(threading.Thread):
    """
    Định kỳ quét DB (hp.read_news) để tìm bài chưa được chiếm bởi luồng nào,
    rồi thêm vào queue của site tương ứng (ưu tiên queue ít hàng chờ hơn).

    Thuật toán phân phối:
      • Bài đã có trong _assigned_ids → bỏ qua (đang được xử lý hoặc đã xếp hàng)
      • Bài có base_url đã có queue   → thêm vào queue đó (workers tự kéo)
      • Bài có base_url chưa có queue → tạo queue mới + spawn worker mới

    Điều kiện dừng:
      • REFILL_EMPTY_STOP lần quét liên tiếp không tìm thấy bài mới
      • Hoặc bị gọi stop() từ bên ngoài

    Khi dừng, set _refill_stop để tất cả workers biết không còn bài mới nào
    và có thể thoát khi queue của mình rỗng.
    """

    def __init__(self, executor: "ThreadPoolExecutor", futures: dict):
        super().__init__(daemon=True, name="RefillThread")
        self._executor     = executor
        self._futures      = futures   # base_url → list[Future] — mutable, dùng chung
        self._stop_req     = threading.Event()
        self._empty_streak = 0

    def stop(self):
        self._stop_req.set()

    def run(self):
        debug_print(
            f"\n{Style.CYAN}🔄 [Refill] Khởi động — quét mỗi {REFILL_INTERVAL}s, "
            f"dừng sau {REFILL_EMPTY_STOP} lần trống.{Style.RESET}"
        )
        while not self._stop_req.is_set():
            # Ngủ, nhưng thức dậy ngay nếu stop() được gọi
            self._stop_req.wait(REFILL_INTERVAL)
            if self._stop_req.is_set():
                break

            added = self._scan_and_distribute()

            if added > 0:
                self._empty_streak = 0
                debug_print(
                    f"{Style.CYAN}🔄 [Refill] +{added} bài mới vào hàng chờ.{Style.RESET}"
                )
            else:
                self._empty_streak += 1
                debug_print(
                    f"{Style.BBLACK}🔄 [Refill] Không có bài mới "
                    f"({self._empty_streak}/{REFILL_EMPTY_STOP}).{Style.RESET}"
                )
                if self._empty_streak >= REFILL_EMPTY_STOP:
                    debug_print(
                        f"{Style.BBLACK}🔄 [Refill] Đạt ngưỡng — dừng refill.{Style.RESET}"
                    )
                    break

        # Báo cho tất cả workers: không còn bài mới, thoát khi queue rỗng
        _refill_stop.set()
        debug_print(f"{Style.BBLACK}🔄 [Refill] Kết thúc.{Style.RESET}")

    # ── Internal ──────────────────────────────────────────────────────────────

    def _scan_and_distribute(self) -> int:
        """
        Đọc DB, tìm bài chưa được chiếm, thêm vào queue site tương ứng.

        Phân phối:
          • Bài thuộc site đã có queue → put() vào queue đó.
            Workers cùng site chia nhau queue chung theo FIFO — load-balance tự nhiên.
          • Bài thuộc site mới → tạo queue + spawn 1 worker mới.
            (Spawn nhiều hơn 1 worker/site mới không hiệu quả vì chỉ có 1 bài lúc này)

        Trả về số bài mới đã thêm vào hàng chờ.
        """
        try:
            with _suppress_stdout():
                all_rows = hp.read_news()
        except Exception as e:
            debug_print(f"{Style.RED}🔄 [Refill] Lỗi đọc DB: {e}{Style.RESET}")
            return 0

        # Snapshot assigned_ids một lần, tránh giữ lock lâu
        with _assigned_lock:
            already = frozenset(_assigned_ids)

        new_rows = [r for r in all_rows if str(r[0]) not in already]
        if not new_rows:
            return 0

        # Sắp xếp: cũ nhất lên trước (nhất quán với lần đọc ban đầu)
        new_rows.sort(key=_news_sort_key)

        added = 0
        for row in new_rows:
            base_url = get_base(str(row[7]) if len(row) > 7 and row[7] else "")
            if not base_url:
                continue

            # Atomic check-and-claim: tránh race giữa hai vòng quét
            with _assigned_lock:
                news_id = str(row[0])
                if news_id in _assigned_ids:
                    continue          # luồng khác vừa claim trước
                _assigned_ids.add(news_id)

            if base_url in _site_queues:
                # Site đã có queue — thêm vào (workers tự kéo theo kiểu FIFO)
                _site_queues[base_url].put(row)
            else:
                # Site mới — tạo queue + spawn worker mới
                q = _queue.Queue()
                q.put(row)
                _site_queues[base_url] = q
                try:
                    f = self._executor.submit(process_website_batch, (base_url, q))
                    self._futures.setdefault(base_url, []).append(f)
                    debug_print(
                        f"{Style.CYAN}🔄 [Refill] Site mới: "
                        f"{base_url.replace('https://','').rstrip('/')} — spawn worker.{Style.RESET}"
                    )
                except RuntimeError:
                    # Executor đã shutdown (Ctrl+C hoặc chương trình đang thoát)
                    # Bỏ claim để không giữ bài mãi trong _assigned_ids
                    with _assigned_lock:
                        _assigned_ids.discard(str(row[0]))
                    del _site_queues[base_url]
                    debug_print(
                        f"{Style.YELLOW}🔄 [Refill] Executor đã đóng — dừng phân phối.{Style.RESET}"
                    )
                    return added   # thoát sớm, không xử lý thêm

            added += 1

        return added


class MainController:
    def run(self):
        if DYNAMIC_REFILL:
            self._run_dynamic()
        else:
            self._run_static()

    # ── Static mode — đọc DB 1 lần, không quét lại ───────────────────────────
    def _run_static(self):
        news = hp.read_news()
        print(f"{Style.BOLD}🚀 TỔNG CỘNG: {len(news)} bài cần xử lý.{Style.RESET}")
        start = time.time()

        news_sorted = sorted(news, key=_news_sort_key)
        count_with_date    = sum(1 for r in news_sorted if len(r) > 12 and r[12])
        count_without_date = len(news_sorted) - count_with_date
        print(f"📅 Thứ tự đăng: cũ nhất trước "
              f"({count_with_date} bài có ngày, {count_without_date} bài không có ngày xếp cuối)")

        site_groups = {}
        for r in news_sorted:
            b = get_base(r[7])
            if b not in site_groups: site_groups[b] = []
            site_groups[b].append(r)

        tasks = []
        print(f"📦 Đang phân phối luồng (Max {MAX_THREADS_PER_SITE} luồng song song cho 1 trường)...")
        for base_url, rows in site_groups.items():
            total_rows = len(rows)
            n_threads = MAX_THREADS_PER_SITE if USE_REST_API else max(1, MAX_THREADS_PER_SITE // 5)
            if total_rows <= n_threads:
                tasks.append((base_url, rows))
            else:
                chunks = [[] for _ in range(n_threads)]
                for i, row in enumerate(rows):
                    chunks[i % n_threads].append(row)
                for chunk in chunks:
                    if chunk:
                        tasks.append((base_url, chunk))

        n_actual = min(len(tasks), MAX_CONCURRENT_WORKERS)
        print(f"⚡ Đã chia thành {len(tasks)} task.")
        print(f"🚀 Khởi chạy {n_actual} workers [{'REST API' if USE_REST_API else 'Selenium'}]...")
        init_dashboard(n_actual)

        _stagger_min = 0.3 if USE_REST_API else 1.5
        _stagger_max = 0.8 if USE_REST_API else 3.0

        with ThreadPoolExecutor(max_workers=n_actual) as executor:
            futures = []
            for i, task in enumerate(tasks):
                futures.append(executor.submit(process_website_batch, task))
                if i < len(tasks) - 1:
                    time.sleep(random.uniform(_stagger_min, _stagger_max))
            for f in futures:
                try: f.result()
                except Exception as e:
                    debug_print(f"{Style.RED}❌ Task lỗi: {e}{Style.RESET}")

        print(f"\n{Style.BOLD}{Style.GREEN}🎉 ĐÃ XONG TOÀN BỘ! (⏱️ {format_duration(time.time() - start)}){Style.RESET}")

    # ── Dynamic mode — workers kéo từ queue, RefillThread thêm bài liên tục ──
    def _run_dynamic(self):
        global _site_queues, _assigned_ids, _refill_stop

        # Reset global state cho mỗi lần chạy
        _site_queues  = {}
        _assigned_ids = set()
        _refill_stop  = threading.Event()

        start = time.time()
        print(f"{Style.BOLD}🚀 [Dynamic] Khởi động với REFILL_INTERVAL={REFILL_INTERVAL}s, "
              f"REFILL_EMPTY_STOP={REFILL_EMPTY_STOP}.{Style.RESET}")

        # ── Đọc lần đầu ──────────────────────────────────────────────────────
        with _suppress_stdout():
            news = hp.read_news()
        news_sorted = sorted(news, key=_news_sort_key)
        print(f"📋 Ban đầu: {len(news_sorted)} bài.")

        # ── Tạo queue riêng cho từng site ────────────────────────────────────
        site_groups: dict = {}
        for r in news_sorted:
            b = get_base(r[7])
            if b not in site_groups: site_groups[b] = []
            site_groups[b].append(r)

        for base_url, rows in site_groups.items():
            q = _queue.Queue()
            for r in rows:
                q.put(r)
                _assigned_ids.add(str(r[0]))
            _site_queues[base_url] = q

        # ── Tính số workers ───────────────────────────────────────────────────
        # Mỗi site: N workers dùng chung 1 queue (load-balance tự nhiên)
        tasks: list = []
        print(f"📦 Phân phối luồng (max {MAX_THREADS_PER_SITE}/site)...")
        for base_url, rows in site_groups.items():
            n_threads = MAX_THREADS_PER_SITE if USE_REST_API else max(1, MAX_THREADS_PER_SITE // 5)
            n_threads = min(n_threads, len(rows))   # không cần nhiều hơn số bài ban đầu
            q = _site_queues[base_url]
            for _ in range(n_threads):
                tasks.append((base_url, q))

        n_actual = min(len(tasks), MAX_CONCURRENT_WORKERS)
        print(f"⚡ {len(tasks)} task — {n_actual} workers đồng thời "
              f"[{'REST API' if USE_REST_API else 'Selenium'}].")
        # Dùng max(1, ...) để init_dashboard không nhận 0 khi DB ban đầu rỗng
        # (trường hợp khởi động trước 1_Lay_bai kịp lưu bài)
        init_dashboard(max(1, n_actual))

        _stagger_min = 0.3 if USE_REST_API else 1.5
        _stagger_max = 0.8 if USE_REST_API else 3.0

        # ── Khởi động executor (KHÔNG dùng context manager để RefillThread
        #    có thể submit thêm worker cho site mới) ──────────────────────────
        # max_workers = MAX_CONCURRENT_WORKERS (workers ban đầu) +
        #               số site tối đa có thể xuất hiện thêm từ RefillThread.
        # Dùng MAX_CONCURRENT_WORKERS làm headroom vì đó là giới hạn song song
        # tổng thể đã được người dùng config — RefillThread không vượt quá đó.
        executor = ThreadPoolExecutor(max_workers=MAX_CONCURRENT_WORKERS * 2)
        futures_dict: dict = {}   # base_url → list[Future]

        for i, (base_url, q) in enumerate(tasks):
            f = executor.submit(process_website_batch, (base_url, q))
            futures_dict.setdefault(base_url, []).append(f)
            if i < len(tasks) - 1:
                time.sleep(random.uniform(_stagger_min, _stagger_max))

        # ── Khởi động RefillThread ───────────────────────────────────────────
        refill = _RefillThread(executor, futures_dict)
        refill.start()

        # ── Chờ RefillThread kết thúc (nó sẽ set _refill_stop) ──────────────
        refill.join()

        # ── Chờ tất cả workers (kể cả worker mới do RefillThread spawn) ─────
        # futures_dict đã được RefillThread bổ sung — thu thập sau join() là an toàn
        all_futures = [f for flist in futures_dict.values() for f in flist]
        for f in all_futures:
            try: f.result()
            except Exception as e:
                debug_print(f"{Style.RED}❌ Task lỗi: {e}{Style.RESET}")

        executor.shutdown(wait=True)
        print(f"\n{Style.BOLD}{Style.GREEN}🎉 ĐÃ XONG TOÀN BỘ! (⏱️ {format_duration(time.time() - start)}){Style.RESET}")

if __name__ == "__main__":
    app = MainController()
    app.run()