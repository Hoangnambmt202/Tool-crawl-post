import requests
import Objectlink as obl
import MenuLink as menulink
from bs4 import BeautifulSoup
import CameraObject as camob
# Function to fetch and parse the webpage
import json
import mysql.connector
import time
import helpers as hp
from difflib import SequenceMatcher
from urllib.parse import urljoin
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from datetime import date
from openpyxl import load_workbook
import time
import re
from urllib.parse import urlparse
FROM_DATE = date(2025, 7, 1)   # chỉ lấy bài từ 01/07/2025 trở về đây
DETAIL_DATE_CSS = "span.post-date"         # ngày bài viết

CHROME_PATH = r"D:\WORKSPACE_CODE\Projects\Web\folder\chrome-win64\chrome-win64\chrome.exe"
CHROME_DRIVER_PATH = r"D:\WORKSPACE_CODE\Projects\Web\folder\chromedriver-win64\chromedriver-win64\chromedriver.exe"

def clean_spaces(s: str) -> str:
    return " ".join((s or "").split()).strip()


def parse_date_from_meta(soup: BeautifulSoup) -> date | None:
    """
    Fallback: lấy ngày từ meta.
    Ưu tiên itemprop=dateCreated (ví dụ: 2025-10-26T06:34:32+0700).
    Hỗ trợ thêm một số meta phổ biến khác.
    """
    if not soup:
        return None

    candidates = []

    # 1) itemprop
    for itemprop in ["dateCreated", "datePublished", "dateModified"]:
        tag = soup.find("meta", attrs={"itemprop": itemprop})
        if tag and tag.get("content"):
            candidates.append(tag["content"])

    # 2) Open Graph / Article
    for prop in ["article:published_time", "article:modified_time"]:
        tag = soup.find("meta", attrs={"property": prop})
        if tag and tag.get("content"):
            candidates.append(tag["content"])

    # 3) name=
    for name in ["pubdate", "publishdate", "date", "dc.date", "dc.date.issued", "dcterms.created", "dcterms.modified"]:
        tag = soup.find("meta", attrs={"name": name})
        if tag and tag.get("content"):
            candidates.append(tag["content"])

    def _to_date(s: str) -> date | None:
        s = (s or "").strip()
        if not s:
            return None

        # Bắt nhanh YYYY-MM-DD
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
        if not m:
            return None
        y, mo, d = map(int, m.groups())
        try:
            return date(y, mo, d)
        except:
            return None

    for s in candidates:
        d = _to_date(s)
        if d:
            return d

    return None

 
 

def parse_public_date_from_uicongkhai(soup: BeautifulSoup) -> date | None:
    """
    Tìm trong table/tr:
      <td>Ngày công bố</td><td>13/10/2025</td>
    """
    root = soup.find("div", class_="UICongKhaiNganSach_Default")
    if not root:
        return None

    for tr in root.find_all("tr"):
        tds = tr.find_all("td")
        if len(tds) >= 2:
            label = clean_spaces(tds[0].get_text(" ", strip=True)).lower()
            if "ngày công bố" in label:
                return parse_vn_date_any(tds[1].get_text(" ", strip=True))

    # fallback: quét toàn root nếu cần
    return None

def parse_vn_date_from_soup(soup: BeautifulSoup) -> date | None:
 
    d = parse_vn_date_from_soup1(soup)                 # ngày hiển thị (nếu có)
    if d is None:
        d = parse_issue_date_from_module34(soup)      # ngày ban hành (nếu có)
    if d is None:
        d = parse_public_date_from_uicongkhai(soup)   # ngày công bố (nếu có)
    if d is None:
        d = parse_date_from_meta(soup)
    return d

def parse_vn_date_from_soup1(soup: BeautifulSoup) -> date | None:
    """
    Parse ngày kiểu: "Thứ hai, 24/11/2025 | 07:36"
    """
    el = soup.find('span',class_='post-date left')
    if not el:
        el = soup.find('div',class_='PostDate')
    # print('el')
    # print(el)
    if not el:
        return None
    text = clean_spaces(el.get_text(" ", strip=True))
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if not m:
        return None
    d, mo, y = map(int, m.groups())
    try:
        return date(y, mo, d)
    except:
        return None
    
def to_int(text, default=None):
    s = (text or "").strip()
    m = re.search(r"\d+", s)          # lấy cụm số đầu tiên
    if not m:
        return default
    return int(m.group())

def get_base(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    p = urlparse(url)
    if not p.scheme or not p.netloc:
        # nếu Excel có URL thiếu scheme, ví dụ: thtienhiep.ninhbinh.edu.vn/...
        p = urlparse("https://" + url)
    return f"{p.scheme}://{p.netloc}/"
def clean_spaces(s: str) -> str:
    return " ".join((s or "").split()).strip()

def parse_vn_date_any(text: str) -> date | None:
    """
    Bắt ngày kiểu dd/mm/yyyy trong một chuỗi bất kỳ.
    """
    text = clean_spaces(text)
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if not m:
        return None
    d, mo, y = map(int, m.groups())
    try:
        return date(y, mo, d)
    except:
        return None

def normalize_download_links_in_content(tag_content):
    print('tag_content', tag_content)
    """
    Chuẩn hoá link tải trong nội dung:

    1) Nếu <a onclick="downloadFile('id','/upload/...')">:
       -> chuyển thành <a href="/upload/..." class="link-download" target="_blank">

    2) Nếu <a href="https://..."> hoặc <a href="/khac/...">:
       -> vẫn giữ href và cũng thêm class link-download + target="_blank"

    Không phụ thuộc đuôi file.
    """
    if not tag_content:
        return

    for a in tag_content.find_all("a"):
        href = (a.get("href") or "").strip()
        onclick = (a.get("onclick") or "").strip()

        file_path = None

        # 1) Ưu tiên bắt từ onclick downloadFile(...)
        if onclick:
            m = re.search(
                r"downloadFile\s*\(\s*(['\"]).*?\1\s*,\s*(['\"])\s*([^'\"\)\s]+)\s*\2\s*\)",
                onclick,
                flags=re.IGNORECASE
            )
            if m:
                file_path = m.group(3).strip()

        # 2) Fallback: lấy từ href (kể cả không phải /upload/)
        if not file_path and href:
            # bỏ các href không hợp lệ/không cần xử lý
            if href.lower().startswith(("javascript:", "#", "mailto:", "tel:")):
                continue
            file_path = href
        print('file_path', file_path)
        # Không có link thì bỏ qua
        if not file_path:
            continue

        # set về chuẩn a href
        a.attrs.pop("onclick", None)
        a["href"] = file_path
        a["target"] = "_blank"

        # class: giữ class cũ + thêm link-download
        cls = a.get("class", [])
        if isinstance(cls, str):
            cls = cls.split()
        if "link-download" not in cls:
            cls.append("link-download")
        a["class"] = cls

        # tuỳ chọn: bỏ style nếu có
        a.attrs.pop("style", None)

def pick_detail_links(soup):
    links = []

    # UI cũ: td.tg-yw4l a
    links += soup.select("td.tg-yw4l a[href]")
    
    # UI mới: td a[title*='Xem chi tiết công khai']
    links += soup.select("td a[title*='Xem chi tiết công khai'][href]")
    print('links')
    # print(links)
    # (tuỳ chọn) nếu title có thể khác hoa/thường
    # BeautifulSoup CSS selector không hỗ trợ case-insensitive chuẩn,
    # nên bạn có thể lọc thêm bằng Python bên dưới.

    return links

def parse_issue_date_from_module34(soup: BeautifulSoup) -> date | None:
    """
    Tìm 'Ngày ban hành' trong bảng thuộc #module34.
    Mẫu hay gặp:
      <th>Ngày ban hành</th><td>27/11/2025</td>
    hoặc có thể nằm trong <td class="td-title">Ngày ban hành</td>
    """
    module = soup.find("div", id="module34")
    if not module:
        return None

    # ưu tiên th (hoặc td) có text "Ngày ban hành"
    label_cells = module.find_all(["th", "td"], string=True)
    for cell in label_cells:
        label = clean_spaces(cell.get_text(" ", strip=True)).lower()
        print(label)
        if "ngày ban hành" in label:
            # lấy ô kế tiếp: thường là <td>...</td>
            nxt = cell.find_next(["td", "th"])
            if nxt:
                d = parse_vn_date_any(nxt.get_text(" ", strip=True))
                if d:
                    return d

            # fallback: nếu cùng hàng <tr> có nhiều ô, tìm ô cuối
            tr = cell.find_parent("tr")
            if tr:
                d = parse_vn_date_any(tr.get_text(" ", strip=True))
                if d:
                    return d

    # fallback cuối: quét toàn module nếu có dd/mm/yyyy
    return parse_vn_date_any(module.get_text(" ", strip=True))

class VspProducts:
    def __init__(self, base, url, cat,target):
        self.base = base
        self.cat = cat
        self.target = target
        self.camlist = camob.Listcam()
        self.url_links = []
        self.menu_links = menulink.MenuLink()

        service = Service(CHROME_DRIVER_PATH)
        chrome_options = Options()
        chrome_options.binary_location = CHROME_PATH
        self.driver = webdriver.Chrome(service=service, options=chrome_options)

        # các biến sẽ được set theo từng dòng excel
        self.url = url
        self.cat_id = ""
        self.url_id = ""
        self.page_link = 1
    def reset_for_row(self):
        # reset queue + menu cho mỗi dòng excel
        self.url_links = []
        self.menu_links = menulink.MenuLink()
        # nếu muốn reset danh sách cam cho mỗi dòng:
        self.camlist = camob.Listcam()

    def get_data(self):
        self.url_links.append(self.url)
        while self.url_links:
            self.extract_data(self.url_links[0])
            self.url_links.pop(0)

    def close(self):
        try:
            self.driver.quit()
        except:
            pass
     
    # case list ArticleList
    def parse_list_ArticleList(self, soup) -> bool:
        ul = soup.find("ul", class_="ArticleList")
        if not ul:
            return False

        items = ul.find_all("li", class_="row")
        if not items:
            return False

        print(f"--> Phát hiện cấu trúc: ArticleList | {len(items)} bài")

        for li in items:
            h2 = li.find("h2", class_="Title")
            if not h2:
                continue

            a = h2.find("a", href=True)
            if not a:
                continue

            name = clean_spaces(a.get_text())
            url  = urljoin(self.base, a["href"])

            if not name or not url:
                continue

            if hp.check_cam_url(self.url_id, url, name):
                if self.camlist.add_cam(
                    camob.CameraObject(0, name, 0, url, "", self.cat_id)
                ):
                    self.url_links.append(url)

        return True

    # case detail ArticleDetailControl
    def parse_detail_ArticleDetailControl(self, soup, o_url) -> bool:
        root = soup.find("div", class_="ArticleDetailControl")
        if not root:
            return False

        print("-> Cấu trúc: ArticleDetailControl (VNPT)")

        tag_content = root.find("div", class_="ArticleContent")
        if not tag_content:
            print("❌ Không tìm thấy ArticleContent")
            return False

        # ngày đăng
        d = parse_vn_date_from_soup(soup)

        normalize_download_links_in_content(tag_content)

        # chuẩn hoá img
        for img in tag_content.find_all("img"):
            src = img.get("src")
            img.attrs = {}
            if src:
                img["src"] = urljoin(self.base, src)

        # xoá rác
        for t in tag_content.find_all(["script", "style", "iframe", "link"]):
            t.decompose()

        for cam in self.camlist.camobs:
            if cam.url == o_url:
                cam.date_publish = d
                cam.short = self.target
                cam.description = str(tag_content)
                if "Công khai" in cam.name:
                    cam.cat_id = "Công khai"

                cam.display_info()
                hp.save_data_cam(self.url_id, cam)
                return True

        return False
       
    def extract_data(self, o_url ):
        time.sleep(1)
        # print(o_url)
        chitiet = 0
        for cam in self.camlist.camobs:
            if cam.url == o_url:
                print('chuan bi doc thong tin san pham')
                chitiet = 1
        try:
            # soup = hp.fetch_webpage( o_url)
            self.driver.get(o_url)
            # Lấy nội dung HTML sau khi JavaScript đã tải xong
            html = self.driver.page_source
            # Phân tích nội dung với BeautifulSoup
            soup = BeautifulSoup(html, 'html.parser')
        except Exception as e:
            print("Error:", e)
            return

        if soup:
            # read module product
            while True:
                # self.driver.get(o_url)
                # Lấy nội dung HTML sau khi JavaScript đã tải xong
                # html = self.driver.page_source
                # Phân tích nội dung với BeautifulSoup
                soup = BeautifulSoup(html, 'html.parser')
                print('dang parse' +  o_url)
                time.sleep(1)
                div_pros = soup.find_all('div', class_="post-item row")
                div_pros = soup.find_all('div', class_=['post-item', 'row'])
                if not div_pros:
                     div_pros = soup.find_all('div', class_=[ 'entry-content' ])
               
                print('doc lai div_pros')
                if(len(div_pros) > 0 and chitiet ==0):
                    for tag_div in div_pros:
                        
                    
                        h3 = tag_div.find('h3' )
                        if not h3:
                            h3 = tag_div.select_one('.entry-title')
                       
                        if h3:
                            tag_a = h3.find('a' )
                            price = 0
                            url = ""
                            name =""
                            photo = ""
                            if tag_a:
                                name=tag_a.text.strip()
                                url =self.base + tag_a.get('href')
                            

                            check = hp.check_cam_url(self.url_id,url,name)
                            # print(check, url, name)
                            if url != '' and name != '' and   check :
                                print("nhap cam")
                                if self.camlist.add_cam(camob.CameraObject(0,name,price, url,photo,self.cat_id)):
                                    self.url_links.append(url)

                tag_title_product = soup.find('h1', class_="title-detail")
                

                print(tag_title_product)
                if chitiet and tag_title_product  :
                    print('chi tiet san pham')
                    tag_content =  soup.find('div',class_='content-detail font-size-text mb-20')
                    tag_divcontent = soup.find('div',class_='article-content')
                    print('tag_divcontent')
                    if tag_divcontent :
                        tag_left = soup.find('div',class_='pull-left mt-5')
                    
                    d = parse_vn_date_from_soup(soup)
                    print('****************date')
                    print(d)
                

                    if  d is not None :
                        print(d < FROM_DATE)
                        # if d < FROM_DATE:
                        #     print(f"Bài {d} < {FROM_DATE} => bỏ qua")
                        #     return

                    photo = ""
                    short = ""
                    if tag_content:

                        for cam in self.camlist.camobs:
                            if cam.url == o_url:
                                cam.date_publish = d
                                cam.short = self.target
                                for tag in tag_content.find_all(True):
                                        if tag.has_attr('style'):
                                            del tag['style']

                                for img_tag in tag_content.find_all('img'):
                                    if 'src' in img_tag.attrs:  # Kiểm tra nếu có thuộc tính 'src'
                                        src_value = img_tag['src']  # Lấy giá trị thuộc tính 'src'
                                        img_tag.attrs = {'src': src_value}  # Chỉ giữ lại thuộc tính 'src'
                                    else:
                                        img_tag.attrs = {}  # Xóa tất cả thuộc tính nếu không có 'src'
                                for tag in tag_content.find_all(True):
                                    if tag.name == 'img':
                                        src = tag.get('src')
                                        tag.attrs = {}
                                        if src:
                                            tag['src'] = src

                                    elif tag.name == 'a':
                                        href = tag.get('href')
                                        cls  = tag.get('class')
                                        target = tag.get('target')

                                        tag.attrs = {}
                                        if href:
                                            tag['href'] = href
                                        if cls:
                                            tag['class'] = cls
                                        if target:
                                            tag['target'] = target

                                    else:
                                        tag.attrs = {}     
                              
                                tags_to_removes = tag_content.find_all ('script' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                tags_to_removes = tag_content.find_all ('link' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                cam.description = str(tag_content)
                                if tag_left:
                                    tag_a_file = tag_left.find('a')
                                    cam.description =  cam.description + '<br/>'+ str(tag_a_file)
                                if 'Công khai' in cam.name:
                                    cam.cat_id = 'Công khai'
                                cam.display_info()
                                hp.save_data_cam(self.url_id,cam)     #dahua 3 #hk 4 #tiandy 5 #ezviz 6 #vsp là 7  #imou 8  #kb 9
                   
                if  chitiet ==0:
                    if self.parse_list_ArticleList(soup):
                        pass
                    tag_ul = soup.find('ul', class_='ArticleList')
                    if not tag_ul:
                        tag_ul = soup.find('ul', class_='down-list')
                    if not tag_ul:
                        tag_ul = soup.find('ul',class_='phal-list row')
                    if tag_ul:
                        tag_ils= tag_ul.find_all('li',class_='row')
                        if not tag_ils:
                            tag_ils= tag_ul.find_all('li')
                        for tag_il in tag_ils:
                            h2 = tag_il.find('h2', class_='Title')
                            if not h2:
                                h2 = tag_il.find('h5' )
                            if not h2:
                                h2 = tag_il.find('h4' )
                            if not h2:
                                h2 =   tag_il.find('div', class_='avatar')
                            if h2:
                                
                                price = 0
                                url = ""
                                name =""
                                photo = ""
                                a3 = h2.find('a')
                                if a3:
                                    name=a3.text.strip()
                                    url =self.base + a3.get('href')
                                    # print('check')
                                    # print(url)
                                    print(name)
                                    # print(self.url_id)
                                    check = hp.check_cam_url(self.url_id,url,name)
                                
                                    if url != '' and name != '' and   check :
                                        print("nhap cam")
                                        if self.camlist.add_cam(camob.CameraObject(0,name,price, url,photo,self.cat_id)):
                                            self.url_links.append(url)
                            
                
                
                    # # truonghop3
                
                    tag_divs= soup.find_all('div',class_='post-title')

                    if not tag_divs:
                        tag_divs= soup.find_all('div',class_='title-news-listType10')
                    if not tag_divs:
                        tag_divs= soup.find_all('div',class_='item-info')
                    # print('tag_divs')
                    # print(tag_divs)
                    for tag_il in tag_divs:
                        h4 = tag_il.find('h4', class_='entry-title' )
                        if not h4:
                            h4 = tag_il.find('h2')
                        if not h4:
                            h4 = tag_il.find('h5')
                        if h4:
                            tag_a = h4.find("a")
                            price = 0
                            url = ""
                            name =""
                            photo = ""
                            if tag_a:
                                name=tag_a.text.strip()
                                url =self.base + tag_a.get('href')
                        

                            check = hp.check_cam_url(self.url_id,url,name)
                            if url != '' and name != '' and   check :
                                print("nhap cam")
                                if self.camlist.add_cam(camob.CameraObject(0,name,price, url,photo,self.cat_id)):
                                    self.url_links.append(url)
                    
                    tag_divs= soup.find_all('a',class_='title-documment')
                    for tag_a in tag_divs:
                        price = 0
                        url = ""
                        name =""
                        photo = ""
                        if tag_a:
                            name=tag_a.text.strip()
                            url =self.base + tag_a.get('href')

                        check = hp.check_cam_url(self.url_id,url,name)
                        if url != '' and name != '' and   check :
                            print("nhap cam")
                            if self.camlist.add_cam(camob.CameraObject(0,name,price, url,photo,self.cat_id)):
                                self.url_links.append(url)
                    


                tag_title_product = soup.find('div', class_="ArticleHeader")
            
                if chitiet and tag_title_product  :
                    print('chi tiet san pham')
                    tag_content =  soup.find('div',class_='ArticleContent')
                    d = parse_vn_date_from_soup(soup)
                    print('****************date')
                    print(d)
                

                    if  d is not None :
                        print(d < FROM_DATE)
                    photo = ""
                    short = ""
                    if tag_content:
                        for cam in self.camlist.camobs:
                            if cam.url == o_url:
                                cam.date_publish = d
                                cam.short = self.target
                                print('normalize')
                                normalize_download_links_in_content(tag_content)
                                for tag in tag_content.find_all(True):
                                        if tag.has_attr('style'):
                                            del tag['style']

                                for img_tag in tag_content.find_all('img'):
                                    if 'src' in img_tag.attrs:  # Kiểm tra nếu có thuộc tính 'src'
                                        src_value = img_tag['src']  # Lấy giá trị thuộc tính 'src'
                                        img_tag.attrs = {'src': src_value}  # Chỉ giữ lại thuộc tính 'src'
                                    else:
                                        img_tag.attrs = {}  # Xóa tất cả thuộc tính nếu không có 'src'
                                for tag in tag_content.find_all(True):
                                    if tag.name == 'img':
                                        src = tag.get('src')
                                        tag.attrs = {}
                                        if src:
                                            tag['src'] = src

                                    elif tag.name == 'a':
                                        href = tag.get('href')
                                        cls  = tag.get('class')
                                        target = tag.get('target')

                                        tag.attrs = {}
                                        if href:
                                            tag['href'] = href
                                        if cls:
                                            tag['class'] = cls
                                        if target:
                                            tag['target'] = target

                                    else:
                                        tag.attrs = {}     
                                tags_to_removes = tag_content.find_all ('script' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                tags_to_removes = tag_content.find_all ('iframe' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                tags_to_removes = tag_content.find_all ('link' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                cam.description = str(tag_content)
                                if 'Công khai' in cam.name:
                                    cam.cat_id = 'Công khai'
                                cam.display_info()
                                hp.save_data_cam(self.url_id,cam) 
                                return
            
                for tag_a in pick_detail_links(soup):
                    name = tag_a.get_text(" ", strip=True)
                    # print('name')
                    # print(name)
                    href = (tag_a.get("href") or "").strip()
                    if not href:
                        continue

                    url = urljoin(self.base, href)   # ✅ dùng urljoin thay vì self.base + href

                    check = hp.check_cam_url(self.url_id, url, name)
                   
                    if url and name and check:
                        print('addcame')
                                                 
                        if self.camlist.add_cam(camob.CameraObject(0, name, 0, url, "", self.cat_id)):
                            self.url_links.append(url)
                print('kt article-content')
                if not tag_title_product:
                    tag_title_product = soup.find('div',class_='article-content')
                    print('article-content')
                if not tag_title_product:
                    tag_title_product = soup.find('table', class_="table")
                    print('table')
                if not tag_title_product:
                    tag_title_product = soup.find('article',class_='download-detail-layout-default')
                    print('download-detail-layout-default')
                    
                if not tag_title_product:
                    tag_title_product = soup.find('article',class_='Article-News')
                    print('article News')
                if not tag_title_product:
                    tag_title_product = soup.find('section', class_="news-detail-default")
                    print('section news-detail-default')
                if not tag_title_product:
                    tag_title_product = soup.find('div', id="module34")
                    print('tim module34')
               
                if not tag_title_product:
                    tag_title_product = soup.find('div', id="module16")
                    print('tim module16')
                
                if chitiet and tag_title_product  :
                    print('chi tiet san pham module 34')
                    tag_content =  tag_title_product
                    # 1) thử ngày đăng bình thường (post-date)
                    d = parse_vn_date_from_soup(soup)
                    print('****************date')
                    print(d)
                    photo = ""
                    short = ""
                    if tag_content:
                        print('trong tag_content')
                        print(o_url)
                        for cam in self.camlist.camobs:
                            if cam.url == o_url:
                                # print(tag_content)
                                if 'Công khai' in cam.name:
                                    cam.cat_id = 'Công khai'
                                if  d is not None :
                                    print(d < FROM_DATE)
                                    # if d < FROM_DATE and cam.cat_id != 'Công khai':
                                    #     print(f"Bài {d} < {FROM_DATE} => bỏ qua")
                                    #     break
                                cam.date_publish = d
                                print('trong for')
                                cam.short = self.target
                                # print('normalize2')
                                print(tag_content)
                                normalize_download_links_in_content(tag_content)
                                print(tag_content)
                                for tag in tag_content.find_all(True):
                                        if tag.has_attr('style'):
                                            del tag['style']

                                for img_tag in tag_content.find_all('img'):
                                    if 'src' in img_tag.attrs:  # Kiểm tra nếu có thuộc tính 'src'
                                        src_value = img_tag['src']  # Lấy giá trị thuộc tính 'src'
                                        img_tag.attrs = {'src': src_value}  # Chỉ giữ lại thuộc tính 'src'
                                    else:
                                        img_tag.attrs = {}  # Xóa tất cả thuộc tính nếu không có 'src'
                                for tag in tag_content.find_all(True):
                                    if tag.name == 'img':
                                        src = tag.get('src')
                                        tag.attrs = {}
                                        if src:
                                            tag['src'] = src

                                    elif tag.name == 'a':
                                        href = tag.get('href')
                                        cls  = tag.get('class')
                                        target = tag.get('target')

                                        tag.attrs = {}
                                        if href:
                                            tag['href'] = href
                                        if cls:
                                            tag['class'] = cls
                                        if target:
                                            tag['target'] = target

                                    else:
                                        tag.attrs = {}          
                                tags_to_removes = tag_content.find_all ('script' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                tags_to_removes = tag_content.find_all ('style' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                tags_to_removes = tag_content.find_all ('iframe' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                tags_to_removes = tag_content.find_all ('link' )
                                for tags_to_remove in tags_to_removes:
                                    tags_to_remove.decompose()
                                cam.description = str(tag_content)
                                print('trong cong khai')
                                if 'Công khai' in cam.name:
                                    cam.cat_id = 'Công khai'
                                cam.display_info()
                                hp.save_data_cam(self.url_id,cam)
                                return
                if chitiet:
                    if self.parse_detail_ArticleDetailControl(soup, o_url):
                        return

                if chitiet == 1:
                    time.sleep(2)
                    return
                btns = self.driver.find_elements(By.CSS_SELECTOR, 'div.col-center a.next')   
                # print(btns)
                # Không còn nút → dừng
                if   btns:
                    b = btns[0]
                    self.driver.execute_script(
                        "arguments[0].scrollIntoView({block:'center'});", b
                    )
                    time.sleep(0.2)
                    self.driver.execute_script("arguments[0].click();", b)
                    time.sleep(2.5)
                if len (btns) == 0:
                    btns = self.driver.find_elements(By.CSS_SELECTOR, 'a.page-link')
                    print('do dai btns')
                    bam = 0
                   
                    for btn in btns:
                        btn_page  = to_int(btn.text, default=-1)
                        cur_page  = int(self.page_link)  # nếu self.page_link đã là số, bỏ int cũng được
                        # print(btn_page)
                        # print(cur_page)
                        if btn_page > cur_page:
                            self.driver.execute_script(
                                "arguments[0].scrollIntoView({block:'center'});", btn
                            )
                            time.sleep(0.2)
                            self.driver.execute_script("arguments[0].click();", btn)
                            time.sleep(1.5)
                            print('bấm rồi')
                            bam = 1
                            self.page_link =  self.page_link + 1  
                            break
                    if bam == 0 or len (btns) == 0:
                        print('ket thuc')
                        break
                    
                      
                    


def run_from_excel(
    excel_path: str,
    sheet_name: str = None,
    start_row: int = 2,
    sleep_between: float = 0.5,
    save_every: int = 1,          # lưu sau mỗi N dòng xử lý
):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # đọc header để map cột theo tên
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v:
            headers[str(v).strip().lower()] = col

    required = ["idurl", "source", "target", "cat_id"]
    for k in required:
        if k not in headers:
            raise ValueError(f"Thiếu cột '{k}' trong Excel. Hiện có: {list(headers.keys())}")

    # cột done (bắt buộc theo yêu cầu mới)
    if "done" not in headers:
        raise ValueError("Thiếu cột 'done' trong Excel (header phải là 'done').")

    done_col = headers["done"]

    bot = None
    processed_since_save = 0

    try:
        for r in range(start_row, ws.max_row + 1):
            idurl  = ws.cell(r, headers["idurl"]).value
            source = ws.cell(r, headers["source"]).value
            target = ws.cell(r, headers["target"]).value
            cat_id = ws.cell(r, headers["cat_id"]).value
            done_v = ws.cell(r, done_col).value

            # nếu không có source thì bỏ qua
            if not source:
                continue

            # chỉ chạy khi done == 0 hoặc trống
            # (chấp nhận "0", 0, "", None)
            done_str = str(done_v).strip() if done_v is not None else ""
            if done_str == "1":
                # đã làm rồi
                continue

            idurl  = str(idurl).strip() if idurl is not None else ""
            source = str(source).strip()
            target = str(target).strip() if target else ""
            cat_id = str(cat_id).strip() if cat_id else ""

            base = get_base(source)

            if bot is None:
                bot = VspProducts(base=base, url=source, cat=cat_id, target=target)

            bot.reset_for_row()
            bot.url    = source
            bot.url_id = idurl
            bot.cat_id = cat_id
            bot.base   = base
            bot.target   = target
            print(f"\n=== ROW {r} | idurl={idurl} | cat_id={cat_id}")
            print(f"source={source}")
            print(f"target={target}")

            try:
                bot.get_data()

                # nếu chạy xong không lỗi -> đánh done = 1
                ws.cell(r, done_col).value = 1

            except Exception as e:
                # nếu lỗi thì KHÔNG đánh done = 1, để lần sau chạy lại
                print(f"[ERROR] ROW {r} => {e}")

            processed_since_save += 1
            if processed_since_save >= save_every:
                wb.save(excel_path)
                processed_since_save = 0

            time.sleep(sleep_between)
            # break

    finally:
        # lưu nốt nếu còn thay đổi chưa save
        try:
            wb.save(excel_path)
        except:
            pass

        if bot:
            bot.close()



# chạy
run_from_excel(r"D:\WORKSPACE_CODE\Projects\Web\folder\websiteninhbinh\danhsachweb - Copy.xlsx", sheet_name=None)
