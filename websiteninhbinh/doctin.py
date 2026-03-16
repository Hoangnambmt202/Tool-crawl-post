import re
import time
from datetime import date
from urllib.parse import urlparse, urljoin

from openpyxl import load_workbook
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException

import MenuLink as menulink
import CameraObject as camob
import helpers as hp
from selenium.webdriver.support.ui import WebDriverWait
import hashlib


# =========================
# CONFIG CHUNG
# =========================
FROM_DATE = date(2025, 7, 1)   # chỉ lấy bài từ 01/07/2025 trở về đây
SLEEP_BETWEEN_ROW = 0.5

# Selector gợi ý (bạn chỉnh theo site):
LIST_ITEM_CSS = "div.post-item.row"       # item trên trang danh sách
LIST_LINK_IN_ITEM_CSS = "h3 a"            # link bài nằm trong item
DETAIL_TITLE_CSS = "h1.title-detail"      # tiêu đề trang chi tiết
DETAIL_CONTENT_CSS = "div.content-detail" # nội dung chi tiết
DETAIL_DATE_CSS = "span.post-date"        # ngày bài viết
PAGER_LINKS_CSS = "div.default-pagination a"  # phân trang JS
LOAD_MORE_BTN_CSS = "next"              # nút load more (nếu có)
MAX_LOAD_MORE_CLICK = 50
import re
from datetime import date
from bs4 import BeautifulSoup

FROM_DATE = date(2025, 7, 1)

CHROME_DRIVER_PATH = r"D:\WORKSPACE_CODE\Projects\Web\folder\chromedriver-win64\chromedriver-win64\chromedriver.exe"
CHROME_BINARY_PATH    = r"D:\WORKSPACE_CODE\Projects\Web\folder\chrome-win64\chrome-win64\chrome.exe"
def get_base(url: str) -> str:
    url = (url or "").strip()
    if not url:
        return ""
    p = urlparse(url)
    if not p.scheme or not p.netloc:
        p = urlparse("https://" + url)
    return f"{p.scheme}://{p.netloc}/"


def clean_spaces(s: str) -> str:
    return " ".join((s or "").split()).strip()


def parse_vn_date_from_soup(soup: BeautifulSoup) -> date | None:
    """
    Parse ngày kiểu: "Thứ hai, 24/11/2025 | 07:36"
    """
    el = soup.select_one(DETAIL_DATE_CSS)
    if not el:
        return None
    text = clean_spaces(el.get_text(" ", strip=True))
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", text)
    if not m:
        return None
    d, mo, y = map(int, m.groups())
    try:
        return date(y, mo, d)
    except:
        return None


def normalize_html_keep_img_src_only(tag_content):
    """
    Bạn đang làm sạch attribute: xóa style/class/href... và chỉ giữ img src.
    """
    if not tag_content:
        return ""

    # xóa style
    for tag in tag_content.find_all(True):
        if tag.has_attr("style"):
            del tag["style"]

    # xóa class
    for tag in tag_content.find_all(True):
        if tag.has_attr("class"):
            del tag["class"]

    # xóa href để tránh link ngoài
    for a in tag_content.find_all("a"):
        if a.has_attr("href"):
            del a["href"]

    # chỉ giữ img src (ưu tiên src; nếu site dùng data-src thì chuyển sang src)
    for img in tag_content.find_all("img"):
        src = img.get("src") or img.get("data-src")
        img.attrs = {}
        if src:
            img["src"] = src

    # xóa script/link
    for bad in tag_content.find_all(["script", "link"]):
        bad.decompose()

    # xóa tất cả attribute còn lại (trừ img src đã xử lý)
    for tag in tag_content.find_all(True):
        if tag.name == "img":
            continue
        tag.attrs = {}

    return str(tag_content)


def click_and_wait_list_change(driver, list_css, click_element, timeout=15):
    """
    Click 1 nút phân trang và chờ list thay đổi (staleness).
    """
    wait = WebDriverWait(driver, timeout)
    old_items = driver.find_elements(By.CSS_SELECTOR, list_css)
    old_first = old_items[0] if old_items else None

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", click_element)
    time.sleep(0.2)
    driver.execute_script("arguments[0].click();", click_element)

    try:
        if old_first:
            wait.until(EC.staleness_of(old_first))
        else:
            time.sleep(1.5)
    except TimeoutException:
        time.sleep(2)


def click_load_more_until_end(driver, button_css=LOAD_MORE_BTN_CSS):
    """
    Click nút load-more MỘT LẦN.
    Trả về:
      > 0  : vẫn còn nút (đã click)
      = 0  : không còn nút → nên dừng
    """
    nav = driver.find_elements(By.ID,'pagination25')
    print(nav)
    btns = driver.find_elements(By.CSS_SELECTOR, button_css)

    # Không còn nút → dừng
    if not btns:
        return 0

    try:
        b = btns[0]
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", b
        )
        time.sleep(0.2)
        driver.execute_script("arguments[0].click();", b)
        time.sleep(2.5)
    except Exception as e:
        print("Không click được load-more:", e)
        return 0

    # Sau khi click xong → trả về số nút còn lại (hoặc 1 để báo còn)
    return len(btns)


def collect_article_urls_from_current_list_page(soup: BeautifulSoup, base: str) -> list[str]:
    urls = []
    for item in soup.select(LIST_ITEM_CSS):
        a = item.select_one(LIST_LINK_IN_ITEM_CSS)
        if not a:
            continue
        href = (a.get("href") or "").strip()
        if not href:
            continue
        full = urljoin(base, href)
        urls.append(full)
    # unique giữ thứ tự
    seen = set()
    out = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out

def list_fingerprint(driver, item_css="div.post-item.row", link_css="h3 a", sample=8):
    items = driver.find_elements(By.CSS_SELECTOR, item_css)
    hrefs = []
    for it in items[:sample]:
        try:
            a = it.find_element(By.CSS_SELECTOR, link_css)
            hrefs.append((a.get_attribute("href") or "").strip() or (a.get_attribute("onclick") or "").strip())
        except:
            hrefs.append("")
    raw = "|".join(hrefs) + f"|count={len(items)}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest(), (hrefs[0] if hrefs else ""), len(items)

def click_next_and_detect_change_v2(driver,
                                   next_css="a.next",
                                   item_css="div.post-item.row",
                                   link_css="h3 a",
                                   timeout=10):
    btns = driver.find_elements(By.CSS_SELECTOR, next_css)
    if not btns:
        return 0

    btn = btns[0]

    # Nếu nút bị disable bằng class/aria
    cls = (btn.get_attribute("class") or "").lower()
    aria = (btn.get_attribute("aria-disabled") or "").lower()
    if "disabled" in cls or aria in ("true", "1"):
        return 0

    # fingerprint trước click
    fp_before, first_before, count_before = list_fingerprint(driver, item_css, link_css)

    # click
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    time.sleep(0.2)
    driver.execute_script("arguments[0].click();", btn)

    # chờ nội dung đổi (không dùng staleness)
    def changed(_):
        fp_after, first_after, count_after = list_fingerprint(driver, item_css, link_css)
        return fp_after != fp_before or first_after != first_before or count_after != count_before

    try:
        WebDriverWait(driver, timeout).until(changed)
        return 1
    except:
        return 0
    
def click_next_and_detect_change(driver, next_css="a.next", item_css="div.post-item.row"):
    next_btns = driver.find_elements(By.CSS_SELECTOR, next_css)
    if not next_btns:
        return 0

    next_btn = next_btns[0]

    # lấy 1 item hiện tại làm mốc
    items = driver.find_elements(By.CSS_SELECTOR, item_css)
    old_first = items[0] if items else None

    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center'});", next_btn
    )
    time.sleep(0.2)
    driver.execute_script("arguments[0].click();", next_btn)

    try:
        if old_first:
            WebDriverWait(driver, 8).until(
                EC.staleness_of(old_first)
            )
            return 1   # có chuyển trang
    except TimeoutException:
        pass

    # không có thay đổi → đã tới trang cuối
    return 0

def click_next_and_detect_change_v2(driver,
                                   next_css="a.next",
                                   item_css="div.post-item.row",
                                   link_css="h3 a",
                                   timeout=10):
    btns = driver.find_elements(By.CSS_SELECTOR, next_css)
    if not btns:
        return 0

    btn = btns[0]

    # Nếu nút bị disable bằng class/aria
    cls = (btn.get_attribute("class") or "").lower()
    aria = (btn.get_attribute("aria-disabled") or "").lower()
    if "disabled" in cls or aria in ("true", "1"):
        return 0

    # fingerprint trước click
    fp_before, first_before, count_before = list_fingerprint(driver, item_css, link_css)

    # click
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    time.sleep(0.2)
    driver.execute_script("arguments[0].click();", btn)

    # chờ nội dung đổi (không dùng staleness)
    def changed(_):
        fp_after, first_after, count_after = list_fingerprint(driver, item_css, link_css)
        return fp_after != fp_before or first_after != first_before or count_after != count_before

    try:
        WebDriverWait(driver, timeout).until(changed)
        return 1
    except:
        return 0
def crawl_js_pagination_collect_all_urls(driver, base: str) -> list[str]:
    """
    Duyệt phân trang JS: click các nút trang / next để gom hết link bài.
    Nếu site không có phân trang -> trả về link trang hiện tại.
    """
    all_urls = []
    visited_click_keys = set()

    while True:
        soup = BeautifulSoup(driver.page_source, "html.parser")
        all_urls.extend(collect_article_urls_from_current_list_page(soup, base))

        pager_els = driver.find_elements(By.CSS_SELECTOR, PAGER_LINKS_CSS)
        if not pager_els:
            break

        # ưu tiên nút Next nếu có
        next_el = None
        for el in pager_els:
            try:
                cls = (el.get_attribute("class") or "").lower()
                txt = clean_spaces(el.text)
                if "active" in cls or "current" in cls or "disabled" in cls:
                    continue
                if txt in [">", "›", "next", "sau", "tiếp"] or "next" in cls:
                    next_el = el
                    break
            except StaleElementReferenceException:
                continue

        if next_el is not None:
            key = (clean_spaces(next_el.text), next_el.get_attribute("href") or "", next_el.get_attribute("class") or "")
            if key in visited_click_keys:
                break
            visited_click_keys.add(key)
            click_and_wait_list_change(driver, LIST_ITEM_CSS, next_el)
            continue

        # nếu không có Next rõ ràng, click lần lượt các nút chưa click
        clicked = False
        for el in pager_els:
            try:
                cls = (el.get_attribute("class") or "").lower()
                txt = clean_spaces(el.text)
                key = (txt, el.get_attribute("href") or "", el.get_attribute("class") or "")
                if key in visited_click_keys:
                    continue
                visited_click_keys.add(key)

                if "active" in cls or "current" in cls or "disabled" in cls:
                    continue

                click_and_wait_list_change(driver, LIST_ITEM_CSS, el)
                clicked = True
                break
            except StaleElementReferenceException:
                continue

        if not clicked:
            break

    # unique giữ thứ tự
    seen = set()
    out = []
    for u in all_urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out


class VspProducts:
    def __init__(self, driver, base: str, cat_id: str, url_id: str):
        self.driver = driver
        self.base = base
        self.cat_id = cat_id
        self.url_id = url_id

        self.camlist = camob.Listcam()
        self.url_links = []
        self.menu_links = menulink.MenuLink()

    def reset_for_row(self, base: str, cat_id: str, url_id: str):
        self.base = base
        self.cat_id = cat_id
        self.url_id = url_id

        self.camlist = camob.Listcam()
        self.url_links = []
        self.menu_links = menulink.MenuLink()

    def extract_detail_and_save_if_ok(self, article_url: str):
        try:
            self.driver.get(article_url)
            time.sleep(1.2)
            soup = BeautifulSoup(self.driver.page_source, "html.parser")
        except Exception as e:
            print("Error detail:", e)
            return

        # check detail title
        if not soup.select_one(DETAIL_TITLE_CSS):
            return

        d = parse_vn_date_from_soup(soup)
        if d is None:
            # Không parse được ngày -> bạn quyết định: bỏ qua hoặc vẫn lưu.
            # Ở đây mình chọn: bỏ qua cho an toàn.
            print("Không đọc được ngày, bỏ qua:", article_url)
            return

        if d < FROM_DATE:
            print(f"Bài {d} < {FROM_DATE} => bỏ qua")
            return

        tag_content = soup.select_one(DETAIL_CONTENT_CSS)
        if not tag_content:
            print("Không thấy content, bỏ qua:", article_url)
            return

        html_content = normalize_html_keep_img_src_only(tag_content)

        # tạo object để lưu
        cam = camob.CameraObject(
            0,
            name=clean_spaces(soup.select_one(DETAIL_TITLE_CSS).get_text(" ", strip=True)),
            price=0,
            url=article_url,
            photo="",
            cat_id=self.cat_id,
            date_publish = d

        )
        cam.description = html_content

        cam.display_info()
        hp.save_data_cam(self.url_id, cam)

    def crawl_one_source_url(self, source_url: str):
        """
        source_url có thể là trang danh sách tin.
        """
        print("OPEN:", source_url)
        self.driver.get(source_url)
        time.sleep(1.2)

        # Nếu có load-more thì click bung hết
        
        while True:
            changed = click_load_more_until_end(
                self.driver,
                button_css="a.next"
               
            )

            print("next changed:", changed)

            soup = BeautifulSoup(self.driver.page_source, "html.parser")
            urls = collect_article_urls_from_current_list_page(soup, self.base)
  
            for u in urls:
                if hp.check_cam_url_pro(self.url_id, u, u):
                    print('save cam')
                    self.extract_detail_and_save_if_ok(u)

            if changed == 0:
                print("Không còn trang tiếp theo → dừng")
                break

def detect_header_row(ws, required_headers, max_scan=20):
    for row in range(1, max_scan + 1):
        headers = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if val:
                headers[str(val).strip().lower()] = col
        if all(h in headers for h in required_headers):
            return row, headers
    raise ValueError("Không tìm thấy dòng header hợp lệ trong Excel")


def build_driver():
    service = Service(CHROME_DRIVER_PATH)
    chrome_options = Options()
    chrome_options.binary_location = CHROME_BINARY_PATH
    # gợi ý: chạy ổn định hơn
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    # chrome_options.add_argument("--headless=new")  # nếu muốn chạy ẩn

    return webdriver.Chrome(service=service, options=chrome_options)


def run_from_excel(excel_path: str, sheet_name: str = None):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    required = ["idurl", "source", "target", "cat_id"]
    header_row, headers = detect_header_row(ws, required)

    driver = build_driver()
    bot = None

    try:
        for r in range(header_row + 1, ws.max_row + 1):
            idurl  = ws.cell(r, headers["idurl"]).value
            source = ws.cell(r, headers["source"]).value
            target = ws.cell(r, headers["target"]).value  # hiện chưa dùng; bạn có thể dùng làm mô tả/đích lưu
            cat_id = ws.cell(r, headers["cat_id"]).value

            if not source:
                continue

            idurl = str(idurl).strip() if idurl is not None else ""
            source = str(source).strip()
            cat_id = str(cat_id).strip() if cat_id is not None else ""
            base = get_base(source)

            print(f"\n=== ROW {r} | url_id={idurl} | cat_id={cat_id}")
            print("source:", source)
            print("base  :", base)
            print("target:", target)

            if bot is None:
                bot = VspProducts(driver=driver, base=base, cat_id=cat_id, url_id=idurl)
            else:
                bot.reset_for_row(base=base, cat_id=cat_id, url_id=idurl)

            bot.crawl_one_source_url(source)

            time.sleep(SLEEP_BETWEEN_ROW)

    finally:
        try:
            driver.quit()
        except:
            pass


if __name__ == "__main__":
    run_from_excel(r"D:\WORKSPACE_CODE\Projects\Web\folder\websiteninhbinh\danhsachweb.xlsx")
